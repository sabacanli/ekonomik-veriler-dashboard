"""
Hazine İç Borç Ödemeleri & Projeksiyon Analizi
─────────────────────────────────────────────────
Ödeme (geçmiş gerçekleşmeler) ve Projeksiyon (gelecek tahminler) dosyalarını
parse edip Excel'e veri + analiz sayfaları ekler.
"""

import json, os, urllib.request
import xlrd
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import (Font, PatternFill, Border, Side, Alignment,
                              numbers)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel

# ── Sabitler ──
AY_SIRA = ['Ocak', 'Şubat', 'Mart', 'Nisan', 'Mayıs', 'Haziran',
           'Temmuz', 'Ağustos', 'Eylül', 'Ekim', 'Kasım', 'Aralık']
AY_MAP = {a: i+1 for i, a in enumerate(AY_SIRA)}

TITLE_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
SECTION_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
HDR_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
LIGHT_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
ACTUAL_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
PROJ_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
WARN_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

thin = Side(style='thin', color='B0B0B0')
brd = Border(left=thin, right=thin, top=thin, bottom=thin)


def load_urls():
    # hazine_ihale_cek.load_urls otomatik keşif yapar + urls.json'u günceller.
    # Böylece ödeme analizi tek başına çalıştırılsa bile en güncel URL'leri alır.
    try:
        import importlib.util
        _p = os.path.join(os.path.dirname(__file__) or '.', 'hazine_ihale_cek.py')
        _spec = importlib.util.spec_from_file_location('hazine_ihale_cek', _p)
        _mod = importlib.util.module_from_spec(_spec)
        _spec.loader.exec_module(_mod)
        return _mod.load_urls()
    except Exception as e:
        print(f"  UYARI: Otomatik kesif atlandi ({e}); urls.json okunuyor.")
        with open(os.path.join(os.path.dirname(__file__) or '.', 'urls.json'), 'r', encoding='utf-8') as f:
            return json.load(f)


def download(url):
    fname = url.split('/')[-1]
    path = os.path.join(os.path.dirname(__file__) or '.', fname)
    urllib.request.urlretrieve(url, path)
    return path


# ═══════════════════════════════════════════════════════════════
# 1) ÖDEME DOSYASI PARSE (Milyar TL)
# ═══════════════════════════════════════════════════════════════
def parse_odemeler(filepath):
    """
    Ödeme .xls: tek sheet, sütunlar B=Yıl, C=Ay, D=Anapara, E=Faiz, F=Toplam, G=Birikimli
    Birim: Milyar TL
    """
    wb = xlrd.open_workbook(filepath)
    ws = wb.sheet_by_index(0)

    rows = []
    current_yil = None
    for r in range(5, ws.nrows):
        yil_val = ws.cell_value(r, 1)
        ay_val = ws.cell_value(r, 2)
        anapara = ws.cell_value(r, 3)
        faiz = ws.cell_value(r, 4)
        toplam = ws.cell_value(r, 5)
        birikimli = ws.cell_value(r, 6)

        if yil_val and isinstance(yil_val, float):
            current_yil = int(yil_val)

        if not current_yil or not ay_val or ay_val not in AY_MAP:
            continue

        # Boş satırları atla (gelecek aylar)
        if anapara == '' and faiz == '' and toplam == '':
            continue

        rows.append({
            'yil': current_yil,
            'ay': ay_val,
            'ay_no': AY_MAP[ay_val],
            'anapara': float(anapara) if anapara else 0,
            'faiz': float(faiz) if faiz else 0,
            'toplam': float(toplam) if toplam else 0,
            'birikimli': float(birikimli) if birikimli else 0,
            'birim': 'Milyar TL',
            'kaynak': 'Gerçekleşme'
        })

    return pd.DataFrame(rows)


# ═══════════════════════════════════════════════════════════════
# 2) PROJEKSİYON DOSYASI PARSE (Milyon TL → Milyar TL'ye çevir)
# ═══════════════════════════════════════════════════════════════
def parse_projeksiyon_son(filepath):
    """
    Son projeksiyon verilerini al (en güncel projeksiyon zamanı satırından).
    Projeksiyon .xlsx: her yıl ayrı sheet.
    Yapı: ANAPARA R7-R18, FAİZ R20-R31
    Sütunlar: D=Proj.Zamanı, E-P = Ocak-Aralık
    Son dolu satır = en güncel projeksiyon.
    Birim: Milyon TL → Milyar TL'ye çevrilir.
    """
    from openpyxl import load_workbook as lw
    wb = lw(filepath, data_only=True)

    results = []

    for sn in wb.sheetnames:
        ws = wb[sn]
        try:
            yil = int(sn)
        except ValueError:
            continue

        # Son projeksiyon satırını bul (ANAPARA: R7-R18, FAİZ: R20-R31)
        anapara_rows = range(7, 19)  # 7-18
        faiz_rows = range(20, 32)    # 20-31

        # Son dolu ANAPARA satırını bul
        last_ana_row = None
        last_ana_proj_ay = None
        for r in anapara_rows:
            proj_ay = ws.cell(r, 4).value  # D sütunu = projeksiyon zamanı
            # Satırda herhangi bir veri var mı?
            has_data = any(ws.cell(r, c).value is not None for c in range(5, 17))
            if proj_ay and has_data:
                last_ana_row = r
                last_ana_proj_ay = proj_ay

        # Son dolu FAİZ satırını bul
        last_faiz_row = None
        for r in faiz_rows:
            proj_ay = ws.cell(r, 4).value
            has_data = any(ws.cell(r, c).value is not None for c in range(5, 17))
            if proj_ay and has_data:
                last_faiz_row = r

        if not last_ana_row or not last_faiz_row:
            continue

        # 12 aylık verileri oku (E=5 → Ocak, P=16 → Aralık)
        for mi, ay in enumerate(AY_SIRA):
            col = 5 + mi  # E(5)=Ocak, F(6)=Şubat, ...
            ana_val = ws.cell(last_ana_row, col).value
            faiz_val = ws.cell(last_faiz_row, col).value

            if ana_val is None and faiz_val is None:
                continue

            ana = float(ana_val) / 1000 if ana_val else 0  # Milyon → Milyar
            fz = float(faiz_val) / 1000 if faiz_val else 0

            results.append({
                'yil': yil,
                'ay': ay,
                'ay_no': mi + 1,
                'anapara': round(ana, 6),
                'faiz': round(fz, 6),
                'toplam': round(ana + fz, 6),
                'birim': 'Milyar TL',
                'kaynak': 'Projeksiyon',
                'proj_zamani': last_ana_proj_ay
            })

    return pd.DataFrame(results)


# ═══════════════════════════════════════════════════════════════
# 3) BİRLEŞTİRİLMİŞ VERİ
# ═══════════════════════════════════════════════════════════════
def combine_odeme_projeksiyon(df_odeme, df_proj):
    """
    Gerçekleşme varsa onu, yoksa projeksiyonu kullan.
    Sonuç: 2024-2026 arası tam aylık tablo.
    """
    # Son ödeme ayını bul
    max_odeme = df_odeme[df_odeme['yil'] == df_odeme['yil'].max()]

    combined = []

    # Mevcut yıl için en son projeksiyon yılını kullan
    proj_yillar = df_proj['yil'].unique() if len(df_proj) > 0 else []

    for yil in sorted(set(list(df_odeme['yil'].unique()) + list(proj_yillar))):
        for ay_no in range(1, 13):
            ay = AY_SIRA[ay_no - 1]

            # Önce gerçekleşme ara
            gercek = df_odeme[(df_odeme['yil'] == yil) & (df_odeme['ay_no'] == ay_no)]
            if len(gercek) > 0:
                row = gercek.iloc[0].to_dict()
                row['kaynak'] = 'Gerçekleşme'
                combined.append(row)
            else:
                # Projeksiyon ara
                proj = df_proj[(df_proj['yil'] == yil) & (df_proj['ay_no'] == ay_no)]
                if len(proj) > 0:
                    row = proj.iloc[0].to_dict()
                    row['kaynak'] = 'Projeksiyon'
                    combined.append(row)

    return pd.DataFrame(combined)


# ═══════════════════════════════════════════════════════════════
# 4) EXCEL'E YAZMA
# ═══════════════════════════════════════════════════════════════
def write_odeme_sheets(output_file):
    """Ana Excel dosyasına Ödeme veri ve analiz sayfalarını ekler."""

    urls = load_urls()

    # İndir ve parse et
    odeme_url = urls.get('Merkezi Yonetim Ic Borc Odemeleri')
    proj_url = urls.get('Merkezi Yonetim Ic Borc Odeme Projeksiyonlari')

    if not odeme_url or not proj_url:
        print("  UYARI: Ödeme/Projeksiyon URL'leri urls.json'da bulunamadı!")
        return

    print("  İç Borç Ödemeleri indiriliyor...")
    odeme_path = download(odeme_url)
    print("  Projeksiyon indiriliyor...")
    proj_path = download(proj_url)

    df_odeme = parse_odemeler(odeme_path)
    df_proj = parse_projeksiyon_son(proj_path)

    print(f"  Ödeme: {len(df_odeme)} satır ({df_odeme['yil'].min()}-{df_odeme['yil'].max()})")
    print(f"  Projeksiyon: {len(df_proj)} satır")

    # Excel'i aç
    wb = load_workbook(output_file)

    # Eski sayfaları sil (varsa)
    for sn in ['İç Borç Ödemeleri', 'Ödeme Analizi']:
        if sn in wb.sheetnames:
            del wb[sn]

    # ══════════════════════════════════════════════════
    # SAYFA 1: İç Borç Ödemeleri - Veri
    # ══════════════════════════════════════════════════
    ws = wb.create_sheet('İç Borç Ödemeleri')

    def cell(r, c, val, font=None, fill=None, align=None, nfmt=None):
        cl = ws.cell(row=r, column=c, value=val)
        cl.border = brd
        if font: cl.font = font
        if fill: cl.fill = fill
        if align: cl.alignment = align
        if nfmt: cl.number_format = nfmt
        return cl

    def merge_fill(r, c1, c2, val, font, fill, align=None, height=None):
        for c in range(c1, c2 + 1):
            cl = ws.cell(row=r, column=c)
            cl.fill = fill
            cl.font = font
            cl.border = brd
            if align: cl.alignment = align
        ws.cell(row=r, column=c1, value=val)
        if align: ws.cell(row=r, column=c1).alignment = align
        if c2 > c1:
            ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
        if height:
            ws.row_dimensions[r].height = height

    COLS = 7
    r = 1

    # Başlık
    merge_fill(r, 1, COLS, "MERKEZİ YÖNETİM İÇ BORÇ ÖDEMELERİ",
               Font(bold=True, color="FFFFFF", size=14), TITLE_FILL,
               Alignment(horizontal="center", vertical="center"), height=36)
    r += 1

    today = datetime.now().strftime("%d.%m.%Y")
    merge_fill(r, 1, COLS,
               f"Son Güncelleme: {today}  •  Birim: Milyar TL  •  Kaynak: T.C. Hazine ve Maliye Bakanlığı",
               Font(italic=True, color="1F4E79", size=10), LIGHT_FILL,
               Alignment(horizontal="center", vertical="center"), height=22)
    r += 1

    # Açıklama
    merge_fill(r, 1, COLS,
               "🟢 Yeşil = Gerçekleşme   🟡 Sarı = Projeksiyon (Son güncellenen projeksiyon verileri)",
               Font(color="333333", size=10), WHITE_FILL,
               Alignment(horizontal="center", vertical="center"), height=22)
    r += 1

    # Header
    headers = ["Yıl", "Ay", "Anapara", "Faiz", "Toplam", "Yıllık Birikimli", "Kaynak"]
    for ci, h in enumerate(headers):
        cell(r, ci+1, h, Font(bold=True, color="FFFFFF", size=10), HDR_FILL,
             Alignment(horizontal="center", vertical="center", wrap_text=True))
    ws.row_dimensions[r].height = 28
    data_header_row = r
    r += 1
    data_start_row = r

    # 2020'den itibaren verileri yaz (geriye doğru çok uzun, son 7 yıl yeterli)
    df_all = df_odeme.copy()
    # Projeksiyondan gelen gelecek verileri ekle
    proj_current = df_proj[df_proj['yil'] >= df_odeme['yil'].max()].copy()

    # Gerçekleşme olmayan ayları projeksiyondan al
    for _, prow in proj_current.iterrows():
        exists = df_all[(df_all['yil'] == prow['yil']) & (df_all['ay_no'] == prow['ay_no'])]
        if len(exists) == 0:
            df_all = pd.concat([df_all, pd.DataFrame([prow])], ignore_index=True)

    df_all = df_all.sort_values(['yil', 'ay_no']).reset_index(drop=True)

    # Son 7 yıl filtrele
    min_yil = max(2020, df_all['yil'].min())
    df_display = df_all[df_all['yil'] >= min_yil].copy()

    # Yıllık birikimli hesapla
    df_display['birikimli_calc'] = df_display.groupby('yil')['toplam'].cumsum()

    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    prev_yil = None
    for _, row in df_display.iterrows():
        is_proj = row.get('kaynak', 'Gerçekleşme') == 'Projeksiyon'
        fill = PROJ_FILL if is_proj else ACTUAL_FILL

        yil_val = int(row['yil']) if row['yil'] != prev_yil else ''
        prev_yil = int(row['yil'])

        cell(r, 1, yil_val, Font(bold=True, size=10), fill, center)
        cell(r, 2, row['ay'], Font(size=10), fill, center)
        cell(r, 3, round(row['anapara'], 2), Font(size=10), fill, right, '#,##0.00')
        cell(r, 4, round(row['faiz'], 2), Font(size=10), fill, right, '#,##0.00')
        cell(r, 5, round(row['toplam'], 2), Font(size=10), fill, right, '#,##0.00')
        bkm = row.get('birikimli', row.get('birikimli_calc', 0))
        cell(r, 6, round(bkm, 2) if bkm else round(row['birikimli_calc'], 2),
             Font(size=10), fill, right, '#,##0.00')
        cell(r, 7, row.get('kaynak', 'Gerçekleşme'),
             Font(italic=True, color="666666", size=9), fill, center)
        r += 1

    data_end_row = r - 1

    # Sütun genişlikleri
    for ci, w in enumerate([8, 10, 16, 16, 16, 18, 14]):
        ws.column_dimensions[get_column_letter(ci+1)].width = w

    ws.sheet_view.showGridLines = False

    # ══════════════════════════════════════════════════
    # SAYFA 2: Ödeme Analizi
    # ══════════════════════════════════════════════════
    ws2 = wb.create_sheet('Ödeme Analizi')

    def cell2(r, c, val, font=None, fill=None, align=None, nfmt=None):
        cl = ws2.cell(row=r, column=c, value=val)
        cl.border = brd
        if font: cl.font = font
        if fill: cl.fill = fill
        if align: cl.alignment = align
        if nfmt: cl.number_format = nfmt
        return cl

    def merge_fill2(r, c1, c2, val, font, fill, align=None, height=None):
        for c in range(c1, c2 + 1):
            cl = ws2.cell(row=r, column=c)
            cl.fill = fill
            cl.font = font
            cl.border = brd
            if align: cl.alignment = align
        ws2.cell(row=r, column=c1, value=val)
        if align: ws2.cell(row=r, column=c1).alignment = align
        if c2 > c1:
            ws2.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
        if height:
            ws2.row_dimensions[r].height = height

    ACOLS = 10
    r = 1

    merge_fill2(r, 1, ACOLS, "İÇ BORÇ ÖDEME ANALİZİ",
                Font(bold=True, color="FFFFFF", size=14), TITLE_FILL,
                Alignment(horizontal="center", vertical="center"), height=36)
    r += 1
    merge_fill2(r, 1, ACOLS,
                f"Gerçekleşme + Projeksiyon Karşılaştırması  •  Birim: Milyar TL  •  {today}",
                Font(italic=True, color="1F4E79", size=10), LIGHT_FILL,
                Alignment(horizontal="center", vertical="center"), height=22)
    r += 2

    # ── ANALİZ 1: YILLIK ÖZET ──
    merge_fill2(r, 1, ACOLS, "1. YILLIK ÖDEME ÖZETİ",
                Font(bold=True, color="FFFFFF", size=12), SECTION_FILL,
                Alignment(horizontal="center", vertical="center"), height=28)
    r += 1

    yillik_headers = ["Yıl", "Toplam Anapara", "Toplam Faiz", "Genel Toplam",
                      "Faiz/Anapara (%)", "Ort. Aylık Ödeme", "Max Ay Ödeme",
                      "Max Ay", "YoY Değişim (%)", "Kaynak"]
    for ci, h in enumerate(yillik_headers):
        cell2(r, ci+1, h, Font(bold=True, color="FFFFFF", size=9), HDR_FILL,
              Alignment(horizontal="center", vertical="center", wrap_text=True))
    ws2.row_dimensions[r].height = 32
    r += 1

    yillar = sorted(df_all[df_all['yil'] >= 2020]['yil'].unique())
    prev_toplam = None

    for yi, yil in enumerate(yillar):
        dy = df_all[df_all['yil'] == yil]
        ana_sum = dy['anapara'].sum()
        faiz_sum = dy['faiz'].sum()
        top_sum = dy['toplam'].sum()
        faiz_pct = (faiz_sum / ana_sum * 100) if ana_sum > 0 else 0
        avg_monthly = top_sum / len(dy) if len(dy) > 0 else 0
        max_row = dy.loc[dy['toplam'].idxmax()] if len(dy) > 0 else None
        max_val = max_row['toplam'] if max_row is not None else 0
        max_ay = max_row['ay'] if max_row is not None else ''
        yoy = ((top_sum - prev_toplam) / prev_toplam * 100) if prev_toplam and prev_toplam > 0 else None
        kaynak = 'Karma' if 'Projeksiyon' in dy['kaynak'].values and 'Gerçekleşme' in dy['kaynak'].values \
                 else dy['kaynak'].iloc[0] if len(dy) > 0 else ''

        fill = LIGHT_FILL if yi % 2 == 0 else WHITE_FILL
        cell2(r, 1, int(yil), Font(bold=True, size=10), fill, center)
        cell2(r, 2, round(ana_sum, 2), Font(size=10), fill, right, '#,##0.00')
        cell2(r, 3, round(faiz_sum, 2), Font(size=10), fill, right, '#,##0.00')
        cell2(r, 4, round(top_sum, 2), Font(bold=True, size=10), fill, right, '#,##0.00')
        cell2(r, 5, round(faiz_pct, 1), Font(size=10), fill, center, '0.0')
        cell2(r, 6, round(avg_monthly, 2), Font(size=10), fill, right, '#,##0.00')
        cell2(r, 7, round(max_val, 2), Font(size=10), fill, right, '#,##0.00')
        cell2(r, 8, max_ay, Font(size=10), fill, center)
        if yoy is not None:
            fnt = Font(color="006100", size=10) if yoy < 0 else Font(color="C00000", size=10)
            cell2(r, 9, round(yoy, 1), fnt, fill, center, '+0.0;-0.0')
        else:
            cell2(r, 9, '-', Font(color="888888", size=10), fill, center)
        cell2(r, 10, kaynak, Font(italic=True, color="666666", size=9), fill, center)

        prev_toplam = top_sum
        r += 1

    r += 1

    # ── ANALİZ 2: 2026 AYLIK PROJEKSİYON vs GERÇEKLEŞME ──
    merge_fill2(r, 1, ACOLS, "2. 2026 AYLIK ÖDEME DETAYI (Gerçekleşme vs Projeksiyon)",
                Font(bold=True, color="FFFFFF", size=12), SECTION_FILL,
                Alignment(horizontal="center", vertical="center"), height=28)
    r += 1

    # Mevcut yılın (2026) en son projeksiyon verisini al
    current_year = datetime.now().year
    proj_cy = df_proj[df_proj['yil'] == current_year].copy()
    odeme_cy = df_odeme[df_odeme['yil'] == current_year].copy()

    m_headers = ["Ay", "Projeksiyon Anapara", "Projeksiyon Faiz", "Projeksiyon Toplam",
                 "Gerçekleşme Anapara", "Gerçekleşme Faiz", "Gerçekleşme Toplam",
                 "Fark (Gerç.-Proj.)", "Sapma (%)", "Durum"]
    for ci, h in enumerate(m_headers):
        cell2(r, ci+1, h, Font(bold=True, color="FFFFFF", size=9), HDR_FILL,
              Alignment(horizontal="center", vertical="center", wrap_text=True))
    ws2.row_dimensions[r].height = 32
    r += 1

    for mi, ay in enumerate(AY_SIRA):
        ay_no = mi + 1

        # Projeksiyon
        p = proj_cy[proj_cy['ay_no'] == ay_no]
        p_ana = p.iloc[0]['anapara'] if len(p) > 0 else None
        p_faiz = p.iloc[0]['faiz'] if len(p) > 0 else None
        p_top = p.iloc[0]['toplam'] if len(p) > 0 else None

        # Gerçekleşme
        g = odeme_cy[odeme_cy['ay_no'] == ay_no]
        g_ana = g.iloc[0]['anapara'] if len(g) > 0 else None
        g_faiz = g.iloc[0]['faiz'] if len(g) > 0 else None
        g_top = g.iloc[0]['toplam'] if len(g) > 0 else None

        has_both = p_top is not None and g_top is not None
        fark = (g_top - p_top) if has_both else None
        sapma = (fark / p_top * 100) if has_both and p_top != 0 else None

        if g_top is not None:
            fill = ACTUAL_FILL
            durum = "✅ Gerçekleşti"
        elif p_top is not None:
            fill = PROJ_FILL
            durum = "📊 Projeksiyon"
        else:
            fill = WHITE_FILL
            durum = "-"

        cell2(r, 1, ay, Font(bold=True, size=10), fill, center)
        cell2(r, 2, round(p_ana, 2) if p_ana is not None else '-', Font(size=10), fill, right, '#,##0.00')
        cell2(r, 3, round(p_faiz, 2) if p_faiz is not None else '-', Font(size=10), fill, right, '#,##0.00')
        cell2(r, 4, round(p_top, 2) if p_top is not None else '-', Font(bold=True, size=10), fill, right, '#,##0.00')
        cell2(r, 5, round(g_ana, 2) if g_ana is not None else '-', Font(size=10), fill, right, '#,##0.00')
        cell2(r, 6, round(g_faiz, 2) if g_faiz is not None else '-', Font(size=10), fill, right, '#,##0.00')
        cell2(r, 7, round(g_top, 2) if g_top is not None else '-', Font(bold=True, size=10), fill, right, '#,##0.00')

        if fark is not None:
            fnt = Font(color="006100", size=10) if fark <= 0 else Font(color="C00000", size=10)
            cell2(r, 8, round(fark, 2), fnt, fill, right, '+#,##0.00;-#,##0.00')
            cell2(r, 9, round(sapma, 1) if sapma is not None else '-', fnt, fill, center, '+0.0;-0.0')
        else:
            cell2(r, 8, '-', Font(color="888888"), fill, center)
            cell2(r, 9, '-', Font(color="888888"), fill, center)

        cell2(r, 10, durum, Font(size=9), fill, center)
        r += 1

    # Toplam satırı
    p_top_sum = proj_cy['toplam'].sum() if len(proj_cy) > 0 else 0
    g_top_sum = odeme_cy['toplam'].sum() if len(odeme_cy) > 0 else 0
    cell2(r, 1, "TOPLAM", Font(bold=True, color="FFFFFF", size=10), SECTION_FILL, center)
    for c in range(2, 5):
        ws2.cell(row=r, column=c).fill = SECTION_FILL
        ws2.cell(row=r, column=c).border = brd
    cell2(r, 4, round(p_top_sum, 2), Font(bold=True, color="FFFFFF", size=10), SECTION_FILL, right, '#,##0.00')
    for c in range(5, 8):
        ws2.cell(row=r, column=c).fill = SECTION_FILL
        ws2.cell(row=r, column=c).border = brd
    cell2(r, 7, round(g_top_sum, 2), Font(bold=True, color="FFFFFF", size=10), SECTION_FILL, right, '#,##0.00')
    for c in range(8, 11):
        ws2.cell(row=r, column=c).fill = SECTION_FILL
        ws2.cell(row=r, column=c).border = brd
    r += 2

    # ── ANALİZ 3: ANAPARA vs FAİZ DAĞILIMI ──
    merge_fill2(r, 1, ACOLS, "3. ANAPARA vs FAİZ DAĞILIMI (Yıllık)",
                Font(bold=True, color="FFFFFF", size=12), SECTION_FILL,
                Alignment(horizontal="center", vertical="center"), height=28)
    r += 1

    af_headers = ["Yıl", "Anapara (Milyar TL)", "Faiz (Milyar TL)", "Genel Toplam",
                  "Anapara Payı (%)", "Faiz Payı (%)", "Faiz/Anapara Oranı",
                  "Faiz Artışı YoY (%)", "", ""]
    for ci, h in enumerate(af_headers):
        cell2(r, ci+1, h, Font(bold=True, color="FFFFFF", size=9), HDR_FILL,
              Alignment(horizontal="center", vertical="center", wrap_text=True))
    ws2.row_dimensions[r].height = 32
    r += 1

    prev_faiz = None
    for yi, yil in enumerate(yillar):
        dy = df_all[df_all['yil'] == yil]
        ana = dy['anapara'].sum()
        fz = dy['faiz'].sum()
        top = ana + fz
        ana_pct = (ana / top * 100) if top > 0 else 0
        fz_pct = (fz / top * 100) if top > 0 else 0
        fz_ana = (fz / ana) if ana > 0 else 0
        fz_yoy = ((fz - prev_faiz) / prev_faiz * 100) if prev_faiz and prev_faiz > 0 else None

        fill = LIGHT_FILL if yi % 2 == 0 else WHITE_FILL
        cell2(r, 1, int(yil), Font(bold=True, size=10), fill, center)
        cell2(r, 2, round(ana, 2), Font(size=10), fill, right, '#,##0.00')
        cell2(r, 3, round(fz, 2), Font(size=10), fill, right, '#,##0.00')
        cell2(r, 4, round(top, 2), Font(bold=True, size=10), fill, right, '#,##0.00')
        cell2(r, 5, round(ana_pct, 1), Font(size=10), fill, center, '0.0')
        cell2(r, 6, round(fz_pct, 1), Font(size=10), fill, center, '0.0')
        cell2(r, 7, round(fz_ana, 2), Font(size=10), fill, center, '0.00')
        if fz_yoy is not None:
            fnt = Font(color="C00000", size=10) if fz_yoy > 0 else Font(color="006100", size=10)
            cell2(r, 8, round(fz_yoy, 1), fnt, fill, center, '+0.0;-0.0')
        else:
            cell2(r, 8, '-', Font(color="888888"), fill, center)

        prev_faiz = fz
        r += 1

    r += 1

    # ── ANALİZ 4: BORÇLANMA vs ÖDEME KARŞILAŞTIRMASI ──
    merge_fill2(r, 1, ACOLS, "4. BORÇLANMA vs ÖDEME KARŞILAŞTIRMASI (Yıllık)",
                Font(bold=True, color="FFFFFF", size=12), SECTION_FILL,
                Alignment(horizontal="center", vertical="center"), height=28)
    r += 1

    # İhale verilerinden yıllık borçlanma toplamını çek
    ws_ihale = wb['Tüm İhaleler']
    ihale_lr = ws_ihale.max_row

    bv_headers = ["Yıl", "Toplam Borçlanma (Milyar TL)", "Toplam Ödeme (Milyar TL)",
                  "Net Borçlanma", "Borçlanma/Ödeme", "Çevrim Oranı (%)",
                  "Anapara Öd.", "Faiz Öd.", "", ""]
    for ci, h in enumerate(bv_headers):
        cell2(r, ci+1, h, Font(bold=True, color="FFFFFF", size=9), HDR_FILL,
              Alignment(horizontal="center", vertical="center", wrap_text=True))
    ws2.row_dimensions[r].height = 32
    r += 1

    # İhale verilerinden yıllık toplam net satış
    ihale_data = {}
    for ir in range(3, ihale_lr + 1):
        yil_val = ws_ihale.cell(ir, 2).value  # B = yıl
        net_val = ws_ihale.cell(ir, 20).value  # T = toplam_net (bin TL)
        if yil_val and net_val:
            try:
                y = int(yil_val)
                n = float(net_val)
                ihale_data[y] = ihale_data.get(y, 0) + n
            except (ValueError, TypeError):
                pass

    # Bin TL → Milyar TL
    for k in ihale_data:
        ihale_data[k] = ihale_data[k] / 1_000_000

    for yi, yil in enumerate(yillar):
        dy = df_all[df_all['yil'] == yil]
        odeme_top = dy['toplam'].sum()
        ana_od = dy['anapara'].sum()
        faiz_od = dy['faiz'].sum()
        borc = ihale_data.get(int(yil), None)

        fill = LIGHT_FILL if yi % 2 == 0 else WHITE_FILL
        cell2(r, 1, int(yil), Font(bold=True, size=10), fill, center)

        if borc is not None:
            cell2(r, 2, round(borc, 2), Font(size=10), fill, right, '#,##0.00')
            net = borc - odeme_top
            ratio = borc / odeme_top if odeme_top > 0 else 0
            cevrim = (odeme_top / borc * 100) if borc > 0 else 0

            cell2(r, 3, round(odeme_top, 2), Font(size=10), fill, right, '#,##0.00')
            fnt = Font(color="006100", bold=True, size=10) if net > 0 else Font(color="C00000", bold=True, size=10)
            cell2(r, 4, round(net, 2), fnt, fill, right, '+#,##0.00;-#,##0.00')
            cell2(r, 5, round(ratio, 2), Font(size=10), fill, center, '0.00')
            cell2(r, 6, round(cevrim, 1), Font(size=10), fill, center, '0.0')
        else:
            cell2(r, 2, 'Veri yok', Font(italic=True, color="888888"), fill, center)
            cell2(r, 3, round(odeme_top, 2), Font(size=10), fill, right, '#,##0.00')
            cell2(r, 4, '-', Font(color="888888"), fill, center)
            cell2(r, 5, '-', Font(color="888888"), fill, center)
            cell2(r, 6, '-', Font(color="888888"), fill, center)

        cell2(r, 7, round(ana_od, 2), Font(size=10), fill, right, '#,##0.00')
        cell2(r, 8, round(faiz_od, 2), Font(size=10), fill, right, '#,##0.00')
        r += 1

    r += 1

    # ── ANALİZ 5: 2026 KALAN ÖDEME PROJEKSİYONU ──
    merge_fill2(r, 1, ACOLS, "5. 2026 KALAN ÖDEME PROJEKSİYONU & YOĞUNLUK",
                Font(bold=True, color="FFFFFF", size=12), SECTION_FILL,
                Alignment(horizontal="center", vertical="center"), height=28)
    r += 1

    # 2026 projeksiyonundan kalan ayları göster
    proj_2026 = df_proj[df_proj['yil'] == current_year].sort_values('ay_no')
    odeme_2026 = df_odeme[df_odeme['yil'] == current_year]
    gerceklesen_aylar = set(odeme_2026['ay_no'].values) if len(odeme_2026) > 0 else set()

    k_headers = ["Ay", "Proj. Anapara", "Proj. Faiz", "Proj. Toplam",
                 "Kalan Toplam (%)", "Yoğunluk", "", "", "", ""]
    for ci, h in enumerate(k_headers):
        cell2(r, ci+1, h, Font(bold=True, color="FFFFFF", size=9), HDR_FILL,
              Alignment(horizontal="center", vertical="center", wrap_text=True))
    ws2.row_dimensions[r].height = 28
    r += 1

    proj_2026_toplam = proj_2026['toplam'].sum() if len(proj_2026) > 0 else 1
    kalan_toplam = 0

    for _, prow in proj_2026.iterrows():
        ay_no = int(prow['ay_no'])
        is_past = ay_no in gerceklesen_aylar
        fill = ACTUAL_FILL if is_past else PROJ_FILL

        pct = (prow['toplam'] / proj_2026_toplam * 100) if proj_2026_toplam > 0 else 0

        # Yoğunluk bar
        bar_len = int(pct / 2)
        bar = "█" * bar_len

        if not is_past:
            kalan_toplam += prow['toplam']

        cell2(r, 1, prow['ay'], Font(bold=True, size=10), fill, center)
        cell2(r, 2, round(prow['anapara'], 2), Font(size=10), fill, right, '#,##0.00')
        cell2(r, 3, round(prow['faiz'], 2), Font(size=10), fill, right, '#,##0.00')
        cell2(r, 4, round(prow['toplam'], 2), Font(bold=True, size=10), fill, right, '#,##0.00')
        cell2(r, 5, round(pct, 1), Font(size=10), fill, center, '0.0')
        cell2(r, 6, bar, Font(color="2E75B6", size=10), fill,
              Alignment(horizontal="left", vertical="center"))

        status = "✅ Ödendi" if is_past else ""
        cell2(r, 7, status, Font(size=9), fill, center)
        for c in range(8, 11):
            ws2.cell(row=r, column=c).fill = fill
            ws2.cell(row=r, column=c).border = brd
        r += 1

    # Kalan toplam
    cell2(r, 1, "KALAN", Font(bold=True, color="FFFFFF", size=10), WARN_FILL, center)
    for c in range(2, 4):
        ws2.cell(row=r, column=c).fill = WARN_FILL
        ws2.cell(row=r, column=c).border = brd
    cell2(r, 4, round(kalan_toplam, 2), Font(bold=True, color="C00000", size=11), WARN_FILL, right, '#,##0.00')
    merge_fill2(r, 5, ACOLS, f"Yıl sonuna kadar ödenecek tutar",
                Font(italic=True, color="C00000", size=10), WARN_FILL,
                Alignment(horizontal="left", vertical="center"))
    r += 2

    # ── ANALİZ 6: DİKKAT ÇEKEN NOKTALAR ──
    merge_fill2(r, 1, ACOLS, "6. DİKKAT ÇEKEN NOKTALAR",
                Font(bold=True, color="FFFFFF", size=12), SECTION_FILL,
                Alignment(horizontal="center", vertical="center"), height=28)
    r += 1

    highlights = []

    # Faiz yükü trendi
    if len(df_all[df_all['yil'] >= 2023]) > 0:
        f23 = df_all[df_all['yil'] == 2023]['faiz'].sum() if 2023 in df_all['yil'].values else 0
        f24 = df_all[df_all['yil'] == 2024]['faiz'].sum() if 2024 in df_all['yil'].values else 0
        f25 = df_all[df_all['yil'] == 2025]['faiz'].sum() if 2025 in df_all['yil'].values else 0
        f26 = df_all[df_all['yil'] == 2026]['faiz'].sum() if 2026 in df_all['yil'].values else 0
        highlights.append(("📈 Faiz Yükü Trendi",
            f"2023: {f23:,.1f} → 2024: {f24:,.1f} → 2025: {f25:,.1f} → 2026: {f26:,.1f} Milyar TL"))

    # En yoğun ödeme ayı (2026)
    if len(proj_2026) > 0:
        max_ay = proj_2026.loc[proj_2026['toplam'].idxmax()]
        highlights.append(("⚠️ 2026 En Yoğun Ödeme Ayı",
            f"{max_ay['ay']}: {max_ay['toplam']:,.2f} Milyar TL "
            f"(Anapara: {max_ay['anapara']:,.2f} + Faiz: {max_ay['faiz']:,.2f})"))

    # Borçlanma vs Ödeme dengesi
    borc_2025 = ihale_data.get(2025, 0)
    odeme_2025 = df_all[df_all['yil'] == 2025]['toplam'].sum()
    if borc_2025 > 0 and odeme_2025 > 0:
        net_2025 = borc_2025 - odeme_2025
        highlights.append(("💰 2025 Net Borçlanma",
            f"Borçlanma: {borc_2025:,.2f} - Ödeme: {odeme_2025:,.2f} = "
            f"Net: {net_2025:+,.2f} Milyar TL"))

    # 2026 projeksiyon vs gerçekleşme sapması
    if len(odeme_2026) > 0 and len(proj_cy) > 0:
        common_months = set(odeme_2026['ay_no'].values) & set(proj_cy['ay_no'].values)
        if common_months:
            g_sum = odeme_2026[odeme_2026['ay_no'].isin(common_months)]['toplam'].sum()
            p_sum = proj_cy[proj_cy['ay_no'].isin(common_months)]['toplam'].sum()
            sapma_pct = (g_sum - p_sum) / p_sum * 100 if p_sum > 0 else 0
            ay_labels = ', '.join(sorted([AY_SIRA[m-1] for m in common_months]))
            highlights.append(("🔍 Projeksiyon Sapması",
                f"{ay_labels}: Projeksiyon {p_sum:,.2f} → Gerçekleşme {g_sum:,.2f} Milyar TL "
                f"(Sapma: {sapma_pct:+.1f}%)"))

    # Anapara/Faiz yapısındaki değişim
    if 2024 in df_all['yil'].values and 2026 in df_all['yil'].values:
        d24 = df_all[df_all['yil'] == 2024]
        d26 = df_all[df_all['yil'] == 2026]
        ratio_24 = d24['faiz'].sum() / d24['anapara'].sum() if d24['anapara'].sum() > 0 else 0
        ratio_26 = d26['faiz'].sum() / d26['anapara'].sum() if d26['anapara'].sum() > 0 else 0
        highlights.append(("📊 Faiz/Anapara Yapısı",
            f"2024: {ratio_24:.2f}x → 2026: {ratio_26:.2f}x "
            f"({'Faiz yükü artıyor' if ratio_26 > ratio_24 else 'Faiz yükü azalıyor'})"))

    stat_label_font = Font(bold=True, color="1F4E79", size=11)
    stat_val_font = Font(color="333333", size=11)

    for hi, (label, text) in enumerate(highlights):
        fill = LIGHT_FILL if hi % 2 == 0 else WHITE_FILL
        cell2(r, 1, label, stat_label_font, fill,
              Alignment(horizontal="left", vertical="center"))
        ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        for c in range(1, 3):
            ws2.cell(row=r, column=c).fill = fill
            ws2.cell(row=r, column=c).border = brd
        cell2(r, 3, text, stat_val_font, fill,
              Alignment(horizontal="left", vertical="center", wrap_text=True))
        ws2.merge_cells(start_row=r, start_column=3, end_row=r, end_column=ACOLS)
        for c in range(3, ACOLS + 1):
            ws2.cell(row=r, column=c).fill = fill
            ws2.cell(row=r, column=c).border = brd
        ws2.row_dimensions[r].height = 32
        r += 1

    # Sütun genişlikleri
    widths = [12, 20, 20, 20, 16, 16, 16, 18, 12, 12]
    for ci, w in enumerate(widths):
        ws2.column_dimensions[get_column_letter(ci+1)].width = w

    ws2.sheet_view.showGridLines = False

    # Kaydet
    wb.save(output_file)
    print(f"  İç Borç Ödeme sayfaları eklendi: {output_file}")
    print(f"  - İç Borç Ödemeleri (veri)")
    print(f"  - Ödeme Analizi (6 analiz tablosu)")


if __name__ == '__main__':
    # Dosya her zaman script'in klasöründe (çalışma dizininden bağımsız).
    write_odeme_sheets(os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                    'hazine_ihale_verileri.xlsx'))
