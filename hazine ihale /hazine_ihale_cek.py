"""
Hazine İhale Verilerini Çekme ve Birleştirme Scripti
=====================================================
HMB sitesinden 6 farklı menkul kıymet ihale verisini çekip
tek bir Excel dosyasında birleştirir.

Sütun Grupları:
  1) Genel Bilgiler     : Senet Türü, Yıl, ISIN, Valör, İtfa, Vade, Kupon
  2) Teklif              : İhaleye gelen toplam teklifler
  3) İhale Sonucu        : İhalede kabul edilen (rekabetçi) tutarlar
  4) Rek. Olmayan - Kamu  : Kamu kurumlarının rekabetçi olmayan teklifleri
  5) Rek. Olmayan - PY    : Piyasa yapıcıların rekabetçi olmayan teklifleri
  6) İhale Sonrası Satış  : İhale sonrası ek satışlar
  7) Toplam Satış         : Tüm kanalların toplamı
  8) Faiz & Fiyat         : Kabul edilen ortalama faiz ve birim fiyat
"""

import requests
import xlrd
import openpyxl
import pandas as pd
import json
import io
import os
import re
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
URLS_FILE = os.path.join(SCRIPT_DIR, "urls.json")

# HMB istatistik sayfasının arka plandaki WordPress REST API'si.
# Sayfanın içeriği (gömülü Excel bağlantıları dahil) buradan JSON olarak gelir.
HMB_PAGE_API = (
    "https://www.hmb.gov.tr/portal/v2/pages?slug=kamu-finansmani-istatistikleri"
)

# urls.json anahtarı -> HMB dosya adı kökü (hash öncesi sabit kısım).
# HMB her ay dosyaları YENİ bir adreste yayınlar:
#   ms.hmb.gov.tr/uploads/YYYY/MM/<ad>-<hash>.xls
# Klasör (YYYY/MM) ve hash her ay değişir; ancak dosya adı kökü sabit kalır.
# Bu kökleri kullanarak güncel URL'yi otomatik buluyoruz, böylece liste eskimez.
URL_STEM_MAP = {
    "TLREF Endeksli Senetler": "TLREFe-Endeksli-Tahviller",
    "TL Cinsi Kira Sertifikalari": "Kira-Sertifikasi",
    "TL Cinsi Kuponsuz Senetler": "TL-Cinsi-Kuponsuz-Senetler",
    "TL Cinsi TUFE Endeksli Senetler": "TUFEye-Endeksli-Tahviller",
    "TL Cinsi Sabit Faizli Kuponlu Senetler": "TL-Cinsi-Sabit-Faizli-Kuponlu-Senetler",
    "TL Cinsi Degisken Faizli Kuponlu Senetler": "TL-Cinsi-Degisken-Faizli-Kuponlu-Tahviller",
    "Merkezi Yonetim Ic Borc Odemeleri": "Merkezi_Yonetim_Ic_Borc_Odemeleri",
    "Merkezi Yonetim Ic Borc Odeme Projeksiyonlari": "Merkezi-Yonetim-Ic-Borc-Odeme-Projeksiyonlari-Aylik",
}


def discover_urls():
    """HMB sayfasından güncel Excel URL'lerini otomatik bulur.

    Resmi istatistik sayfasının WordPress API içeriğini çeker, içindeki tüm
    Excel bağlantılarını toplar ve her seri için dosya adının sabit kökünden
    en güncel (YYYY/MM en yüksek) URL'yi eşler.
    Ağ/parse hatasında boş sözlük döner; çağıran taraf cache'e (urls.json) düşer.
    """
    headers = {"User-Agent": "Mozilla/5.0", "Accept": "application/json"}
    try:
        r = requests.get(HMB_PAGE_API, headers=headers, timeout=30)
        r.raise_for_status()
        content = r.json()[0]["content"]["rendered"]
    except Exception as e:
        print(f"  UYARI: Otomatik URL keşfi basarisiz ({e}); urls.json kullanilacak.")
        return {}

    # Sayfadaki tüm Excel bağlantıları (href içinde, kapanış tırnağı ile sınırlı)
    all_urls = re.findall(
        r'href="(https://ms\.hmb\.gov\.tr/uploads/\d{4}/\d{2}/[^"]+?\.xlsx?)"',
        content,
    )
    seen = set()
    uniq = [u for u in all_urls if not (u in seen or seen.add(u))]

    def _ym(url):
        m = re.search(r"/uploads/(\d{4})/(\d{2})/", url)
        return (int(m.group(1)), int(m.group(2))) if m else (0, 0)

    found = {}
    for key, stem in URL_STEM_MAP.items():
        # Dosya adı tam olarak "<kök>-<hash>" ile başlamalı (kısmi çakışmaları önler;
        # ör. "Kira-Sertifikasi-" -> "Altina-Dayali-Kira-Sertifikasi"yi YAKALAMAZ).
        cands = [u for u in uniq if u.rsplit("/", 1)[-1].startswith(stem + "-")]
        if cands:
            found[key] = max(cands, key=_ym)  # birden fazlaysa en güncel klasör
    return found


def load_urls():
    """Güncel Excel URL'lerini yükle (önce otomatik keşif, sonra cache).

    1) HMB sayfasından güncel URL'leri keşfetmeyi dener.
    2) Bulunanları urls.json'a yazar (cache + diğer scriptler için fallback).
    3) Keşfedilemeyen seriler için urls.json'daki son bilinen değeri korur.
    Böylece 'güncelle' her zaman en yeni ayları çeker; site erişilemezse bile
    çalışmaya devam eder.
    """
    # Son bilinen URL'ler (cache)
    try:
        with open(URLS_FILE, "r", encoding="utf-8") as f:
            cached = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        cached = {}

    discovered = discover_urls()

    merged = dict(cached)
    changed = []
    for key, url in discovered.items():
        if cached.get(key) != url:
            changed.append(key)
        merged[key] = url

    if discovered:
        missing = [k for k in URL_STEM_MAP if k not in discovered]
        print(f"  Otomatik kesif: {len(discovered)}/{len(URL_STEM_MAP)} URL bulundu"
              + (f", {len(changed)} guncellendi." if changed else ", degisiklik yok."))
        for k in changed:
            print(f"      yeni -> {k}: {merged[k].rsplit('/', 1)[-1]}")
        if missing:
            print(f"  (Bulunamayan {len(missing)} seri cache'den kullanilacak: "
                  + ", ".join(missing) + ")")
        if changed:
            try:
                with open(URLS_FILE, "w", encoding="utf-8") as f:
                    json.dump(merged, f, ensure_ascii=False, indent=2)
                print("  urls.json guncellendi.")
            except Exception as e:
                print(f"  UYARI: urls.json yazilamadi: {e}")

    if not merged:
        raise RuntimeError("Hicbir URL bulunamadi (ne keşif ne cache).")
    return merged

# ── Sütun tanımları ──────────────────────────────────────────────
# Her sütunun (grup_adı, sütun_adı) şeklinde tuple'ı
COLUMNS = [
    # Grup 1: Genel Bilgiler
    ("Genel Bilgiler", "Senet Türü"),
    ("Genel Bilgiler", "Yıl"),
    ("Genel Bilgiler", "ISIN Kodu"),
    ("Genel Bilgiler", "Valör Tarihi"),
    ("Genel Bilgiler", "İtfa Tarihi"),
    ("Genel Bilgiler", "Vade (Gün)"),
    ("Genel Bilgiler", "Kupon Dönemi"),
    ("Genel Bilgiler", "Kupon Oranı (%)"),
    # Grup 2: Teklif Edilen
    ("Teklif Edilen Tutar", "Nominal (Bin TL)"),
    ("Teklif Edilen Tutar", "Net (Bin TL)"),
    # Grup 3: İhale Sonucu (Kabul Edilen)
    ("İhale Kabul Edilen Tutar", "Nominal (Bin TL)"),
    ("İhale Kabul Edilen Tutar", "Net (Bin TL)"),
    # Grup 4: Rekabetçi Olmayan - Kamu
    ("Rek. Olmayan Teklif - Kamu", "Nominal (Bin TL)"),
    ("Rek. Olmayan Teklif - Kamu", "Net (Bin TL)"),
    # Grup 5: Rekabetçi Olmayan - PY
    ("Rek. Olmayan Teklif - Piyasa Yapıcı", "Nominal (Bin TL)"),
    ("Rek. Olmayan Teklif - Piyasa Yapıcı", "Net (Bin TL)"),
    # Grup 6: İhale Sonrası Satış
    ("İhale Sonrası Satış", "Nominal (Bin TL)"),
    ("İhale Sonrası Satış", "Net (Bin TL)"),
    # Grup 7: Toplam Satış
    ("Toplam Satış", "Nominal (Bin TL)"),
    ("Toplam Satış", "Net (Bin TL)"),
    # Grup 8: Faiz & Fiyat
    ("Kabul Edilen Faiz (%)", "Ort. Dönem (Basit)"),
    ("Kabul Edilen Faiz (%)", "Ort. Yıllık Bileşik"),
    ("Fiyat", "İhraç Birim Fiyatı (Ort.)"),
]

# DataFrame için düz sütun adları (iç kullanım)
FLAT_COLS = [
    "senet_turu", "yil", "isin", "valor", "itfa", "vade", "kupon_donemi", "kupon_orani",
    "teklif_nom", "teklif_net",
    "kabul_nom", "kabul_net",
    "rek_kamu_nom", "rek_kamu_net",
    "rek_py_nom", "rek_py_net",
    "ihale_sonrasi_nom", "ihale_sonrasi_net",
    "toplam_nom", "toplam_net",
    "faiz_donem", "faiz_bilesik",
    "birim_fiyat",
]


def download_file(url):
    print(f"  İndiriliyor: {url.split('/')[-1]}")
    r = requests.get(url, timeout=30)
    r.raise_for_status()
    return r.content


# ── XLS parser (yıl bazlı sheet'ler) ────────────────────────────
def parse_xls_multi_sheet(content, senet_turu):
    wb = xlrd.open_workbook(file_contents=content)
    all_rows = []

    for sheet_name in wb.sheet_names():
        ws = wb.sheet_by_name(sheet_name)
        if ws.nrows < 6:
            continue

        yil = sheet_name.strip()
        header3 = [str(ws.cell_value(3, j)).strip().upper() for j in range(ws.ncols)]
        header4 = [str(ws.cell_value(4, j)).strip().upper() for j in range(ws.ncols)]
        header5 = [str(ws.cell_value(5, j)).strip().upper() for j in range(ws.ncols)] if ws.nrows > 5 else []

        cmap = _build_col_map(header3, header4, ws.ncols)

        # Veri başlangıcı: row 5 "Nominal/Net" alt başlığıysa row 6'dan başla
        data_start = 5
        if header5 and any(h in ("NOMİNAL", "NOMINAL", "NOMİNAL (BİN TL)") for h in header5):
            data_start = 6

        for i in range(data_start, ws.nrows):
            isin_col = cmap.get("isin", 1)
            isin_val = str(ws.cell_value(i, isin_col)).strip()
            if not isin_val:
                continue

            row = _extract(ws, i, cmap, senet_turu, yil)
            if row:
                all_rows.append(row)

    return all_rows


def _build_col_map(h3, h4, ncols):
    cm = {}

    # Row 3 header'larını sırayla tara
    # ÖNEMLİ: "Rekabetçi Olmayan Teklif" hem "REKABETÇİ" hem "TEKLİF" içerir.
    # Bu yüzden REKABETÇİ kontrolü TEKLİF'ten ÖNCE yapılmalı.
    for j in range(ncols):
        t = h3[j] if j < len(h3) else ""
        if not t:
            continue

        if "ISIN" in t:
            cm["isin"] = j
        elif any(k in t for k in ("VALÖR", "VALOR")):
            cm["valor"] = j
        elif any(k in t for k in ("İTFA", "ITFA")):
            cm["itfa"] = j
        elif "VADE" in t and "İTFA" not in t and "ITFA" not in t:
            cm["vade"] = j
        elif "KUPON DÖNEMİ" in t or "KUPON DÖNEM" in t:
            cm["kupon_donemi"] = j
        elif "KUPON ORANI" in t:
            cm["kupon_orani"] = j
        elif any(k in t for k in ("REKABETÇİ", "REKABET")):
            # "Rekabetçi Olmayan Teklif" - bu TEKLİF'ten ÖNCE yakalanmalı
            pass  # Row 4'ten Kamu/PY pozisyonlarıyla çözeceğiz
        elif any(k in t for k in ("TEKLİF", "TEKLIF")):
            # "Teklif Edilen Tutar" - ihaleye gelen teklif toplamı
            cm["teklif_nom"] = j
            cm["teklif_net"] = j + 1
        elif "KABUL" in t and "FAİZ" not in t and "FAIZ" not in t:
            # "İhalede Kabul Edilen Tutar" - rekabetçi ihale sonucu
            cm["kabul_nom"] = j
            cm["kabul_net"] = j + 1
        elif any(k in t for k in ("İHALE SONRASI", "IHALE SONRASI")):
            cm["ihale_sonrasi_nom"] = j
            cm["ihale_sonrasi_net"] = j + 1
        elif "TOPLAM" in t:
            cm["toplam_nom"] = j
            cm["toplam_net"] = j + 1
        elif "FAİZ" in t or "FAIZ" in t:
            # "İhalede Kabul Edilen Faiz (%)"
            cm["faiz_donem"] = j
            cm["faiz_bilesik"] = j + 1
        elif any(k in t for k in ("BİRİM", "BIRIM", "İHRAÇ", "IHRAC")):
            cm["birim_fiyat"] = j

    # Row 4: Kamu / PY pozisyonları (Rekabetçi Olmayan alt başlıkları)
    for j in range(ncols):
        t4 = h4[j] if j < len(h4) else ""
        if "KAMU" in t4:
            cm["rek_kamu_nom"] = j
            cm["rek_kamu_net"] = j + 1
        elif any(k in t4 for k in ("PİYASA", "PIYASA")):
            cm["rek_py_nom"] = j
            cm["rek_py_net"] = j + 1

    if "isin" not in cm:
        cm["isin"] = 1

    # Birim fiyat: genelde en son sütun
    if "birim_fiyat" not in cm:
        cm["birim_fiyat"] = ncols - 1

    return cm


def _safe(ws, row, col):
    try:
        if col is None or col >= ws.ncols:
            return None
        val = ws.cell_value(row, col)
        if val == "" or val is None:
            return None
        if ws.cell_type(row, col) == xlrd.XL_CELL_DATE:
            try:
                return xlrd.xldate_as_datetime(val, ws.book.datemode).strftime("%d.%m.%Y")
            except:
                return val
        return val
    except:
        return None


def _extract(ws, ri, cm, senet_turu, yil):
    isin = _safe(ws, ri, cm.get("isin"))
    if not isin:
        return None
    return dict(zip(FLAT_COLS, [
        senet_turu,
        yil,
        isin,
        _safe(ws, ri, cm.get("valor")),
        _safe(ws, ri, cm.get("itfa")),
        _safe(ws, ri, cm.get("vade")),
        _safe(ws, ri, cm.get("kupon_donemi")),
        _safe(ws, ri, cm.get("kupon_orani")),
        _safe(ws, ri, cm.get("teklif_nom")),
        _safe(ws, ri, cm.get("teklif_net")),
        _safe(ws, ri, cm.get("kabul_nom")),
        _safe(ws, ri, cm.get("kabul_net")),
        _safe(ws, ri, cm.get("rek_kamu_nom")),
        _safe(ws, ri, cm.get("rek_kamu_net")),
        _safe(ws, ri, cm.get("rek_py_nom")),
        _safe(ws, ri, cm.get("rek_py_net")),
        _safe(ws, ri, cm.get("ihale_sonrasi_nom")),
        _safe(ws, ri, cm.get("ihale_sonrasi_net")),
        _safe(ws, ri, cm.get("toplam_nom")),
        _safe(ws, ri, cm.get("toplam_net")),
        _safe(ws, ri, cm.get("faiz_donem")),
        _safe(ws, ri, cm.get("faiz_bilesik")),
        _safe(ws, ri, cm.get("birim_fiyat")),
    ]))


# ── Kira Sertifikası parser (tek sheet) ─────────────────────────
def parse_kira_sertifikasi(content):
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb[wb.sheetnames[0]]
    all_rows = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        isin = str(row[0]).strip()
        if not isin or isin.upper() == "ISIN KODU":
            continue
        para = str(row[1]).strip() if row[1] else ""
        if "TL" not in para.upper():
            continue

        valor = row[3]
        itfa = row[4]
        yil = ""
        if isinstance(valor, datetime):
            yil = str(valor.year)
            valor = valor.strftime("%d.%m.%Y")
        if isinstance(itfa, datetime):
            itfa = itfa.strftime("%d.%m.%Y")

        all_rows.append(dict(zip(FLAT_COLS, [
            "TL Cinsi Kira Sertifikalari",
            yil,
            isin,
            valor,
            itfa,
            row[5] if len(row) > 5 else None,   # vade
            row[7] if len(row) > 7 else None,   # kupon dönemi
            row[6] if len(row) > 6 else None,   # kupon oranı (kira oranı)
            None, None,                          # teklif
            row[8] if len(row) > 8 else None,   # kabul nom (nominal miktar)
            row[9] if len(row) > 9 else None,   # kabul net
            None, None,                          # rek kamu
            None, None,                          # rek py
            None, None,                          # ihale sonrası
            None, None,                          # toplam
            None, None,                          # faiz
            row[10] if len(row) > 10 else None,  # birim fiyat
        ])))

    return all_rows


# ── Excel formatlama ─────────────────────────────────────────────
GROUP_COLORS = {
    "Genel Bilgiler":                       "1F4E79",  # koyu mavi
    "Teklif Edilen Tutar":                  "2E75B6",  # mavi
    "İhale Kabul Edilen Tutar":             "548235",  # yeşil
    "Rek. Olmayan Teklif - Kamu":           "BF8F00",  # altın
    "Rek. Olmayan Teklif - Piyasa Yapıcı":  "C55A11",  # turuncu
    "İhale Sonrası Satış":                  "7030A0",  # mor
    "Toplam Satış":                         "385723",  # koyu yeşil
    "Kabul Edilen Faiz (%)":                "843C0C",  # kahve
    "Fiyat":                                "404040",  # gri
}


def format_sheet(ws_out, df):
    """İki satırlık grup + alt başlık header'ı oluştur ve formatla."""
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Satır 1: Grup başlıkları (merge edilecek) ──
    col_idx = 1
    group_spans = []
    current_group = None
    start_col = 1

    for i, (grp, sub) in enumerate(COLUMNS):
        if grp != current_group:
            if current_group is not None:
                group_spans.append((current_group, start_col, col_idx - 1))
            current_group = grp
            start_col = col_idx
        col_idx += 1
    group_spans.append((current_group, start_col, col_idx - 1))

    for grp, sc, ec in group_spans:
        color = GROUP_COLORS.get(grp, "333333")
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        font = Font(bold=True, color="FFFFFF", size=11)
        for c in range(sc, ec + 1):
            cell = ws_out.cell(row=1, column=c)
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        ws_out.cell(row=1, column=sc, value=grp)
        if ec > sc:
            ws_out.merge_cells(start_row=1, start_column=sc, end_row=1, end_column=ec)

    # ── Satır 2: Alt sütun başlıkları ──
    for i, (grp, sub) in enumerate(COLUMNS):
        c = i + 1
        color = GROUP_COLORS.get(grp, "333333")
        # Daha açık ton (alt başlık için)
        lighter = _lighten_hex(color)
        fill = PatternFill(start_color=lighter, end_color=lighter, fill_type="solid")
        cell = ws_out.cell(row=2, column=c, value=sub)
        cell.fill = fill
        cell.font = Font(bold=True, color="1A1A1A", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # ── Veri satırları (satır 3'ten itibaren) ──
    for row_idx in range(len(df)):
        for col_i, flat_col in enumerate(FLAT_COLS):
            c = col_i + 1
            val = df.iloc[row_idx][flat_col]
            # Yıl ve vade'yi tam sayı olarak yaz (formüller için)
            if flat_col == "yil" and pd.notna(val):
                val = int(val)
            elif flat_col == "vade" and pd.notna(val):
                val = int(val)
            elif pd.isna(val) if isinstance(val, float) else False:
                val = None
            cell = ws_out.cell(row=row_idx + 3, column=c, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # Sayısal sütunlar için format
            if flat_col.endswith(("_nom", "_net")):
                cell.number_format = '#,##0.00'
            elif flat_col in ("faiz_donem", "faiz_bilesik", "kupon_orani"):
                cell.number_format = '0.00'
            elif flat_col == "birim_fiyat":
                cell.number_format = '0.000'

    # ── Sütun genişlikleri ──
    widths = {
        "senet_turu": 36, "yil": 6, "isin": 16, "valor": 12, "itfa": 12,
        "vade": 10, "kupon_donemi": 12, "kupon_orani": 12,
    }
    default_width = 18
    for i, col in enumerate(FLAT_COLS):
        ws_out.column_dimensions[get_column_letter(i + 1)].width = widths.get(col, default_width)

    # Freeze: ilk 2 satır (header) ve ilk 3 sütun (senet türü, yıl, ISIN)
    ws_out.freeze_panes = "D3"

    # Satır yükseklikleri
    ws_out.row_dimensions[1].height = 28
    ws_out.row_dimensions[2].height = 36


def _lighten_hex(hex_color):
    """Hex rengi açık ton yap (header alt satırı için)."""
    r = int(hex_color[0:2], 16)
    g = int(hex_color[2:4], 16)
    b = int(hex_color[4:6], 16)
    factor = 0.55
    r = int(r + (255 - r) * factor)
    g = int(g + (255 - g) * factor)
    b = int(b + (255 - b) * factor)
    return f"{r:02X}{g:02X}{b:02X}"


# ── Özet Sayfası ──────────────────────────────────────────────────
def create_ozet_sheet(wb, df):
    """İlk sayfa olarak Özet raporu oluştur."""
    ws = wb.create_sheet("Özet", 0)
    ws.sheet_properties.tabColor = "C00000"

    thin = Side(style="thin", color="999999")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)
    no_brd = Border()

    # Stiller
    TITLE_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    SECTION_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    HDR_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HIGHLIGHT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    LIGHT_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
    UP_FONT = Font(bold=True, color="006100", size=10)
    DOWN_FONT = Font(bold=True, color="C00000", size=10)

    def cell(r, c, val, font=None, fill=None, align=None, nfmt=None, border=True):
        cl = ws.cell(row=r, column=c, value=val)
        if border:
            cl.border = brd
        if font:
            cl.font = font
        if fill:
            cl.fill = fill
        if align:
            cl.alignment = align
        if nfmt:
            cl.number_format = nfmt
        return cl

    def merge_fill(r, c1, c2, val, font, fill, align=None, height=None):
        for c in range(c1, c2 + 1):
            cl = ws.cell(row=r, column=c)
            cl.fill = fill
            cl.font = font
            cl.border = brd
            if align:
                cl.alignment = align
        ws.cell(row=r, column=c1, value=val)
        if align:
            ws.cell(row=r, column=c1).alignment = align
        if c2 > c1:
            ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
        if height:
            ws.row_dimensions[r].height = height

    # ── Veri hazırlığı ──
    dn = df.copy()
    for c in ['faiz_bilesik', 'vade', 'toplam_net', 'teklif_nom', 'kabul_nom',
              'rek_kamu_net', 'rek_py_net', 'ihale_sonrasi_net']:
        dn[c] = pd.to_numeric(dn[c], errors='coerce')
    dn['yil'] = pd.to_numeric(dn['yil'], errors='coerce')

    # Ay bilgisi çıkar (valor DD.MM.YYYY formatında)
    dn['ay'] = dn['valor'].apply(
        lambda x: int(str(x).split('.')[1]) if pd.notna(x) and '.' in str(x) else 0)

    d4 = dn[dn['yil'] == 2024]
    d5 = dn[dn['yil'] == 2025]
    d6 = dn[dn['yil'] == 2026]

    # 2026'da geçen ayları bul → 2025 aynı dönemi filtrele
    now = datetime.now()
    ytd_max_ay = now.month - 1 if now.day <= 5 else now.month  # ay başıysa önceki aya kadar
    ytd_max_ay = max(d6['ay'].max(), 1) if len(d6) > 0 else now.month - 1
    d5_ayni = d5[d5['ay'] <= ytd_max_ay]  # 2025 aynı dönem

    def stats(d):
        fb = d['faiz_bilesik'].dropna()
        vd = d['vade'].dropna()
        tn = d['toplam_net'].dropna()
        tk = d['teklif_nom'].dropna().sum()
        kb = d['kabul_nom'].dropna().sum()
        return {
            'n': len(d),
            'faiz': fb.mean() if len(fb) else 0,
            'vade': vd.mean() if len(vd) else 0,
            'avg_net': tn.mean() if len(tn) else 0,
            'sum_net': tn.sum(),
            'tk': tk / kb if kb > 0 else 0,
        }

    s4, s5, s6 = stats(d4), stats(d5), stats(d6)
    s5a = stats(d5_ayni)  # 2025 aynı dönem

    def fmt_tl(v):
        """Bin TL → okunabilir format."""
        if pd.isna(v) or v == 0:
            return "0 TL"
        if abs(v) >= 1_000_000:
            return f"{v / 1_000_000:,.2f} Milyar TL"
        if abs(v) >= 1_000:
            return f"{v / 1_000:,.1f} Milyon TL"
        return f"{v:,.0f} Bin TL"

    # ── Ana cümle ──
    sentence = (
        f"Hazine 2024 yılının tamamında ortalama %{s4['faiz']:.2f} faizle, "
        f"{int(s4['vade'])} gün vadeyle, toplamda {fmt_tl(s4['sum_net'])} borçlanırken, "
        f"2025 yılının tamamında bu rakamlar sırasıyla %{s5['faiz']:.2f}, "
        f"{int(s5['vade'])} gün ve {fmt_tl(s5['sum_net'])} olarak gerçekleşirken; "
        f"2026 yılında şuana kadar %{s6['faiz']:.2f}, "
        f"{int(s6['vade'])} gün ve {fmt_tl(s6['sum_net'])}'dir. "
        f"(Geçen yıl aynı dönemde %{s5a['faiz']:.2f}, "
        f"{int(s5a['vade'])} gün ve {fmt_tl(s5a['sum_net'])})"
    )

    # ════════════════════ LAYOUT ════════════════════
    COLS = 8
    r = 1

    # ── Başlık ──
    merge_fill(r, 1, COLS, "HAZİNE İHALE VERİLERİ — ÖZET RAPORU",
               Font(bold=True, color="FFFFFF", size=16), TITLE_FILL,
               Alignment(horizontal="center", vertical="center"), height=42)
    r += 1

    # Alt başlık
    today = datetime.now().strftime("%d.%m.%Y")
    merge_fill(r, 1, COLS,
               f"Son Güncelleme: {today}  •  Toplam {len(df)} İhale Kaydı  •  Kaynak: T.C. Hazine ve Maliye Bakanlığı",
               Font(italic=True, color="1F4E79", size=10),
               PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
               Alignment(horizontal="center", vertical="center"), height=24)
    r += 2

    # ── Ana paragraf ──
    merge_fill(r, 1, COLS, sentence,
               Font(color="333333", size=12), HIGHLIGHT_FILL,
               Alignment(horizontal="left", vertical="center", wrap_text=True), height=65)
    r += 2

    # ══════════ DÖNEMSEL KARŞILAŞTIRMA TABLOSU ══════════
    merge_fill(r, 1, COLS, "DÖNEMSEL KARŞILAŞTIRMA",
               Font(bold=True, color="FFFFFF", size=12), SECTION_FILL,
               Alignment(horizontal="center", vertical="center"), height=28)
    r += 1

    # Header
    comp_h = ["Dönem", "İhale Sayısı", "Ort. Faiz (%)", "Ort. Vade (Gün)",
              "İhale Başı Ort. (Bin TL)", "Toplam Net (Bin TL)", "Teklif/Kabul", "Yıllık Değişim"]
    for ci, h in enumerate(comp_h):
        cell(r, ci + 1, h,
             Font(bold=True, color="FFFFFF", size=10), HDR_FILL,
             Alignment(horizontal="center", vertical="center", wrap_text=True))
    ws.row_dimensions[r].height = 32
    r += 1

    # Veri satırları
    ay_adi = {1: "Ocak", 2: "Şubat", 3: "Mart", 4: "Nisan", 5: "Mayıs", 6: "Haziran",
              7: "Temmuz", 8: "Ağustos", 9: "Eylül", 10: "Ekim", 11: "Kasım", 12: "Aralık"}
    donem_label = f"Ocak-{ay_adi.get(ytd_max_ay, '')}"
    periods = [
        ("2024 (Tam Yıl)", s4, None),
        ("2025 (Tam Yıl)", s5, s4),
        (f"2026 ({donem_label})", s6, s5a),
        (f"2025 ({donem_label})", s5a, None),
    ]

    for pi, (label, s, prev) in enumerate(periods):
        fill = LIGHT_FILL if pi % 2 == 0 else None
        cell(r, 1, label, Font(bold=True, size=10), fill,
             Alignment(horizontal="left", vertical="center"))
        cell(r, 2, s['n'], nfmt='#,##0', fill=fill,
             align=Alignment(horizontal="center", vertical="center"))
        cell(r, 3, round(s['faiz'], 2), nfmt='0.00', fill=fill,
             align=Alignment(horizontal="center", vertical="center"))
        cell(r, 4, int(s['vade']), nfmt='#,##0', fill=fill,
             align=Alignment(horizontal="center", vertical="center"))
        cell(r, 5, round(s['avg_net'], 0), nfmt='#,##0', fill=fill,
             align=Alignment(horizontal="center", vertical="center"))
        cell(r, 6, round(s['sum_net'], 0), nfmt='#,##0', fill=fill,
             align=Alignment(horizontal="center", vertical="center"))
        cell(r, 7, round(s['tk'], 2), nfmt='0.00"x"', fill=fill,
             align=Alignment(horizontal="center", vertical="center"))
        # Değişim
        if prev:
            chg = s['faiz'] - prev['faiz']
            direction = "↓" if chg < 0 else "↑"
            chg_text = f"Faiz {direction} {abs(chg):.2f} puan"
            fnt = DOWN_FONT if chg < 0 else UP_FONT
            cell(r, 8, chg_text, fnt, fill,
                 Alignment(horizontal="center", vertical="center"))
        else:
            cell(r, 8, "Baz Dönem", Font(italic=True, color="888888", size=10), fill,
                 Alignment(horizontal="center", vertical="center"))
        r += 1

    r += 1

    # ══════════ DİKKAT ÇEKEN İSTATİSTİKLER ══════════
    merge_fill(r, 1, COLS, "DİKKAT ÇEKEN İSTATİSTİKLER",
               Font(bold=True, color="FFFFFF", size=12), SECTION_FILL,
               Alignment(horizontal="center", vertical="center"), height=28)
    r += 1

    # Stat satırları
    stat_label_font = Font(bold=True, color="1F4E79", size=11)
    stat_val_font = Font(color="333333", size=11)

    highlights = []

    # 1) Faiz trendi — aynı dönem karşılaştırması
    faiz_chg = s6['faiz'] - s5a['faiz']
    direction = "düştü" if faiz_chg < 0 else "yükseldi"
    highlights.append(("📉 Faiz Trendi (Aynı Dönem)",
        f"2025 {donem_label}: %{s5a['faiz']:.2f} → 2026 {donem_label}: %{s6['faiz']:.2f} "
        f"({faiz_chg:+.2f} puan, faizler {direction})"))

    # 2) Vade stratejisi — aynı dönem
    vade_chg = s6['vade'] - s5a['vade']
    direction = "uzadı" if vade_chg > 0 else "kısaldı"
    highlights.append(("📅 Vade Stratejisi",
        f"Ort. vade 2025 {donem_label}: {int(s5a['vade'])} gün → "
        f"2026: {int(s6['vade'])} gün ({direction}, {vade_chg:+.0f} gün)"))

    # 3) En büyük ihale
    idx_max = dn['toplam_net'].idxmax()
    if pd.notna(idx_max):
        top = dn.loc[idx_max]
        highlights.append(("🏆 Tarihin En Büyük İhalesi",
            f"{top['senet_turu']} — {top['valor']} — {fmt_tl(top['toplam_net'])}"))

    # 4) 2026 senet dağılımı
    if len(d6) > 0:
        dist = d6.groupby('senet_turu')['toplam_net'].sum().sort_values(ascending=False)
        top_senet = dist.index[0]
        top_pct = dist.iloc[0] / dist.sum() * 100
        highlights.append(("📊 2026 Ağırlıklı Senet Türü",
            f"{top_senet} (%{top_pct:.1f} pay)"))

    # 5) Talep gücü — aynı dönem
    tk_chg = s6['tk'] - s5a['tk']
    direction = "arttı" if tk_chg > 0 else "azaldı"
    highlights.append(("💪 Talep Gücü (Teklif/Kabul)",
        f"2025 {donem_label}: {s5a['tk']:.2f}x → 2026: {s6['tk']:.2f}x "
        f"(talep {direction}, {tk_chg:+.2f}x)"))

    # 6) 2026 yıllık projeksiyon
    if len(d6) > 0 and ytd_max_ay > 0:
        monthly_avg = s6['sum_net'] / ytd_max_ay
        annual_proj = monthly_avg * 12
        highlights.append(("🔮 2026 Yıllık Projeksiyon",
            f"İlk {ytd_max_ay} ayda {fmt_tl(s6['sum_net'])} → "
            f"Yıl sonu tahmini: ~{fmt_tl(annual_proj)} "
            f"(2025 toplamı: {fmt_tl(s5['sum_net'])})"))

    # 7) 2024 → 2025 tam yıl karşılaştırma
    faiz_2425 = s5['faiz'] - s4['faiz']
    net_chg = (s5['sum_net'] - s4['sum_net']) / s4['sum_net'] * 100 if s4['sum_net'] else 0
    highlights.append(("⚡ 2024 → 2025 Değişim",
        f"Faiz: %{s4['faiz']:.2f} → %{s5['faiz']:.2f} ({faiz_2425:+.2f} puan), "
        f"Toplam borçlanma: {fmt_tl(s4['sum_net'])} → {fmt_tl(s5['sum_net'])} "
        f"({net_chg:+.1f}%)"))

    # 8) Dönemsel borçlanma hacmi karşılaştırma
    if s5a['sum_net'] > 0:
        hacim_chg = (s6['sum_net'] - s5a['sum_net']) / s5a['sum_net'] * 100
        highlights.append(("📈 Dönemsel Borçlanma Hacmi",
            f"2025 {donem_label}: {fmt_tl(s5a['sum_net'])} → "
            f"2026 {donem_label}: {fmt_tl(s6['sum_net'])} "
            f"({hacim_chg:+.1f}% değişim)"))

    for hi, (label, text) in enumerate(highlights):
        fill = LIGHT_FILL if hi % 2 == 0 else None
        cell(r, 1, label, stat_label_font, fill,
             Alignment(horizontal="left", vertical="center"))
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        for c in range(1, 3):
            ws.cell(row=r, column=c).fill = fill or PatternFill()
            ws.cell(row=r, column=c).border = brd
        cell(r, 3, text, stat_val_font, fill,
             Alignment(horizontal="left", vertical="center", wrap_text=True))
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=COLS)
        for c in range(3, COLS + 1):
            ws.cell(row=r, column=c).fill = fill or PatternFill()
            ws.cell(row=r, column=c).border = brd
        ws.row_dimensions[r].height = 28
        r += 1

    # ── Sütun genişlikleri ──
    widths = [28, 16, 16, 16, 22, 22, 14, 28]
    for ci, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(ci + 1)].width = w

    ws.sheet_view.showGridLines = False
    print(f"  Özet sayfası oluşturuldu.")


# ── Ana akış ─────────────────────────────────────────────────────
def main():
    print("=" * 60)
    print("Hazine İhale Verileri Çekme Aracı")
    print("=" * 60)

    all_data = []
    URLS = load_urls()

    # Ödeme/Projeksiyon URL'lerini ihale parse'dan ayır
    SKIP_KEYS = {"Merkezi Yonetim Ic Borc Odemeleri",
                 "Merkezi Yonetim Ic Borc Odeme Projeksiyonlari"}

    for senet_turu, url in URLS.items():
        if senet_turu in SKIP_KEYS:
            continue
        print(f"\n>>> {senet_turu}")
        try:
            content = download_file(url)
            if senet_turu == "TL Cinsi Kira Sertifikalari":
                rows = parse_kira_sertifikasi(content)
            else:
                rows = parse_xls_multi_sheet(content, senet_turu)
            print(f"  {len(rows)} ihale kaydı bulundu.")
            all_data.extend(rows)
        except Exception as e:
            print(f"  HATA: {e}")
            import traceback
            traceback.print_exc()

    df = pd.DataFrame(all_data, columns=FLAT_COLS)

    # Yıl ve vade sütunlarını sayıya çevir (formüller için kritik)
    df["yil"] = pd.to_numeric(df["yil"], errors="coerce")
    df["vade"] = pd.to_numeric(df["vade"], errors="coerce")

    # Tarih sütunlarını string yap
    for col in ["valor", "itfa"]:
        df[col] = df[col].apply(lambda x: str(x) if x else "")

    # Çıktı her zaman script'in klasörüne yazılır (çalışma dizininden bağımsız —
    # otomasyon repo kökünden çalıştırdığında dosya yanlış yere düşmesin).
    output_file = os.path.join(SCRIPT_DIR, "hazine_ihale_verileri.xlsx")
    print(f"\n>>> Excel dosyasına yazılıyor: {output_file}")

    wb_out = openpyxl.Workbook()
    # Varsayılan sheet'i sil
    wb_out.remove(wb_out.active)

    # 1) Tüm İhaleler sheet'i
    ws_all = wb_out.create_sheet("Tüm İhaleler")
    format_sheet(ws_all, df)

    # 2) Her senet türü ayrı sheet
    for senet_turu in URLS.keys():
        df_f = df[df["senet_turu"] == senet_turu].reset_index(drop=True)
        if df_f.empty:
            continue
        sheet_name = senet_turu[:31]
        ws = wb_out.create_sheet(sheet_name)
        format_sheet(ws, df_f)

    # Özet sayfası (en başa)
    create_ozet_sheet(wb_out, df)

    wb_out.save(output_file)

    print(f"\n{'=' * 60}")
    print(f"Toplam {len(df)} ihale kaydı başarıyla çekildi.")
    print(f"Dosya: {output_file}")
    print(f"")
    print(f"Sütun Grupları:")
    print(f"  1. Genel Bilgiler          → Senet Türü, Yıl, ISIN, Tarih, Vade, Kupon")
    print(f"  2. Teklif Edilen Tutar      → İhaleye gelen toplam teklifler")
    print(f"  3. İhale Kabul Edilen Tutar → Rekabetçi ihalede kabul edilen")
    print(f"  4. Rek. Olmayan - Kamu      → Kamu kurumları rekabetçi olmayan")
    print(f"  5. Rek. Olmayan - PY        → Piyasa yapıcılar rekabetçi olmayan")
    print(f"  6. İhale Sonrası Satış      → İhale sonrası ek satışlar")
    print(f"  7. Toplam Satış             → Tüm kanalların toplamı")
    print(f"  8. Faiz & Fiyat             → Ort. basit, bileşik faiz + birim fiyat")
    print(f"{'=' * 60}")

    # Analizleri ekle
    print("\n>>> Analizler ekleniyor...")
    from hazine_analiz import write_analysis
    write_analysis(output_file)

    # İç Borç Ödeme & Projeksiyon sayfalarını ekle
    print("\n>>> İç Borç Ödeme verileri ekleniyor...")
    from hazine_odeme import write_odeme_sheets
    write_odeme_sheets(output_file)


if __name__ == "__main__":
    main()
