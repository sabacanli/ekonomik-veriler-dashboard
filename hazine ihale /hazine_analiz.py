"""
Hazine İhale Analiz Modülü - Excel Formülleri ile Yaşayan Tablolar
==================================================================
Tüm analiz tabloları SUMIF/COUNTIFS/AVERAGEIFS/SUMPRODUCT gibi
Excel formülleriyle oluşturulur. Veri güncellendiğinde analizler
otomatik olarak yeniden hesaplanır.
"""

import os

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Dosya her zaman script'in klasöründe (çalışma dizininden bağımsız).
_VARSAYILAN_XLSX = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                "hazine_ihale_verileri.xlsx")

# ── Sabitler ─────────────────────────────────────────────────────
DS = "Tüm İhaleler"  # veri sheet adı

# Veri sheet'indeki sütun harfleri
C = dict(
    senet="A", yil="B", isin="C", valor="D", itfa="E", vade="F",
    kupon_d="G", kupon_o="H",
    tek_n="I", tek_net="J", kab_n="K", kab_net="L",
    rk_n="M", rk_net="N", rp_n="O", rp_net="P",
    is_n="Q", is_net="R", top_n="S", top_net="T",
    f_d="U", f_b="V", bf="W",
)

KUPON_TYPES = [
    ("İskontolu", "*Kuponsuz*"),
    ("Sabit Kuponlu", "*Sabit*"),
    ("TÜFE Endeksli", "*TUFE*"),
    ("Değişken Faizli (FRN)", "*Degisken*"),
    ("TLREF Endeksli", "*TLREF*"),
    ("Kira Sertifikası", "*Kira*"),
]

VADE_BANDS = [
    ("0-182 gün (Kısa)", 0, 182),
    ("183-364 gün", 183, 364),
    ("365-728 gün (1-2 yıl)", 365, 728),
    ("729-1460 gün (2-4 yıl)", 729, 1460),
    ("1461-2555 gün (4-7 yıl)", 1461, 2555),
    ("2556+ gün (7+ yıl)", 2556, 99999),
]

AY_ADLARI = {
    1: "Ocak", 2: "Şubat", 3: "Mart", 4: "Nisan", 5: "Mayıs", 6: "Haziran",
    7: "Temmuz", 8: "Ağustos", 9: "Eylül", 10: "Ekim", 11: "Kasım", 12: "Aralık",
}

SATISKANALLARI = [
    ("İhale Kabul (Rekabetçi)", "kab_net"),
    ("Rek. Olmayan - Kamu", "rk_net"),
    ("Rek. Olmayan - Piyasa Yapıcı", "rp_net"),
    ("İhale Sonrası Satış", "is_net"),
]

# ── Stiller ──────────────────────────────────────────────────────
THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TITLE_FONT = Font(bold=True, color="1F4E79", size=13)
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
HDR_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
SUB_FONT = Font(bold=True, color="1A1A1A", size=10)
SUB_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
TOT_FONT = Font(bold=True, color="FFFFFF", size=10)
TOT_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
POS_FONT = Font(color="006100", bold=True)
NEG_FONT = Font(color="9C0006", bold=True)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


# ── Yardımcılar ──────────────────────────────────────────────────
def R(key, lr):
    """Veri sheet'inde mutlak sütun aralığı: 'Tüm İhaleler'!$X$3:$X$lr"""
    return f"'{DS}'!${C[key]}$3:${C[key]}${lr}"


def _c(ws, r, c, val, font=None, fill=None, nfmt=None, align=CENTER):
    """Hücreye değer/formül yaz ve formatla."""
    cell = ws.cell(row=r, column=c, value=val)
    cell.border = BORDER
    cell.alignment = align
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if nfmt:
        cell.number_format = nfmt
    return cell


def _hdr_row(ws, r, headers, font=HDR_FONT, fill=HDR_FILL):
    """Header satırı yaz."""
    for ci, h in enumerate(headers):
        _c(ws, r, ci + 1, h, font=font, fill=fill)


def _set_widths(ws, widths):
    """Sütun genişliklerini ayarla."""
    for ci, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(ci + 1)].width = w


def _chg_formula(r, col_new, col_old):
    """Değişim yüzdesi formülü: (new-old)/ABS(old)"""
    cn = get_column_letter(col_new)
    co = get_column_letter(col_old)
    return f'=IFERROR(({cn}{r}-{co}{r})/ABS({co}{r}),"")'


# ── ANALİZ 1: Yıl Bazlı Genel Özet ─────────────────────────────
def _write_yil_ozet(ws, years, lr, start_row):
    r = start_row
    _c(ws, r, 1, "1. Yıl Bazlı Genel Özet", font=TITLE_FONT, align=LEFT)
    r += 1

    # Header: Metrik | 2023 | 2024 | ... | Değişim
    headers = ["Metrik"] + [int(y) for y in years]
    if len(years) >= 2:
        headers.append(f"{int(years[-2])}→{int(years[-1])} Değişim")
    _hdr_row(ws, r, headers)
    r += 1

    yr = R("yil", lr)  # yıl aralığı

    # Her metrik için formüller
    metrics = [
        ("İhale Sayısı", "#,##0"),
        ("Toplam Borçlanma Net (Bin TL)", "#,##0"),
        ("Ortalama Faiz (%) Yıllık Bileşik", "0.00"),
        ("Ortalama Vadesi (gün)", "#,##0"),
        ("Ortalama İhraç Fiyatı", "0.000"),
        ("Teklif/Kabul Oranı (Nominal)", "0.00\"x\""),
        ("Toplam Teklif Nominal (Bin TL)", "#,##0"),
        ("Toplam Kabul Nominal (Bin TL)", "#,##0"),
        ("Ort. İhale Başına Borçlanma Net (Bin TL)", "#,##0"),
    ]

    for mi, (label, nfmt) in enumerate(metrics):
        _c(ws, r, 1, label, align=LEFT)

        for yi, y in enumerate(years):
            col = yi + 2
            cl = get_column_letter(col)
            ycell = f"{cl}${start_row + 1}"  # yıl hücresi referansı

            if mi == 0:  # İhale Sayısı
                f = f"=COUNTIF({yr},{ycell})"
            elif mi == 1:  # Toplam Net
                f = f"=SUMIF({yr},{ycell},{R('top_net', lr)})"
            elif mi == 2:  # Ort Faiz
                f = f'=IFERROR(AVERAGEIFS({R("f_b", lr)},{yr},{ycell},{R("f_b", lr)},"<>"),"")'
            elif mi == 3:  # Ort Vade
                f = f'=IFERROR(AVERAGEIFS({R("vade", lr)},{yr},{ycell},{R("vade", lr)},"<>"),"")'
            elif mi == 4:  # Ort Fiyat
                f = f'=IFERROR(AVERAGEIFS({R("bf", lr)},{yr},{ycell},{R("bf", lr)},"<>"),"")'
            elif mi == 5:  # Teklif/Kabul
                f = f"=IFERROR(SUMIF({yr},{ycell},{R('tek_n', lr)})/SUMIF({yr},{ycell},{R('kab_n', lr)}),\"\")"
            elif mi == 6:  # Toplam Teklif
                f = f"=SUMIF({yr},{ycell},{R('tek_n', lr)})"
            elif mi == 7:  # Toplam Kabul
                f = f"=SUMIF({yr},{ycell},{R('kab_n', lr)})"
            elif mi == 8:  # Ort İhale Başına
                # Bu sheet'teki satırları referansla: toplam_net / ihale_sayısı
                net_row = start_row + 2 + 1  # 2. metrik satırı (toplam net)
                cnt_row = start_row + 2       # 1. metrik satırı (ihale sayısı)
                f = f"=IFERROR({cl}{net_row}/{cl}{cnt_row},\"\")"

            _c(ws, r, col, f, nfmt=nfmt)

        # Değişim sütunu
        if len(years) >= 2:
            chg_col = len(years) + 2
            _c(ws, r, chg_col, _chg_formula(r, chg_col - 1, chg_col - 2), nfmt="0.0%")
        r += 1

    return r


# ── ANALİZ 2: Kupon Türüne Göre Borçlanma Dağılımı ──────────────
def _write_kupon_dagilim(ws, years, lr, start_row):
    r = start_row
    _c(ws, r, 1, "2. Kupon Türüne Göre Borçlanma Dağılımı (Toplam Satış Net, Bin TL)",
       font=TITLE_FONT, align=LEFT)
    r += 1

    headers = ["Kupon Türü"]
    for y in years:
        headers.extend([f"{int(y)} Sayı", f"{int(y)} Net", f"{int(y)} Pay"])
    _hdr_row(ws, r, headers, font=SUB_FONT, fill=SUB_FILL)
    r += 1

    yr = R("yil", lr)
    sr = R("senet", lr)
    tr = R("top_net", lr)
    total_start_row = r + len(KUPON_TYPES)  # TOPLAM satırı

    for ki, (label, pattern) in enumerate(KUPON_TYPES):
        _c(ws, r, 1, label, align=LEFT)
        for yi, y in enumerate(years):
            bc = 2 + yi * 3  # base column for this year

            # Yıl hücresi: header satırındaki yıl sütunundan referans
            # Sayı sütunundaki yılı kullan
            yr_cell = f"${get_column_letter(bc)}${start_row + 1}"
            # Hmm, header'da yıl değil "2023 Sayı" var. Daha iyi: TOPLAM satırından veya
            # direkt yıl değeri kullan
            y_val = int(y)

            # Sayı
            _c(ws, r, bc,
               f'=COUNTIFS({yr},{y_val},{sr},"{pattern}")',
               nfmt="#,##0")
            # Net
            _c(ws, r, bc + 1,
               f'=SUMIFS({tr},{yr},{y_val},{sr},"{pattern}")',
               nfmt="#,##0")
            # Pay = bu satırın net / TOPLAM satırın net
            net_cell = f"{get_column_letter(bc + 1)}{r}"
            total_net_cell = f"{get_column_letter(bc + 1)}{total_start_row}"
            _c(ws, r, bc + 2,
               f'=IFERROR({net_cell}/{total_net_cell},"")',
               nfmt="0.0%")
        r += 1

    # TOPLAM satırı
    _c(ws, r, 1, "TOPLAM", font=TOT_FONT, fill=TOT_FILL, align=LEFT)
    for yi, y in enumerate(years):
        bc = 2 + yi * 3
        y_val = int(y)
        _c(ws, r, bc,
           f"=COUNTIF({yr},{y_val})",
           font=TOT_FONT, fill=TOT_FILL, nfmt="#,##0")
        _c(ws, r, bc + 1,
           f"=SUMIF({yr},{y_val},{tr})",
           font=TOT_FONT, fill=TOT_FILL, nfmt="#,##0")
        _c(ws, r, bc + 2, 1.0, font=TOT_FONT, fill=TOT_FILL, nfmt="0.0%")
    r += 1

    return r


# ── ANALİZ 3: Kupon Türüne Göre Ortalama Faiz ───────────────────
def _write_kupon_faiz(ws, years, lr, start_row):
    r = start_row
    _c(ws, r, 1, "3. Kupon Türüne Göre Ortalama Faiz (%) Yıllık Bileşik",
       font=TITLE_FONT, align=LEFT)
    r += 1

    headers = ["Kupon Türü"]
    for y in years:
        headers.extend([f"{int(y)} Ort. Faiz", f"{int(y)} Ort. Vade"])
    _hdr_row(ws, r, headers, font=SUB_FONT, fill=SUB_FILL)
    r += 1

    yr = R("yil", lr)
    sr = R("senet", lr)
    fr = R("f_b", lr)
    vr = R("vade", lr)

    # Kira Sertifikası hariç (faiz verisi yok)
    for label, pattern in KUPON_TYPES:
        if "Kira" in pattern:
            continue
        _c(ws, r, 1, label, align=LEFT)
        for yi, y in enumerate(years):
            bc = 2 + yi * 2
            y_val = int(y)
            _c(ws, r, bc,
               f'=IFERROR(AVERAGEIFS({fr},{yr},{y_val},{sr},"{pattern}",{fr},"<>"),"")',
               nfmt="0.00")
            _c(ws, r, bc + 1,
               f'=IFERROR(AVERAGEIFS({vr},{yr},{y_val},{sr},"{pattern}",{vr},"<>"),"")',
               nfmt="#,##0")
        r += 1

    return r


# ── ANALİZ 4: Aylık Borçlanma Dağılımı ──────────────────────────
def _write_aylik(ws, years, lr, start_row):
    r = start_row
    _c(ws, r, 1, "4. Aylık Borçlanma Dağılımı (Toplam Satış Net, Bin TL)",
       font=TITLE_FONT, align=LEFT)
    r += 1

    headers = ["Ay"]
    for y in years:
        headers.extend([f"{int(y)} Sayı", f"{int(y)} Net", f"{int(y)} Faiz"])
    _hdr_row(ws, r, headers)
    r += 1

    yr = R("yil", lr)
    dr = R("valor", lr)  # valör tarihi (DD.MM.YYYY format text)
    tr = R("top_net", lr)
    fr = R("f_b", lr)
    total_row = r + 12  # TOPLAM satırı (12 ay sonra)

    for ay in range(1, 13):
        _c(ws, r, 1, AY_ADLARI[ay], align=LEFT)
        # Ay pattern: "??.01.*" formatında wildcard
        ay_pattern = f"??.{ay:02d}.*"

        for yi, y in enumerate(years):
            bc = 2 + yi * 3
            y_val = int(y)

            # Sayı
            _c(ws, r, bc,
               f'=COUNTIFS({yr},{y_val},{dr},"{ay_pattern}")',
               nfmt="#,##0")
            # Net
            _c(ws, r, bc + 1,
               f'=SUMIFS({tr},{yr},{y_val},{dr},"{ay_pattern}")',
               nfmt="#,##0")
            # Faiz
            _c(ws, r, bc + 2,
               f'=IFERROR(AVERAGEIFS({fr},{yr},{y_val},{dr},"{ay_pattern}",{fr},"<>"),"")',
               nfmt="0.00")
        r += 1

    # TOPLAM satırı
    _c(ws, r, 1, "TOPLAM", font=TOT_FONT, fill=TOT_FILL, align=LEFT)
    for yi, y in enumerate(years):
        bc = 2 + yi * 3
        y_val = int(y)
        _c(ws, r, bc,
           f"=COUNTIF({yr},{y_val})",
           font=TOT_FONT, fill=TOT_FILL, nfmt="#,##0")
        _c(ws, r, bc + 1,
           f"=SUMIF({yr},{y_val},{tr})",
           font=TOT_FONT, fill=TOT_FILL, nfmt="#,##0")
        _c(ws, r, bc + 2,
           f'=IFERROR(AVERAGEIFS({fr},{yr},{y_val},{fr},"<>"),"")',
           font=TOT_FONT, fill=TOT_FILL, nfmt="0.00")
    r += 1

    return r


# ── ANALİZ 5: Vade Dağılımı ─────────────────────────────────────
def _write_vade_dagilim(ws, years, lr, start_row):
    r = start_row
    _c(ws, r, 1, "5. Vade Dağılımı (Toplam Satış Net, Bin TL)",
       font=TITLE_FONT, align=LEFT)
    r += 1

    headers = ["Vade Aralığı"]
    for y in years:
        headers.extend([f"{int(y)} Sayı", f"{int(y)} Net", f"{int(y)} Pay", f"{int(y)} Ort.Faiz"])
    _hdr_row(ws, r, headers)
    r += 1

    yr = R("yil", lr)
    vr = R("vade", lr)
    tr = R("top_net", lr)
    fr = R("f_b", lr)

    # Her yılın toplam net'ini hesaplayan satır lazım - en alta koyacağız
    total_row = r + len(VADE_BANDS)

    for label, lo, hi in VADE_BANDS:
        _c(ws, r, 1, label, align=LEFT)
        for yi, y in enumerate(years):
            bc = 2 + yi * 4
            y_val = int(y)

            # Sayı
            _c(ws, r, bc,
               f"=COUNTIFS({yr},{y_val},{vr},\">=\"&{lo},{vr},\"<=\"&{hi})",
               nfmt="#,##0")
            # Net
            _c(ws, r, bc + 1,
               f"=SUMIFS({tr},{yr},{y_val},{vr},\">=\"&{lo},{vr},\"<=\"&{hi})",
               nfmt="#,##0")
            # Pay
            net_cell = f"{get_column_letter(bc + 1)}{r}"
            tot_cell = f"{get_column_letter(bc + 1)}{total_row}"
            _c(ws, r, bc + 2,
               f'=IFERROR({net_cell}/{tot_cell},"")',
               nfmt="0.0%")
            # Ort Faiz
            _c(ws, r, bc + 3,
               f'=IFERROR(AVERAGEIFS({fr},{yr},{y_val},{vr},">="&{lo},{vr},"<="&{hi},{fr},"<>"),"")',
               nfmt="0.00")
        r += 1

    # TOPLAM satırı
    _c(ws, r, 1, "TOPLAM", font=TOT_FONT, fill=TOT_FILL, align=LEFT)
    for yi, y in enumerate(years):
        bc = 2 + yi * 4
        y_val = int(y)
        _c(ws, r, bc,
           f"=COUNTIF({yr},{y_val})",
           font=TOT_FONT, fill=TOT_FILL, nfmt="#,##0")
        _c(ws, r, bc + 1,
           f"=SUMIF({yr},{y_val},{tr})",
           font=TOT_FONT, fill=TOT_FILL, nfmt="#,##0")
        _c(ws, r, bc + 2, 1.0, font=TOT_FONT, fill=TOT_FILL, nfmt="0.0%")
        _c(ws, r, bc + 3,
           f'=IFERROR(AVERAGEIFS({fr},{yr},{y_val},{fr},"<>"),"")',
           font=TOT_FONT, fill=TOT_FILL, nfmt="0.00")
    r += 1

    return r


# ── ANALİZ 6: Borçlanma Maliyeti ve Talep Analizi ───────────────
def _write_maliyet_talep(ws, years, lr, start_row):
    r = start_row
    _c(ws, r, 1, "6. Borçlanma Maliyeti ve Talep Analizi",
       font=TITLE_FONT, align=LEFT)
    r += 1

    headers = ["Metrik"] + [int(y) for y in years]
    if len(years) >= 2:
        headers.extend([f"Değişim ({int(years[-2])}→{int(years[-1])})", "Değişim (%)"])
    _hdr_row(ws, r, headers)
    r += 1

    yr = R("yil", lr)

    metrics = [
        ("Ağırlıklı Ort. Faiz (Net Bazlı, %)", "wavg_faiz", "0.00"),
        ("Toplam Net Borçlanma (Bin TL)", "sum", "top_net", "#,##0"),
        ("Teklif/Kabul Oranı", "ratio", None, "0.00\"x\""),
        ("Ort. İhale Başına Net (Bin TL)", "avg_per", None, "#,##0"),
        ("Toplam Rek. Kamu Net (Bin TL)", "sum", "rk_net", "#,##0"),
        ("Toplam Rek. PY Net (Bin TL)", "sum", "rp_net", "#,##0"),
        ("Toplam İhale Sonrası Net (Bin TL)", "sum", "is_net", "#,##0"),
    ]

    for mi, metric_def in enumerate(metrics):
        label = metric_def[0]
        mtype = metric_def[1]
        _c(ws, r, 1, label, align=LEFT)

        for yi, y in enumerate(years):
            col = yi + 2
            cl = get_column_letter(col)
            y_val = int(y)

            if mtype == "wavg_faiz":
                nfmt = metric_def[2]
                # Ağırlıklı ortalama: SUMPRODUCT(yil=year * faiz * net) / SUMIF(yil,year,net)
                f = (f"=IFERROR(SUMPRODUCT(({yr}={y_val})*({R('f_b', lr)})*({R('top_net', lr)}))"
                     f"/SUMIF({yr},{y_val},{R('top_net', lr)}),\"\")")
            elif mtype == "sum":
                col_key = metric_def[2]
                nfmt = metric_def[3]
                f = f"=SUMIF({yr},{y_val},{R(col_key, lr)})"
            elif mtype == "ratio":
                nfmt = metric_def[3]
                f = f"=IFERROR(SUMIF({yr},{y_val},{R('tek_n', lr)})/SUMIF({yr},{y_val},{R('kab_n', lr)}),\"\")"
            elif mtype == "avg_per":
                nfmt = metric_def[3]
                # toplam_net / ihale_sayısı - referans aynı sheet'teki satırlara
                net_row = start_row + 2 + 1  # 2. metrik (Toplam Net) satırı
                cnt_formula = f"COUNTIF({yr},{y_val})"
                net_formula = f"SUMIF({yr},{y_val},{R('top_net', lr)})"
                f = f"=IFERROR({net_formula}/{cnt_formula},\"\")"

            _c(ws, r, col, f, nfmt=nfmt)

        # Değişim sütunları
        if len(years) >= 2:
            nfmt_m = metric_def[2] if mtype == "wavg_faiz" else metric_def[3]
            abs_col = len(years) + 2
            pct_col = abs_col + 1
            prev_cl = get_column_letter(abs_col - 1)
            prev2_cl = get_column_letter(abs_col - 2)
            # Mutlak değişim
            _c(ws, r, abs_col,
               f"=IFERROR({prev_cl}{r}-{prev2_cl}{r},\"\")",
               nfmt=nfmt_m)
            # Yüzde değişim
            _c(ws, r, pct_col,
               _chg_formula(r, abs_col - 1, abs_col - 2),
               nfmt="0.0%")
        r += 1

    return r


# ── ANALİZ 7: Yatırımcı / Satış Kanalı Profili ──────────────────
def _write_yatirimci(ws, years, lr, start_row):
    r = start_row
    _c(ws, r, 1, "7. Yatırımcı / Satış Kanalı Dağılımı (Net, Bin TL)",
       font=TITLE_FONT, align=LEFT)
    r += 1

    headers = ["Kanal"]
    for y in years:
        headers.extend([f"{int(y)} Net", f"{int(y)} Pay"])
    _hdr_row(ws, r, headers, font=SUB_FONT, fill=SUB_FILL)
    r += 1

    yr = R("yil", lr)
    total_row = r + len(SATISKANALLARI)

    for label, col_key in SATISKANALLARI:
        _c(ws, r, 1, label, align=LEFT)
        for yi, y in enumerate(years):
            bc = 2 + yi * 2
            y_val = int(y)
            # Net
            _c(ws, r, bc,
               f"=SUMIF({yr},{y_val},{R(col_key, lr)})",
               nfmt="#,##0")
            # Pay
            net_cell = f"{get_column_letter(bc)}{r}"
            tot_cell = f"{get_column_letter(bc)}{total_row}"
            _c(ws, r, bc + 1,
               f'=IFERROR({net_cell}/{tot_cell},"")',
               nfmt="0.0%")
        r += 1

    # TOPLAM
    _c(ws, r, 1, "TOPLAM", font=TOT_FONT, fill=TOT_FILL, align=LEFT)
    for yi, y in enumerate(years):
        bc = 2 + yi * 2
        y_val = int(y)
        _c(ws, r, bc,
           f"=SUMIF({yr},{y_val},{R('top_net', lr)})",
           font=TOT_FONT, fill=TOT_FILL, nfmt="#,##0")
        _c(ws, r, bc + 1, 1.0, font=TOT_FONT, fill=TOT_FILL, nfmt="0.0%")
    r += 1

    return r


# ── ANALİZ 8: Çeyreklik Faiz Trendi ─────────────────────────────
def _write_ceyrek_faiz(ws, years_all, lr, start_row):
    """Çeyreklik faiz trendi - 2020'den itibaren."""
    r = start_row
    _c(ws, r, 1, "8. Çeyreklik Faiz Trendi (2020+)",
       font=TITLE_FONT, align=LEFT)
    r += 1

    headers = ["Yıl-Çeyrek", "İhale Sayısı", "Ort. Bileşik Faiz (%)",
               "Ağırlıklı Ort. Faiz (%)", "Ort. Vade (gün)", "Toplam Net (Bin TL)"]
    _hdr_row(ws, r, headers)
    r += 1

    yr = R("yil", lr)
    dr = R("valor", lr)
    fr = R("f_b", lr)
    vr = R("vade", lr)
    tr = R("top_net", lr)

    q_months = {
        "Q1": ["01", "02", "03"],
        "Q2": ["04", "05", "06"],
        "Q3": ["07", "08", "09"],
        "Q4": ["10", "11", "12"],
    }

    for y in [yy for yy in years_all if yy >= 2020]:
        for q_name, months in q_months.items():
            _c(ws, r, 1, f"{int(y)}-{q_name}")
            y_val = int(y)

            # 3 aylık pattern: birleştirerek COUNTIFS topla
            m_patterns = [f"??.{m}.*" for m in months]

            # Sayı: COUNTIFS(yil,y,valor,"??.01.*") + COUNTIFS(yil,y,valor,"??.02.*") + ...
            count_parts = [f'COUNTIFS({yr},{y_val},{dr},"{p}")' for p in m_patterns]
            _c(ws, r, 2, f"={'+'.join(count_parts)}", nfmt="#,##0")

            # Ort Faiz: SUMPRODUCT ile 3 ay birden
            # (yil=y) * ((MID(valor,4,2)="01") + (MID(valor,4,2)="02") + (MID(valor,4,2)="03")) * (faiz<>"")
            mid_conds = "+".join([f'(MID({dr},4,2)="{m}")' for m in months])
            _c(ws, r, 3,
               f'=IFERROR(SUMPRODUCT(({yr}={y_val})*({mid_conds})*({fr})*(({fr}<>"")*1))'
               f'/SUMPRODUCT(({yr}={y_val})*({mid_conds})*(({fr}<>"")*1)),"")',
               nfmt="0.00")

            # Ağırlıklı Ort Faiz
            _c(ws, r, 4,
               f'=IFERROR(SUMPRODUCT(({yr}={y_val})*({mid_conds})*({fr})*({tr}))'
               f'/SUMPRODUCT(({yr}={y_val})*({mid_conds})*({tr})),"")',
               nfmt="0.00")

            # Ort Vade
            _c(ws, r, 5,
               f'=IFERROR(SUMPRODUCT(({yr}={y_val})*({mid_conds})*({vr})*(({vr}<>"")*1))'
               f'/SUMPRODUCT(({yr}={y_val})*({mid_conds})*(({vr}<>"")*1)),"")',
               nfmt="#,##0")

            # Toplam Net
            sum_parts = [f'SUMIFS({tr},{yr},{y_val},{dr},"{p}")' for p in m_patterns]
            _c(ws, r, 6, f"={'+'.join(sum_parts)}", nfmt="#,##0")

            r += 1

    return r


# ── ANALİZ 9: En Büyük 20 İhale (LARGE + INDEX) ────────────────
def _write_top20(ws, lr, start_row):
    """Top 20 ihale - LARGE ve INDEX/MATCH formülleriyle."""
    r = start_row
    _c(ws, r, 1, "9. Tarihteki En Büyük 20 İhale (Toplam Net Bazlı)",
       font=TITLE_FONT, align=LEFT)
    r += 1

    headers = ["Sıra", "Toplam Net (Bin TL)", "Senet Türü", "ISIN",
               "Valör", "Yıl", "Vade (Gün)", "Faiz Bileşik (%)", "Birim Fiyat"]
    _hdr_row(ws, r, headers)
    r += 1

    tr = R("top_net", lr)

    for k in range(1, 21):
        _c(ws, r, 1, k)
        # Toplam Net (k. en büyük)
        _c(ws, r, 2,
           f"=IFERROR(LARGE({tr},{k}),\"\")",
           nfmt="#,##0")

        # Diğer sütunlar: INDEX(sütun, MATCH(LARGE değeri, toplam_net, 0))
        large_ref = f"B{r}"  # LARGE sonucunu referansla
        for ci, col_key in enumerate(["senet", "isin", "valor", "yil", "vade", "f_b", "bf"]):
            col_rng = R(col_key, lr)
            nfmt = None
            if col_key == "f_b":
                nfmt = "0.00"
            elif col_key == "bf":
                nfmt = "0.000"
            elif col_key == "vade":
                nfmt = "#,##0"
            _c(ws, r, 3 + ci,
               f"=IFERROR(INDEX({col_rng},MATCH({large_ref},{tr},0)),\"\")",
               nfmt=nfmt)
        r += 1

    return r


# ── Sheet Oluşturucular ──────────────────────────────────────────
def _create_genel_ozet(wb, years, lr):
    ws = wb.create_sheet("Analiz - Genel Özet")
    ws.sheet_properties.tabColor = "1F4E79"

    r = _write_yil_ozet(ws, years, lr, 1)
    r = _write_kupon_dagilim(ws, years, lr, r + 2)
    r = _write_kupon_faiz(ws, years, lr, r + 2)

    _set_widths(ws, [42] + [20] * 20)
    ws.freeze_panes = "B3"


def _create_aylik_vade(wb, years, lr):
    ws = wb.create_sheet("Analiz - Aylık & Vade")
    ws.sheet_properties.tabColor = "548235"

    r = _write_aylik(ws, years, lr, 1)
    r = _write_vade_dagilim(ws, years, lr, r + 2)

    _set_widths(ws, [26] + [18] * 20)
    ws.freeze_panes = "B3"


def _create_maliyet_talep(wb, years, lr):
    ws = wb.create_sheet("Analiz - Maliyet & Talep")
    ws.sheet_properties.tabColor = "C55A11"

    r = _write_maliyet_talep(ws, years, lr, 1)
    r = _write_yatirimci(ws, years, lr, r + 2)

    _set_widths(ws, [42] + [22] * 16)
    ws.freeze_panes = "B3"


def _create_detay(wb, years_all, lr):
    ws = wb.create_sheet("Analiz - Detay")
    ws.sheet_properties.tabColor = "7030A0"

    r = _write_ceyrek_faiz(ws, years_all, lr, 1)
    r = _write_top20(ws, lr, r + 2)

    _set_widths(ws, [16, 24, 36, 18, 14, 8, 12, 16, 12])
    ws.freeze_panes = "B3"


# ── Ana Giriş Noktası ────────────────────────────────────────────
def write_analysis(filepath=_VARSAYILAN_XLSX):
    print("Analiz formülleri ekleniyor...")
    wb = load_workbook(filepath)
    ws_data = wb[DS]
    lr = ws_data.max_row

    # Yılları bul
    years_all = set()
    for r in range(3, lr + 1):
        y = ws_data.cell(r, 2).value
        if y is not None:
            try:
                years_all.add(int(y))
            except (ValueError, TypeError):
                pass
    years_all = sorted(years_all)
    recent = [y for y in years_all if y >= 2023]

    # Mevcut analiz sheet'lerini sil
    for name in list(wb.sheetnames):
        if name.startswith("Analiz"):
            del wb[name]

    _create_genel_ozet(wb, recent, lr)
    _create_aylik_vade(wb, recent, lr)
    _create_maliyet_talep(wb, recent, lr)
    _create_detay(wb, years_all, lr)

    wb.save(filepath)

    print(f"  Formül tabanlı analizler eklendi: {filepath}")
    print("  - Analiz - Genel Özet      (1. Yıl Özet, 2. Kupon Dağılım, 3. Kupon Faiz)")
    print("  - Analiz - Aylık & Vade    (4. Aylık Dağılım, 5. Vade Dağılımı)")
    print("  - Analiz - Maliyet & Talep (6. Maliyet Analizi, 7. Yatırımcı Profili)")
    print("  - Analiz - Detay           (8. Çeyreklik Faiz Trendi, 9. Top 20 İhale)")
    print("  Tüm tablolar Excel formülleriyle yaşıyor - veri güncellenince otomatik hesaplanır.")


if __name__ == "__main__":
    write_analysis()
