"""
Altın Tahvili & Altına Dayalı Kira Sertifikası İhale Verilerini Birleştirme
============================================================================
HMB sitesinden indirilen iki Excel dosyasını parse edip
tek bir Excel dosyasında birleştirir.

Valör tarihindeki altın ons fiyatı (USD) ve USDTRY kuru çekilerek
gram altın miktarının USD ve TRY karşılıkları hesaplanır.

Formül:
  USD = (gram / 31.1035) × altın_ons_fiyat_usd
  TRY = USD × USDTRY
"""

import openpyxl
import pandas as pd
import yfinance as yf
import os
from datetime import datetime, timedelta
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

TROY_OUNCE_GRAM = 31.1035

# Kaynak dosyalar
ALTIN_KIRA_FILE = os.path.join(SCRIPT_DIR, "altin_kira_sertifikasi.xlsx")
ALTIN_TAHVIL_FILE = os.path.join(SCRIPT_DIR, "altin_tahvili.xlsx")
OUTPUT_FILE = os.path.join(SCRIPT_DIR, "altin_ihale_verileri.xlsx")

# Sütun tanımları (grup, alt başlık)
COLUMNS = [
    ("Genel Bilgiler", "Senet Türü"),
    ("Genel Bilgiler", "Yıl"),
    ("Genel Bilgiler", "ISIN Kodu"),
    ("Genel Bilgiler", "Talep Toplama Tarihi"),
    ("Genel Bilgiler", "Valör Tarihi"),
    ("Genel Bilgiler", "İtfa Tarihi"),
    ("Genel Bilgiler", "Vade (Gün)"),
    ("Genel Bilgiler", "Kupon/Kira Dönemi"),
    ("Genel Bilgiler", "Kupon/Kira Oranı (%)"),
    ("İhraç Bilgileri", "Kabul Edilen Altın Miktarı (Gram)"),
    ("İhraç Bilgileri", "Toplam İhraç Adedi"),
    ("Piyasa Verileri", "Altın Ons Fiyatı (USD)"),
    ("Piyasa Verileri", "USDTRY Kuru"),
    ("Hesaplanan Değerler", "Altın Miktarı (Ons)"),
    ("Hesaplanan Değerler", "USD Karşılığı"),
    ("Hesaplanan Değerler", "TRY Karşılığı"),
]

FLAT_COLS = [
    "senet_turu", "yil", "isin", "talep_tarihi", "valor", "itfa",
    "vade", "kupon_donemi", "kupon_orani", "altin_miktar", "ihrac_adedi",
    "altin_ons_fiyat", "usdtry", "altin_ons", "usd_karsilik", "try_karsilik",
]

GROUP_COLORS = {
    "Genel Bilgiler": "1F4E79",
    "İhraç Bilgileri": "548235",
    "Piyasa Verileri": "BF8F00",
    "Hesaplanan Değerler": "C55A11",
}


def _fmt_date(val):
    """datetime → DD.MM.YYYY string."""
    if isinstance(val, datetime):
        return val.strftime("%d.%m.%Y")
    return val


def parse_altin_excel(filepath, senet_turu_label):
    """Altın tahvil/kira sertifikası Excel dosyasını parse et.
    2020+ sheet'leri 10 sütun, 2017-2019 sheet'leri 14 sütun yapısında."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    all_rows = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row < 3:
            continue

        yil = sheet_name.strip()
        ncols = ws.max_column

        # Header'dan yapıyı belirle
        h2 = [str(ws.cell(2, c).value or '').upper() for c in range(1, ncols + 1)]

        # 14 sütunlu eski format (2017-2019): Etap sütunu var
        is_old_format = ncols >= 14 and any('ETAP' in h for h in h2)

        for r in range(3, ws.max_row + 1):
            if is_old_format:
                # Eski format: col1=ihraç dönemi, col2=ISIN, col3=Etap, col4=Senet Türü,
                # col5=Talep Başlangıç, col6=Talep Bitiş, col7=Valör, col8=İtfa,
                # col9=Fiziki Altın Dönemi, col10=Vade, col11=Kupon Dönemi,
                # col12=Kupon Oranı, col13=Altın Miktar, col14=Adet
                isin = ws.cell(r, 2).value
                if not isin or not str(isin).strip():
                    continue
                isin = str(isin).strip()
                if isin.startswith("*") or isin.upper().startswith("NOT") or isin.upper().startswith("ISIN"):
                    continue

                senet_val = ws.cell(r, 4).value
                if not senet_val:
                    continue

                talep = ws.cell(r, 5).value  # talep başlangıç tarihi
                valor = ws.cell(r, 7).value
                itfa = ws.cell(r, 8).value
                vade = ws.cell(r, 10).value
                kupon_donemi = ws.cell(r, 11).value
                kupon_orani = ws.cell(r, 12).value
                altin_miktar = ws.cell(r, 13).value
                ihrac_adedi = ws.cell(r, 14).value
            else:
                # Yeni format (2020+): 10 sütun
                isin = ws.cell(r, 1).value
                if not isin or not str(isin).strip():
                    continue
                isin = str(isin).strip()
                if isin.startswith("*") or isin.upper().startswith("NOT") or isin.upper().startswith("ISIN"):
                    continue

                senet_val = ws.cell(r, 2).value
                if not senet_val:
                    continue

                talep = ws.cell(r, 3).value
                valor = ws.cell(r, 4).value
                itfa = ws.cell(r, 5).value
                vade = ws.cell(r, 6).value
                kupon_donemi = ws.cell(r, 7).value
                kupon_orani = ws.cell(r, 8).value
                altin_miktar = ws.cell(r, 9).value
                ihrac_adedi = ws.cell(r, 10).value

            # Altın miktarı sayısal olmalı
            if not isinstance(altin_miktar, (int, float)):
                continue

            all_rows.append({
                "senet_turu": senet_turu_label,
                "yil": yil,
                "isin": isin,
                "talep_tarihi": _fmt_date(talep),
                "valor": _fmt_date(valor),
                "itfa": _fmt_date(itfa),
                "vade": vade,
                "kupon_donemi": kupon_donemi,
                "kupon_orani": kupon_orani,
                "altin_miktar": altin_miktar,
                "ihrac_adedi": ihrac_adedi,
            })

    return all_rows


def fetch_market_data():
    """Yahoo Finance'ten altın ons fiyatı ve USDTRY tarihsel verisini çek."""
    end_date = (datetime.now() + timedelta(days=1)).strftime('%Y-%m-%d')

    print("   Altın ons fiyatları çekiliyor (GC=F)...")
    gold = yf.download('GC=F', start='2017-01-01', end=end_date, progress=False)

    print("   USDTRY kuru çekiliyor (USDTRY=X)...")
    usdtry = yf.download('USDTRY=X', start='2017-01-01', end=end_date, progress=False)

    # MultiIndex sütunları düzleştir
    if isinstance(gold.columns, pd.MultiIndex):
        gold.columns = gold.columns.get_level_values(0)
    if isinstance(usdtry.columns, pd.MultiIndex):
        usdtry.columns = usdtry.columns.get_level_values(0)

    print(f"   → Altın: {len(gold)} gün ({gold.index[0].date()} - {gold.index[-1].date()})")
    print(f"   → USDTRY: {len(usdtry)} gün ({usdtry.index[0].date()} - {usdtry.index[-1].date()})")

    return gold['Close'], usdtry['Close']


def get_price_on_date(series, target_date, lookback_days=5):
    """Belirli tarihteki fiyatı bul. Tatilse önceki iş gününü al."""
    for offset in range(lookback_days + 1):
        dt = target_date - timedelta(days=offset)
        ts = pd.Timestamp(dt)
        if ts in series.index:
            return float(series[ts])
    return None


def parse_date(d):
    try:
        return datetime.strptime(str(d), "%d.%m.%Y")
    except Exception:
        return None


def _lighten_hex(hex_color):
    r = int(hex_color[0:2], 16)
    g = int(hex_color[2:4], 16)
    b = int(hex_color[4:6], 16)
    factor = 0.55
    r = int(r + (255 - r) * factor)
    g = int(g + (255 - g) * factor)
    b = int(b + (255 - b) * factor)
    return f"{r:02X}{g:02X}{b:02X}"


def format_sheet(ws_out, df, columns=None, flat_cols=None):
    """İki satırlık grup + alt başlık header'ı oluştur ve formatla."""
    if columns is None:
        columns = COLUMNS
    if flat_cols is None:
        flat_cols = FLAT_COLS

    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Satır 1: Grup başlıkları (merge)
    col_idx = 1
    group_spans = []
    current_group = None
    start_col = 1

    for i, (grp, sub) in enumerate(columns):
        if grp != current_group:
            if current_group is not None:
                group_spans.append((current_group, start_col, col_idx - 1))
            current_group = grp
            start_col = col_idx
        col_idx += 1
    group_spans.append((current_group, start_col, col_idx - 1))

    for grp, sc, ec in group_spans:
        color = GROUP_COLORS.get(grp, "333333")
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        font = Font(bold=True, color="FFFFFF", size=11)
        for c in range(sc, ec + 1):
            cell = ws_out.cell(row=1, column=c)
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        ws_out.cell(row=1, column=sc, value=grp)
        if ec > sc:
            ws_out.merge_cells(start_row=1, start_column=sc, end_row=1, end_column=ec)

    # Satır 2: Alt sütun başlıkları
    for i, (grp, sub) in enumerate(columns):
        c = i + 1
        color = GROUP_COLORS.get(grp, "333333")
        lighter = _lighten_hex(color)
        fill = PatternFill(start_color=lighter, end_color=lighter, fill_type="solid")
        cell = ws_out.cell(row=2, column=c, value=sub)
        cell.fill = fill
        cell.font = Font(bold=True, color="1A1A1A", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # Veri satırları (satır 3'ten itibaren)
    for row_idx in range(len(df)):
        for col_i, flat_col in enumerate(flat_cols):
            c = col_i + 1
            val = df.iloc[row_idx][flat_col]

            if flat_col in ("yil", "vade") and pd.notna(val):
                try:
                    val = int(val)
                except (ValueError, TypeError):
                    pass
            elif isinstance(val, float) and pd.isna(val):
                val = None

            cell = ws_out.cell(row=row_idx + 3, column=c, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # Sayısal format
            if flat_col in ("altin_miktar", "ihrac_adedi"):
                cell.number_format = '#,##0'
            elif flat_col == "kupon_orani":
                cell.number_format = '0.00'
            elif flat_col == "altin_ons_fiyat":
                cell.number_format = '#,##0.00'
            elif flat_col == "usdtry":
                cell.number_format = '0.0000'
            elif flat_col == "altin_ons":
                cell.number_format = '#,##0.00'
            elif flat_col in ("usd_karsilik", "try_karsilik"):
                cell.number_format = '#,##0'

    # Sütun genişlikleri
    widths = {
        "senet_turu": 30, "yil": 6, "isin": 18, "talep_tarihi": 14,
        "valor": 14, "itfa": 14, "vade": 10, "kupon_donemi": 14,
        "kupon_orani": 14, "altin_miktar": 22, "ihrac_adedi": 20,
        "altin_ons_fiyat": 18, "usdtry": 14, "altin_ons": 16,
        "usd_karsilik": 20, "try_karsilik": 22,
    }
    for i, col in enumerate(flat_cols):
        ws_out.column_dimensions[get_column_letter(i + 1)].width = widths.get(col, 16)

    ws_out.freeze_panes = "D3"
    ws_out.row_dimensions[1].height = 28
    ws_out.row_dimensions[2].height = 36


def create_ozet_sheet(wb, df):
    """İlk sayfa olarak yıl bazlı Özet raporu oluştur."""
    ws = wb.create_sheet("Özet", 0)
    ws.sheet_properties.tabColor = "C00000"

    thin = Side(style="thin", color="999999")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)

    TITLE_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HDR_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    SECTION_FILL = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
    LIGHT_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
    TOTAL_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    def cell(r, c, val, font=None, fill=None, align=None, nfmt=None):
        cl = ws.cell(row=r, column=c, value=val)
        cl.border = brd
        if font:
            cl.font = font
        if fill:
            cl.fill = fill
        if align:
            cl.alignment = align
        if nfmt:
            cl.number_format = nfmt
        return cl

    def merge_fill(r, c1, c2, val, font, fill, align=None, height=None):
        for c in range(c1, c2 + 1):
            cl = ws.cell(row=r, column=c)
            cl.fill = fill
            cl.font = font
            cl.border = brd
            if align:
                cl.alignment = align
        ws.cell(row=r, column=c1, value=val)
        if align:
            ws.cell(row=r, column=c1).alignment = align
        if c2 > c1:
            ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
        if height:
            ws.row_dimensions[r].height = height

    center = Alignment(horizontal="center", vertical="center")
    left_al = Alignment(horizontal="left", vertical="center")

    # Veri hazırlığı
    dn = df.copy()
    for c in ['altin_miktar', 'usd_karsilik', 'try_karsilik', 'altin_ons', 'altin_ons_fiyat', 'usdtry']:
        dn[c] = pd.to_numeric(dn[c], errors='coerce')
    dn['yil'] = pd.to_numeric(dn['yil'], errors='coerce')

    yillar = sorted(dn['yil'].dropna().unique().astype(int))

    # ── BAŞLIK ──
    merge_fill(1, 1, 9, "ALTIN TAHVİL & KİRA SERTİFİKASI İHALE ÖZETİ",
               Font(bold=True, color="FFFFFF", size=14), TITLE_FILL, center, 35)

    merge_fill(2, 1, 9, f"Güncelleme: {datetime.now().strftime('%d.%m.%Y')}",
               Font(italic=True, color="FFFFFF", size=10), TITLE_FILL, center, 20)

    # ── ÖZET CÜMLESİ ──
    dn_pre = df.copy()
    for _c in ['kupon_orani', 'vade', 'altin_miktar', 'usd_karsilik', 'try_karsilik']:
        dn_pre[_c] = pd.to_numeric(dn_pre[_c], errors='coerce')
    dn_pre['yil'] = pd.to_numeric(dn_pre['yil'], errors='coerce')
    dn_pre['ay'] = dn_pre['valor'].apply(
        lambda x: int(str(x).split('.')[1]) if pd.notna(x) and '.' in str(x) else 0)

    d4 = dn_pre[dn_pre['yil'] == 2024]
    d5 = dn_pre[dn_pre['yil'] == 2025]
    d6 = dn_pre[dn_pre['yil'] == 2026]

    ytd_max_ay = int(d6['ay'].max()) if len(d6) > 0 and d6['ay'].max() > 0 else 0
    d5_ayni = d5[d5['ay'] <= ytd_max_ay] if ytd_max_ay > 0 else d5.iloc[0:0]

    def _gstats(d):
        vd = d['vade'].dropna()
        gram = d['altin_miktar'].dropna().sum()
        tr = d['try_karsilik'].dropna().sum()
        mask = d['kupon_orani'].notna() & d['try_karsilik'].notna() & (d['try_karsilik'] > 0)
        if mask.any():
            w_kup = (d.loc[mask, 'kupon_orani'] * d.loc[mask, 'try_karsilik']).sum() / d.loc[mask, 'try_karsilik'].sum()
        else:
            kp = d['kupon_orani'].dropna()
            w_kup = kp.mean() if len(kp) else 0
        return {
            'n': len(d),
            'kupon': w_kup,
            'vade': vd.mean() if len(vd) else 0,
            'gram': gram,
            'try': tr,
        }

    g4, g5, g6, g5a = _gstats(d4), _gstats(d5), _gstats(d6), _gstats(d5_ayni)

    def _fmt_tl_milyar(v):
        if pd.isna(v) or v == 0:
            return "0 TL"
        if abs(v) >= 1e9:
            return f"{v / 1e9:,.2f} Milyar TL"
        if abs(v) >= 1e6:
            return f"{v / 1e6:,.1f} Milyon TL"
        return f"{v:,.0f} TL"

    def _fmt_ton(g):
        if pd.isna(g) or g == 0:
            return "0 kg"
        t = g / 1_000_000  # gram → ton
        if t >= 1:
            return f"{t:,.2f} ton"
        return f"{g / 1000:,.1f} kg"

    ozet_cumle = (
        f"Hazine altın tahvil ve kira sertifikası ihracında; "
        f"2024 yılının tamamında ortalama %{g4['kupon']:.2f} kupon oranıyla, "
        f"{int(g4['vade'])} gün vadeyle, toplamda {_fmt_ton(g4['gram'])} altına karşılık "
        f"{_fmt_tl_milyar(g4['try'])} ihraç ederken, "
        f"2025 yılının tamamında bu rakamlar sırasıyla %{g5['kupon']:.2f}, "
        f"{int(g5['vade'])} gün, {_fmt_ton(g5['gram'])} ve {_fmt_tl_milyar(g5['try'])} olarak gerçekleşirken; "
        f"2026 yılında şuana kadar %{g6['kupon']:.2f}, "
        f"{int(g6['vade'])} gün, {_fmt_ton(g6['gram'])} ve {_fmt_tl_milyar(g6['try'])}'dir. "
        f"(Geçen yıl aynı dönemde %{g5a['kupon']:.2f}, "
        f"{int(g5a['vade'])} gün, {_fmt_ton(g5a['gram'])} ve {_fmt_tl_milyar(g5a['try'])})"
    )

    OZET_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    no_brd = Border()
    for c in range(1, 10):
        cl = ws.cell(row=3, column=c)
        cl.fill = OZET_FILL
        cl.border = no_brd
        cl.font = Font(bold=False, color="1A1A1A", size=11)
        cl.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.cell(row=3, column=1, value=ozet_cumle)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=9)
    ws.row_dimensions[3].height = 75

    # ── BÖLÜM 1: YIL BAZLI GENEL ÖZET ──
    row = 5
    merge_fill(row, 1, 9, "YIL BAZLI GENEL ÖZET",
               Font(bold=True, color="FFFFFF", size=12), SECTION_FILL, center, 28)

    row = 6
    headers = ["Yıl", "İhale Sayısı", "Toplam Altın (Gram)", "Toplam Altın (Ons)",
               "Ort. Ons Fiyatı ($)", "Ort. USDTRY", "USD Karşılığı", "TRY Karşılığı",
               "Gram Başına TRY"]
    for ci, h in enumerate(headers):
        cell(row, ci + 1, h, Font(bold=True, color="FFFFFF", size=10), HDR_FILL, center)

    row = 7
    grand = {'n': 0, 'gram': 0, 'ons': 0, 'usd': 0, 'try': 0}
    for yil in yillar:
        dy = dn[dn['yil'] == yil]
        n = len(dy)
        gram = dy['altin_miktar'].sum()
        ons = dy['altin_ons'].sum()
        usd = dy['usd_karsilik'].sum()
        try_val = dy['try_karsilik'].sum()
        avg_ons = dy['altin_ons_fiyat'].mean()
        avg_fx = dy['usdtry'].mean()
        gram_try = try_val / gram if gram > 0 else 0

        fill = LIGHT_FILL if (yil % 2 == 0) else None
        cell(row, 1, yil, Font(bold=True, size=10), fill, center)
        cell(row, 2, n, None, fill, center)
        cell(row, 3, gram, None, fill, center, '#,##0')
        cell(row, 4, ons, None, fill, center, '#,##0')
        cell(row, 5, avg_ons, None, fill, center, '#,##0')
        cell(row, 6, avg_fx, None, fill, center, '0.00')
        cell(row, 7, usd, None, fill, center, '#,##0')
        cell(row, 8, try_val, None, fill, center, '#,##0')
        cell(row, 9, gram_try, None, fill, center, '0.00')

        grand['n'] += n
        grand['gram'] += gram
        grand['ons'] += ons
        grand['usd'] += usd
        grand['try'] += try_val
        row += 1

    # Toplam satırı
    cell(row, 1, "TOPLAM", Font(bold=True, size=11), TOTAL_FILL, center)
    cell(row, 2, grand['n'], Font(bold=True), TOTAL_FILL, center)
    cell(row, 3, grand['gram'], Font(bold=True), TOTAL_FILL, center, '#,##0')
    cell(row, 4, grand['ons'], Font(bold=True), TOTAL_FILL, center, '#,##0')
    cell(row, 5, None, Font(bold=True), TOTAL_FILL, center)
    cell(row, 6, None, Font(bold=True), TOTAL_FILL, center)
    cell(row, 7, grand['usd'], Font(bold=True), TOTAL_FILL, center, '#,##0')
    cell(row, 8, grand['try'], Font(bold=True), TOTAL_FILL, center, '#,##0')
    grand_gram_try = grand['try'] / grand['gram'] if grand['gram'] > 0 else 0
    cell(row, 9, grand_gram_try, Font(bold=True), TOTAL_FILL, center, '0.00')
    row += 2

    # ── BÖLÜM 2: SENET TÜRÜ BAZLI ÖZET ──
    for senet_turu in ["Altın Tahvili", "Altına Dayalı Kira Sertifikası"]:
        ds = dn[dn['senet_turu'] == senet_turu]
        if ds.empty:
            continue

        merge_fill(row, 1, 9, f"{senet_turu.upper()} - YIL BAZLI",
                   Font(bold=True, color="FFFFFF", size=12), SECTION_FILL, center, 28)
        row += 1

        for ci, h in enumerate(headers):
            cell(row, ci + 1, h, Font(bold=True, color="FFFFFF", size=10), HDR_FILL, center)
        row += 1

        sub_total = {'n': 0, 'gram': 0, 'ons': 0, 'usd': 0, 'try': 0}
        for yil in yillar:
            dy = ds[ds['yil'] == yil]
            if dy.empty:
                continue
            n = len(dy)
            gram = dy['altin_miktar'].sum()
            ons = dy['altin_ons'].sum()
            usd = dy['usd_karsilik'].sum()
            try_val = dy['try_karsilik'].sum()
            avg_ons = dy['altin_ons_fiyat'].mean()
            avg_fx = dy['usdtry'].mean()
            gram_try = try_val / gram if gram > 0 else 0

            fill = LIGHT_FILL if (yil % 2 == 0) else None
            cell(row, 1, yil, Font(bold=True, size=10), fill, center)
            cell(row, 2, n, None, fill, center)
            cell(row, 3, gram, None, fill, center, '#,##0')
            cell(row, 4, ons, None, fill, center, '#,##0')
            cell(row, 5, avg_ons, None, fill, center, '#,##0')
            cell(row, 6, avg_fx, None, fill, center, '0.00')
            cell(row, 7, usd, None, fill, center, '#,##0')
            cell(row, 8, try_val, None, fill, center, '#,##0')
            cell(row, 9, gram_try, None, fill, center, '0.00')

            sub_total['n'] += n
            sub_total['gram'] += gram
            sub_total['ons'] += ons
            sub_total['usd'] += usd
            sub_total['try'] += try_val
            row += 1

        cell(row, 1, "TOPLAM", Font(bold=True, size=11), TOTAL_FILL, center)
        cell(row, 2, sub_total['n'], Font(bold=True), TOTAL_FILL, center)
        cell(row, 3, sub_total['gram'], Font(bold=True), TOTAL_FILL, center, '#,##0')
        cell(row, 4, sub_total['ons'], Font(bold=True), TOTAL_FILL, center, '#,##0')
        cell(row, 5, None, Font(bold=True), TOTAL_FILL, center)
        cell(row, 6, None, Font(bold=True), TOTAL_FILL, center)
        cell(row, 7, sub_total['usd'], Font(bold=True), TOTAL_FILL, center, '#,##0')
        cell(row, 8, sub_total['try'], Font(bold=True), TOTAL_FILL, center, '#,##0')
        sub_gram_try = sub_total['try'] / sub_total['gram'] if sub_total['gram'] > 0 else 0
        cell(row, 9, sub_gram_try, Font(bold=True), TOTAL_FILL, center, '0.00')
        row += 2

    # Sütun genişlikleri
    col_widths = [10, 12, 22, 20, 18, 14, 22, 22, 16]
    for i, w in enumerate(col_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    ws.freeze_panes = "A5"


def main():
    print("Altın İhale Verileri Toplanıyor...")

    # 1) Parse
    print("\n1) Excel dosyaları parse ediliyor...")
    kira_rows = parse_altin_excel(ALTIN_KIRA_FILE, "Altına Dayalı Kira Sertifikası")
    print(f"   Kira Sertifikaları: {len(kira_rows)} satır")

    tahvil_rows = parse_altin_excel(ALTIN_TAHVIL_FILE, "Altın Tahvili")
    print(f"   Altın Tahvilleri: {len(tahvil_rows)} satır")

    # 2) Piyasa verileri
    print("\n2) Piyasa verileri çekiliyor...")
    gold_prices, usdtry_rates = fetch_market_data()

    # 3) Birleştir ve hesapla
    print("\n3) USD ve TRY karşılıkları hesaplanıyor...")
    all_rows = kira_rows + tahvil_rows

    missing_dates = []
    for row in all_rows:
        valor_dt = parse_date(row['valor'])
        gram = row.get('altin_miktar')

        if valor_dt and gram and isinstance(gram, (int, float)):
            gold_price = get_price_on_date(gold_prices, valor_dt.date())
            usdtry_rate = get_price_on_date(usdtry_rates, valor_dt.date())

            if gold_price and usdtry_rate:
                ons = gram / TROY_OUNCE_GRAM
                usd = ons * gold_price
                try_val = usd * usdtry_rate

                row['altin_ons_fiyat'] = round(gold_price, 2)
                row['usdtry'] = round(usdtry_rate, 4)
                row['altin_ons'] = round(ons, 2)
                row['usd_karsilik'] = round(usd, 0)
                row['try_karsilik'] = round(try_val, 0)
            else:
                missing_dates.append(row['valor'])
                for k in ('altin_ons_fiyat', 'usdtry', 'altin_ons', 'usd_karsilik', 'try_karsilik'):
                    row[k] = None
        else:
            for k in ('altin_ons_fiyat', 'usdtry', 'altin_ons', 'usd_karsilik', 'try_karsilik'):
                row[k] = None

    if missing_dates:
        print(f"   ⚠ {len(missing_dates)} tarih için piyasa verisi bulunamadı: {missing_dates}")

    df = pd.DataFrame(all_rows, columns=FLAT_COLS)

    # Valör tarihine göre sırala (en yeniden eskiye)
    df['_sort'] = df['valor'].apply(lambda d: parse_date(d) or datetime(1900, 1, 1))
    df = df.sort_values('_sort', ascending=False).drop(columns=['_sort']).reset_index(drop=True)

    print(f"\n   Toplam: {len(df)} ihale kaydı")

    total_usd = pd.to_numeric(df['usd_karsilik'], errors='coerce').dropna().sum()
    total_try = pd.to_numeric(df['try_karsilik'], errors='coerce').dropna().sum()
    print(f"   Toplam USD karşılığı: ${total_usd:,.0f}")
    print(f"   Toplam TRY karşılığı: ₺{total_try:,.0f}")

    # 4) Excel yaz
    print(f"\n4) Excel dosyası oluşturuluyor...")
    wb = openpyxl.Workbook()

    ws_data = wb.active
    ws_data.title = "Altın İhale Verileri"
    ws_data.sheet_properties.tabColor = "BF8F00"
    format_sheet(ws_data, df)

    df_kira = df[df['senet_turu'] == "Altına Dayalı Kira Sertifikası"].reset_index(drop=True)
    df_tahvil = df[df['senet_turu'] == "Altın Tahvili"].reset_index(drop=True)

    ws_kira = wb.create_sheet("Kira Sertifikaları")
    ws_kira.sheet_properties.tabColor = "548235"
    format_sheet(ws_kira, df_kira)

    ws_tahvil = wb.create_sheet("Altın Tahvilleri")
    ws_tahvil.sheet_properties.tabColor = "C55A11"
    format_sheet(ws_tahvil, df_tahvil)

    # Özet sayfası (en başa)
    create_ozet_sheet(wb, df)

    wb.save(OUTPUT_FILE)
    print(f"\n✓ Kayıt tamamlandı: {OUTPUT_FILE}")
    print(f"  - Özet: Yıl bazlı genel + senet türü özeti")
    print(f"  - Altın İhale Verileri: {len(df)} satır (tümü)")
    print(f"  - Kira Sertifikaları: {len(df_kira)} satır")
    print(f"  - Altın Tahvilleri: {len(df_tahvil)} satır")


if __name__ == "__main__":
    main()
