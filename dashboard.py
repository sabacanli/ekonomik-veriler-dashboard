"""
Ekonomik Veriler Dashboard
==========================
Türkiye ekonomik verilerini tek panelden takip etmek için
Streamlit tabanlı interaktif dashboard.

Modüller:
  1. BDDK Haftalık Bankacılık Verileri
  2. Hazine İhale Verileri
  3. TCMB Doğrudan Alım İşlemleri
  4. TCMB Haftalık Menkul Kıymet Stokları
"""

import streamlit as st
import pandas as pd
import os
import re
import subprocess
import sys
from pathlib import Path
from datetime import datetime
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import io
import base64
from PIL import Image
import json
import openpyxl
import glob as glob_module

# ── Sayfa Ayarları ─────────────────────────────────────────
st.set_page_config(
    page_title="Ekonomik Veriler Dashboard",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# API anahtarını Streamlit secrets'tan al (varsa) ve alt-süreç fetch scriptlerine
# ortam değişkeni olarak aktar. Secret yoksa scriptler kendi varsayılan anahtarını kullanır.
try:
    _api_key = st.secrets["EVDS_API_KEY"]
    os.environ["EVDS_API_KEY"] = _api_key
    os.environ["TCMB_API_KEY"] = _api_key
except Exception:
    pass

BASE_DIR = Path(__file__).parent


def isin_to_itfa(code):
    """DİBS ISIN kodundan itfa tarihini çıkarır.

    Format: TR + 1 harf + GGAAYY + sonek. Örn: TRT051033T12 -> 05.10.2033.
    Kaynak veride İtfa Tarihi yalnızca 2014 sonrası dolu olduğundan, eksik
    tarihler bu fonksiyonla tamamlanır. İtfa tarihi bilinen 832 kayıtla
    %100 doğrulandı (sıfır hata).
    """
    if not isinstance(code, str):
        return pd.NaT
    m = re.match(r"^TR[A-Z](\d{2})(\d{2})(\d{2})", code.strip())
    if not m:
        return pd.NaT
    dd, mm, yy = int(m.group(1)), int(m.group(2)), int(m.group(3))
    year = 1900 + yy if yy >= 90 else 2000 + yy
    try:
        return pd.Timestamp(year=year, month=mm, day=dd)
    except Exception:
        return pd.NaT


# ── Plotly Tema ───────────────────────────────────────────
# Bloomberg terminal esintili renk paleti (amber öncelikli)
CHART_COLORS = ['#FF9E1B', '#4C9AFF', '#26C281', '#FF5A5F', '#B98AFF',
                '#37C7D4', '#F45D9C', '#8FD14F', '#FFC93C', '#7A8CFF']
CHART_LAYOUT = dict(
    font=dict(family="'IBM Plex Sans', -apple-system, sans-serif", size=12, color="#AEB9C7"),
    plot_bgcolor="rgba(0,0,0,0)",
    paper_bgcolor="rgba(0,0,0,0)",
    margin=dict(l=40, r=20, t=40, b=40),
    legend=dict(
        bgcolor="rgba(21,26,35,0.7)",
        bordercolor="#2A3240",
        borderwidth=1,
        font=dict(size=11, color="#AEB9C7"),
    ),
    xaxis=dict(gridcolor="rgba(255,255,255,0.06)", linecolor="#2A3240", zerolinecolor="#2A3240"),
    yaxis=dict(gridcolor="rgba(255,255,255,0.06)", linecolor="#2A3240", zerolinecolor="#2A3240"),
)

def apply_chart_style(fig):
    """Plotly figürüne profesyonel stil uygular."""
    fig.update_layout(**CHART_LAYOUT)
    fig.update_layout(colorway=CHART_COLORS)
    # Title yoksa boş string ata (undefined görünmesin)
    current_title = fig.layout.title.text if fig.layout.title else None
    if current_title:
        fig.update_layout(title_font=dict(size=15, color="#E6EDF3", family="'IBM Plex Sans', sans-serif"))
    else:
        fig.update_layout(title_text="")
    fig.update_xaxes(showgrid=True, gridwidth=1)
    fig.update_yaxes(showgrid=True, gridwidth=1)
    return fig


def styled_chart(fig, **kwargs):
    """Plotly grafiğini profesyonel stilde, başlıksız gösterir."""
    apply_chart_style(fig)
    # Plotly'nin "undefined" title bug'ı için config
    config = kwargs.pop("config", {"displaylogo": False})
    st.plotly_chart(fig, use_container_width=True, config=config, **kwargs)

# ── Özel CSS ───────────────────────────────────────────────
st.markdown("""
<style>
    /* ── Google Fonts ── */
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');

    /* ── Global ── */
    html, body, [class*="css"] {
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
    }
    .block-container {
        padding-top: 2rem !important;
        padding-bottom: 2rem !important;
        max-width: 1200px;
    }
    /* Hide streamlit branding */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header[data-testid="stHeader"] {
        background: linear-gradient(135deg, #0f172a 0%, #1e3a5f 100%);
        height: 4px;
    }

    /* ── Sidebar ── */
    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #0f172a 0%, #1a2744 40%, #1e3a5f 100%);
        border-right: none;
    }
    [data-testid="stSidebar"] * {
        color: #cbd5e1 !important;
    }
    [data-testid="stSidebar"] .sidebar-brand {
        color: #f8fafc !important;
    }
    [data-testid="stSidebar"] [data-testid="stRadio"] label {
        background: rgba(255,255,255,0.04);
        border-radius: 10px;
        padding: 0.6rem 1rem !important;
        margin-bottom: 4px;
        transition: all 0.2s ease;
        border: 1px solid transparent;
        font-size: 0.9rem !important;
        font-weight: 500 !important;
    }
    [data-testid="stSidebar"] [data-testid="stRadio"] label:hover {
        background: rgba(255,255,255,0.1);
        border-color: rgba(99,179,237,0.3);
    }
    [data-testid="stSidebar"] [data-testid="stRadio"] label[data-checked="true"],
    [data-testid="stSidebar"] [data-testid="stRadio"] label:has(input:checked) {
        background: linear-gradient(135deg, rgba(59,130,246,0.25), rgba(99,179,237,0.15)) !important;
        border-color: rgba(99,179,237,0.5) !important;
        color: #f0f9ff !important;
        font-weight: 600 !important;
    }
    [data-testid="stSidebar"] [data-testid="stRadio"] div[role="radiogroup"] > label > div:first-child {
        display: none;
    }
    .sidebar-brand {
        font-size: 1.5rem;
        font-weight: 800;
        letter-spacing: -0.5px;
        padding: 0.5rem 0 0.3rem;
        background: linear-gradient(135deg, #60a5fa, #38bdf8);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
    }
    .sidebar-version {
        font-size: 0.7rem;
        color: #64748b !important;
        padding: 1rem 0 0;
        border-top: 1px solid rgba(255,255,255,0.06);
        text-align: center;
    }

    /* ── Ana Başlık ── */
    .main-header {
        font-size: 1.75rem;
        font-weight: 800;
        color: #0f172a;
        padding: 0 0 0.75rem;
        margin-bottom: 1.5rem;
        border-bottom: 2px solid #e2e8f0;
        letter-spacing: -0.5px;
        display: flex;
        align-items: center;
        gap: 0.5rem;
    }
    .main-header-accent {
        font-size: 1.75rem;
        font-weight: 800;
        background: linear-gradient(135deg, #1e3a5f, #3b82f6);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        padding: 0 0 0.75rem;
        margin-bottom: 1.5rem;
        border-bottom: 2px solid #e2e8f0;
        letter-spacing: -0.5px;
    }

    /* ── Stat Kartları (Ana Sayfa) ── */
    .stat-card {
        background: white;
        border-radius: 16px;
        padding: 1.5rem 1rem;
        text-align: center;
        box-shadow: 0 1px 3px rgba(0,0,0,0.04), 0 4px 12px rgba(0,0,0,0.06);
        border: 1px solid #f1f5f9;
        transition: all 0.25s ease;
        position: relative;
        overflow: hidden;
    }
    .stat-card::before {
        content: '';
        position: absolute;
        top: 0; left: 0; right: 0;
        height: 3px;
        background: linear-gradient(90deg, #3b82f6, #38bdf8);
    }
    .stat-card:hover {
        transform: translateY(-3px);
        box-shadow: 0 4px 6px rgba(0,0,0,0.05), 0 10px 24px rgba(0,0,0,0.1);
    }
    .stat-value {
        font-size: 2.2rem;
        font-weight: 800;
        color: #0f172a;
        line-height: 1;
    }
    .stat-label {
        font-size: 0.78rem;
        color: #94a3b8;
        margin-top: 0.5rem;
        font-weight: 500;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }

    /* ── Modül Kartları ── */
    .module-card {
        background: white;
        border-radius: 16px;
        padding: 1.5rem;
        border: 1px solid #e2e8f0;
        margin-bottom: 0.75rem;
        transition: all 0.25s ease;
        cursor: pointer;
        position: relative;
        overflow: hidden;
    }
    .module-card::after {
        content: '';
        position: absolute;
        left: 0; top: 0; bottom: 0;
        width: 4px;
        background: linear-gradient(180deg, #3b82f6, #1e3a5f);
        border-radius: 0 4px 4px 0;
        opacity: 0;
        transition: opacity 0.2s;
    }
    .module-card:hover {
        border-color: #bfdbfe;
        box-shadow: 0 4px 12px rgba(59,130,246,0.1);
        transform: translateX(4px);
    }
    .module-card:hover::after {
        opacity: 1;
    }
    .module-title {
        font-size: 1rem;
        font-weight: 700;
        color: #0f172a;
        margin-bottom: 0.4rem;
    }
    .module-desc {
        font-size: 0.82rem;
        color: #64748b;
        line-height: 1.5;
    }

    /* ── Update Info ── */
    .update-info {
        background: linear-gradient(135deg, #f0fdf4, #dcfce7);
        border-radius: 12px;
        padding: 0.75rem 1rem;
        font-size: 0.82rem;
        color: #166534;
        margin-bottom: 1rem;
        border: 1px solid #bbf7d0;
        font-weight: 500;
    }

    /* ── Tabs ── */
    .stTabs [data-baseweb="tab-list"] {
        gap: 4px;
        background: #f8fafc;
        border-radius: 12px;
        padding: 4px;
        border: 1px solid #e2e8f0;
    }
    .stTabs [data-baseweb="tab"] {
        border-radius: 8px;
        padding: 8px 20px;
        font-weight: 500;
        font-size: 0.85rem;
    }
    .stTabs [aria-selected="true"] {
        background: white !important;
        box-shadow: 0 1px 3px rgba(0,0,0,0.08);
        font-weight: 600 !important;
    }

    /* ── Buttons ── */
    .stButton > button {
        border-radius: 10px;
        font-weight: 600;
        font-size: 0.85rem;
        padding: 0.5rem 1.25rem;
        border: 1px solid #e2e8f0;
        transition: all 0.2s ease;
    }
    .stButton > button:hover {
        border-color: #3b82f6;
        color: #3b82f6;
        box-shadow: 0 2px 8px rgba(59,130,246,0.15);
    }

    /* ── Expanders ── */
    [data-testid="stExpander"] {
        border: 1px solid #e2e8f0 !important;
        border-radius: 12px !important;
        overflow: hidden;
    }
    [data-testid="stExpander"] summary {
        font-weight: 600;
        font-size: 0.9rem;
    }

    /* ── Metrics ── */
    [data-testid="stMetric"] {
        background: white;
        border-radius: 12px;
        padding: 1rem;
        border: 1px solid #f1f5f9;
        box-shadow: 0 1px 3px rgba(0,0,0,0.04);
    }
    [data-testid="stMetric"] label {
        font-size: 0.75rem !important;
        text-transform: uppercase;
        letter-spacing: 0.5px;
        color: #94a3b8 !important;
        font-weight: 600 !important;
    }
    [data-testid="stMetric"] [data-testid="stMetricValue"] {
        font-size: 1.5rem !important;
        font-weight: 800 !important;
        color: #0f172a !important;
    }

    /* ── DataFrames ── */
    [data-testid="stDataFrame"] {
        border-radius: 12px;
        overflow: hidden;
        border: 1px solid #e2e8f0;
    }

    /* ── Info/Warning/Error boxes ── */
    [data-testid="stAlert"] {
        border-radius: 12px;
        font-size: 0.85rem;
    }

    /* ── Section Divider ── */
    .section-divider {
        height: 1px;
        background: linear-gradient(90deg, transparent, #e2e8f0, transparent);
        margin: 2rem 0;
    }

    /* ── KPI Row ── */
    .kpi-container {
        display: flex;
        gap: 0.75rem;
        margin-bottom: 1.5rem;
    }
    .kpi-card {
        flex: 1;
        background: white;
        border-radius: 14px;
        padding: 1rem 1.2rem;
        border: 1px solid #f1f5f9;
        box-shadow: 0 1px 3px rgba(0,0,0,0.03);
    }
    .kpi-label {
        font-size: 0.7rem;
        font-weight: 600;
        color: #94a3b8;
        text-transform: uppercase;
        letter-spacing: 0.5px;
        margin-bottom: 0.25rem;
    }
    .kpi-value {
        font-size: 1.3rem;
        font-weight: 800;
        color: #0f172a;
    }
    .kpi-value.positive { color: #16a34a; }
    .kpi-value.negative { color: #dc2626; }
    .kpi-sub {
        font-size: 0.72rem;
        color: #94a3b8;
        margin-top: 0.15rem;
    }

    /* ── Welcome Hero ── */
    .hero-container {
        background: linear-gradient(135deg, #0f172a 0%, #1e3a5f 50%, #1e40af 100%);
        border-radius: 20px;
        padding: 2.5rem 2rem;
        margin-bottom: 2rem;
        color: white;
        position: relative;
        overflow: hidden;
    }
    .hero-container::before {
        content: '';
        position: absolute;
        top: -50%; right: -20%;
        width: 400px; height: 400px;
        background: radial-gradient(circle, rgba(59,130,246,0.15) 0%, transparent 70%);
        border-radius: 50%;
    }
    .hero-container::after {
        content: '';
        position: absolute;
        bottom: -30%; left: 10%;
        width: 300px; height: 300px;
        background: radial-gradient(circle, rgba(56,189,248,0.1) 0%, transparent 70%);
        border-radius: 50%;
    }
    .hero-title {
        font-size: 2rem;
        font-weight: 800;
        margin-bottom: 0.5rem;
        letter-spacing: -0.5px;
        position: relative;
        z-index: 1;
    }
    .hero-subtitle {
        font-size: 0.95rem;
        color: #94a3b8;
        max-width: 600px;
        line-height: 1.6;
        position: relative;
        z-index: 1;
    }

    /* ── Download Button ── */
    [data-testid="stDownloadButton"] > button {
        border-radius: 10px;
        font-weight: 600;
        border: 1px solid #e2e8f0;
    }
</style>
""", unsafe_allow_html=True)


# ── Bloomberg Terminal Teması (koyu) — üstteki açık temayı ezer ──
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@300;400;500;600;700&family=IBM+Plex+Mono:wght@400;500;600;700&display=swap');

:root{
  --bg:#0B0E14; --panel:#151A23; --panel2:#1B212C; --border:#232A36;
  --amber:#FF9E1B; --blue:#4C9AFF; --green:#26C281; --red:#FF5A5F;
  --text:#E6EDF3; --muted:#8B98A9;
}

/* Yazı tipleri */
html, body, [class*="css"], .stApp, .stMarkdown, p, span, div, label, input, button, select, textarea {
  font-family:'IBM Plex Sans', -apple-system, BlinkMacSystemFont, sans-serif;
}
/* Sayısal/veri alanları monospace (terminal hissi) */
[data-testid="stMetricValue"], [data-testid="stMetricDelta"], .stat-value, .kpi-value,
[data-testid="stDataFrame"], code, .update-info { font-family:'IBM Plex Mono', monospace !important; }

/* Uygulama arka planı */
.stApp, [data-testid="stAppViewContainer"], section.main, .main { background:var(--bg) !important; }
[data-testid="stHeader"]{ background:transparent !important; }
.block-container{ max-width:1280px; }
h1,h2,h3,h4,h5,h6{ color:var(--text) !important; }
.stMarkdown, .stMarkdown p, .stMarkdown li{ color:#C9D2DD !important; }

/* Kenar çubuğu */
[data-testid="stSidebar"]{ background:#0E131C !important; border-right:1px solid var(--border) !important; }
[data-testid="stSidebar"] *{ color:#AEB9C7 !important; }
.sidebar-brand{ background:none !important; -webkit-text-fill-color:var(--amber) !important;
  color:var(--amber) !important; font-family:'IBM Plex Mono',monospace !important; letter-spacing:1px; }
[data-testid="stSidebar"] [data-testid="stRadio"] label{ background:rgba(255,255,255,0.03) !important; border:1px solid transparent !important; }
[data-testid="stSidebar"] [data-testid="stRadio"] label:hover{ background:rgba(255,158,27,0.08) !important; border-color:rgba(255,158,27,0.25) !important; }
[data-testid="stSidebar"] [data-testid="stRadio"] label:has(input:checked){
  background:rgba(255,158,27,0.14) !important; border-color:rgba(255,158,27,0.5) !important;
  box-shadow:inset 3px 0 0 var(--amber); }
[data-testid="stSidebar"] [data-testid="stRadio"] label:has(input:checked) *{ color:#FFE7C2 !important; }

/* Ana başlık — amber vurgu çubuğu */
.main-header{ color:var(--text) !important; border-bottom:1px solid var(--border) !important;
  padding-left:.7rem; box-shadow:inset 4px 0 0 var(--amber); }
.main-header-accent{ -webkit-text-fill-color:var(--amber) !important; background:none !important;
  color:var(--amber) !important; border-bottom:1px solid var(--border) !important; }

/* Paneller / kartlar */
.stat-card,.module-card,[data-testid="stMetric"],.kpi-card{
  background:var(--panel) !important; border:1px solid var(--border) !important; box-shadow:none !important; }
.stat-value,.kpi-value,[data-testid="stMetric"] [data-testid="stMetricValue"]{ color:var(--text) !important; }
.stat-label,.kpi-label,[data-testid="stMetric"] label{ color:var(--muted) !important; }
.stat-card::before{ background:var(--amber) !important; }
.module-title{ color:var(--text) !important; }
.module-desc{ color:var(--muted) !important; }
[data-testid="stMetric"]{ border-left:3px solid var(--amber) !important; }

/* Güncelleme bilgisi */
.update-info{ background:rgba(38,194,129,0.08) !important; border:1px solid rgba(38,194,129,0.22) !important; color:#7EE2B8 !important; }

/* Sekmeler */
.stTabs [data-baseweb="tab-list"]{ background:var(--panel) !important; border:1px solid var(--border) !important; }
.stTabs [data-baseweb="tab"]{ color:var(--muted) !important; }
.stTabs [aria-selected="true"]{ background:var(--panel2) !important; color:var(--amber) !important; box-shadow:none !important; }

/* Butonlar */
.stButton>button,[data-testid="stDownloadButton"]>button{
  background:var(--panel2) !important; color:var(--text) !important; border:1px solid var(--border) !important; }
.stButton>button:hover,[data-testid="stDownloadButton"]>button:hover{
  border-color:var(--amber) !important; color:var(--amber) !important; box-shadow:0 0 0 1px rgba(255,158,27,0.3) !important; }

/* Açılır paneller */
[data-testid="stExpander"]{ background:var(--panel) !important; border:1px solid var(--border) !important; }
[data-testid="stExpander"] summary,[data-testid="stExpander"] summary *{ color:var(--text) !important; }

/* Tablolar */
[data-testid="stDataFrame"]{ border:1px solid var(--border) !important; }

/* Uyarı/bilgi kutuları → koyu panel + mavi vurgu */
[data-testid="stAlert"]{ background:var(--panel) !important; border:1px solid var(--border) !important;
  border-left:3px solid var(--blue) !important; }
[data-testid="stAlert"] *{ color:#D6DEE8 !important; }

/* Girişler (selectbox, slider, multiselect) */
[data-baseweb="select"]>div, [data-baseweb="input"]>div{ background:var(--panel2) !important; border-color:var(--border) !important; }

/* Görseller → beyaz kart (açık zeminli PNG grafikler koyu temada düzgün dursun) */
[data-testid="stImage"] img{ background:#fff !important; padding:12px !important; border-radius:10px !important; }

/* Hero (ana sayfa) */
.hero-container{ background:linear-gradient(135deg,#0E131C 0%, #141B27 60%, #1a2536 100%) !important;
  border:1px solid var(--border) !important; }
.hero-title{ color:var(--text) !important; }
.hero-subtitle{ color:var(--muted) !important; }

/* Bölüm ayıracı */
.section-divider{ background:linear-gradient(90deg,transparent,var(--border),transparent) !important; }
</style>
""", unsafe_allow_html=True)


# ── Yardımcı Fonksiyonlar ─────────────────────────────────

def get_file_mod_time(filepath):
    """Dosya son değiştirilme zamanını döndürür."""
    if os.path.exists(filepath):
        ts = os.path.getmtime(filepath)
        return datetime.fromtimestamp(ts).strftime("%d.%m.%Y %H:%M")
    return "Bilinmiyor"


def get_download_button(filepath, label=None):
    """Dosya indirme butonu oluşturur."""
    if not os.path.exists(filepath):
        return
    fname = os.path.basename(filepath)
    if label is None:
        label = f"📥 {fname}"
    with open(filepath, "rb") as f:
        data = f.read()

    ext = Path(filepath).suffix.lower()
    mime_map = {
        ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ".xls": "application/vnd.ms-excel",
        ".csv": "text/csv",
        ".png": "image/png",
        ".pdf": "application/pdf",
        ".json": "application/json",
    }
    mime = mime_map.get(ext, "application/octet-stream")
    st.download_button(label=label, data=data, file_name=fname, mime=mime)


def show_image_with_download(img_path, caption=""):
    """Resmi gösterip indirme butonu ekler."""
    if not os.path.exists(img_path):
        st.warning(f"Dosya bulunamadı: {os.path.basename(img_path)}")
        return
    img = Image.open(img_path)
    st.image(img, caption=caption, use_container_width=True)
    get_download_button(img_path, f"📥 İndir: {os.path.basename(img_path)}")


def run_script(script_path, cwd=None, timeout_sec=300):
    """Python scriptini çalıştırır ve sonucu session_state'e kaydeder."""
    # Path'i absolute yap (relative path sorununu önle)
    script_path = os.path.abspath(script_path)
    if cwd is None:
        cwd = os.path.dirname(script_path)
    script_name = os.path.basename(script_path)
    python_exe = sys.executable
    with st.spinner(f"⏳ Çalıştırılıyor: {script_name}... (Bu işlem birkaç dakika sürebilir)"):
        try:
            result = subprocess.run(
                [python_exe, script_path],
                capture_output=True, text=True, cwd=cwd, timeout=timeout_sec
            )
            if result.returncode == 0:
                st.session_state["last_script_result"] = ("success", script_name, result.stdout)
                st.success(f"✅ {script_name} başarıyla tamamlandı!")
                with st.expander("📋 Çıktı Detayları", expanded=False):
                    if result.stdout:
                        st.code(result.stdout[-3000:], language="text")
                    if result.stderr:
                        st.code(result.stderr[-3000:], language="text")
                return True
            else:
                st.session_state["last_script_result"] = ("error", script_name, result.stderr)
                st.error(f"❌ {script_name} hata ile sonlandı! (return code: {result.returncode})")
                with st.expander("🔍 Hata Detayları", expanded=True):
                    if result.stderr:
                        st.code(result.stderr[-3000:], language="text")
                    if result.stdout:
                        st.text("STDOUT:")
                        st.code(result.stdout[-2000:], language="text")
                return False
        except subprocess.TimeoutExpired:
            st.error(f"⏰ {script_name} zaman aşımına uğradı ({timeout_sec//60} dakika).")
            return False
        except Exception as e:
            st.error(f"💥 Çalıştırma hatası: {e}")
            return False


def read_excel_preview(filepath, sheet_name=0, nrows=50):
    """Excel dosyasından önizleme okur."""
    try:
        df = pd.read_excel(filepath, sheet_name=sheet_name, nrows=nrows)
        return df
    except Exception:
        return None


# ── Sidebar ────────────────────────────────────────────────

with st.sidebar:
    st.markdown('<div class="sidebar-brand">Ekonomik Veriler</div>', unsafe_allow_html=True)
    st.markdown(
        '<p style="font-size:0.78rem; color:#64748b !important; margin:-0.5rem 0 1.5rem; font-weight:400;">'
        'Piyasa analiz platformu</p>',
        unsafe_allow_html=True,
    )

    modules = {
        "ana_sayfa": "Ana Sayfa",
        "tcmb_stok": "TCMB Haftalık Stok",
        "hazine": "Hazine İhale Verileri",
        "tcmb_alim": "TCMB Doğrudan Alım",
        "bddk": "BDDK Bankacılık Verileri",
        "cari_acik": "Cari Açık (EVDS)",
        "net_rezerv": "TCMB Net Rezerv",
        "enflasyon": "TÜFE Enflasyon",
        "kredi": "Kredi Faizleri",
        "mevduat": "Mevduat Faizleri",
        "butce": "Bütçe Dengesi",
        "nakit": "Hazine Nakit Gerçekleşmeleri",
        "dth": "Yabancı Para Hareketi",
    }

    selected = st.radio(
        "Modül Seçin",
        list(modules.keys()),
        format_func=lambda x: modules[x],
        label_visibility="collapsed",
    )

    st.markdown(
        f'<div class="sidebar-version">'
        f'{datetime.now().strftime("%d.%m.%Y %H:%M")}<br>'
        f'<span style="color:#475569 !important;">v2.0</span></div>',
        unsafe_allow_html=True,
    )


# ── Kredi/Mevduat ortak yardımcıları ──
def _km_pct(v):
    return "—" if pd.isna(v) else f"{v:.2f}".replace(".", ",")


@st.cache_data
def _load_km_sheet(path, sheet):
    d = pd.read_excel(path, sheet_name=sheet)
    d["tarih"] = pd.to_datetime(d["tarih"], errors="coerce")
    return d.dropna(subset=["tarih"]).sort_values("tarih").reset_index(drop=True)


def _km_update_bar(fetch_script, data_file, key):
    col_u1, col_u2 = st.columns([1, 4])
    with col_u1:
        if st.button("🔄 Verileri Güncelle", key=key, use_container_width=True):
            if run_script(str(fetch_script), timeout_sec=180):
                st.cache_data.clear()
    with col_u2:
        if data_file.exists():
            st.markdown(
                f'<div class="update-info">Son güncelleme: {get_file_mod_time(data_file)}</div>',
                unsafe_allow_html=True,
            )
        else:
            st.warning("Henüz veri çekilmemiş. Yandaki butona tıklayın.")


def _try_publish(rel_paths, message):
    """Verilen dosyaları commit + push etmeyi dener.
    Yerelde (git kimliği + keychain) çalışır → bulut güncellenir.
    Bulutta (push yetkisi yok) sessizce başarısız olur → sadece oturum içi yenileme.
    Statik sitenin veri paketleri (site/data) de aynı commit'te tazelenir."""
    env = {**os.environ, "GIT_TERMINAL_PROMPT": "0"}
    try:
        try:
            subprocess.run([sys.executable, str(BASE_DIR / "site_export.py")],
                           cwd=str(BASE_DIR), capture_output=True, timeout=120)
            rel_paths = [*rel_paths, "site/data"]
        except Exception:
            pass
        subprocess.run(["git", "add", *rel_paths], cwd=str(BASE_DIR),
                       capture_output=True, timeout=30, env=env)
        subprocess.run(["git", "commit", "-m", message], cwd=str(BASE_DIR),
                       capture_output=True, timeout=30, env=env)
        # Bulut (GitHub Actions) bu arada commit atmış olabilir — önce senkronla.
        subprocess.run(["git", "pull", "--rebase", "origin", "main"], cwd=str(BASE_DIR),
                       capture_output=True, timeout=60, env=env)
        p = subprocess.run(["git", "push", "origin", "main"], cwd=str(BASE_DIR),
                           capture_output=True, timeout=90, env=env)
        return p.returncode == 0
    except Exception:
        return False


# ══════════════════════════════════════════════════════════
# ANA SAYFA
# ══════════════════════════════════════════════════════════

if selected == "ana_sayfa":
    # Hero section
    st.markdown("""<div class="hero-container">
        <div class="hero-title">Ekonomik Veriler Dashboard</div>
        <div class="hero-subtitle">Türkiye ekonomik verilerini tek panelden takip edin.
        TCMB, Hazine ve BDDK verilerini analiz edin, raporlar oluşturun.</div>
    </div>""", unsafe_allow_html=True)

    # ── Modül bazlı güncel analiz kartları (canlı, veriden üretilir) ──
    _HOME_AY = {1: "Ocak", 2: "Şubat", 3: "Mart", 4: "Nisan", 5: "Mayıs", 6: "Haziran",
                7: "Temmuz", 8: "Ağustos", 9: "Eylül", 10: "Ekim", 11: "Kasım", 12: "Aralık"}

    def _ht(v, d=1, sign=False):
        if pd.isna(v):
            return "—"
        fmt = f"{{:+,.{d}f}}" if sign else f"{{:,.{d}f}}"
        return fmt.format(v).replace(",", "\x00").replace(".", ",").replace("\x00", ".")

    @st.cache_data
    def home_summaries(_bust):
        out = []
        # ── 1) TÜFE Enflasyon ──
        try:
            g = pd.read_excel(BASE_DIR / "enflasyon" / "enflasyon.xlsx", sheet_name="Genel")
            g["tarih"] = pd.to_datetime(g["tarih"]); L = g.iloc[-1]
            base = (f"{_HOME_AY[L['tarih'].month]} {L['tarih'].year} itibarıyla yıllık enflasyon "
                    f"<b>%{_ht(L['yillik'], 2)}</b>, aylık <b>%{_ht(L['aylik'], 2)}</b>; "
                    f"yılbaşından beri %{_ht(L['ytd'], 2)}.")
            analiz = ""
            try:
                d_yil = float(L["yillik"] - g.iloc[-2]["yillik"])
                yon = "geriledi" if d_yil < 0 else "yükseldi"
                trend = "dezenflasyon sürüyor" if d_yil < 0 else "enflasyon yeniden hızlandı"
                avg3 = float(g["aylik"].tail(3).mean())
                analiz = (f" Yıllık oran önceki aya göre <b>{_ht(abs(d_yil), 2)} puan {yon}</b> — {trend}; "
                          f"son 3 ayın ortalama aylık artışı %{_ht(avg3, 2)}.")
            except Exception:
                pass
            out.append(("📈 TÜFE Enflasyon", base + analiz))
        except Exception:
            pass
        # ── 2) Bütçe Dengesi ──
        try:
            b = pd.read_excel(BASE_DIR / "butce" / "butce.xlsx", sheet_name="Aylik")
            b["tarih"] = pd.to_datetime(b["tarih"]); b = b.sort_values("tarih"); L = b.iloc[-1]
            cy, cm = int(L["yil"]), int(L["ay"])
            ytd = b[(b.yil == cy) & (b.ay <= cm)]["denge"].sum()
            yon = "açık" if L["denge"] < 0 else "fazla"
            yon_ytd = "açık" if ytd < 0 else "fazla"
            base = (f"{_HOME_AY[cm]} {cy} ayında merkezi yönetim bütçesi <b>{_ht(abs(L['denge'])/1000)} milyar TL {yon}</b> verdi; "
                    f"yılbaşından beri kümülatif {yon_ytd} <b>{_ht(abs(ytd)/1000)} milyar TL</b>.")
            analiz = ""
            try:
                ytd_prev = b[(b.yil == cy - 1) & (b.ay <= cm)]["denge"].sum()
                if ytd_prev != 0:
                    pct = (abs(ytd) - abs(ytd_prev)) / abs(ytd_prev) * 100
                    yon2 = "genişledi" if abs(ytd) > abs(ytd_prev) else "daraldı"
                    analiz = (f" Açık geçen yılın aynı dönemine ({_ht(abs(ytd_prev)/1000)} milyar TL) göre "
                              f"<b>%{_ht(abs(pct), 0)} {yon2}</b>.")
            except Exception:
                pass
            out.append(("🏛️ Bütçe Dengesi", base + analiz))
        except Exception:
            pass
        # ── 3) Hazine Nakit Gerçekleşmeleri ──
        try:
            nk = pd.read_excel(BASE_DIR / "hazine nakit" / "nakit.xlsx", sheet_name="Aylik")
            nk["tarih"] = pd.to_datetime(nk["tarih"]); nk = nk.sort_values("tarih"); L = nk.iloc[-1]
            cy, cm = int(L["yil"]), int(L["ay"])
            ytd = nk[(nk.yil == cy) & (nk.ay <= cm)]["nakit_denge"].sum()
            yon = "açık" if L["nakit_denge"] < 0 else "fazla"
            yon_ytd = "açık" if ytd < 0 else "fazla"
            base = (f"{_HOME_AY[cm]} {cy} ayında Hazine nakit dengesi <b>{_ht(abs(L['nakit_denge'])/1000)} milyar TL {yon}</b> verdi; "
                    f"yılbaşından beri kümülatif {yon_ytd} <b>{_ht(abs(ytd)/1000)} milyar TL</b>.")
            analiz = ""
            try:
                ytd_prev = nk[(nk.yil == cy - 1) & (nk.ay <= cm)]["nakit_denge"].sum()
                ib = nk[(nk.yil == cy) & (nk.ay <= cm)]["ic_borclanma_net"].sum()
                if ytd_prev != 0:
                    pct = (abs(ytd) - abs(ytd_prev)) / abs(ytd_prev) * 100
                    yon2 = "daha yüksek" if abs(ytd) > abs(ytd_prev) else "daha düşük"
                    analiz = (f" Bu açık geçen yılın aynı dönemine ({_ht(abs(ytd_prev)/1000)} milyar TL) göre "
                              f"<b>%{_ht(abs(pct), 0)} {yon2}</b>; ağırlıkla iç borçlanmayla finanse edildi "
                              f"(net {_ht(ib/1000)} milyar TL).")
            except Exception:
                pass
            out.append(("🪙 Hazine Nakit Gerçekleşmeleri", base + analiz))
        except Exception:
            pass
        # ── 3b) Yabancı Para Hareketi (DTH) ──
        try:
            dh = pd.read_excel(BASE_DIR / "yabanci para hareketi" / "dth.xlsx", sheet_name="Haftalik")
            dh["tarih"] = pd.to_datetime(dh["tarih"]); dh = dh.sort_values("tarih"); L = dh.iloc[-1]
            yon = "arttı" if L["yerlesik_toplam"] >= 0 else "azaldı"
            base = (f"{L['tarih'].strftime('%d.%m.%Y')} haftasında yerleşiklerin YP mevduatı (parite düzeltilmiş) "
                    f"<b>{_ht(abs(L['yerlesik_toplam'])/1000, 1)} milyar USD</b> {yon} "
                    f"(tüzel {_ht(L['tuzel_kisiler']/1000, 1, sign=True)}, gerçek {_ht(L['gercek_kisiler']/1000, 1, sign=True)} milyar).")
            analiz = ""
            try:
                s4 = float(dh["yerlesik_toplam"].tail(4).sum())
                yflow = float(dh[dh["tarih"].dt.year == int(L["tarih"].year)]["yerlesik_toplam"].sum())
                w4 = "artış" if s4 >= 0 else "azalış"
                wy = "artış (dolarizasyon)" if yflow >= 0 else "azalış (de-dolarizasyon)"
                analiz = (f" Son 4 haftada kümülatif <b>{_ht(abs(s4)/1000)} milyar USD {w4}</b>; "
                          f"yılbaşından beri {_ht(abs(yflow)/1000)} milyar USD {wy}.")
            except Exception:
                pass
            out.append(("💱 Yabancı Para Hareketi", base + analiz))
        except Exception:
            pass
        # ── 4) TCMB Rezervleri (analist çerçevesi: brüt döviz + altın + net UR) ──
        try:
            r = pd.read_excel(BASE_DIR / "net rezerv" / "net_rezerv.xlsx")
            r["tarih"] = pd.to_datetime(r["tarih"]); r = r.sort_values("tarih"); L = r.iloc[-1]
            if "net_ur" in r.columns:
                ri = r.set_index("tarih")[["dis_varliklar", "altin", "net_ur"]]
                wd = ri.resample("W-FRI").last().dropna(how="all").diff().iloc[-1]
                ys = r[r["tarih"] >= pd.Timestamp(L["tarih"].year, 1, 1)]
                ytd_nur = (float(L["net_ur"]) - float(ys.iloc[0]["net_ur"])) / 1000 if len(ys) else None
                out.append(("💵 TCMB Rezervleri",
                    f"{L['tarih'].strftime('%d.%m.%Y')} itibarıyla brüt döviz rezervleri "
                    f"<b>{_ht(L['dis_varliklar']/1000)} milyar USD</b> (haftalık {_ht(wd['dis_varliklar']/1000, 1, sign=True)}), "
                    f"altın {_ht(L['altin']/1000)} milyar USD. Net uluslararası rezervler (swap dahil) "
                    f"<b>{_ht(L['net_ur']/1000)} milyar USD</b> (haftalık {_ht(wd['net_ur']/1000, 1, sign=True)}); "
                    f"yıl başından beri {_ht(ytd_nur, 1, sign=True)} milyar USD."))
            else:
                out.append(("💵 TCMB Rezervleri",
                    f"{L['tarih'].strftime('%d.%m.%Y')} itibarıyla brüt dış varlıklar "
                    f"<b>{_ht(L['dis_varliklar']/1000)} milyar USD</b>."))
        except Exception:
            pass
        # ── 5) Cari Denge ──
        try:
            c = pd.read_excel(BASE_DIR / "cari acik" / "cari_acik_son.xlsx")
            col = [x for x in c.columns if "Cari" in str(x)][0]
            L = c.iloc[-1]
            base = f"Son dönem ({L['Tarih']}) cari işlemler dengesi <b>{_ht(L[col], 0)} milyon USD</b>."
            analiz = ""
            try:
                vals = c[col].astype(float).tolist()
                if len(vals) >= 8:
                    last4 = sum(vals[-4:]); prev4 = sum(vals[-8:-4])
                    kel = "açık" if last4 < 0 else "fazla"
                    yon2 = "genişledi" if abs(last4) > abs(prev4) else "daraldı"
                    analiz = (f" Son dört çeyreğin (12 aylık) toplamı <b>{_ht(abs(last4)/1000)} milyar USD {kel}</b>; "
                              f"bir yıl öncesine göre {yon2}.")
            except Exception:
                pass
            out.append(("🌍 Cari Denge", base + analiz))
        except Exception:
            pass
        # ── 6) Kredi Faizleri ──
        try:
            ka = pd.read_excel(BASE_DIR / "kredi mevduat" / "kredi_mevduat.xlsx", sheet_name="Kredi_Akim")
            ka["tarih"] = pd.to_datetime(ka["tarih"]); ka = ka.sort_values("tarih"); L = ka.iloc[-1]
            base = (f"Yeni kredi faizleri ({L['tarih'].strftime('%d.%m.%Y')}): İhtiyaç <b>%{_ht(L.get('İhtiyaç Kredisi'), 2)}</b>, "
                    f"Konut %{_ht(L.get('Konut Kredisi'), 2)}, Ticari %{_ht(L.get('Ticari Krediler'), 2)}.")
            analiz = ""
            try:
                if len(ka) >= 5:
                    P = ka.iloc[-5]
                    di = float(L["İhtiyaç Kredisi"] - P["İhtiyaç Kredisi"])
                    dk = float(L["Konut Kredisi"] - P["Konut Kredisi"])
                    wi = "geriledi" if di < 0 else "yükseldi"
                    wk = "yükseldi" if dk > 0 else "geriledi"
                    analiz = (f" Son 4 haftada ihtiyaç kredisi faizi <b>{_ht(abs(di), 2)} puan {wi}</b>, "
                              f"konut {_ht(abs(dk), 2)} puan {wk}.")
            except Exception:
                pass
            out.append(("🏦 Kredi Faizleri", base + analiz))
        except Exception:
            pass
        # ── 7) Mevduat Faizleri ──
        try:
            ma = pd.read_excel(BASE_DIR / "kredi mevduat" / "kredi_mevduat.xlsx", sheet_name="Mevduat_Akim")
            ma["tarih"] = pd.to_datetime(ma["tarih"]); ma = ma.sort_values("tarih"); L = ma.iloc[-1]
            base = (f"TL mevduat faizi toplam <b>%{_ht(L.get('Toplam'), 2)}</b>; 3 ay %{_ht(L.get('3 Aya Kadar Vadeli'), 2)}, "
                    f"1 yıl %{_ht(L.get('1 Yıla Kadar Vadeli'), 2)} ({L['tarih'].strftime('%d.%m.%Y')}).")
            analiz = ""
            try:
                if len(ma) >= 5:
                    dt = float(L["Toplam"] - ma.iloc[-5]["Toplam"])
                    if abs(dt) < 0.10:
                        analiz = " Son 4 haftada büyük ölçüde <b>yatay</b> seyretti."
                    else:
                        w = "yükseldi" if dt > 0 else "geriledi"
                        analiz = f" Son 4 haftada toplam mevduat faizi <b>{_ht(abs(dt), 2)} puan {w}</b>."
            except Exception:
                pass
            out.append(("💰 Mevduat Faizleri", base + analiz))
        except Exception:
            pass
        # ── 7b) BDDK Bankacılık ──
        try:
            import bddk_analiz as _ba
            _tlb, _usdb, _ = _ba.load_latest(BASE_DIR / "bddk_data")
            if _tlb is not None:
                _t1, _t2, _son, _ = _ba.hesapla(_tlb, _usdb)

                def _g(rows, lbl):
                    return next((r for r in rows if r["label"] == lbl), {})

                def _pp(v):
                    return "—" if v is None else f"{v:+.1f}%".replace(".", ",")

                _kr = _g(_t1, "Toplam Krediler (TL)")
                _mv = _g(_t2, "Toplam Mevduat (TL Cinsi)")
                _tlm = _g(_t2, "TL Mevduat")
                _ypm = _g(_t2, "YP Mevduat (USD)")
                out.append(("📑 BDDK Bankacılık",
                    f"{_son.strftime('%d.%m.%Y')} haftası: toplam krediler (TL) haftalık <b>{_pp(_kr.get('hafta'))}</b> "
                    f"(yıllık {_pp(_kr.get('yillik'))}), toplam mevduat <b>{_pp(_mv.get('hafta'))}</b> "
                    f"(yıllık {_pp(_mv.get('yillik'))}). TL mevduat {_pp(_tlm.get('hafta'))}, "
                    f"YP mevduat (USD) {_pp(_ypm.get('hafta'))}."))
        except Exception:
            pass
        # ── 7c) Hazine İhaleleri & TCMB Doğrudan Alım ──
        def _wavg_home(vals, weights):
            m = vals.notna() & weights.notna() & (weights > 0)
            return float((vals[m] * weights[m]).sum() / weights[m].sum()) if m.any() else None

        try:
            hz = pd.read_excel(BASE_DIR / "hazine ihale " / "hazine_ihale_verileri.xlsx",
                               sheet_name="Tüm İhaleler", header=[0, 1])
            hz.columns = [" / ".join(str(x) for x in c) for c in hz.columns]
            cv = "Genel Bilgiler / Valör Tarihi"
            hz[cv] = pd.to_datetime(hz[cv], format="%d.%m.%Y", errors="coerce")
            hz = hz.dropna(subset=[cv])
            L_t = hz[cv].max(); cy = int(L_t.year)
            ytd = hz[hz[cv].dt.year == cy]
            satis = float(ytd["Toplam Satış / Net (Bin TL)"].sum()) / 1e6
            s3 = hz[hz[cv] >= L_t - pd.DateOffset(months=3)]
            f3 = _wavg_home(s3["Kabul Edilen Faiz (%) / Ort. Yıllık Bileşik"],
                            s3["Toplam Satış / Nominal (Bin TL)"])
            out.append(("🏦 Hazine İhaleleri",
                f"{cy} yılında iç borçlanma ihalelerinde toplam <b>{_ht(satis)} milyar TL</b> (net) satış "
                f"({len(ytd)} ihale, son: {L_t.strftime('%d.%m.%Y')}). Son 3 ayın satış ağırlıklı "
                f"ortalama bileşik faizi <b>%{_ht(f3, 2)}</b>."))
        except Exception:
            pass
        try:
            ta = pd.read_excel(BASE_DIR / "tcmb dogrudan alım" / "tcmb_dogrudan_alim.xlsx",
                               sheet_name="Doğrudan Alım İşlemleri")
            ta["İşlem Tarihi"] = pd.to_datetime(ta["İşlem Tarihi"], errors="coerce")
            ta = ta.dropna(subset=["İşlem Tarihi"])
            L_t = ta["İşlem Tarihi"].max(); cy = int(L_t.year)
            ytd = ta[ta["İşlem Tarihi"].dt.year == cy]
            alim = float(ytd["Kazanan Tutar (Nominal)"].sum()) / 1e6
            s3 = ta[ta["İşlem Tarihi"] >= L_t - pd.DateOffset(months=3)]
            f3 = _wavg_home(s3["Ortalama Bileşik Faiz"], s3["Kazanan Tutar (Nominal)"])
            out.append(("🎯 TCMB Doğrudan Alım",
                f"{cy} yılında doğrudan alım ihalelerinde toplam <b>{_ht(alim)} milyar TL</b> (nominal) "
                f"işlem ({len(ytd)} işlem, son: {L_t.strftime('%d.%m.%Y')}). Son 3 ayın ortalama "
                f"bileşik faizi <b>%{_ht(f3, 2)}</b>."))
        except Exception:
            pass
        # ── 8) Yabancı Menkul Kıymet Yatırımı ──
        try:
            _td = BASE_DIR / "tcmb haftalık stok" / "output"

            def _lv(n):
                fp = _td / f"raw_{n}.csv"
                if not fp.exists():
                    return None, None
                d = pd.read_csv(fp)
                return (float(d.iloc[-1]["value"]), d.iloc[-1]["date"]) if len(d) else (None, None)

            def _flow(n):
                fp = _td / f"raw_{n}.csv"
                if not fp.exists():
                    return None
                d = pd.read_csv(fp)
                d["date"] = pd.to_datetime(d["date"], errors="coerce")
                return d.dropna(subset=["date"])

            _hs, _ld = _lv("Hisse_Stok")
            _ds, _ = _lv("DIBS_Stok")
            _hd, _ = _lv("Hisse_Degisim")
            _dd, _ = _lv("DIBS_Degisim")
            if _hs is not None or _ds is not None:
                _stok = (_hs or 0) + (_ds or 0)
                base = (f"{_ld} itibarıyla yurt dışı yerleşiklerin Türkiye menkul kıymet stoku "
                        f"<b>{_ht(_stok/1000)} milyar USD</b> (Hisse {_ht(_hs/1000)}, DİBS {_ht(_ds/1000)} milyar).")
                analiz = ""
                try:
                    hx = pd.read_excel(_td / "hareket.xlsx", sheet_name="Haftalik")
                    hx["tarih"] = pd.to_datetime(hx["tarih"]); hx = hx.sort_values("tarih")
                    hL = hx.iloc[-1]
                    s4h = float(hx["toplam"].tail(4).sum())
                    hy = hx[hx["tarih"].dt.year == int(hL["tarih"].year)]
                    yflow = float(hy["toplam"].sum())
                    _isim = {"hisse": "Hisse", "dibs_kesin": "DİBS kesin", "dibs_dolayli": "DİBS dolaylı",
                             "ost": "ÖST", "eurobond": "Eurobond"}
                    _ycomp = {k: float(hy[k].sum()) for k in _isim}
                    _lider = max(_ycomp, key=lambda k: _ycomp[k])
                    _dibs_hafta = float(hL[["dibs_kesin", "dibs_dolayli"]].sum())
                    analiz = (f" Bu hafta toplam net yabancı hareketi <b>{_ht(hL['toplam']/1000, 1, sign=True)} milyar USD</b> "
                              f"(Hisse {_ht(hL['hisse'], 0, sign=True)}, DİBS {_ht(_dibs_hafta, 0, sign=True)}, "
                              f"ÖST {_ht(hL['ost'], 0, sign=True)}, Eurobond {_ht(hL['eurobond'], 0, sign=True)} milyon); "
                              f"son 4 haftada <b>{_ht(s4h/1000, 1, sign=True)} milyar USD</b>. "
                              f"Yılbaşından beri {_ht(yflow/1000, 1, sign=True)} milyar USD giriş — en büyük katkı "
                              f"<b>{_isim[_lider]} ({_ht(_ycomp[_lider]/1000, 1, sign=True)} milyar)</b>.")
                except Exception:
                    try:
                        fh = _flow("Hisse_Degisim"); fd = _flow("DIBS_Degisim")
                        if fh is not None and fd is not None:
                            s4 = float(fh["value"].tail(4).sum() + fd["value"].tail(4).sum())
                            _yr = int(pd.to_datetime(_ld).year)
                            yflow = float(fh[fh["date"].dt.year == _yr]["value"].sum()
                                          + fd[fd["date"].dt.year == _yr]["value"].sum())
                            w4 = "giriş" if s4 >= 0 else "çıkış"
                            wy = "giriş" if yflow >= 0 else "çıkış"
                            analiz = (f" Son 4 haftada kümülatif net <b>{_ht(abs(s4), 0)} milyon USD {w4}</b>; "
                                      f"yılbaşından beri {_ht(abs(yflow), 0)} milyon USD {wy}.")
                    except Exception:
                        pass
                out.append(("📊 Yabancı Menkul Kıymet Yatırımı", base + analiz))
        except Exception:
            pass
        return out

    _paths = [BASE_DIR / "enflasyon" / "enflasyon.xlsx", BASE_DIR / "butce" / "butce.xlsx",
              BASE_DIR / "hazine nakit" / "nakit.xlsx", BASE_DIR / "yabanci para hareketi" / "dth.xlsx",
              BASE_DIR / "net rezerv" / "net_rezerv.xlsx", BASE_DIR / "kredi mevduat" / "kredi_mevduat.xlsx",
              BASE_DIR / "cari acik" / "cari_acik_son.xlsx",
              *sorted((BASE_DIR / "bddk_data").glob("bddk_*.xls*"))]
    _bust = "|".join(str(int(p.stat().st_mtime)) for p in _paths if p.exists())
    cards = home_summaries(_bust)

    st.markdown("#### Güncel Görünüm")
    if cards:
        cols = st.columns(2)
        for i, (title, sent) in enumerate(cards):
            with cols[i % 2]:
                st.markdown(f"""<div class="module-card">
                    <div class="module-title">{title}</div>
                    <div class="module-desc">{sent}</div>
                </div>""", unsafe_allow_html=True)
    else:
        st.info("Analizler için soldaki modüllerden verileri güncelleyin.")


# ══════════════════════════════════════════════════════════
# TCMB HAFTALIK STOK
# ══════════════════════════════════════════════════════════

elif selected == "tcmb_stok":
    st.markdown('<div class="main-header">Menkul Kıymetlerde Net Yabancı Hareketi ve Stok</div>', unsafe_allow_html=True)

    output_dir = BASE_DIR / "tcmb haftalık stok" / "output"
    script_path = BASE_DIR / "tcmb haftalık stok" / "tcmb_data.py"

    # Son güncelleme bilgisi
    tablo_path = output_dir / "tablo_son_haftalar.png"
    st.markdown(
        f'<div class="update-info">📅 Son veri güncellemesi: '
        f'{get_file_mod_time(tablo_path)}</div>',
        unsafe_allow_html=True,
    )

    # Güncelleme butonu
    col_btn1, col_btn2 = st.columns([1, 4])
    with col_btn1:
        if st.button("🔄 Verileri Güncelle", key="tcmb_stok_update"):
            if run_script(str(script_path)):
                st.cache_data.clear()

    # ── Otomatik Özet Bilgisi ──
    # Tüm veriler YURT DIŞI yerleşiklerin Türkiye menkul kıymet yatırımıdır.
    # raw_Hisse_Stok / raw_DIBS_Stok = STOK (portföy düzeyi),
    # output/hareket.xlsx = haftalık NET YABANCI HAREKETİ bileşenleri (akım):
    #   hisse (M7), dibs_kesin (M8), dibs_dolayli (M9+M10+M11), ost (M12),
    #   eurobond (M23+M24+M25+M26), menkul_toplam (Eurobond hariç), toplam.
    @st.cache_data
    def build_tcmb_stok_summary(out_dir, cache_key):
        def _last(name):
            fp = Path(out_dir) / f"raw_{name}.csv"
            if not fp.exists():
                return None, None
            d = pd.read_csv(fp)
            if d.empty:
                return None, None
            return float(d.iloc[-1]["value"]), d.iloc[-1]["date"]

        hisse_stok, dt1 = _last("Hisse_Stok")
        dibs_stok, dt2 = _last("DIBS_Stok")
        hisse_deg, _ = _last("Hisse_Degisim")
        dibs_deg, _ = _last("DIBS_Degisim")
        if hisse_stok is None and dibs_stok is None:
            return None
        hisse_stok = hisse_stok or 0.0
        dibs_stok = dibs_stok or 0.0
        hisse_deg = hisse_deg or 0.0
        dibs_deg = dibs_deg or 0.0
        return {"last_date": dt1 or dt2, "stok": hisse_stok + dibs_stok,
                "hisse_stok": hisse_stok, "dibs_stok": dibs_stok,
                "akim": hisse_deg + dibs_deg, "hisse_deg": hisse_deg, "dibs_deg": dibs_deg}

    @st.cache_data
    def load_hareket(path, cache_key):
        hh = pd.read_excel(path, sheet_name="Haftalik")
        hh["tarih"] = pd.to_datetime(hh["tarih"], errors="coerce")
        hh = hh.dropna(subset=["tarih"]).sort_values("tarih").reset_index(drop=True)
        hh["dibs_toplam"] = hh[["dibs_kesin", "dibs_dolayli"]].sum(axis=1, min_count=1)
        hh["toplam_4h"] = hh["toplam"].rolling(4).sum()
        hh["menkul_4h"] = hh["menkul_toplam"].rolling(4).sum()
        return hh

    csv_paths = list(output_dir.glob("raw_*.csv"))
    _cache_key = max((p.stat().st_mtime for p in csv_paths), default=0)
    stok_ozet = build_tcmb_stok_summary(str(output_dir), _cache_key)

    hareket_path = output_dir / "hareket.xlsx"
    H = None
    if hareket_path.exists():
        try:
            H = load_hareket(str(hareket_path), int(hareket_path.stat().st_mtime))
        except Exception:
            H = None

    # Türkçe sayı biçimleri (taban birim: Milyon USD)
    def _tr0(v):
        return f"{v:,.0f}".replace(",", ".")

    def _trs(v):
        return f"{v:+,.0f}".replace(",", ".")

    def _mia(v):
        return f"{v/1000:+,.1f}".replace(",", "\x00").replace(".", ",").replace("\x00", ".")

    if stok_ozet:
        ozet = (f"Yurt dışı yerleşiklerin Türkiye menkul kıymet stoku {stok_ozet['last_date']} itibarıyla toplam "
                f"**{_tr0(stok_ozet['stok'])} milyon USD** (Hisse {_tr0(stok_ozet['hisse_stok'])}, "
                f"DİBS {_tr0(stok_ozet['dibs_stok'])} milyon USD).")
        if H is not None and not H.empty:
            Lh = H.iloc[-1]
            yon = "giriş" if Lh["toplam"] >= 0 else "çıkış"
            ozet += (f" **{Lh['tarih'].strftime('%d.%m.%Y')}** haftasında toplam net yabancı hareketi "
                     f"**{_mia(Lh['toplam'])} milyar USD {yon}**: Hisse {_trs(Lh['hisse'])}, "
                     f"DİBS {_trs(Lh['dibs_toplam'])} (kesin {_trs(Lh['dibs_kesin'])}, dolaylı {_trs(Lh['dibs_dolayli'])}), "
                     f"ÖST {_trs(Lh['ost'])}, Eurobond {_trs(Lh['eurobond'])} milyon USD. "
                     f"Son 4 haftada kümülatif **{_mia(Lh['toplam_4h'])} milyar USD**.")
        st.info(f"📋 **Özet:** {ozet}")
        st.caption("Tüm veriler **yurt dışı yerleşiklerin** Türkiye menkul kıymet yatırımıdır (Kaynak: TCMB EVDS).")
        cm1, cm2, cm3, cm4 = st.columns(4)
        with cm1:
            st.metric("Toplam Stok (Milyon USD)", _tr0(stok_ozet["stok"]),
                      _trs(stok_ozet["akim"]))
        with cm2:
            st.metric("Hisse Senedi Stok", _tr0(stok_ozet["hisse_stok"]),
                      _trs(stok_ozet["hisse_deg"]))
        with cm3:
            st.metric("DİBS Stok", _tr0(stok_ozet["dibs_stok"]),
                      _trs(stok_ozet["dibs_deg"]))
        with cm4:
            if H is not None and not H.empty:
                st.metric("Net Yabancı Hareketi (hafta)", _trs(H.iloc[-1]["toplam"]),
                          help="Hisse + DİBS (kesin+dolaylı) + ÖST + Eurobond — Milyon USD.")
            else:
                st.metric("Net Yabancı Hareketi (hafta)", "—")
        st.caption("Δ = bu haftaki net akım, yalnız kesin alım (Hisse M7 / DİBS M8) — Milyon USD")

    # ── Net Yabancı Hareketi Grafikleri (haftalık akım) ──
    if H is not None and len(H) >= 4:
        hafta_n = st.slider("Gösterilecek hafta sayısı", 12, 104, 27, key="tcmb_hafta_n")
        hp = H.tail(hafta_n)
        Lh = H.iloc[-1]

        # Türkçe etiket biçimleri
        def _t1(v):
            return f"{v:,.1f}".replace(",", "\x00").replace(".", ",").replace("\x00", ".")

        def _t0(v):
            return f"{v:,.0f}".replace(",", ".")

        # Son bar segment etiketi (yığın sırasına göre uç noktaya yazar)
        def _bar_etiket(fig, x, items):
            pos = neg = 0.0
            for y, txt, color in items:
                if pd.isna(y):
                    continue
                if y >= 0:
                    pos += y; yy = pos; shift = 12
                else:
                    neg += y; yy = neg; shift = -14
                fig.add_annotation(x=x, y=yy, text=txt, showarrow=False, yshift=shift,
                                   font=dict(size=13, color=color))

        # 1) Toplam net yabancı hareketi (Eurobond dahil) + 4 haftalık hareketli toplam
        st.subheader("Türkiye'nin Menkul Kıymetlerinde Tahmini Net Yabancı Hareketi")
        st.caption("Hisse + DİBS (kesin+dolaylı) + ÖST + Eurobond — Milyar USD · Kesikli çizgi: 4 haftalık hareketli toplam")
        f1 = go.Figure()
        f1.add_bar(x=hp["tarih"], y=hp["toplam"] / 1000, name="Hisse + DİBS + ÖST + Eurobond",
                   marker_color="#4C9AFF")
        f1.add_scatter(x=hp["tarih"], y=hp["toplam_4h"] / 1000, name="4-Haftalık Hareketli Toplam",
                       mode="lines", line=dict(color="#9AA4B2", width=2, dash="dash"))
        f1.update_traces(hovertemplate="%{x|%d.%m.%Y}<br>%{y:.1f} milyar USD<extra>%{fullData.name}</extra>")
        f1.update_layout(height=400, separators=",.", legend_title_text="",
                         legend=dict(orientation="h", yanchor="bottom", y=-0.35))
        _bar_etiket(f1, Lh["tarih"], [(Lh["toplam"] / 1000, _t1(Lh["toplam"] / 1000), "#4C9AFF")])
        f1.add_annotation(x=Lh["tarih"], y=Lh["toplam_4h"] / 1000, text=_t1(Lh["toplam_4h"] / 1000),
                          showarrow=False, yshift=14, font=dict(size=13, color="#9AA4B2"))
        styled_chart(f1)

        # 2) Menkul kıymetler (Eurobond hariç) + 4 haftalık hareketli toplam
        st.subheader("Menkul Kıymetlerde Net Yabancı Hareketi (Eurobond Hariç)")
        st.caption("Hisse + DİBS (kesin+dolaylı) + ÖST — Milyar USD · Kesikli çizgi: 4 haftalık hareketli toplam")
        f2 = go.Figure()
        f2.add_bar(x=hp["tarih"], y=hp["menkul_toplam"] / 1000, name="Hisse + DİBS + ÖST",
                   marker_color="#4C9AFF")
        f2.add_scatter(x=hp["tarih"], y=hp["menkul_4h"] / 1000, name="4-Haftalık Hareketli Toplam",
                       mode="lines", line=dict(color="#9AA4B2", width=2, dash="dash"))
        f2.update_traces(hovertemplate="%{x|%d.%m.%Y}<br>%{y:.1f} milyar USD<extra>%{fullData.name}</extra>")
        f2.update_layout(height=400, separators=",.", legend_title_text="",
                         legend=dict(orientation="h", yanchor="bottom", y=-0.35))
        _bar_etiket(f2, Lh["tarih"], [(Lh["menkul_toplam"] / 1000, _t1(Lh["menkul_toplam"] / 1000), "#4C9AFF")])
        f2.add_annotation(x=Lh["tarih"], y=Lh["menkul_4h"] / 1000, text=_t1(Lh["menkul_4h"] / 1000),
                          showarrow=False, yshift=14, font=dict(size=13, color="#9AA4B2"))
        styled_chart(f2)

        # 3) Hisse
        st.subheader("Hisse Senedi — Net Yabancı Hareketi")
        st.caption("Milyon USD")
        f3 = go.Figure()
        f3.add_bar(x=hp["tarih"], y=hp["hisse"], name="Hisse", marker_color="#ED7D31")
        f3.update_traces(hovertemplate="%{x|%d.%m.%Y}<br>%{y:,.0f} milyon USD<extra></extra>")
        f3.update_layout(height=360, separators=",.", showlegend=False)
        _bar_etiket(f3, Lh["tarih"], [(Lh["hisse"], _t0(Lh["hisse"]), "#ED7D31")])
        styled_chart(f3)

        # 4) DİBS — Kesin & Dolaylı Alım (yığılmış)
        st.subheader("DİBS — Net Yabancı Hareketi (Kesin & Dolaylı Alım)")
        st.caption("Milyon USD · Dolaylı alım = ters repo + teminat + ödünç")
        f4 = go.Figure()
        f4.add_bar(x=hp["tarih"], y=hp["dibs_kesin"], name="DİBS Kesin Alım", marker_color="#3D7BE0")
        f4.add_bar(x=hp["tarih"], y=hp["dibs_dolayli"], name="DİBS Dolaylı Alım", marker_color="#6FD1FF")
        f4.update_traces(hovertemplate="%{x|%d.%m.%Y}<br>%{y:,.0f} milyon USD<extra>%{fullData.name}</extra>")
        f4.update_layout(height=380, separators=",.", barmode="relative", legend_title_text="",
                         legend=dict(orientation="h", yanchor="bottom", y=-0.35))
        _bar_etiket(f4, Lh["tarih"], [(Lh["dibs_kesin"], _t0(Lh["dibs_kesin"]), "#3D7BE0"),
                                      (Lh["dibs_dolayli"], _t0(Lh["dibs_dolayli"]), "#6FD1FF")])
        styled_chart(f4)

        # 5) ÖST (Özel Sektör Tahvili)
        st.subheader("Özel Sektör Tahvili — Net Yabancı Hareketi")
        st.caption("Milyon USD · Genel yönetim dışındaki sektörlerce ihraç edilen borçlanma senetleri")
        f5o = go.Figure()
        f5o.add_bar(x=hp["tarih"], y=hp["ost"], name="ÖST", marker_color="#9AA4B2")
        f5o.update_traces(hovertemplate="%{x|%d.%m.%Y}<br>%{y:,.0f} milyon USD<extra></extra>")
        f5o.update_layout(height=360, separators=",.", showlegend=False)
        _bar_etiket(f5o, Lh["tarih"], [(Lh["ost"], _t0(Lh["ost"]), "#9AA4B2")])
        styled_chart(f5o)

        # 6) Eurobond
        st.subheader("Türkiye Eurobond — Net Yabancı Hareketi")
        st.caption("Milyon USD · Yurt dışında ihraç edilen genel yönetim, banka, şirket ve diğer finansal kuruluş borçlanma araçları dahil")
        f5 = go.Figure()
        f5.add_bar(x=hp["tarih"], y=hp["eurobond"], name="Eurobond", marker_color="#4CAF7D")
        f5.update_traces(hovertemplate="%{x|%d.%m.%Y}<br>%{y:,.0f} milyon USD<extra></extra>")
        f5.update_layout(height=360, separators=",.", showlegend=False)
        _bar_etiket(f5, Lh["tarih"], [(Lh["eurobond"], _t0(Lh["eurobond"]), "#4CAF7D")])
        styled_chart(f5)

        # 7) Yıllık net yabancı hareketi — bileşen bazında (son 3 yıl)
        st.subheader("Menkul Kıymetlerde Yıllık Net Yabancı Hareketi")
        _cy = int(H["tarih"].dt.year.max())
        st.caption(f"Milyar USD · Bileşen bazında, son 3 yıl ({_cy} yılbaşından bugüne)")
        Hy = H.copy()
        Hy["yil"] = Hy["tarih"].dt.year
        ycomp = (Hy.groupby("yil")[["dibs_kesin", "eurobond", "hisse", "dibs_dolayli", "ost"]]
                 .sum().tail(3) / 1000)
        _ylab = [str(int(y)) for y in ycomp.index]
        f7 = go.Figure()
        for _nm, _col, _renk in [
            ("DİBS Kesin Alım", "dibs_kesin", "#3D7BE0"),
            ("Eurobond + Döviz Sukuk", "eurobond", "#4CAF7D"),
            ("Hisse", "hisse", "#ED7D31"),
            ("DİBS Dolaylı Alım", "dibs_dolayli", "#6FD1FF"),
            ("ÖST", "ost", "#9AA4B2"),
        ]:
            _vals = ycomp[_col].tolist()
            f7.add_bar(x=_ylab, y=_vals, name=_nm, marker_color=_renk,
                       text=[_t1(v) for v in _vals], textposition="inside",
                       insidetextfont=dict(size=12, color="#FFFFFF"),
                       hovertemplate="%{x}<br>" + _nm + ": %{y:.1f} milyar USD<extra></extra>")
        f7.update_layout(height=460, separators=",.", barmode="relative", legend_title_text="",
                         legend=dict(orientation="h", yanchor="bottom", y=-0.22), bargap=0.45)
        styled_chart(f7)

        # Detay tablo — son 8 hafta
        with st.expander("📋 Son 8 Hafta Detay (Milyon USD)", expanded=False):
            t8 = H.tail(8).iloc[::-1]
            disp = pd.DataFrame({
                "Hafta": t8["tarih"].dt.strftime("%d.%m.%Y"),
                "Hisse": t8["hisse"].apply(_trs),
                "DİBS Kesin": t8["dibs_kesin"].apply(_trs),
                "DİBS Dolaylı": t8["dibs_dolayli"].apply(_trs),
                "ÖST": t8["ost"].apply(_trs),
                "Eurobond": t8["eurobond"].apply(_trs),
                "Menkul (Eurobond hariç)": t8["menkul_toplam"].apply(_trs),
                "Toplam": t8["toplam"].apply(_trs),
            })
            st.dataframe(disp, hide_index=True, use_container_width=True)
        get_download_button(str(hareket_path), "📥 Net Yabancı Hareketi (.xlsx)")
    else:
        st.info("Net yabancı hareketi verisi henüz yok — '🔄 Verileri Güncelle'ye tıklayın.")

    # ── Ham Veriler (stok & akım serileri) ──
    with st.expander("📁 Ham Veriler — Stok & Akım Serileri", expanded=False):
        csv_files = sorted(output_dir.glob("raw_*.csv")) if output_dir.exists() else []

        if csv_files:
            selected_csv = st.selectbox(
                "Veri Serisi Seçin",
                csv_files,
                format_func=lambda x: x.stem.replace("raw_", "").replace("_", " "),
                key="tcmb_csv",
            )
            df = pd.read_csv(selected_csv)
            df["date"] = pd.to_datetime(df["date"])

            col_m1, col_m2, col_m3 = st.columns(3)
            with col_m1:
                st.metric("Toplam Kayıt", f"{len(df):,}")
            with col_m2:
                st.metric("Son Tarih", df["date"].max().strftime("%d.%m.%Y"))
            with col_m3:
                last_change = df["change"].iloc[-1] if "change" in df.columns else 0
                st.metric("Son Hafta Değişim", f"{last_change:,.0f} M USD")

            # İnteraktif grafik
            fig = go.Figure()
            fig.add_trace(go.Scatter(
                x=df["date"], y=df["value"],
                mode="lines", name="Değer (Milyon USD)",
                line=dict(color="#1F4E79", width=2),
            ))
            fig.update_layout(
                title=selected_csv.stem.replace("raw_", "").replace("_", " "),
                xaxis_title="Tarih", yaxis_title="Milyon USD",
                template="plotly_white", height=450,
            )
            styled_chart(fig)

            st.dataframe(
                df.tail(20).sort_values("date", ascending=False),
                use_container_width=True,
                hide_index=True,
            )
            get_download_button(str(selected_csv))
        else:
            st.info("Henüz ham veri dosyası yok. Verileri güncelleyin.")

    # PDF Rapor
    st.markdown("---")
    pdf_path = output_dir / "TCMB_Menkul_Kiymet_Raporu.pdf"
    if pdf_path.exists():
        st.subheader("📄 PDF Rapor")
        get_download_button(str(pdf_path), "📥 TCMB Menkul Kıymet Raporu (PDF)")


# ══════════════════════════════════════════════════════════
# HAZİNE İHALE VERİLERİ
# ══════════════════════════════════════════════════════════

elif selected == "hazine":
    st.markdown('<div class="main-header">Hazine İhale Verileri</div>', unsafe_allow_html=True)

    hazine_dir = BASE_DIR / "hazine ihale "
    ihale_file = hazine_dir / "hazine_ihale_verileri.xlsx"
    script_cek = hazine_dir / "hazine_ihale_cek.py"
    script_analiz = hazine_dir / "hazine_analiz.py"
    script_odeme = hazine_dir / "hazine_odeme.py"

    st.markdown(
        f'<div class="update-info">📅 İhale verileri son güncelleme: '
        f'{get_file_mod_time(ihale_file)}</div>',
        unsafe_allow_html=True,
    )

    # Güncelleme butonları
    col_b1, col_b2, col_b3, col_b4 = st.columns(4)
    with col_b1:
        if st.button("🔄 İhale Verilerini Çek", key="hazine_cek"):
            if run_script(str(script_cek)):
                st.cache_data.clear()
    with col_b2:
        if st.button("📊 Analiz Oluştur", key="hazine_analiz"):
            if run_script(str(script_analiz)):
                st.cache_data.clear()
    with col_b3:
        if st.button("💳 Ödeme Analizi", key="hazine_odeme"):
            if run_script(str(script_odeme)):
                st.cache_data.clear()

    # ── Hazine Özet Raporu (Özet sayfasındaki açıklama cümleleri) ──
    @st.cache_data
    def load_hazine_ozet_text(path):
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        ws = wb["Özet"]
        rows = {}
        for r in range(1, min(ws.max_row + 1, 25)):
            v = ws.cell(r, 1).value
            v3 = ws.cell(r, 3).value
            if v:
                rows[r] = str(v)
            if v3:
                rows[(r, 3)] = str(v3)
        wb.close()
        return rows

    if ihale_file.exists():
        try:
            ozet_h = load_hazine_ozet_text(str(ihale_file))
            # R4: Ana açıklama cümlesi
            ana_cumle = ozet_h.get(4, "")
            # Altın ihale özet cümlesi (hazine_dir / altin_ihale_verileri.xlsx Özet R3)
            altin_ozet_cumle = ""
            altin_ihale_top_file = hazine_dir / "altin_ihale_verileri.xlsx"
            if altin_ihale_top_file.exists():
                try:
                    wb_a = openpyxl.load_workbook(str(altin_ihale_top_file), read_only=True, data_only=True)
                    if "Özet" in wb_a.sheetnames:
                        altin_ozet_cumle = wb_a["Özet"].cell(3, 1).value or ""
                    wb_a.close()
                except Exception:
                    pass
            if ana_cumle or altin_ozet_cumle:
                combined = ana_cumle
                if altin_ozet_cumle:
                    combined = f"{combined}\n\n{altin_ozet_cumle}" if combined else altin_ozet_cumle
                st.info(f"📋 **Özet:** {combined}")
            # R14-R21: Dikkat çeken istatistikler
            istatistikler = []
            for r in range(14, 22):
                label = ozet_h.get(r, "")
                detail = ozet_h.get((r, 3), "")
                if label and detail:
                    istatistikler.append(f"**{label}** {detail}")
                elif label:
                    istatistikler.append(f"**{label}**")
            if istatistikler:
                with st.expander("📊 Dikkat Çeken İstatistikler", expanded=False):
                    for ist in istatistikler:
                        st.markdown(f"- {ist}")
        except Exception:
            pass

    # Sekmeler
    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "📋 İhale Verileri",
        "📊 Analiz Grafikleri",
        "🥇 Altın Tahvil",
        "💳 Borç Ödemeleri",
        "📁 Dosyalar",
    ])

    # Tab 1: İhale Verileri
    with tab1:
        if ihale_file.exists():
            try:
                df = pd.read_excel(ihale_file, sheet_name="Tüm İhaleler", header=[0, 1])
                # Flatten multi-index columns
                df.columns = [
                    f"{a} - {b}" if "Unnamed" not in str(a) and "Unnamed" not in str(b)
                    else (b if "Unnamed" not in str(b) else a)
                    for a, b in df.columns
                ]
            except Exception:
                df = pd.read_excel(ihale_file)

            if df is not None and not df.empty:
                # Kısa sütun isim haritası
                short_names = {
                    "Genel Bilgiler - Senet Türü": "Senet Türü",
                    "Genel Bilgiler - Yıl": "Yıl",
                    "Genel Bilgiler - ISIN Kodu": "ISIN",
                    "Genel Bilgiler - Valör Tarihi": "Valör",
                    "Genel Bilgiler - İtfa Tarihi": "İtfa",
                    "Genel Bilgiler - Vade (Gün)": "Vade (Gün)",
                    "Genel Bilgiler - Kupon Dönemi": "Kupon Dönemi",
                    "Genel Bilgiler - Kupon Oranı (%)": "Kupon (%)",
                    "Teklif Edilen Tutar - Nominal (Bin TL)": "Teklif Nom.",
                    "Teklif Edilen Tutar - Net (Bin TL)": "Teklif Net",
                    "İhale Kabul Edilen Tutar - Nominal (Bin TL)": "Kabul Nom.",
                    "İhale Kabul Edilen Tutar - Net (Bin TL)": "Kabul Net",
                    "Toplam Satış - Nominal (Bin TL)": "Top. Satış Nom.",
                    "Toplam Satış - Net (Bin TL)": "Top. Satış Net",
                    "Kabul Edilen Faiz (%) - Ort. Dönem (Basit)": "Faiz Basit (%)",
                    "Kabul Edilen Faiz (%) - Ort. Yıllık Bileşik": "Faiz Bileşik (%)",
                    "Fiyat - İhraç Birim Fiyatı (Ort.)": "Birim Fiyat",
                }
                df = df.rename(columns=short_names)

                # Tarih sütunlarını formatla
                # ÖNEMLİ: Kaynak tarihler GG.AA.YYYY (gün önce). dayfirst=True olmazsa
                # pandas ay-önce (ABD) varsayar; günü 12'den büyük tarihler (13, 15...)
                # geçersiz ay olur ve None'a düşer, ≤12 olanlar ise gün/ay ters çevrilir.
                # İtfa'yı önce hesapla; Valör yedeği (İtfa − Vade) buna dayanıyor.
                _itfa_dt = None
                if "İtfa" in df.columns:
                    _itfa_dt = pd.to_datetime(df["İtfa"], errors="coerce", dayfirst=True)
                    # Kaynakta boş kalan (çoğunlukla 2014 öncesi) itfa tarihlerini
                    # ISIN kodundan tamamla.
                    if "ISIN" in df.columns:
                        _itfa_dt = _itfa_dt.fillna(df["ISIN"].apply(isin_to_itfa))
                    df["İtfa"] = _itfa_dt.dt.strftime("%d.%m.%Y")
                if "Valör" in df.columns:
                    _valor_dt = pd.to_datetime(df["Valör"], errors="coerce", dayfirst=True)
                    # Kaynakta BOZUK (ör. yıl "206" gibi yazım hatası) veya EKSİK valör
                    # için yedek: Valör = İtfa − Vade(Gün). Vade, tanım gereği valör ile
                    # itfa arasındaki gün sayısıdır; bu kimlik kesindir. Yalnızca
                    # ayrıştırılamayan satırları doldurur, geçerli valörlere dokunmaz.
                    _vade_col = "Vade (Gün)" if "Vade (Gün)" in df.columns else None
                    if _itfa_dt is not None and _vade_col is not None:
                        _vade_num = pd.to_numeric(df[_vade_col], errors="coerce")
                        _derived = _itfa_dt - pd.to_timedelta(_vade_num, unit="D")
                        _valor_dt = _valor_dt.fillna(_derived)
                    df["Valör"] = _valor_dt.dt.strftime("%d.%m.%Y")

                st.subheader(f"Toplam {len(df):,} İhale Kaydı")

                # Sütun referansları
                senet_col = "Senet Türü" if "Senet Türü" in df.columns else df.columns[0]
                yil_col = "Yıl" if "Yıl" in df.columns else df.columns[1]

                # Filtreler
                col_f1, col_f2, col_f3 = st.columns(3)
                with col_f1:
                    senet_options = ["Tümü"] + sorted(df[senet_col].dropna().unique().tolist())
                    senet_filter = st.selectbox("Senet Türü", senet_options, key="hz_senet")
                with col_f2:
                    _yil_num = pd.to_numeric(df[yil_col], errors="coerce").dropna().astype(int)
                    yil_vals = [str(y) for y in sorted(_yil_num.unique().tolist(), reverse=True)]
                    yil_options = ["Tümü"] + yil_vals
                    yil_filter = st.selectbox("Yıl", yil_options, key="hz_yil")

                filtered = df.copy()
                if senet_filter != "Tümü":
                    filtered = filtered[filtered[senet_col] == senet_filter]
                if yil_filter != "Tümü":
                    filtered = filtered[pd.to_numeric(filtered[yil_col], errors="coerce") == int(yil_filter)]

                # ── Görüntü için biçimlendirme (binlik ayraç + Yıl'dan .0 kaldır) ──
                # Not: Hesaplamalar 'filtered' üzerinden gider; bu kopya yalnızca gösterim içindir.
                display_df = filtered.copy()
                # Yıl: 2026.0 -> 2026 (binlik ayraçsız tam sayı)
                if yil_col in display_df.columns:
                    display_df[yil_col] = pd.to_numeric(display_df[yil_col], errors="coerce").apply(
                        lambda v: f"{int(v)}" if pd.notna(v) else ""
                    )
                # Tutar sütunları: binlik ayraçlı (virgül), ondalıksız
                _amount_cols = set()
                for _c in ["Teklif Nom.", "Teklif Net", "Kabul Nom.", "Kabul Net",
                           "Top. Satış Nom.", "Top. Satış Net"]:
                    if _c in display_df.columns:
                        _amount_cols.add(_c)
                for _c in display_df.columns:
                    if "Bin TL" in str(_c):
                        _amount_cols.add(_c)
                for _c in _amount_cols:
                    display_df[_c] = pd.to_numeric(display_df[_c], errors="coerce").apply(
                        lambda v: f"{v:,.0f}" if pd.notna(v) else ""
                    )
                # Vade (Gün): binlik ayraçlı tam sayı
                if "Vade (Gün)" in display_df.columns:
                    display_df["Vade (Gün)"] = pd.to_numeric(display_df["Vade (Gün)"], errors="coerce").apply(
                        lambda v: f"{int(round(v)):,}" if pd.notna(v) else ""
                    )
                # Oran / fiyat sütunları: 2 ondalık
                for _c in ["Kupon (%)", "Faiz Basit (%)", "Faiz Bileşik (%)", "Birim Fiyat"]:
                    if _c in display_df.columns:
                        display_df[_c] = pd.to_numeric(display_df[_c], errors="coerce").apply(
                            lambda v: f"{v:,.2f}" if pd.notna(v) else ""
                        )

                st.dataframe(display_df, use_container_width=True, height=500, hide_index=True)

                # Özet grafikler
                st.markdown("---")
                st.subheader("Özet Grafikler")

                # ── Senet Türlerini Kupon Türüne Gruplama ──
                KUPON_GRUP_MAP = {
                    "TL Cinsi Kuponsuz Senetler": "Kuponsuz",
                    "TL Cinsi Sabit Faizli Kuponlu Senetler": "Sabit Faizli Kuponlu",
                    "TL Cinsi Degisken Faizli Kuponlu Senetler": "Değişken Faizli Kuponlu",
                    "TL Cinsi TUFE Endeksli Senetler": "TÜFE Endeksli",
                    "TL Cinsi Kira Sertifikalari": "Kira Sertifikası",
                    "TLREF Endeksli Senetler": "TLREF Endeksli",
                }
                # Bilinmeyenler için orijinal adı koru
                filtered = filtered.copy()
                filtered["_kupon_turu"] = filtered[senet_col].map(KUPON_GRUP_MAP).fillna(filtered[senet_col])

                # ── İtfa Tarihini ISIN Kodundan Türet ──
                # İtfa tarihi ISIN kodundan hesaplanır (modül düzeyindeki isin_to_itfa).
                # Kaynak veride İtfa Tarihi yalnızca 2014 sonrası dolu; ISIN ile tüm
                # kayıtlar için güvenilir hesaplanır (832 bilinen tarihle %100 doğrulandı).
                isin_col = "ISIN" if "ISIN" in filtered.columns else None
                if isin_col:
                    filtered["_itfa_dt"] = filtered[isin_col].apply(isin_to_itfa)
                else:
                    filtered["_itfa_dt"] = pd.NaT

                BUGUN = pd.Timestamp(datetime.now().date())

                # ── Görünüm Seçimi: Güncel Açık Stok vs Tarihsel Toplam ──
                has_itfa = filtered["_itfa_dt"].notna().any()
                if has_itfa:
                    view_mode = st.radio(
                        "Hesaplama görünümü",
                        ["📌 Güncel Açık Stok (vadesi gelmemiş)", "📚 Tarihsel Toplam (tüm ihaleler)"],
                        horizontal=True,
                        key="hz_view_mode",
                        help=(
                            "Güncel Açık Stok: yalnızca itfa tarihi bugünden sonra olan, "
                            "yani hâlâ piyasada olan senetler. İtfa tarihi ISIN kodundan hesaplanır.\n\n"
                            "Tarihsel Toplam: 1998'den bugüne tüm ihaleler (vadesi dolmuşlar dahil)."
                        ),
                    )
                else:
                    view_mode = "📚 Tarihsel Toplam (tüm ihaleler)"

                acik_stok = view_mode.startswith("📌")
                if acik_stok:
                    view_df = filtered[filtered["_itfa_dt"] > BUGUN].copy()
                    _itfa_olan = filtered[filtered["_itfa_dt"] <= BUGUN]
                    st.caption(
                        f"🟢 **Güncel açık stok** görünümü: {len(view_df):,} aktif ihale gösteriliyor. "
                        f"Vadesi dolmuş (itfa olmuş) {len(_itfa_olan):,} senet hariç tutuldu "
                        f"(itfa tarihi ISIN kodundan hesaplandı, baz tarih: {BUGUN.strftime('%d.%m.%Y')})."
                    )
                else:
                    view_df = filtered.copy()
                    st.caption(
                        f"🔵 **Tarihsel toplam** görünümü: 1998'den bugüne {len(view_df):,} ihalenin tamamı "
                        "(vadesi dolmuşlar dahil) toplanıyor."
                    )

                col_g1, col_g2 = st.columns(2)

                with col_g1:
                    # Yaşayan (itfa olmamış) menkullerin TUTARINA göre dağılım
                    _pie_satis_col = "Top. Satış Net" if "Top. Satış Net" in view_df.columns else None
                    if _pie_satis_col:
                        _pie_src = view_df.copy()
                        _pie_src[_pie_satis_col] = pd.to_numeric(_pie_src[_pie_satis_col], errors="coerce")
                        dist = _pie_src.groupby("_kupon_turu")[_pie_satis_col].sum() / 1_000_000  # Milyar TL
                        dist = dist[dist > 0].sort_values(ascending=False)
                        _pie_baslik = (
                            "Kupon Türüne Göre Açık Stok (Milyar TL)" if acik_stok
                            else "Kupon Türüne Göre Dağılım (Milyar TL)"
                        )
                        fig = px.pie(
                            values=dist.values, names=dist.index,
                            title=_pie_baslik,
                            color_discrete_sequence=px.colors.qualitative.Set2,
                        )
                        fig.update_traces(
                            textposition="inside",
                            texttemplate="%{label}<br>%{percent}",
                            hovertemplate="%{label}<br>%{value:,.1f} Milyar TL<br>%{percent}<extra></extra>",
                        )
                    else:
                        dist = view_df["_kupon_turu"].value_counts()
                        fig = px.pie(
                            values=dist.values, names=dist.index,
                            title="Kupon Türüne Göre Dağılım (İhale Sayısı)",
                            color_discrete_sequence=px.colors.qualitative.Set2,
                        )
                    fig.update_layout(height=400)
                    styled_chart(fig)

                with col_g2:
                    # "Tüm İhraçlar": bu grafik yıl filtresinden BAĞIMSIZ (tek yıl seçilince
                    # tek bar kalmasın diye). Yalnızca senet türü filtresine uyar.
                    _yearly_src = df.copy()
                    if senet_filter != "Tümü":
                        _yearly_src = _yearly_src[_yearly_src[senet_col] == senet_filter]
                    yearly = _yearly_src.groupby(yil_col).size().reset_index(name="İhale Sayısı")
                    # Yılı tam sayı + string yap; eksen kategorik olsun ki "2.026" / "2025.6"
                    # gibi binlik-virgül ve ondalık tik'ler oluşmasın.
                    yearly[yil_col] = pd.to_numeric(yearly[yil_col], errors="coerce")
                    yearly = yearly.dropna(subset=[yil_col]).sort_values(yil_col)
                    yearly[yil_col] = yearly[yil_col].astype(int).astype(str)
                    fig = px.bar(
                        yearly, x=yil_col, y="İhale Sayısı",
                        title="Yıllara Göre İhale Sayısı (Tüm İhraçlar)",
                        color_discrete_sequence=["#1F4E79"],
                    )
                    fig.update_xaxes(type="category")
                    fig.update_layout(height=400)
                    styled_chart(fig)

                # ── Kupon Türüne Göre Detaylı Özet Tablo ──
                _tablo_baslik = "Güncel Açık Stok Özeti" if acik_stok else "Tarihsel Toplam Özeti"
                st.markdown(f"#### Kupon Türüne Göre Rakamsal Özet — {_tablo_baslik}")
                st.caption(
                    f"Görünüme göre {len(view_df):,} kayıt üzerinden hesaplandı. "
                    "Ortalama faiz/vade ağırlıklı ortalamadır (Toplam Net Satışa göre ağırlıklandırılmış)."
                )

                # Veri hazırlığı: sayısal sütunlar
                num_satis_col = "Top. Satış Net" if "Top. Satış Net" in view_df.columns else None
                num_faiz_col = "Faiz Bileşik (%)" if "Faiz Bileşik (%)" in view_df.columns else None
                num_vade_col = "Vade (Gün)" if "Vade (Gün)" in view_df.columns else None

                # Sayısala çevir
                _fdf = view_df.copy()
                for c in [num_satis_col, num_faiz_col, num_vade_col]:
                    if c is not None:
                        _fdf[c] = pd.to_numeric(_fdf[c], errors="coerce")

                def _weighted_mean(g, vcol, wcol):
                    s = g[[vcol, wcol]].dropna()
                    if s.empty or s[wcol].sum() == 0:
                        return None
                    return (s[vcol] * s[wcol]).sum() / s[wcol].sum()

                tbl_rows = []
                groups = _fdf.groupby("_kupon_turu")
                total_satis = _fdf[num_satis_col].sum() if num_satis_col else 0
                total_count = len(_fdf)

                for kt, g in groups:
                    sat = g[num_satis_col].sum() if num_satis_col else 0
                    row = {
                        "Kupon Türü": kt,
                        "İhale Sayısı": len(g),
                        "Pay (Sayı %)": (len(g) / total_count * 100) if total_count else 0,
                        "Toplam Net Satış (Milyar TL)": sat / 1_000_000 if num_satis_col else None,
                        "Pay (Tutar %)": (sat / total_satis * 100) if (num_satis_col and total_satis) else None,
                        "Ağ. Ort. Faiz (%)": _weighted_mean(g, num_faiz_col, num_satis_col) if (num_faiz_col and num_satis_col) else None,
                        "Ağ. Ort. Vade (Gün)": _weighted_mean(g, num_vade_col, num_satis_col) if (num_vade_col and num_satis_col) else None,
                    }
                    tbl_rows.append(row)

                # Toplam satırı
                tbl_rows.append({
                    "Kupon Türü": "TOPLAM",
                    "İhale Sayısı": total_count,
                    "Pay (Sayı %)": 100.0,
                    "Toplam Net Satış (Milyar TL)": total_satis / 1_000_000 if num_satis_col else None,
                    "Pay (Tutar %)": 100.0 if num_satis_col else None,
                    "Ağ. Ort. Faiz (%)": _weighted_mean(_fdf, num_faiz_col, num_satis_col) if (num_faiz_col and num_satis_col) else None,
                    "Ağ. Ort. Vade (Gün)": _weighted_mean(_fdf, num_vade_col, num_satis_col) if (num_vade_col and num_satis_col) else None,
                })

                kupon_summary_df = pd.DataFrame(tbl_rows)
                # En büyük tutardan küçüğe sırala (TOPLAM en altta kalsın)
                _toplam_row = kupon_summary_df[kupon_summary_df["Kupon Türü"] == "TOPLAM"]
                _detay_rows = kupon_summary_df[kupon_summary_df["Kupon Türü"] != "TOPLAM"]
                if num_satis_col:
                    _detay_rows = _detay_rows.sort_values("Toplam Net Satış (Milyar TL)", ascending=False)
                else:
                    _detay_rows = _detay_rows.sort_values("İhale Sayısı", ascending=False)
                kupon_summary_df = pd.concat([_detay_rows, _toplam_row], ignore_index=True)

                # Görselleştirme için formatla
                disp = kupon_summary_df.copy()
                disp["İhale Sayısı"] = disp["İhale Sayısı"].apply(lambda v: f"{int(v):,}".replace(",", "."))
                disp["Pay (Sayı %)"] = disp["Pay (Sayı %)"].apply(lambda v: f"%{v:.1f}" if pd.notna(v) else "—")
                disp["Toplam Net Satış (Milyar TL)"] = disp["Toplam Net Satış (Milyar TL)"].apply(
                    lambda v: f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".") if pd.notna(v) else "—"
                )
                disp["Pay (Tutar %)"] = disp["Pay (Tutar %)"].apply(lambda v: f"%{v:.1f}" if pd.notna(v) else "—")
                disp["Ağ. Ort. Faiz (%)"] = disp["Ağ. Ort. Faiz (%)"].apply(lambda v: f"%{v:.2f}" if pd.notna(v) else "—")
                disp["Ağ. Ort. Vade (Gün)"] = disp["Ağ. Ort. Vade (Gün)"].apply(lambda v: f"{int(round(v)):,}".replace(",", ".") if pd.notna(v) else "—")

                # Streamlit dataframe ile göster, son satır (TOPLAM) bold
                st.dataframe(
                    disp, use_container_width=True, hide_index=True,
                    column_config={
                        "Kupon Türü": st.column_config.TextColumn(width="medium"),
                        "İhale Sayısı": st.column_config.TextColumn(width="small"),
                        "Pay (Sayı %)": st.column_config.TextColumn(width="small"),
                        "Toplam Net Satış (Milyar TL)": st.column_config.TextColumn(width="medium"),
                        "Pay (Tutar %)": st.column_config.TextColumn(width="small"),
                        "Ağ. Ort. Faiz (%)": st.column_config.TextColumn(width="small"),
                        "Ağ. Ort. Vade (Gün)": st.column_config.TextColumn(width="small"),
                    },
                )

                with st.expander("ℹ️ Hesaplama Yöntemi ve İtfa Mantığı"):
                    st.markdown(
                        "**Görünümler:**\n"
                        "- **📌 Güncel Açık Stok**: Yalnızca **vadesi henüz gelmemiş** (itfa tarihi bugünden sonra olan) "
                        "senetler toplanır. 'Şu an piyasada ne kadar borç var?' sorusunu yanıtlar.\n"
                        "- **📚 Tarihsel Toplam**: 1998'den bugüne **tüm ihaleler** (vadesi dolmuşlar dahil) toplanır. "
                        "'Tarih boyunca ne kadar ihraç edildi?' sorusunu yanıtlar.\n\n"
                        "**İtfa tarihi nasıl bulunuyor?** Kaynak veride İtfa Tarihi sütunu yalnızca 2014 sonrası ihalelerde "
                        "dolu. Bu yüzden itfa tarihi **ISIN kodundan** hesaplanıyor: DİBS ISIN formatı `TR` + 1 harf + "
                        "`GGAAYY` + sonek şeklindedir. Örn. `TRT051033T12` → **05.10.2033** itfa tarihi. "
                        "Bu yöntem, itfa tarihi bilinen 832 kayıtla **%100 doğrulandı** (sıfır hata).\n\n"
                        "**Sütunlar:**\n"
                        "- **İhale Sayısı**: O kupon türündeki ihale kaydı sayısı.\n"
                        "- **Pay (Sayı %)**: O kupon türünün toplam ihale sayısı içindeki yüzdesi.\n"
                        "- **Toplam Net Satış (Milyar TL)**: 'Toplam Satış - Net (Bin TL)' sütunundaki değerlerin toplamı, "
                        "milyar TL'ye çevrilmiştir (÷1.000.000).\n"
                        "- **Pay (Tutar %)**: Toplam net satış içindeki yüzdesi.\n"
                        "- **Ağ. Ort. Faiz (%)**: 'Faiz Bileşik (%)' sütununun Toplam Net Satışa göre **ağırlıklı ortalaması** "
                        "(her ihalenin payı kadar etkili). Formül: Σ(faiz × net_satış) / Σ(net_satış).\n"
                        "- **Ağ. Ort. Vade (Gün)**: 'Vade (Gün)' sütununun aynı yöntemle ağırlıklı ortalaması.\n"
                        "- Kira Sertifikası gibi senetlerde 'Toplam Satış - Net' değeri kaynakta boş gelebilir; "
                        "bu durumda tutar 0 ve faiz/vade '—' gösterilir."
                    )

                # Faiz trendi
                faiz_col = "Faiz Bileşik (%)" if "Faiz Bileşik (%)" in df.columns else None
                valor_col = "Valör" if "Valör" in df.columns else None

                if faiz_col and valor_col:
                    fdf = filtered[[valor_col, faiz_col, senet_col]].dropna()
                    # Valör artık "GG.AA.YYYY" string'i; dayfirst=True olmadan pandas
                    # ay-önce (ABD) varsayar ve 07.01 → 5 Tem, 11.02 → 5 Kas gibi
                    # GERÇEK OLMAYAN gelecek tarihler üretir.
                    fdf[valor_col] = pd.to_datetime(
                        fdf[valor_col], errors="coerce", dayfirst=True
                    )
                    fdf = fdf.dropna()
                    if not fdf.empty:
                        fig = px.scatter(
                            fdf, x=valor_col, y=faiz_col, color=senet_col,
                            title="Faiz Oranı Trendi",
                            labels={valor_col: "Valör Tarihi", faiz_col: "Yıllık Bileşik Faiz (%)"},
                        )
                        fig.update_layout(height=450, template="plotly_white")
                        styled_chart(fig)

                get_download_button(str(ihale_file))
        else:
            st.info("İhale verileri henüz çekilmemiş. 'İhale Verilerini Çek' butonuna tıklayın.")

    # ── Yardımcı: ihale verisini oku ve hazırla ──
    @st.cache_data
    def load_ihale_raw(path):
        df = pd.read_excel(path, sheet_name="Tüm İhaleler", header=[0, 1])
        df.columns = [
            f"{a} - {b}" if "Unnamed" not in str(a) and "Unnamed" not in str(b)
            else (b if "Unnamed" not in str(b) else a)
            for a, b in df.columns
        ]
        renames = {
            "Genel Bilgiler - Senet Türü": "senet_turu",
            "Genel Bilgiler - Yıl": "yil",
            "Genel Bilgiler - ISIN Kodu": "isin",
            "Genel Bilgiler - Valör Tarihi": "valor",
            "Genel Bilgiler - İtfa Tarihi": "itfa",
            "Genel Bilgiler - Vade (Gün)": "vade",
            "Genel Bilgiler - Kupon Dönemi": "kupon_donemi",
            "Genel Bilgiler - Kupon Oranı (%)": "kupon_orani",
            "Teklif Edilen Tutar - Nominal (Bin TL)": "teklif_nom",
            "Teklif Edilen Tutar - Net (Bin TL)": "teklif_net",
            "İhale Kabul Edilen Tutar - Nominal (Bin TL)": "kabul_nom",
            "İhale Kabul Edilen Tutar - Net (Bin TL)": "kabul_net",
            "Rek. Olmayan Teklif - Kamu - Nominal (Bin TL)": "rek_kamu_nom",
            "Rek. Olmayan Teklif - Kamu - Net (Bin TL)": "rek_kamu_net",
            "Rek. Olmayan Teklif - Piyasa Yapıcı - Nominal (Bin TL)": "rek_py_nom",
            "Rek. Olmayan Teklif - Piyasa Yapıcı - Net (Bin TL)": "rek_py_net",
            "İhale Sonrası Satış - Nominal (Bin TL)": "ihale_sonrasi_nom",
            "İhale Sonrası Satış - Net (Bin TL)": "ihale_sonrasi_net",
            "Toplam Satış - Nominal (Bin TL)": "top_nom",
            "Toplam Satış - Net (Bin TL)": "top_net",
            "Kabul Edilen Faiz (%) - Ort. Dönem (Basit)": "faiz_basit",
            "Kabul Edilen Faiz (%) - Ort. Yıllık Bileşik": "faiz_bilesik",
            "Fiyat - İhraç Birim Fiyatı (Ort.)": "birim_fiyat",
        }
        df = df.rename(columns=renames)
        df["yil"] = pd.to_numeric(df["yil"], errors="coerce")
        for c in ["vade", "teklif_nom", "teklif_net", "kabul_nom", "kabul_net",
                   "rek_kamu_nom", "rek_kamu_net", "rek_py_nom", "rek_py_net",
                   "ihale_sonrasi_nom", "ihale_sonrasi_net", "top_nom", "top_net",
                   "faiz_basit", "faiz_bilesik", "birim_fiyat", "kupon_orani"]:
            df[c] = pd.to_numeric(df[c], errors="coerce")
        # Ham veri "GG.AA.YYYY" (gün-önce) formatında; dayfirst=True şart, aksi halde
        # aylık/çeyreklik dağılım ve tarih gösterimleri bozulur.
        # İtfa'yı ISIN kodundan, eksik/bozuk Valör'ü (ör. yıl "206" yazım hatası)
        # İtfa − Vade(Gün) kimliğinden tamamla.
        df["itfa"] = pd.to_datetime(df["itfa"], errors="coerce", dayfirst=True)
        if "isin" in df.columns:
            df["itfa"] = df["itfa"].fillna(df["isin"].apply(isin_to_itfa))
        df["valor"] = pd.to_datetime(df["valor"], errors="coerce", dayfirst=True)
        df["valor"] = df["valor"].fillna(
            df["itfa"] - pd.to_timedelta(df["vade"], unit="D")
        )
        df["ay"] = df["valor"].dt.month
        df["ceyrek"] = df["valor"].dt.quarter
        return df

    # Tab 2: Analizler (ham veriden hesaplanmış)
    with tab2:
        if ihale_file.exists():
            try:
                raw = load_ihale_raw(str(ihale_file))
                years = sorted(raw["yil"].dropna().unique())
                recent_years = [y for y in years if y >= 2023]

                analiz_sec = st.selectbox("Analiz Seçin", [
                    "1. Yıl Bazlı Genel Özet",
                    "2. Kupon Türüne Göre Borçlanma Dağılımı",
                    "3. Kupon Türüne Göre Ort. Faiz ve Vade",
                    "4. Aylık Borçlanma Dağılımı",
                    "5. Vade Bandı Analizi",
                    "6. Borçlanma Maliyeti ve Talep Analizi",
                    "7. Satış Kanalı Dağılımı",
                    "8. Çeyreklik Faiz Trendi",
                    "9. En Büyük 20 İhale",
                ], key="hz_analiz_sec")

                # ── 1. Yıl Bazlı Genel Özet ──
                if analiz_sec.startswith("1."):
                    st.subheader("Yıl Bazlı Genel Özet")
                    rows = []
                    for y in recent_years:
                        yd = raw[raw["yil"] == y]
                        rows.append({
                            "Yıl": int(y),
                            "İhale Sayısı": len(yd),
                            "Toplam Net Borçlanma (Milyar TL)": yd["top_net"].sum() / 1e6 if yd["top_net"].sum() else 0,
                            "Ort. Faiz (%) Bileşik": yd["faiz_bilesik"].mean(),
                            "Ort. Vade (Gün)": yd["vade"].mean(),
                            "Ort. İhraç Fiyatı": yd["birim_fiyat"].mean(),
                            "Teklif/Kabul (Nom.)": yd["teklif_nom"].sum() / yd["kabul_nom"].sum() if yd["kabul_nom"].sum() else 0,
                            "İhale Başı Ort. Net (Milyar TL)": (yd["top_net"].sum() / len(yd) / 1e6) if len(yd) else 0,
                        })
                    ozet = pd.DataFrame(rows)
                    st.dataframe(ozet.style.format({
                        "Toplam Net Borçlanma (Milyar TL)": "{:,.1f}",
                        "Ort. Faiz (%) Bileşik": "{:.2f}",
                        "Ort. Vade (Gün)": "{:.0f}",
                        "Ort. İhraç Fiyatı": "{:.2f}",
                        "Teklif/Kabul (Nom.)": "{:.2f}",
                        "İhale Başı Ort. Net (Milyar TL)": "{:,.2f}",
                    }), use_container_width=True, hide_index=True)

                    # Grafik: Yıllık toplam borçlanma ve faiz
                    col_g1, col_g2 = st.columns(2)
                    with col_g1:
                        fig = px.bar(ozet, x="Yıl", y="Toplam Net Borçlanma (Milyar TL)",
                                     title="Yıllık Toplam Net Borçlanma",
                                     color_discrete_sequence=["#1F4E79"])
                        fig.update_layout(height=350, template="plotly_white")
                        styled_chart(fig)
                    with col_g2:
                        fig = px.line(ozet, x="Yıl", y="Ort. Faiz (%) Bileşik",
                                      title="Yıllık Ortalama Bileşik Faiz",
                                      markers=True, color_discrete_sequence=["#E53935"])
                        fig.update_layout(height=350, template="plotly_white")
                        styled_chart(fig)

                # ── 2. Kupon Türüne Göre Borçlanma Dağılımı ──
                elif analiz_sec.startswith("2."):
                    st.subheader("Kupon Türüne Göre Borçlanma Dağılımı")
                    kupon_map = {
                        "Kuponsuz": "TL Cinsi Kuponsuz Senetler",
                        "Sabit Kuponlu": "TL Cinsi Sabit Faizli Kuponlu",
                        "TÜFE Endeksli": "TL Cinsi TUFE Endeksli",
                        "Değişken Faizli": "TL Cinsi Degisken Faizli",
                        "TLREF Endeksli": "TLREF Endeksli Senetler",
                        "Kira Sertifikası": "TL Cinsi Kira Sertifikalari",
                    }
                    rows = []
                    for y in recent_years:
                        yd = raw[raw["yil"] == y]
                        row = {"Yıl": int(y)}
                        for kname, kfilter in kupon_map.items():
                            kd = yd[yd["senet_turu"].str.contains(kfilter, case=False, na=False)]
                            row[f"{kname} Sayı"] = len(kd)
                            row[f"{kname} Net (Milyar)"] = kd["top_net"].sum() / 1e6
                        rows.append(row)
                    kupon_df = pd.DataFrame(rows)
                    st.dataframe(kupon_df.style.format(
                        {c: "{:,.1f}" for c in kupon_df.columns if "Milyar" in str(c)}
                    ), use_container_width=True, hide_index=True)

                    # Pie chart
                    y_sec = st.selectbox("Yıl Seçin", recent_years, index=len(recent_years)-1, key="kupon_yil")
                    yd = raw[raw["yil"] == y_sec]
                    pie_data = []
                    for kname, kfilter in kupon_map.items():
                        kd = yd[yd["senet_turu"].str.contains(kfilter, case=False, na=False)]
                        net = kd["top_net"].sum()
                        if net > 0:
                            pie_data.append({"Tür": kname, "Net (Milyar TL)": net / 1e6})
                    if pie_data:
                        fig = px.pie(pd.DataFrame(pie_data), values="Net (Milyar TL)", names="Tür",
                                     title=f"{int(y_sec)} - Kupon Türüne Göre Borçlanma Payı",
                                     color_discrete_sequence=px.colors.qualitative.Set2)
                        fig.update_layout(height=400)
                        styled_chart(fig)

                # ── 3. Kupon Türüne Göre Ort. Faiz ve Vade ──
                elif analiz_sec.startswith("3."):
                    st.subheader("Kupon Türüne Göre Ortalama Faiz ve Vade")
                    kupon_map = {
                        "Kuponsuz": "TL Cinsi Kuponsuz Senetler",
                        "Sabit Kuponlu": "TL Cinsi Sabit Faizli Kuponlu",
                        "TÜFE Endeksli": "TL Cinsi TUFE Endeksli",
                        "Değişken Faizli": "TL Cinsi Degisken Faizli",
                        "TLREF Endeksli": "TLREF Endeksli Senetler",
                    }
                    rows = []
                    for y in recent_years:
                        yd = raw[raw["yil"] == y]
                        for kname, kfilter in kupon_map.items():
                            kd = yd[yd["senet_turu"].str.contains(kfilter, case=False, na=False)]
                            if len(kd) > 0:
                                rows.append({
                                    "Yıl": int(y), "Kupon Türü": kname,
                                    "Ort. Faiz (%)": kd["faiz_bilesik"].mean(),
                                    "Ort. Vade (Gün)": kd["vade"].mean(),
                                })
                    fv_df = pd.DataFrame(rows)
                    st.dataframe(fv_df.style.format({
                        "Ort. Faiz (%)": "{:.2f}", "Ort. Vade (Gün)": "{:.0f}",
                    }), use_container_width=True, hide_index=True)

                    # Grouped bar chart
                    if not fv_df.empty:
                        fig = px.bar(fv_df, x="Kupon Türü", y="Ort. Faiz (%)", color="Yıl",
                                     barmode="group", title="Kupon Türüne Göre Yıllık Ort. Faiz",
                                     color_discrete_sequence=px.colors.qualitative.Set1)
                        fig.update_layout(height=400, template="plotly_white")
                        styled_chart(fig)

                # ── 4. Aylık Borçlanma Dağılımı ──
                elif analiz_sec.startswith("4."):
                    st.subheader("Aylık Borçlanma Dağılımı")
                    ay_adlari = {1:"Ocak",2:"Şubat",3:"Mart",4:"Nisan",5:"Mayıs",6:"Haziran",
                                 7:"Temmuz",8:"Ağustos",9:"Eylül",10:"Ekim",11:"Kasım",12:"Aralık"}
                    rows = []
                    for m in range(1, 13):
                        row = {"Ay": ay_adlari[m]}
                        for y in recent_years:
                            md = raw[(raw["yil"] == y) & (raw["ay"] == m)]
                            row[f"{int(y)} Sayı"] = len(md)
                            row[f"{int(y)} Net (Milyar)"] = md["top_net"].sum() / 1e6
                            row[f"{int(y)} Faiz (%)"] = md["faiz_bilesik"].mean() if len(md) > 0 else None
                        rows.append(row)
                    aylik = pd.DataFrame(rows)
                    fmt = {c: "{:,.1f}" for c in aylik.columns if "Milyar" in str(c)}
                    fmt.update({c: "{:.2f}" for c in aylik.columns if "Faiz" in str(c)})
                    st.dataframe(aylik.style.format(fmt, na_rep="-"), use_container_width=True, hide_index=True)

                    # Aylık grafik
                    y_sec = st.selectbox("Yıl Seçin", recent_years, index=len(recent_years)-1, key="aylik_yil")
                    net_col = f"{int(y_sec)} Net (Milyar)"
                    if net_col in aylik.columns:
                        fig = px.bar(aylik, x="Ay", y=net_col,
                                     title=f"{int(y_sec)} Aylık Net Borçlanma (Milyar TL)",
                                     color_discrete_sequence=["#1F4E79"])
                        fig.update_layout(height=350, template="plotly_white",
                                          xaxis=dict(categoryorder="array", categoryarray=list(ay_adlari.values())))
                        styled_chart(fig)

                # ── 5. Vade Bandı Analizi ──
                elif analiz_sec.startswith("5."):
                    st.subheader("Vade Bandı Analizi")
                    vade_bands = [
                        ("0-182 gün (Kısa)", 0, 182),
                        ("183-364 gün", 183, 364),
                        ("365-728 gün (1-2 yıl)", 365, 728),
                        ("729-1460 gün (2-4 yıl)", 729, 1460),
                        ("1461-2555 gün (4-7 yıl)", 1461, 2555),
                        ("2556+ gün (7+ yıl)", 2556, 99999),
                    ]
                    rows = []
                    for label, lo, hi in vade_bands:
                        row = {"Vade Aralığı": label}
                        for y in recent_years:
                            vd = raw[(raw["yil"] == y) & (raw["vade"] >= lo) & (raw["vade"] <= hi)]
                            total_net = raw[raw["yil"] == y]["top_net"].sum()
                            row[f"{int(y)} Sayı"] = len(vd)
                            row[f"{int(y)} Net (Milyar)"] = vd["top_net"].sum() / 1e6
                            row[f"{int(y)} Pay (%)"] = (vd["top_net"].sum() / total_net * 100) if total_net else 0
                            row[f"{int(y)} Ort.Faiz"] = vd["faiz_bilesik"].mean() if len(vd) > 0 else None
                        rows.append(row)
                    vade_df = pd.DataFrame(rows)
                    fmt = {c: "{:,.1f}" for c in vade_df.columns if "Milyar" in str(c)}
                    fmt.update({c: "{:.1f}" for c in vade_df.columns if "Pay" in str(c)})
                    fmt.update({c: "{:.2f}" for c in vade_df.columns if "Faiz" in str(c)})
                    st.dataframe(vade_df.style.format(fmt, na_rep="-"), use_container_width=True, hide_index=True)

                    # Vade dağılımı stacked bar
                    y_sec = st.selectbox("Yıl Seçin", recent_years, index=len(recent_years)-1, key="vade_yil")
                    pay_col = f"{int(y_sec)} Pay (%)"
                    if pay_col in vade_df.columns:
                        fig = px.bar(vade_df, x="Vade Aralığı", y=pay_col,
                                     title=f"{int(y_sec)} Vade Dağılımı (Net Borçlanma Payı %)",
                                     color_discrete_sequence=["#2E7D32"])
                        fig.update_layout(height=350, template="plotly_white")
                        styled_chart(fig)

                # ── 6. Borçlanma Maliyeti ve Talep Analizi ──
                elif analiz_sec.startswith("6."):
                    st.subheader("Borçlanma Maliyeti ve Talep Analizi")
                    rows = []
                    for y in recent_years:
                        yd = raw[raw["yil"] == y]
                        total_net = yd["top_net"].sum()
                        total_kabul_nom = yd["kabul_nom"].sum()
                        # Ağırlıklı ortalama faiz
                        mask = yd["faiz_bilesik"].notna() & yd["top_net"].notna() & (yd["top_net"] > 0)
                        w_faiz = (yd.loc[mask, "faiz_bilesik"] * yd.loc[mask, "top_net"]).sum() / yd.loc[mask, "top_net"].sum() if yd.loc[mask, "top_net"].sum() else 0
                        rows.append({
                            "Yıl": int(y),
                            "Ağırlıklı Ort. Faiz (%)": w_faiz,
                            "Toplam Net Borçlanma (Milyar TL)": total_net / 1e6,
                            "Teklif/Kabul Oranı": yd["teklif_nom"].sum() / total_kabul_nom if total_kabul_nom else 0,
                            "Ort. İhale Başı Net (Milyar TL)": total_net / len(yd) / 1e6 if len(yd) else 0,
                            "Rek. Kamu Net (Milyar TL)": yd["rek_kamu_net"].sum() / 1e6,
                            "Rek. PY Net (Milyar TL)": yd["rek_py_net"].sum() / 1e6,
                            "İhale Sonrası Net (Milyar TL)": yd["ihale_sonrasi_net"].sum() / 1e6,
                        })
                    maliyet = pd.DataFrame(rows)
                    fmt = {c: "{:,.1f}" for c in maliyet.columns if "Milyar" in str(c)}
                    fmt["Ağırlıklı Ort. Faiz (%)"] = "{:.2f}"
                    fmt["Teklif/Kabul Oranı"] = "{:.2f}"
                    st.dataframe(maliyet.style.format(fmt), use_container_width=True, hide_index=True)

                # ── 7. Satış Kanalı Dağılımı ──
                elif analiz_sec.startswith("7."):
                    st.subheader("Satış Kanalı Dağılımı")
                    rows = []
                    for y in recent_years:
                        yd = raw[raw["yil"] == y]
                        total = yd["top_net"].sum()
                        channels = {
                            "İhale Kabul (Rekabetçi)": yd["kabul_net"].sum(),
                            "Rek. Olmayan - Kamu": yd["rek_kamu_net"].sum(),
                            "Rek. Olmayan - Piyasa Yapıcı": yd["rek_py_net"].sum(),
                            "İhale Sonrası Satış": yd["ihale_sonrasi_net"].sum(),
                        }
                        for ch_name, ch_net in channels.items():
                            rows.append({
                                "Yıl": int(y), "Kanal": ch_name,
                                "Net (Milyar TL)": ch_net / 1e6,
                                "Pay (%)": (ch_net / total * 100) if total else 0,
                            })
                    kanal = pd.DataFrame(rows)
                    st.dataframe(kanal.style.format({
                        "Net (Milyar TL)": "{:,.1f}", "Pay (%)": "{:.1f}",
                    }), use_container_width=True, hide_index=True)

                    # Stacked bar
                    fig = px.bar(kanal, x="Yıl", y="Net (Milyar TL)", color="Kanal",
                                 title="Yıllık Satış Kanalı Dağılımı (Net, Milyar TL)",
                                 color_discrete_sequence=px.colors.qualitative.Set2)
                    fig.update_layout(height=400, template="plotly_white", barmode="stack")
                    styled_chart(fig)

                # ── 8. Çeyreklik Faiz Trendi ──
                elif analiz_sec.startswith("8."):
                    st.subheader("Çeyreklik Faiz Trendi (2020+)")
                    rows = []
                    for y in sorted(raw["yil"].dropna().unique()):
                        if y < 2020:
                            continue
                        for q in range(1, 5):
                            qd = raw[(raw["yil"] == y) & (raw["ceyrek"] == q)]
                            if len(qd) == 0:
                                continue
                            total_net = qd["top_net"].sum()
                            mask = qd["faiz_bilesik"].notna() & qd["top_net"].notna() & (qd["top_net"] > 0)
                            w_faiz = (qd.loc[mask, "faiz_bilesik"] * qd.loc[mask, "top_net"]).sum() / qd.loc[mask, "top_net"].sum() if qd.loc[mask, "top_net"].sum() else 0
                            rows.append({
                                "Dönem": f"{int(y)}-Q{q}",
                                "İhale Sayısı": len(qd),
                                "Ort. Bileşik Faiz (%)": qd["faiz_bilesik"].mean(),
                                "Ağırlıklı Ort. Faiz (%)": w_faiz,
                                "Ort. Vade (Gün)": qd["vade"].mean(),
                                "Toplam Net (Milyar TL)": total_net / 1e6,
                            })
                    ceyrek = pd.DataFrame(rows)
                    st.dataframe(ceyrek.style.format({
                        "Ort. Bileşik Faiz (%)": "{:.2f}",
                        "Ağırlıklı Ort. Faiz (%)": "{:.2f}",
                        "Ort. Vade (Gün)": "{:.0f}",
                        "Toplam Net (Milyar TL)": "{:,.1f}",
                    }), use_container_width=True, hide_index=True)

                    if not ceyrek.empty:
                        fig = go.Figure()
                        fig.add_trace(go.Scatter(
                            x=ceyrek["Dönem"], y=ceyrek["Ağırlıklı Ort. Faiz (%)"],
                            mode="lines+markers", name="Ağırlıklı Ort. Faiz (%)",
                            line=dict(color="#E53935", width=2.5),
                        ))
                        fig.add_trace(go.Bar(
                            x=ceyrek["Dönem"], y=ceyrek["Toplam Net (Milyar TL)"],
                            name="Toplam Net (Milyar TL)", marker_color="#1F4E79", opacity=0.5,
                            yaxis="y2",
                        ))
                        fig.update_layout(
                            title="Çeyreklik Faiz Trendi ve Borçlanma Hacmi",
                            yaxis=dict(title="Faiz (%)", side="left"),
                            yaxis2=dict(title="Net Borçlanma (Milyar TL)", side="right", overlaying="y"),
                            height=450, template="plotly_white", legend=dict(x=0, y=1.1, orientation="h"),
                        )
                        styled_chart(fig)

                # ── 9. En Büyük 20 İhale ──
                elif analiz_sec.startswith("9."):
                    st.subheader("Tarihteki En Büyük 20 İhale (Toplam Net Bazında)")
                    top20 = raw.nlargest(20, "top_net")[
                        ["senet_turu", "isin", "valor", "yil", "vade", "faiz_bilesik", "birim_fiyat", "top_net"]
                    ].copy()
                    top20["valor"] = top20["valor"].dt.strftime("%d.%m.%Y")
                    top20["top_net_milyar"] = top20["top_net"] / 1e6
                    top20 = top20.rename(columns={
                        "senet_turu": "Senet Türü", "isin": "ISIN", "valor": "Valör",
                        "yil": "Yıl", "vade": "Vade (Gün)", "faiz_bilesik": "Faiz Bileşik (%)",
                        "birim_fiyat": "Birim Fiyat", "top_net_milyar": "Toplam Net (Milyar TL)",
                    }).drop(columns=["top_net"])
                    top20.insert(0, "Sıra", range(1, 21))
                    st.dataframe(top20.style.format({
                        "Faiz Bileşik (%)": "{:.2f}",
                        "Birim Fiyat": "{:.3f}",
                        "Toplam Net (Milyar TL)": "{:,.2f}",
                        "Vade (Gün)": "{:.0f}",
                    }), use_container_width=True, hide_index=True)

            except Exception as e:
                st.error(f"Analiz oluşturulamadı: {e}")
                import traceback
                st.code(traceback.format_exc())

    # ── Yardımcı: altın Excel okuma ──
    def read_altin_yearly_excel(filepath):
        """altin_tahvili.xlsx / altin_kira_sertifikasi.xlsx okur.
        Her sheet bir yıl, R1=başlık, R2=header, R3+=veri.
        2017-2019: 14 sütun (None+Etap+ekstra tarih+Fiziki Altın alanları)
        2020+: 10 sütun (standart yapı)."""
        wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
        all_rows = []
        standard_headers = None  # İlk 10-sütunlu sheet'ten alınacak
        for sname in wb.sheetnames:
            ws = wb[sname]
            raw_headers = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
            headers = [h.replace("\n", " ").strip() if h else None for h in raw_headers]
            non_null = [h for h in headers if h]
            ncols_raw = len(non_null)
            if ncols_raw == 0:
                continue
            col_indices = [i + 1 for i, h in enumerate(headers) if h]
            is_old_format = ncols_raw > 10

            # Standart header'ları 10-sütunlu sheet'ten al
            if standard_headers is None and not is_old_format:
                standard_headers = non_null[:10]

            for r in range(3, ws.max_row + 1):
                raw_row = [ws.cell(r, c).value for c in col_indices]
                if not raw_row[0] or str(raw_row[0]).strip() == "":
                    continue
                s = str(raw_row[0]).strip()
                if s.startswith("*") or s.startswith("NOT:") or "Bilgi amaçlıdır" in s:
                    continue
                if "İhraç Dönemi" in s or "ihraç dönemi" in s.lower():
                    continue

                if is_old_format:
                    # Eski format: ISIN(0), Etap(1), Senet Türü(2),
                    # Talep Başlangıç(3), Talep Bitiş(4), Valör(5), İtfa(6),
                    # Fiziki Altın Dönemi(7), Vade(8), Kupon/Kira Dönemi(9),
                    # Kupon/Kira Oranı(10), Altın Miktarı(11), Adet(12)
                    normalized = [
                        raw_row[0],   # ISIN Kodu
                        raw_row[2],   # Senet Türü
                        raw_row[4],   # Talep Bitiş -> Talep Toplama Tarihi
                        raw_row[5],   # Valör Tarihi
                        raw_row[6],   # İtfa Tarihi
                        raw_row[8],   # Vadesi (gün)
                        raw_row[9],   # Kupon/Kira Dönemi
                        raw_row[10],  # Kupon/Kira Oranı
                        raw_row[11],  # Altın Miktarı
                        raw_row[12] if len(raw_row) > 12 else None,
                    ]
                    all_rows.append(normalized)
                else:
                    all_rows.append(raw_row[:10])
        wb.close()
        if not all_rows:
            return pd.DataFrame()
        # Eğer hiç 10-sütunlu sheet yoksa (tüm yıllar eski formatsa) header'ları oluştur
        if standard_headers is None:
            standard_headers = [
                "ISIN Kodu", "Senet Türü", "Talep Toplama Tarihi",
                "Valör Tarihi", "İtfa Tarihi", "Vadesi (gün)",
                "Kupon/Kira Dönemi", "Kupon/Kira Oranı (%)",
                "Kabul Edilen Altın Miktarı (Gram)", "Toplam Adet"
            ]
        df = pd.DataFrame(all_rows, columns=standard_headers[:len(all_rows[0])])
        for c in df.columns:
            if "tarih" in str(c).lower():
                try:
                    df[c] = pd.to_datetime(df[c], errors="coerce", dayfirst=True).dt.strftime("%d.%m.%Y")
                except Exception:
                    pass
        df = df.dropna(how="all")
        return df

    def read_altin_ihale_combined(filepath):
        """altin_ihale_verileri.xlsx okur - multi-header sheet'ler."""
        wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
        results = {}
        for sname in wb.sheetnames:
            ws = wb[sname]
            if sname == "Özet":
                # Özet sayfasını özel oku
                rows = []
                for r in range(1, ws.max_row + 1):
                    row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
                    rows.append(row)
                results["Özet"] = rows
                continue
            # Veri sayfaları: R1=grup header, R2=sütun header
            headers_r2 = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
            headers_r2 = [h if h else "" for h in headers_r2]
            ncols = len([h for h in headers_r2 if h])
            if ncols == 0:
                continue
            data_rows = []
            for r in range(3, ws.max_row + 1):
                row = [ws.cell(r, c).value for c in range(1, ncols + 1)]
                if any(v is not None and str(v).strip() for v in row):
                    data_rows.append(row)
            if data_rows:
                df = pd.DataFrame(data_rows, columns=headers_r2[:ncols])
                # Tarih sütunlarını formatla
                for c in df.columns:
                    if "tarih" in str(c).lower():
                        try:
                            df[c] = pd.to_datetime(df[c], errors="coerce", dayfirst=True).dt.strftime("%d.%m.%Y")
                        except Exception:
                            pass
                results[sname] = df
        wb.close()
        return results

    # Tab 3: Altın Tahvil
    with tab3:
        altin_tahvil_file = hazine_dir / "altin_tahvili.xlsx"
        altin_ihale_file = hazine_dir / "altin_ihale_verileri.xlsx"
        altin_kira_file = hazine_dir / "altin_kira_sertifikasi.xlsx"

        # ── Yardımcı: openpyxl ile Özet okuma ──
        def read_ozet_tables(filepath):
            """altin_ihale_verileri.xlsx Özet sayfasından tabloları okur."""
            wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
            ws = wb["Özet"]
            tables = []
            current_title = ""
            r = 1
            while r <= ws.max_row:
                cell_val = ws.cell(r, 1).value
                if cell_val is None:
                    r += 1
                    continue
                s = str(cell_val).strip()
                # Header satırı mı? ("Yıl" ile başlıyor)
                if s == "Yıl":
                    headers = []
                    for c in range(1, ws.max_column + 1):
                        h = ws.cell(r, c).value
                        if h:
                            headers.append(str(h))
                    ncols = len(headers)
                    data = []
                    r += 1
                    while r <= ws.max_row:
                        row_data = [ws.cell(r, c).value for c in range(1, ncols + 1)]
                        if row_data[0] is None or str(row_data[0]).strip() == "":
                            break
                        data.append(row_data)
                        r += 1
                    tables.append({"title": current_title, "headers": headers, "data": data})
                    current_title = ""
                else:
                    # Başlık satırı olabilir
                    cell2 = ws.cell(r, 2).value
                    if cell2 is None:
                        current_title = s
                    r += 1
            wb.close()
            return tables

        # ── Altın İhale Özet ──
        if altin_ihale_file.exists():
            try:
                altin_data = read_altin_ihale_combined(altin_ihale_file)

                # Özet sayfası
                with st.expander("🥇 Altın İhale Özeti", expanded=True):
                    try:
                        ozet_tables = read_ozet_tables(str(altin_ihale_file))
                        for tbl in ozet_tables:
                            if tbl["title"]:
                                st.markdown(f"**{tbl['title']}**")
                            df = pd.DataFrame(tbl["data"], columns=tbl["headers"])
                            for c in df.columns:
                                df[c] = pd.to_numeric(df[c], errors="ignore")
                            # Sayısal sütunları formatla (string'e çevir)
                            for c in df.columns:
                                if c == "Yıl":
                                    df[c] = df[c].astype(str)
                                    continue
                                if df[c].dtype in ("float64", "int64"):
                                    if any(k in c for k in ["Fiyat", "USDTRY", "Başına", "Oranı"]):
                                        df[c] = df[c].apply(lambda x: f"{x:,.2f}" if pd.notna(x) else "-")
                                    else:
                                        df[c] = df[c].apply(lambda x: f"{x:,.0f}" if pd.notna(x) else "-")
                            st.dataframe(df, use_container_width=True, hide_index=True)

                        # Grafik: İlk tablo (Genel Özet) verisiyle
                        if ozet_tables:
                            gdf = pd.DataFrame(ozet_tables[0]["data"], columns=ozet_tables[0]["headers"])
                            for c in gdf.columns:
                                gdf[c] = pd.to_numeric(gdf[c], errors="coerce")
                            gdf = gdf[gdf["Yıl"].notna() & (gdf["Yıl"] > 0)]  # TOPLAM satırını çıkar
                            if "Toplam Altın (Ons)" in gdf.columns:
                                col1, col2 = st.columns(2)
                                with col1:
                                    fig = px.bar(gdf, x="Yıl", y="Toplam Altın (Ons)",
                                                 title="Yıllık Toplam Altın İhracı (Ons)",
                                                 color_discrete_sequence=["#C6A700"])
                                    fig.update_layout(height=350, template="plotly_white")
                                    styled_chart(fig)
                                with col2:
                                    if "TRY Karşılığı" in gdf.columns:
                                        gdf["TRY (Milyar)"] = gdf["TRY Karşılığı"] / 1e9
                                        fig = px.bar(gdf, x="Yıl", y="TRY (Milyar)",
                                                     title="Yıllık TRY Karşılığı (Milyar TL)",
                                                     color_discrete_sequence=["#1F4E79"])
                                        fig.update_layout(height=350, template="plotly_white")
                                        styled_chart(fig)
                    except Exception as e:
                        st.warning(f"Özet okunamadı: {e}")

                # Altın İhale Verileri (birleştirilmiş)
                for sheet_name in ["Altın İhale Verileri", "Altın Tahvilleri", "Kira Sertifikaları"]:
                    if sheet_name in altin_data:
                        with st.expander(f"🥇 {sheet_name}", expanded=False):
                            df = altin_data[sheet_name]
                            st.dataframe(df, use_container_width=True, height=400, hide_index=True)

                get_download_button(str(altin_ihale_file), "📥 Altın İhale Verileri Excel")
            except Exception as e:
                st.error(f"Altın ihale verileri okunamadı: {e}")
                import traceback
                st.code(traceback.format_exc())

        st.markdown("---")

        # ── Altın Tahvili (yıllık sayfalar) ──
        if altin_tahvil_file.exists():
            with st.expander("🥇 Altın Tahvili (Tüm Yıllar)", expanded=False):
                try:
                    df_at = read_altin_yearly_excel(altin_tahvil_file)
                    if not df_at.empty:
                        st.dataframe(df_at, use_container_width=True, height=400, hide_index=True)

                        # Yıllık özet grafik
                        if "Vadesi (gün)" in df_at.columns:
                            df_at["Vadesi (gün)"] = pd.to_numeric(df_at["Vadesi (gün)"], errors="coerce")
                        ihrac_col = [c for c in df_at.columns if "Kabul Edilen" in str(c)]
                        if ihrac_col:
                            df_at[ihrac_col[0]] = pd.to_numeric(df_at[ihrac_col[0]], errors="coerce")
                            # Yılı valör tarihinden çıkar
                            valor_col = [c for c in df_at.columns if "Valör" in str(c)]
                            if valor_col:
                                df_at["_yil"] = pd.to_datetime(df_at[valor_col[0]], format="%d.%m.%Y", errors="coerce").dt.year
                                yearly = df_at.groupby("_yil").agg(
                                    ihale_sayi=("_yil", "count"),
                                    toplam_altin=(ihrac_col[0], "sum"),
                                ).reset_index()
                                yearly["toplam_altin_ton"] = yearly["toplam_altin"] / 1e6  # gram -> ton
                                col1, col2 = st.columns(2)
                                with col1:
                                    fig = px.bar(yearly, x="_yil", y="ihale_sayi",
                                                 title="Yıllık Altın Tahvil İhale Sayısı",
                                                 labels={"_yil": "Yıl", "ihale_sayi": "İhale Sayısı"},
                                                 color_discrete_sequence=["#FF8F00"])
                                    fig.update_layout(height=350, template="plotly_white")
                                    styled_chart(fig)
                                with col2:
                                    fig = px.bar(yearly, x="_yil", y="toplam_altin_ton",
                                                 title="Yıllık Toplam Altın Miktarı (Ton)",
                                                 labels={"_yil": "Yıl", "toplam_altin_ton": "Ton"},
                                                 color_discrete_sequence=["#C6A700"])
                                    fig.update_layout(height=350, template="plotly_white")
                                    styled_chart(fig)

                    get_download_button(str(altin_tahvil_file), "📥 Altın Tahvili Excel")
                except Exception as e:
                    st.warning(f"Dosya okunamadı: {e}")

        # ── Altın Kira Sertifikası (yıllık sayfalar) ──
        if altin_kira_file.exists():
            with st.expander("🥇 Altına Dayalı Kira Sertifikaları (Tüm Yıllar)", expanded=False):
                try:
                    df_ks = read_altin_yearly_excel(altin_kira_file)
                    if not df_ks.empty:
                        st.dataframe(df_ks, use_container_width=True, height=400, hide_index=True)

                        # Yıllık grafik
                        ihrac_col = [c for c in df_ks.columns if "Kabul Edilen" in str(c)]
                        valor_col = [c for c in df_ks.columns if "Valör" in str(c)]
                        if ihrac_col and valor_col:
                            df_ks[ihrac_col[0]] = pd.to_numeric(df_ks[ihrac_col[0]], errors="coerce")
                            df_ks["_yil"] = pd.to_datetime(df_ks[valor_col[0]], format="%d.%m.%Y", errors="coerce").dt.year
                            yearly = df_ks.groupby("_yil").agg(
                                ihale_sayi=("_yil", "count"),
                                toplam_altin=(ihrac_col[0], "sum"),
                            ).reset_index()
                            yearly["toplam_altin_ton"] = yearly["toplam_altin"] / 1e6
                            fig = px.bar(yearly, x="_yil", y="toplam_altin_ton",
                                         title="Yıllık Kira Sertifikası Altın Miktarı (Ton)",
                                         labels={"_yil": "Yıl", "toplam_altin_ton": "Ton"},
                                         color_discrete_sequence=["#2E7D32"])
                            fig.update_layout(height=350, template="plotly_white")
                            styled_chart(fig)

                    get_download_button(str(altin_kira_file), "📥 Altın Kira Sertifikası Excel")
                except Exception as e:
                    st.warning(f"Dosya okunamadı: {e}")

    # Tab 4: Borç Ödemeleri (hazine_ihale_verileri.xlsx içindeki sayfalardan)
    with tab4:
        if ihale_file.exists():
            try:
                wb = openpyxl.load_workbook(str(ihale_file), read_only=True, data_only=True)

                # ── İç Borç Ödemeleri tablosu ──
                if "İç Borç Ödemeleri" in wb.sheetnames:
                    st.subheader("Merkezi Yönetim İç Borç Ödemeleri (Milyar TL)")
                    ws = wb["İç Borç Ödemeleri"]
                    # Header satır 4, veri satır 5+
                    headers = [ws.cell(4, c).value for c in range(1, ws.max_column + 1) if ws.cell(4, c).value]
                    data_rows = []
                    for r in range(5, ws.max_row + 1):
                        row = [ws.cell(r, c).value for c in range(1, len(headers) + 1)]
                        if any(v is not None for v in row):
                            data_rows.append(row)
                    odeme_df = pd.DataFrame(data_rows, columns=headers)
                    # Yıl sütununu doldur (merged cells)
                    if "Yıl" in odeme_df.columns:
                        odeme_df["Yıl"] = odeme_df["Yıl"].ffill()

                    with st.expander("📊 Aylık Borç Ödemeleri", expanded=True):
                        st.dataframe(odeme_df, use_container_width=True, height=500, hide_index=True)

                    # Yıllık toplam grafik
                    if "Toplam" in odeme_df.columns and "Yıl" in odeme_df.columns:
                        odeme_df["Toplam"] = pd.to_numeric(odeme_df["Toplam"], errors="coerce")
                        yillik_toplam = odeme_df.groupby("Yıl")["Toplam"].sum().reset_index()
                        fig = px.bar(yillik_toplam, x="Yıl", y="Toplam",
                                     title="Yıllık İç Borç Ödemeleri (Milyar TL)",
                                     color_discrete_sequence=["#C62828"])
                        fig.update_layout(height=350, template="plotly_white")
                        styled_chart(fig)

                # ── Ödeme Analizi tablosu ──
                if "Ödeme Analizi" in wb.sheetnames:
                    ws = wb["Ödeme Analizi"]

                    # Tablo 1: Yıllık Ödeme Özeti (R5-R12)
                    st.subheader("Yıllık Ödeme Özeti")
                    h1 = [ws.cell(5, c).value for c in range(1, 11)]
                    h1 = [h for h in h1 if h]
                    t1_rows = []
                    for r in range(6, 13):
                        row = [ws.cell(r, c).value for c in range(1, len(h1) + 1)]
                        if any(v is not None for v in row):
                            t1_rows.append(row)
                    if t1_rows:
                        yillik_ozet = pd.DataFrame(t1_rows, columns=h1)
                        for c in yillik_ozet.columns:
                            if c not in ["Yıl", "Max Ay", "Kaynak"]:
                                yillik_ozet[c] = pd.to_numeric(yillik_ozet[c], errors="coerce")
                        st.dataframe(yillik_ozet.style.format({
                            c: "{:,.2f}" for c in yillik_ozet.columns
                            if c not in ["Yıl", "Max Ay", "Kaynak"] and yillik_ozet[c].dtype != object
                        }, na_rep="-"), use_container_width=True, hide_index=True)

                        # Grafik: Yıllık toplam ödeme
                        if "Genel Toplam" in yillik_ozet.columns:
                            fig = px.bar(yillik_ozet, x="Yıl", y="Genel Toplam",
                                         title="Yıllık İç Borç Ödemeleri - Genel Toplam (Milyar TL)",
                                         color="Kaynak" if "Kaynak" in yillik_ozet.columns else None,
                                         color_discrete_map={"Gerçekleşme": "#2E7D32", "Karma": "#FF8F00", "Projeksiyon": "#1565C0"})
                            fig.update_layout(height=350, template="plotly_white")
                            styled_chart(fig)

                    # Tablo 2: 2026 Aylık Ödeme Detayı (R15-R28)
                    st.subheader("2026 Aylık Ödeme Detayı (Gerçekleşme vs Projeksiyon)")
                    h2 = [ws.cell(15, c).value for c in range(1, 11)]
                    h2 = [h for h in h2 if h]
                    t2_rows = []
                    for r in range(16, 29):
                        row = [ws.cell(r, c).value for c in range(1, len(h2) + 1)]
                        if any(v is not None for v in row):
                            t2_rows.append(row)
                    if t2_rows:
                        aylik_2026 = pd.DataFrame(t2_rows, columns=h2)
                        st.dataframe(aylik_2026, use_container_width=True, hide_index=True)

                        # Grafik
                        for c in aylik_2026.columns:
                            if c not in ["Ay", "Durum"]:
                                aylik_2026[c] = pd.to_numeric(aylik_2026[c].replace("-", None), errors="coerce")
                        proj_col = "Projeksiyon Toplam" if "Projeksiyon Toplam" in aylik_2026.columns else None
                        gercek_col = "Gerçekleşme Toplam" if "Gerçekleşme Toplam" in aylik_2026.columns else None
                        if proj_col or gercek_col:
                            fig = go.Figure()
                            ay_order = ["Ocak","Şubat","Mart","Nisan","Mayıs","Haziran",
                                        "Temmuz","Ağustos","Eylül","Ekim","Kasım","Aralık"]
                            if proj_col:
                                fig.add_trace(go.Bar(x=aylik_2026["Ay"], y=aylik_2026[proj_col],
                                                     name="Projeksiyon", marker_color="#1565C0", opacity=0.6))
                            if gercek_col:
                                fig.add_trace(go.Bar(x=aylik_2026["Ay"], y=aylik_2026[gercek_col],
                                                     name="Gerçekleşme", marker_color="#2E7D32"))
                            fig.update_layout(
                                title="2026 Aylık Borç Ödemeleri - Projeksiyon vs Gerçekleşme",
                                barmode="overlay", height=400, template="plotly_white",
                                xaxis=dict(categoryorder="array", categoryarray=ay_order),
                            )
                            styled_chart(fig)

                    # Tablo 3: Anapara vs Faiz (R31-R35+)
                    st.subheader("Anapara vs Faiz Dağılımı (Yıllık)")
                    h3 = [ws.cell(31, c).value for c in range(1, 11)]
                    h3 = [h for h in h3 if h]
                    t3_rows = []
                    for r in range(32, ws.max_row + 1):
                        row = [ws.cell(r, c).value for c in range(1, len(h3) + 1)]
                        if any(v is not None for v in row):
                            t3_rows.append(row)
                        else:
                            break
                    if t3_rows:
                        anapara_faiz = pd.DataFrame(t3_rows, columns=h3)
                        for c in anapara_faiz.columns:
                            if c != "Yıl":
                                anapara_faiz[c] = pd.to_numeric(anapara_faiz[c].replace("-", None), errors="coerce")
                        st.dataframe(anapara_faiz.style.format({
                            c: "{:,.2f}" for c in anapara_faiz.columns
                            if c != "Yıl" and anapara_faiz[c].dtype != object
                        }, na_rep="-"), use_container_width=True, hide_index=True)

                        # Stacked bar: Anapara + Faiz
                        ap_col = [c for c in anapara_faiz.columns if "Anapara" in c and "Milyar" in c]
                        fz_col = [c for c in anapara_faiz.columns if "Faiz" in c and "Milyar" in c]
                        if ap_col and fz_col:
                            fig = go.Figure()
                            fig.add_trace(go.Bar(x=anapara_faiz["Yıl"], y=anapara_faiz[ap_col[0]],
                                                 name="Anapara", marker_color="#1F4E79"))
                            fig.add_trace(go.Bar(x=anapara_faiz["Yıl"], y=anapara_faiz[fz_col[0]],
                                                 name="Faiz", marker_color="#C62828"))
                            fig.update_layout(
                                title="Yıllık Anapara vs Faiz Ödemeleri (Milyar TL)",
                                barmode="stack", height=400, template="plotly_white",
                            )
                            styled_chart(fig)

                wb.close()
            except Exception as e:
                st.error(f"Borç ödeme verileri okunamadı: {e}")
                import traceback
                st.code(traceback.format_exc())

        # Ek: Orijinal dosyalar
        odeme_file = hazine_dir / "Merkezi_Yonetim_Ic_Borc_Odemeleri-4451657a30c677d9.xls"
        proj_file = hazine_dir / "Merkezi-Yonetim-Ic-Borc-Odeme-Projeksiyonlari-Aylik-9bab85356957d509.xlsx"
        st.markdown("---")
        st.subheader("Kaynak Dosyalar")
        col_dl1, col_dl2 = st.columns(2)
        with col_dl1:
            if odeme_file.exists():
                get_download_button(str(odeme_file), "📥 Borç Ödemeleri (XLS)")
        with col_dl2:
            if proj_file.exists():
                get_download_button(str(proj_file), "📥 Ödeme Projeksiyonları (XLSX)")

    # Tab 5: Tüm Dosyalar
    with tab5:
        st.subheader("Hazine Modülü Dosyaları")
        all_files = sorted(hazine_dir.glob("*.xlsx")) + sorted(hazine_dir.glob("*.xls"))
        for f in all_files:
            if f.name.startswith("~$"):
                continue
            col_d1, col_d2, col_d3 = st.columns([3, 2, 1])
            with col_d1:
                st.text(f.name)
            with col_d2:
                st.text(get_file_mod_time(f))
            with col_d3:
                get_download_button(str(f), "📥 İndir")


# ══════════════════════════════════════════════════════════
# TCMB DOĞRUDAN ALIM
# ══════════════════════════════════════════════════════════

elif selected == "tcmb_alim":
    st.markdown('<div class="main-header">TCMB Doğrudan Alım İşlemleri</div>', unsafe_allow_html=True)

    alim_dir = BASE_DIR / "tcmb dogrudan alım"
    alim_file = alim_dir / "tcmb_dogrudan_alim.xlsx"
    script_guncelle = alim_dir / "guncelle.py"

    st.markdown(
        f'<div class="update-info">📅 Son güncelleme: '
        f'{get_file_mod_time(alim_file)}</div>',
        unsafe_allow_html=True,
    )

    col_b1, col_b2 = st.columns([1, 4])
    with col_b1:
        if st.button("🔄 Verileri Güncelle", key="tcmb_alim_update"):
            if run_script(str(script_guncelle)):
                st.cache_data.clear()

    # ── Ham veri yükle (tüm sekmeler ortak kullanacak) ──
    @st.cache_data
    def load_tcmb_raw(path):
        df = pd.read_excel(path, sheet_name="Doğrudan Alım İşlemleri")
        df["İşlem Tarihi"] = pd.to_datetime(df["İşlem Tarihi"], errors="coerce")
        df["Valör"] = pd.to_datetime(df["Valör"], errors="coerce")
        df["Vade"] = pd.to_datetime(df["Vade"], errors="coerce")
        df["Yıl"] = df["İşlem Tarihi"].dt.year
        df["Ay"] = df["İşlem Tarihi"].dt.month
        df["Vade (Gün)"] = (df["Vade"] - df["Valör"]).dt.days
        return df

    @st.cache_data
    def load_tcmb_ozet_text(path):
        """Portföy Özeti sayfasındaki özet ve uyarı metinlerini oku."""
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        ws = wb["Portföy Özeti"]
        ozet = {}
        for r in range(1, ws.max_row + 1):
            v = ws.cell(r, 1).value
            if v:
                ozet[r] = str(v)
        wb.close()
        return ozet

    if alim_file.exists():
        try:
            raw = load_tcmb_raw(str(alim_file))
        except Exception as e:
            raw = pd.DataFrame()
            st.error(f"Veri yüklenemedi: {e}")

        # ── Portföy Özeti bilgilerini oku ──
        ozet_rows = {}
        try:
            ozet_rows = load_tcmb_ozet_text(str(alim_file))
        except Exception:
            pass

        # Özet metin ve uyarıları göster
        ozet_text = ozet_rows.get(21, "")
        uyarilar = []
        for r in sorted(ozet_rows.keys()):
            if r >= 26:
                uyarilar.append(ozet_rows[r])

        if ozet_text:
            st.info(f"📋 **Portföy Özeti:** {ozet_text}")
        if uyarilar:
            with st.expander("⚠️ Uyarılar ve Tespitler", expanded=False):
                for u in uyarilar:
                    if u.startswith("●"):
                        st.error(u)
                    elif u.startswith("▲"):
                        st.warning(u)
                    elif u.startswith("◆"):
                        st.info(u)
                    else:
                        st.write(u)

        if not raw.empty:
            # ── Portföy Özeti metrikleri (ham veriden hesapla) ──
            today = raw["İşlem Tarihi"].max()
            aktif = raw[raw["Vade"] > today].copy()
            toplam_stok_nom = aktif["Kazanan Tutar (Nominal)"].sum()
            toplam_stok_net = aktif["Kazanan Tutar (Net)"].sum()
            # Ağırlıklı ort. bileşik faiz (aktif portföy)
            w_faiz = aktif["Kazanan Tutar (Nominal)"] * aktif["Ortalama Bileşik Faiz"]
            ort_getiri = w_faiz.sum() / aktif["Kazanan Tutar (Nominal)"].sum() if aktif["Kazanan Tutar (Nominal)"].sum() > 0 else 0
            portfoy_ihale = len(aktif)

            cm1, cm2, cm3, cm4, cm5 = st.columns(5)
            with cm1:
                st.metric("Referans Tarihi", today.strftime("%d.%m.%Y"))
            with cm2:
                st.metric("Portföy Stoku (Milyar TL)", f"{toplam_stok_nom/1e6:,.1f}")
            with cm3:
                st.metric("Net Stok (Milyar TL)", f"{toplam_stok_net/1e6:,.1f}")
            with cm4:
                st.metric("Portföy Getirisi (%)", f"%{ort_getiri:,.2f}")
            with cm5:
                st.metric("Aktif İhale Sayısı", f"{portfoy_ihale:,}")

            tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8 = st.tabs([
                "📋 İşlem Verileri",
                "📊 Yıllık Borçlanma",
                "📅 Yıllık İtfa Profili",
                "🔄 Borçlanma vs İtfa",
                "📐 Vade Dağılımı",
                "🏷️ ISIN Analizi",
                "📈 Aylık Faiz Trendi",
                "📁 İndir",
            ])

            # ═══════ TAB 1: İŞLEM VERİLERİ ═══════
            with tab1:
                st.subheader(f"Toplam {len(raw):,} İşlem Kaydı")
                cm1, cm2, cm3 = st.columns(3)
                with cm1:
                    st.metric("İlk İşlem", raw["İşlem Tarihi"].min().strftime("%d.%m.%Y"))
                with cm2:
                    st.metric("Son İşlem", raw["İşlem Tarihi"].max().strftime("%d.%m.%Y"))
                with cm3:
                    st.metric("Toplam İşlem", f"{len(raw):,}")

                # Gösterim için tarih formatlı kopya
                df_show = raw.drop(columns=["Yıl", "Ay", "Vade (Gün)"], errors="ignore").copy()
                for c in ["İşlem Tarihi", "Valör", "Vade"]:
                    if c in df_show.columns:
                        df_show[c] = df_show[c].dt.strftime("%d.%m.%Y")
                st.dataframe(df_show, use_container_width=True, height=500, hide_index=True)

                # Yıllık toplam borçlanma grafiği
                yearly_nom = raw.groupby("Yıl")["Kazanan Tutar (Nominal)"].sum().reset_index()
                yearly_nom.columns = ["Yıl", "Toplam Nominal (Bin TL)"]
                fig = px.bar(yearly_nom, x="Yıl", y="Toplam Nominal (Bin TL)",
                             title="Yıllık Toplam Borçlanma (Nominal, Bin TL)",
                             color_discrete_sequence=["#1F4E79"], template="plotly_white")
                fig.update_layout(height=400)
                styled_chart(fig, key="tcmb_yillik_borclanma_chart")

                # Aylık trend grafiği
                monthly = raw.groupby([raw["İşlem Tarihi"].dt.to_period("M")])["Kazanan Tutar (Nominal)"].sum().reset_index()
                monthly.columns = ["Ay", "Nominal"]
                monthly["Ay"] = monthly["Ay"].astype(str)
                fig2 = px.line(monthly, x="Ay", y="Nominal",
                               title="Aylık Borçlanma Trendi (Nominal, Bin TL)",
                               template="plotly_white")
                fig2.update_layout(height=350, xaxis_tickangle=-45)
                styled_chart(fig2, key="tcmb_aylik_trend_chart")

            # ═══════ TAB 2: YILLIK BORÇLANMA ANALİZİ ═══════
            with tab2:
                st.subheader("📊 Yıllık Borçlanma Analizi (İşlem Tarihine Göre)")

                yillik = raw.groupby("Yıl").agg(
                    ihale_sayisi=("Kazanan Tutar (Nominal)", "count"),
                    toplam_teklif=("Teklif Tutarı (Nominal)", "sum"),
                    toplam_borc_nom=("Kazanan Tutar (Nominal)", "sum"),
                    toplam_borc_net=("Kazanan Tutar (Net)", "sum"),
                ).reset_index()

                # Karşılanma oranı
                yillik["karsilanma"] = (yillik["toplam_borc_nom"] / yillik["toplam_teklif"] * 100).round(2)

                # Ağırlıklı ortalama faizler
                def wavg(group, val_col, wt_col):
                    d = group[[val_col, wt_col]].dropna()
                    if d[wt_col].sum() == 0:
                        return None
                    return (d[val_col] * d[wt_col]).sum() / d[wt_col].sum()

                wf_basit = raw.groupby("Yıl").apply(lambda g: wavg(g, "Ortalama Basit Faiz", "Kazanan Tutar (Nominal)")).reset_index()
                wf_basit.columns = ["Yıl", "ort_basit_faiz"]
                wf_bilesik = raw.groupby("Yıl").apply(lambda g: wavg(g, "Ortalama Bileşik Faiz", "Kazanan Tutar (Nominal)")).reset_index()
                wf_bilesik.columns = ["Yıl", "ort_bilesik_faiz"]
                wf_vade = raw.groupby("Yıl").apply(lambda g: wavg(g, "Vade (Gün)", "Kazanan Tutar (Nominal)")).reset_index()
                wf_vade.columns = ["Yıl", "ort_vade"]

                yillik = yillik.merge(wf_basit, on="Yıl").merge(wf_bilesik, on="Yıl").merge(wf_vade, on="Yıl")

                # TOPLAM satırı
                toplam = pd.DataFrame([{
                    "Yıl": "TOPLAM",
                    "ihale_sayisi": yillik["ihale_sayisi"].sum(),
                    "toplam_teklif": yillik["toplam_teklif"].sum(),
                    "toplam_borc_nom": yillik["toplam_borc_nom"].sum(),
                    "toplam_borc_net": yillik["toplam_borc_net"].sum(),
                    "karsilanma": (yillik["toplam_borc_nom"].sum() / yillik["toplam_teklif"].sum() * 100) if yillik["toplam_teklif"].sum() > 0 else 0,
                    "ort_basit_faiz": None,
                    "ort_bilesik_faiz": None,
                    "ort_vade": None,
                }])
                yillik = pd.concat([yillik, toplam], ignore_index=True)

                yillik.columns = ["Yıl", "İhale Sayısı", "Toplam Teklif (Nominal, Bin TL)",
                                  "Toplam Borçlanma (Nominal, Bin TL)", "Toplam Borçlanma (Net, Bin TL)",
                                  "Karşılanma Oranı (%)", "Ağ.Ort. Basit Faiz (%)",
                                  "Ağ.Ort. Bileşik Faiz (%)", "Ağ.Ort. Vade (Gün)"]

                # Formatlama
                for c in yillik.columns:
                    if c == "Yıl":
                        yillik[c] = yillik[c].astype(str)
                        continue
                    if yillik[c].dtype in ("float64", "int64", "Int64"):
                        if "Faiz" in c or "Oranı" in c:
                            yillik[c] = yillik[c].apply(lambda x: f"{x:,.2f}" if pd.notna(x) else "-")
                        elif "Vade" in c:
                            yillik[c] = yillik[c].apply(lambda x: f"{x:,.0f}" if pd.notna(x) else "-")
                        elif "Sayısı" in c:
                            yillik[c] = yillik[c].apply(lambda x: f"{int(x):,}" if pd.notna(x) else "-")
                        else:
                            yillik[c] = yillik[c].apply(lambda x: f"{x:,.0f}" if pd.notna(x) else "-")
                st.dataframe(yillik, use_container_width=True, hide_index=True)

                # Grafik: Yıllık borçlanma + faiz dual-axis
                yillik_num = raw.groupby("Yıl").agg(
                    borc=("Kazanan Tutar (Nominal)", "sum"),
                ).reset_index()
                wf_b = raw.groupby("Yıl").apply(lambda g: wavg(g, "Ortalama Bileşik Faiz", "Kazanan Tutar (Nominal)")).reset_index()
                wf_b.columns = ["Yıl", "Faiz"]
                yillik_num = yillik_num.merge(wf_b, on="Yıl")
                yillik_num["borc_milyar"] = yillik_num["borc"] / 1e6

                fig = go.Figure()
                fig.add_trace(go.Bar(x=yillik_num["Yıl"], y=yillik_num["borc_milyar"],
                                     name="Borçlanma (Milyar TL)", marker_color="#1F4E79", yaxis="y"))
                fig.add_trace(go.Scatter(x=yillik_num["Yıl"], y=yillik_num["Faiz"],
                                         name="Ağ.Ort. Bileşik Faiz (%)", mode="lines+markers",
                                         marker_color="#E74C3C", yaxis="y2"))
                fig.update_layout(
                    title="Yıllık Borçlanma ve Faiz Trendi",
                    yaxis=dict(title="Borçlanma (Milyar TL)", side="left"),
                    yaxis2=dict(title="Bileşik Faiz (%)", side="right", overlaying="y"),
                    template="plotly_white", height=450, legend=dict(orientation="h", y=-0.15),
                )
                styled_chart(fig, key="tcmb_yillik_borc_faiz")

            # ═══════ TAB 3: YILLIK İTFA PROFİLİ ═══════
            with tab3:
                st.subheader("📅 Yıllık İtfa Profili (Vade Tarihine Göre)")

                raw_itfa = raw.copy()
                raw_itfa["İtfa Yılı"] = raw_itfa["Vade"].dt.year
                itfa = raw_itfa.groupby("İtfa Yılı").agg(
                    itfa_nom=("Kazanan Tutar (Nominal)", "sum"),
                    itfa_net=("Kazanan Tutar (Net)", "sum"),
                    ihale_sayisi=("Kazanan Tutar (Nominal)", "count"),
                ).reset_index()

                wf_itfa = raw_itfa.groupby("İtfa Yılı").apply(
                    lambda g: wavg(g, "Ortalama Bileşik Faiz", "Kazanan Tutar (Nominal)")
                ).reset_index()
                wf_itfa.columns = ["İtfa Yılı", "ort_faiz"]
                itfa = itfa.merge(wf_itfa, on="İtfa Yılı")

                # Vadesi geçmiş / gelmemiş ayırımı
                itfa["Durum"] = itfa["İtfa Yılı"].apply(
                    lambda y: "İtfa Olmuş" if y <= today.year else "Vadesi Gelmemiş"
                )

                itfa.columns = ["İtfa Yılı", "İtfa Nominal (Bin TL)", "İtfa Net (Bin TL)",
                                "İhale Sayısı", "Ağ.Ort. Bileşik Faiz (%)", "Durum"]

                for c in itfa.columns:
                    if c in ("İtfa Yılı", "Durum"):
                        itfa[c] = itfa[c].astype(str)
                    elif "Faiz" in c:
                        itfa[c] = itfa[c].apply(lambda x: f"{x:,.2f}" if pd.notna(x) else "-")
                    elif "Sayısı" in c:
                        itfa[c] = itfa[c].apply(lambda x: f"{int(x):,}" if pd.notna(x) else "-")
                    else:
                        itfa[c] = itfa[c].apply(lambda x: f"{x:,.0f}" if pd.notna(x) else "-")
                st.dataframe(itfa, use_container_width=True, hide_index=True)

                # İtfa profili grafiği
                itfa_num = raw_itfa.groupby("İtfa Yılı")["Kazanan Tutar (Nominal)"].sum().reset_index()
                itfa_num.columns = ["Yıl", "İtfa (Bin TL)"]
                itfa_num["İtfa (Milyar TL)"] = itfa_num["İtfa (Bin TL)"] / 1e6
                colors = ["#2ECC71" if y <= today.year else "#E67E22" for y in itfa_num["Yıl"]]
                fig = go.Figure()
                fig.add_trace(go.Bar(x=itfa_num["Yıl"], y=itfa_num["İtfa (Milyar TL)"],
                                     marker_color=colors, name="İtfa Tutarı"))
                fig.update_layout(title="Yıllık İtfa Profili (Milyar TL)",
                                  xaxis_title="İtfa Yılı", yaxis_title="Milyar TL",
                                  template="plotly_white", height=400)
                # Bugünkü yılı dikey çizgiyle işaretle
                fig.add_vline(x=today.year, line_dash="dash", line_color="red",
                              annotation_text="Bugün")
                styled_chart(fig, key="tcmb_itfa_profili")

            # ═══════ TAB 4: BORÇLANMA vs İTFA ═══════
            with tab4:
                st.subheader("🔄 Yıllık Borçlanma vs İtfa Karşılaştırması")

                # Borçlanma (işlem yılına göre)
                borc_yil = raw.groupby("Yıl")["Kazanan Tutar (Nominal)"].sum().reset_index()
                borc_yil.columns = ["Yıl", "Borçlanma (Nominal)"]
                borc_net = raw.groupby("Yıl")["Kazanan Tutar (Net)"].sum().reset_index()
                borc_net.columns = ["Yıl", "Borçlanma (Net)"]

                # İtfa (vade yılına göre)
                raw_itfa2 = raw.copy()
                raw_itfa2["İtfa Yılı"] = raw_itfa2["Vade"].dt.year
                itfa_yil = raw_itfa2.groupby("İtfa Yılı")["Kazanan Tutar (Nominal)"].sum().reset_index()
                itfa_yil.columns = ["Yıl", "İtfa (Nominal)"]

                # Tüm yıllar
                all_years = sorted(set(borc_yil["Yıl"].tolist() + itfa_yil["Yıl"].tolist()))
                bvi = pd.DataFrame({"Yıl": all_years})
                bvi = bvi.merge(borc_yil, on="Yıl", how="left").merge(itfa_yil, on="Yıl", how="left")
                bvi = bvi.merge(borc_net, on="Yıl", how="left")
                bvi = bvi.fillna(0)
                bvi["Net Pozisyon"] = bvi["Borçlanma (Nominal)"] - bvi["İtfa (Nominal)"]
                bvi["İtfa/Borçlanma (%)"] = (bvi["İtfa (Nominal)"] / bvi["Borçlanma (Nominal)"] * 100).where(bvi["Borçlanma (Nominal)"] > 0, 0).round(2)
                bvi["Kümülatif Net"] = bvi["Net Pozisyon"].cumsum()

                bvi_show = bvi.copy()
                for c in bvi_show.columns:
                    if c == "Yıl":
                        bvi_show[c] = bvi_show[c].astype(str)
                    elif "%" in c:
                        bvi_show[c] = bvi_show[c].apply(lambda x: f"{x:,.2f}" if pd.notna(x) else "-")
                    else:
                        bvi_show[c] = bvi_show[c].apply(lambda x: f"{x:,.0f}" if pd.notna(x) else "-")
                st.dataframe(bvi_show, use_container_width=True, hide_index=True)

                # Grafik
                fig = go.Figure()
                fig.add_trace(go.Bar(x=bvi["Yıl"], y=bvi["Borçlanma (Nominal)"] / 1e6,
                                     name="Borçlanma (Milyar TL)", marker_color="#1F4E79"))
                fig.add_trace(go.Bar(x=bvi["Yıl"], y=bvi["İtfa (Nominal)"] / 1e6,
                                     name="İtfa (Milyar TL)", marker_color="#E74C3C"))
                fig.add_trace(go.Scatter(x=bvi["Yıl"], y=bvi["Kümülatif Net"] / 1e6,
                                         name="Kümülatif Net (Milyar TL)",
                                         mode="lines+markers", marker_color="#27AE60", yaxis="y2"))
                fig.update_layout(
                    title="Borçlanma vs İtfa (Milyar TL)",
                    barmode="group",
                    yaxis=dict(title="Milyar TL", side="left"),
                    yaxis2=dict(title="Kümülatif Net (Milyar TL)", side="right", overlaying="y"),
                    template="plotly_white", height=450,
                    legend=dict(orientation="h", y=-0.15),
                )
                styled_chart(fig, key="tcmb_borc_vs_itfa")

            # ═══════ TAB 5: VADE DAĞILIM ANALİZİ ═══════
            with tab5:
                st.subheader("📐 Vade Dağılım Analizi (İşlem Yılına Göre)")

                def vade_bucket(gun):
                    if pd.isna(gun):
                        return "Bilinmiyor"
                    if gun <= 365:
                        return "0-1 Yıl"
                    elif gun <= 730:
                        return "1-2 Yıl"
                    elif gun <= 1095:
                        return "2-3 Yıl"
                    elif gun <= 1825:
                        return "3-5 Yıl"
                    elif gun <= 2555:
                        return "5-7 Yıl"
                    else:
                        return "7+ Yıl"

                raw_vade = raw.copy()
                raw_vade["Vade Grubu"] = raw_vade["Vade (Gün)"].apply(vade_bucket)
                bucket_order = ["0-1 Yıl", "1-2 Yıl", "2-3 Yıl", "3-5 Yıl", "5-7 Yıl", "7+ Yıl"]

                vade_pivot = raw_vade.pivot_table(
                    index="Yıl", columns="Vade Grubu",
                    values="Kazanan Tutar (Nominal)", aggfunc="sum", fill_value=0
                ).reindex(columns=bucket_order, fill_value=0)
                vade_pivot["Toplam"] = vade_pivot.sum(axis=1)

                # Yüzde tablosu
                vade_pct = vade_pivot.div(vade_pivot["Toplam"], axis=0).drop(columns=["Toplam"]) * 100

                # Gösterim tablosu (nominal + yüzde birleşik)
                vade_show = vade_pivot.copy()
                for c in vade_show.columns:
                    vade_show[c] = vade_show[c].apply(lambda x: f"{x:,.0f}" if pd.notna(x) else "-")
                vade_show.index = vade_show.index.astype(str)
                st.markdown("**Nominal Tutarlar (Bin TL)**")
                st.dataframe(vade_show, use_container_width=True)

                # Yüzde tablosu
                vade_pct_show = vade_pct.copy()
                for c in vade_pct_show.columns:
                    vade_pct_show[c] = vade_pct_show[c].apply(lambda x: f"%{x:.1f}" if pd.notna(x) else "-")
                vade_pct_show.index = vade_pct_show.index.astype(str)
                st.markdown("**Yüzde Dağılım (%)**")
                st.dataframe(vade_pct_show, use_container_width=True)

                # Stacked bar chart
                vade_melt = vade_pct.reset_index().melt(id_vars="Yıl", var_name="Vade Grubu", value_name="Oran (%)")
                fig = px.bar(vade_melt, x="Yıl", y="Oran (%)", color="Vade Grubu",
                             title="Vade Dağılımı (% - Yıllara Göre)",
                             color_discrete_sequence=px.colors.sequential.Blues_r,
                             template="plotly_white")
                fig.update_layout(barmode="stack", height=450)
                styled_chart(fig, key="tcmb_vade_dagilim")

                # Son yıl pasta grafiği
                son_yil = raw["Yıl"].max()
                son_yil_data = vade_pct.loc[son_yil] if son_yil in vade_pct.index else None
                if son_yil_data is not None:
                    fig2 = px.pie(values=son_yil_data.values, names=son_yil_data.index,
                                  title=f"{son_yil} Yılı Vade Dağılımı",
                                  color_discrete_sequence=px.colors.sequential.Blues_r)
                    fig2.update_layout(height=400)
                    styled_chart(fig2, key="tcmb_vade_pie")

            # ═══════ TAB 6: ISIN ANALİZİ ═══════
            with tab6:
                st.subheader("🏷️ En Çok Kullanılan ISIN Analizi")

                isin_agg = raw.groupby("Tanım (ISIN)").agg(
                    ihale_sayisi=("Kazanan Tutar (Nominal)", "count"),
                    toplam_nom=("Kazanan Tutar (Nominal)", "sum"),
                    toplam_net=("Kazanan Tutar (Net)", "sum"),
                    ilk_islem=("İşlem Tarihi", "min"),
                    son_islem=("İşlem Tarihi", "max"),
                    vade_tarihi=("Vade", "max"),
                ).reset_index().sort_values("ihale_sayisi", ascending=False).head(50)

                isin_show = isin_agg.copy()
                isin_show.columns = ["ISIN Kodu", "İhale Sayısı", "Toplam Nominal (Bin TL)",
                                     "Toplam Net (Bin TL)", "İlk İşlem", "Son İşlem", "Vade Tarihi"]
                for c in ["İlk İşlem", "Son İşlem", "Vade Tarihi"]:
                    isin_show[c] = isin_show[c].dt.strftime("%d.%m.%Y")
                for c in ["Toplam Nominal (Bin TL)", "Toplam Net (Bin TL)"]:
                    isin_show[c] = isin_show[c].apply(lambda x: f"{x:,.0f}")
                isin_show["İhale Sayısı"] = isin_show["İhale Sayısı"].apply(lambda x: f"{x:,}")
                st.dataframe(isin_show, use_container_width=True, height=500, hide_index=True)

                # Top 15 ISIN grafiği
                top15 = isin_agg.head(15).copy()
                fig = px.bar(top15, x="Tanım (ISIN)", y="toplam_nom",
                             title="En Çok İşlem Gören 15 ISIN (Nominal Tutar)",
                             color_discrete_sequence=["#1F4E79"], template="plotly_white")
                fig.update_layout(height=400, xaxis_tickangle=-45,
                                  xaxis_title="ISIN", yaxis_title="Toplam Nominal (Bin TL)")
                styled_chart(fig, key="tcmb_isin_bar")

            # ═══════ TAB 7: AYLIK FAİZ TRENDİ ═══════
            with tab7:
                st.subheader("📈 Aylık Ağırlıklı Ortalama Faiz Trendi")

                aylik = raw.groupby(["Yıl", "Ay"]).agg(
                    ihale_sayisi=("Kazanan Tutar (Nominal)", "count"),
                    toplam_nom=("Kazanan Tutar (Nominal)", "sum"),
                ).reset_index()

                wf_ay_basit = raw.groupby(["Yıl", "Ay"]).apply(
                    lambda g: wavg(g, "Ortalama Basit Faiz", "Kazanan Tutar (Nominal)")
                ).reset_index()
                wf_ay_basit.columns = ["Yıl", "Ay", "ort_basit"]

                wf_ay_bilesik = raw.groupby(["Yıl", "Ay"]).apply(
                    lambda g: wavg(g, "Ortalama Bileşik Faiz", "Kazanan Tutar (Nominal)")
                ).reset_index()
                wf_ay_bilesik.columns = ["Yıl", "Ay", "ort_bilesik"]

                aylik = aylik.merge(wf_ay_basit, on=["Yıl", "Ay"]).merge(wf_ay_bilesik, on=["Yıl", "Ay"])
                aylik["Dönem"] = aylik["Yıl"].astype(str) + "-" + aylik["Ay"].astype(str).str.zfill(2)

                aylik_show = aylik[["Dönem", "ihale_sayisi", "toplam_nom", "ort_basit", "ort_bilesik"]].copy()
                aylik_show.columns = ["Dönem", "İhale Sayısı", "Toplam Nominal (Bin TL)",
                                      "Ağ.Ort. Basit Faiz (%)", "Ağ.Ort. Bileşik Faiz (%)"]
                for c in aylik_show.columns:
                    if c == "Dönem":
                        continue
                    elif "Faiz" in c:
                        aylik_show[c] = aylik_show[c].apply(lambda x: f"{x:,.2f}" if pd.notna(x) else "-")
                    elif "Sayısı" in c:
                        aylik_show[c] = aylik_show[c].apply(lambda x: f"{int(x):,}" if pd.notna(x) else "-")
                    else:
                        aylik_show[c] = aylik_show[c].apply(lambda x: f"{x:,.0f}" if pd.notna(x) else "-")
                st.dataframe(aylik_show, use_container_width=True, height=500, hide_index=True)

                # Faiz trendi grafiği
                fig = go.Figure()
                fig.add_trace(go.Scatter(x=aylik["Dönem"], y=aylik["ort_basit"],
                                         name="Basit Faiz (%)", mode="lines",
                                         line=dict(color="#1F4E79")))
                fig.add_trace(go.Scatter(x=aylik["Dönem"], y=aylik["ort_bilesik"],
                                         name="Bileşik Faiz (%)", mode="lines",
                                         line=dict(color="#E74C3C")))
                fig.update_layout(title="Aylık Ağırlıklı Ortalama Faiz Trendi",
                                  xaxis_title="Dönem", yaxis_title="Faiz (%)",
                                  template="plotly_white", height=450,
                                  xaxis_tickangle=-45,
                                  legend=dict(orientation="h", y=-0.2))
                styled_chart(fig, key="tcmb_aylik_faiz")

                # Hacim ve Faiz dual-axis
                fig2 = go.Figure()
                fig2.add_trace(go.Bar(x=aylik["Dönem"], y=aylik["toplam_nom"] / 1e6,
                                      name="Borçlanma (Milyar TL)", marker_color="#1F4E79", yaxis="y"))
                fig2.add_trace(go.Scatter(x=aylik["Dönem"], y=aylik["ort_bilesik"],
                                          name="Bileşik Faiz (%)", mode="lines+markers",
                                          marker_color="#E74C3C", yaxis="y2"))
                fig2.update_layout(
                    title="Aylık Borçlanma Hacmi ve Faiz",
                    yaxis=dict(title="Milyar TL", side="left"),
                    yaxis2=dict(title="Bileşik Faiz (%)", side="right", overlaying="y"),
                    template="plotly_white", height=450,
                    xaxis_tickangle=-45,
                    legend=dict(orientation="h", y=-0.2),
                )
                styled_chart(fig2, key="tcmb_hacim_faiz")

            # ═══════ TAB 8: İNDİR ═══════
            with tab8:
                get_download_button(str(alim_file), "📥 TCMB Doğrudan Alım Excel İndir")

    else:
        st.info("Veri dosyası bulunamadı. 'Verileri Güncelle' butonuna tıklayın.")


# ══════════════════════════════════════════════════════════
# BDDK BANKACILIK VERİLERİ
# ══════════════════════════════════════════════════════════

elif selected == "bddk":
    st.markdown('<div class="main-header">BDDK Bankacılık Verileri</div>', unsafe_allow_html=True)

    scraper_dir = BASE_DIR / "bddk veri çekme"     # yerel scraper klasörü (Selenium)
    published_dir = BASE_DIR / "bddk_data"          # buluta yayınlanan kopya (repoda)
    script_tl = scraper_dir / "enhanced_manual_scraper.py"
    script_usd = scraper_dir / "enhanced_manual_scraperUSD.py"
    publish_script = BASE_DIR / "bddk_yayinla.py"
    is_local = script_tl.exists()                   # scraper varsa yerel sürümdeyiz

    # Görüntü kaynağı: yerelde tam scraper klasörü, bulutta yayınlanan snapshot
    source_dir = scraper_dir if is_local else published_dir

    def load_bddk_files():
        if not source_dir.exists():
            return []
        return sorted(
            list(source_dir.glob("bddk_*.xlsx")) + list(source_dir.glob("bddk_*.xls")),
            key=lambda x: x.stat().st_mtime, reverse=True,
        )

    new_file_path = None

    if is_local:
        st.info("BDDK gelişmiş bülteninden haftalık bankacılık verilerini çekin (TL/USD). "
                "Çektikten sonra **☁️ Buluta Yayınla** ile iş arkadaşlarınızın bulut sürümünde de görünür olur.")
        col_b1, col_b2, col_b3, col_b4 = st.columns(4)
        with col_b1:
            run_tl = st.button("🔄 TL Çek", key="bddk_tl", use_container_width=True)
        with col_b2:
            run_usd = st.button("🔄 USD Çek", key="bddk_usd", use_container_width=True)
        with col_b3:
            publish = st.button("☁️ Buluta Yayınla", key="bddk_publish", use_container_width=True)
        with col_b4:
            if st.button("📂 Klasörü Aç", key="bddk_open_folder", use_container_width=True):
                try:
                    subprocess.run(["open", str(scraper_dir)])
                    st.toast("Klasör açıldı", icon="📂")
                except Exception as e:
                    st.error(f"Açılamadı: {e}")

        if run_tl or run_usd:
            files_before = set(f.name for f in load_bddk_files())
            if run_script(str(script_tl) if run_tl else str(script_usd), timeout_sec=600):
                st.cache_data.clear()
                new_files = [f for f in load_bddk_files() if f.name not in files_before]
                if new_files:
                    new_file_path = new_files[0]
                    st.success(f"✅ Yeni dosya oluşturuldu: **{new_file_path.name}**")

        if publish:
            if run_script(str(publish_script), timeout_sec=180):
                st.cache_data.clear()
                st.success("☁️ Buluta yayınlandı! Bulut sürümü birkaç dakika içinde güncellenecek.")
    else:
        st.info("Haftalık BDDK bankacılık verileri (TL/USD). Veriler yerelde çekilip buraya yayınlanır; "
                "aşağıda en güncel kopya listelenir.")

    # Mevcut Excel dosyaları
    bddk_files = load_bddk_files()

    if bddk_files:
        st.markdown(
            f'<div class="update-info">Son veri: '
            f'{get_file_mod_time(bddk_files[0])} &nbsp;·&nbsp; '
            f'Toplam {len(bddk_files)} dosya</div>',
            unsafe_allow_html=True,
        )

    # ── Haftalık Analiz Tabloları (ısı haritası) ──
    # Excel indirmeye gerek kalmadan iki özet tablo burada üretilir.
    import bddk_analiz as _ba

    @st.cache_data
    def _bddk_tablolar(src_dir, cache_key):
        tl_b, usd_b, kaynak = _ba.load_latest(src_dir)
        if tl_b is None:
            return None
        t1, t2, son, onceki = _ba.hesapla(tl_b, usd_b)
        return {"t1": t1, "t2": t2, "son": son.strftime("%d.%m.%Y"),
                "onceki": onceki.strftime("%d.%m.%Y") if onceki is not None else "", "kaynak": kaynak}

    _bkey = max((f.stat().st_mtime for f in bddk_files), default=0)
    _sonuc = _bddk_tablolar(str(source_dir), _bkey) if bddk_files else None

    if _sonuc:
        def _bul(rows, lbl):
            for r in rows:
                if r["label"] == lbl:
                    return r
            return {}

        def _p(v):
            return "—" if v is None else f"{v:+.1f}%".replace(".", ",")

        _kr = _bul(_sonuc["t1"], "Toplam Krediler (TL)")
        _mv = _bul(_sonuc["t2"], "Toplam Mevduat (TL Cinsi)")
        _mk = _bul(_sonuc["t1"], "Toplam Menkul Değerler")
        st.info(
            f"📋 **{_sonuc['son']}** haftası: Toplam krediler (TL) haftalık **{_p(_kr.get('hafta'))}** "
            f"(YtD {_p(_kr.get('ytd'))}, yıllık {_p(_kr.get('yillik'))}); toplam mevduat **{_p(_mv.get('hafta'))}** "
            f"(YtD {_p(_mv.get('ytd'))}); menkul değerler {_p(_mk.get('hafta'))}. "
            f"Önceki hafta: {_sonuc['onceki']}."
        )

        def _tablo_goster(rows, baslik):
            st.subheader(baslik)
            st.caption(f"{_sonuc['son']} itibarıyla · Toplam: Milyon TL (USD satırları Milyon USD) · "
                       "Sağ 4 kolon: banka grubu bazında haftalık değişim")
            cols = ["Kalem", "Toplam", "Haftalık", "YtD", "Yıllık",
                    "Kamu", "Yerli Özel", "Yabancı Özel", "Katılım"]
            num_map = {"Haftalık": "hafta", "YtD": "ytd", "Yıllık": "yillik",
                       "Kamu": "kamu", "Yerli Özel": "yerli",
                       "Yabancı Özel": "yabanci", "Katılım": "katilim"}

            def _fmt_top(v):
                return "—" if v is None else f"{v:,.0f}".replace(",", ".")

            def _fmt_pct(v):
                return "—" if v is None else f"{v:.1f}%".replace(".", ",")

            disp = pd.DataFrame({
                "Kalem": [" " * 4 * r["indent"] + r["label"] for r in rows],
                "Toplam": [_fmt_top(r["toplam"]) for r in rows],
                **{c: [_fmt_pct(r[k]) for r in rows] for c, k in num_map.items()},
            })[cols]

            def _stil(_):
                css = pd.DataFrame("", index=disp.index, columns=disp.columns)
                for c, k in num_map.items():
                    for i in disp.index:
                        rgb = _ba.renk(rows[i][k], _ba.OLCEK[k])
                        if rgb:
                            css.loc[i, c] = (f"background-color: rgb{rgb}; color:#111418; "
                                             "font-weight:600; text-align:center;")
                for i in disp.index:
                    if rows[i]["bold"]:
                        css.loc[i, "Kalem"] += "font-weight:700; color:#FFD28A;"
                        css.loc[i, "Toplam"] += "font-weight:700;"
                return css

            st.dataframe(disp.style.apply(_stil, axis=None), hide_index=True,
                         use_container_width=True, height=len(rows) * 35 + 42)

        _tablo_goster(_sonuc["t1"], "📊 Menkul Değerler · Krediler · Bankalardan Alacaklar")
        _tablo_goster(_sonuc["t2"], "💰 Mevduat · Diğer Bilanço Kalemleri")
        st.caption("Kaynak: BDDK Haftalık Bülten (gelişmiş) · YtD = önceki yıl sonuna göre, "
                   "Yıllık = 52 hafta öncesine göre değişim · KMH kalemleri çekilen veri setinde yok.")
    else:
        st.info("Analiz tabloları için önce veri çekilmeli (yerelde 'TL Çek' + 'USD Çek').")

    st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

    # İndirilen dosyalar listesi (isteğe bağlı — tablolar için indirme GEREKMEZ)
    if bddk_files:
        with st.expander(f"📁 Excel Dosyaları ({len(bddk_files)}) — indirme isteğe bağlı", expanded=False):
            st.caption("📥 'İndir' butonuna tıkladığınızda dosya tarayıcınızın varsayılan indirme klasörüne (genellikle `~/Downloads`) inecektir.")

            for idx, f in enumerate(bddk_files):
                col_d1, col_d2, col_d3 = st.columns([4, 2, 1])
                # En yeni dosya vurgulansın
                highlight = (new_file_path is not None and f.name == new_file_path.name)
                new_badge = ('<span style="background:#fef3c7; color:#92400e; padding:2px 8px; '
                             'border-radius:6px; font-size:0.7rem; font-weight:600; margin-right:8px;">YENİ</span>') if highlight else ''
                with col_d1:
                    # TL/USD ayrımı için renk
                    tag = ""
                    if "_TL_" in f.name:
                        tag = '<span style="background:#dbeafe; color:#1e40af; padding:2px 8px; border-radius:6px; font-size:0.7rem; font-weight:600; margin-right:8px;">TL</span>'
                    elif "_USD_" in f.name:
                        tag = '<span style="background:#dcfce7; color:#166534; padding:2px 8px; border-radius:6px; font-size:0.7rem; font-weight:600; margin-right:8px;">USD</span>'
                    size_kb = f.stat().st_size / 1024
                    st.markdown(
                        f'{new_badge}{tag}<span style="font-family:monospace; font-size:0.85rem;">{f.name}</span> '
                        f'<span style="color:#94a3b8; font-size:0.75rem;">({size_kb:.0f} KB)</span>',
                        unsafe_allow_html=True,
                    )
                with col_d2:
                    st.markdown(
                        f'<span style="color:#64748b; font-size:0.85rem;">{get_file_mod_time(f)}</span>',
                        unsafe_allow_html=True,
                    )
                with col_d3:
                    get_download_button(str(f), "📥 İndir")
    else:
        st.info("Henüz BDDK verisi çekilmemiş. Yukarıdaki butonlarla veri çekebilirsiniz.")


# ══════════════════════════════════════════════════════════
# CARI AÇIK (EVDS)
# ══════════════════════════════════════════════════════════

elif selected == "cari_acik":
    st.markdown('<div class="main-header">Cari Açık Verileri (TCMB EVDS)</div>', unsafe_allow_html=True)

    cari_dir = BASE_DIR / "cari acik"
    fetch_script = cari_dir / "evds_fetch.py"
    data_file = cari_dir / "cari_acik_son.xlsx"

    st.info(
        "TCMB EVDS sisteminden çeyreklik (Q) cari işlemler verileri. "
        "Veriler **milyon USD** cinsindendir (negatif = açık, pozitif = fazla)."
    )

    # Üst bar: güncelleme + dosya bilgisi
    col_u1, col_u2 = st.columns([1, 4])
    with col_u1:
        if st.button("🔄 Verileri Güncelle", key="cari_update", use_container_width=True):
            if run_script(str(fetch_script), timeout_sec=120):
                st.cache_data.clear()
    with col_u2:
        if data_file.exists():
            st.markdown(
                f'<div class="update-info">Son güncelleme: {get_file_mod_time(data_file)}</div>',
                unsafe_allow_html=True,
            )
        else:
            st.warning("Henüz veri çekilmemiş. Yandaki butona tıklayın.")

    if not data_file.exists():
        st.stop()

    # Veriyi yükle
    @st.cache_data
    def load_cari_data(path):
        df = pd.read_excel(path)
        # Tarih bileşenleri
        df["Yil"] = df["Tarih"].apply(lambda s: int(str(s).split("-")[0]))
        df["Ceyrek"] = df["Tarih"].apply(lambda s: str(s).split("-")[1])
        df["Sira"] = df["Yil"] + df["Ceyrek"].apply(lambda q: (int(q[1]) - 1) * 0.25)
        return df.sort_values("Sira").reset_index(drop=True)

    df = load_cari_data(str(data_file))

    # Excel'deki sütun adları ASCII (Cari Islemler Hesabi). Görüntüleme için Türkçe etiket eşleşmesi:
    COL_CIH = "Cari Islemler Hesabi"
    COL_FH = "Finans Hesabi"
    COL_NHN = "Net Hata ve Noksan"
    LABEL_CIH = "Cari İşlemler Hesabı"
    LABEL_FH = "Finans Hesabı"
    LABEL_NHN = "Net Hata ve Noksan"

    # ─── Üst KPI'lar (Aylık veri varsa onu kullan, yoksa çeyreklik) ───
    odm_raw_file = cari_dir / "odemeler_dengesi_raw.xlsx"

    AY_ADLARI = {1: "Ocak", 2: "Şubat", 3: "Mart", 4: "Nisan", 5: "Mayıs", 6: "Haziran",
                 7: "Temmuz", 8: "Ağustos", 9: "Eylül", 10: "Ekim", 11: "Kasım", 12: "Aralık"}

    if odm_raw_file.exists():
        # Aylık veriyi kullan - PDF tablosuyla tutarlı
        @st.cache_data
        def load_monthly_data(path):
            mdf = pd.read_excel(path)
            mdf["Tarih_dt"] = pd.to_datetime(mdf["Tarih"], format="%Y-%m", errors="coerce")
            mdf = mdf.sort_values("Tarih_dt").reset_index(drop=True)
            return mdf

        mdf = load_monthly_data(str(odm_raw_file))
        son_ay_dt = mdf["Tarih_dt"].iloc[-1]
        son_ay_label = f"{AY_ADLARI[son_ay_dt.month]} {son_ay_dt.year}"

        # Son ay
        son_ay_v = mdf["Cari İşlemler Dengesi"].iloc[-1]
        # Önceki yıl aynı ay
        onc_ay_dt = son_ay_dt.replace(year=son_ay_dt.year - 1)
        onc_match = mdf[mdf["Tarih_dt"] == onc_ay_dt]
        onc_ay_v = onc_match["Cari İşlemler Dengesi"].iloc[0] if not onc_match.empty else None

        # Son 12 ay
        son12_v = mdf.tail(12)["Cari İşlemler Dengesi"].sum()
        onc12_v = mdf.iloc[-24:-12]["Cari İşlemler Dengesi"].sum() if len(mdf) >= 24 else None

        # Bu yıl birikimli
        bu_yil = mdf[(mdf["Tarih_dt"].dt.year == son_ay_dt.year) &
                      (mdf["Tarih_dt"].dt.month <= son_ay_dt.month)]
        bu_yil_v = bu_yil["Cari İşlemler Dengesi"].sum()
        onc_yil = mdf[(mdf["Tarih_dt"].dt.year == son_ay_dt.year - 1) &
                       (mdf["Tarih_dt"].dt.month <= son_ay_dt.month)]
        onc_yil_v = onc_yil["Cari İşlemler Dengesi"].sum() if not onc_yil.empty else None

        col_k1, col_k2, col_k3 = st.columns(3)
        with col_k1:
            st.metric(
                f"Aylık ({son_ay_label})",
                f"{son_ay_v:,.0f} M$",
                f"{son_ay_v - onc_ay_v:,.0f}" if onc_ay_v is not None else None,
                delta_color="normal",
                help=f"{son_ay_label} dönemi cari işlemler dengesi",
            )
        with col_k2:
            st.metric(
                f"{son_ay_dt.year} Yılı (Ocak - {AY_ADLARI[son_ay_dt.month]})",
                f"{bu_yil_v:,.0f} M$",
                f"{bu_yil_v - onc_yil_v:,.0f}" if onc_yil_v is not None else None,
                delta_color="normal",
                help=f"{son_ay_dt.year} yılı yıl başından bu yana birikimli toplam",
            )
        with col_k3:
            st.metric(
                "Son 12 Ay",
                f"{son12_v:,.0f} M$",
                f"{son12_v - onc12_v:,.0f}" if onc12_v is not None else None,
                delta_color="normal",
                help=f"Son 12 ayın (yıllıklandırılmış) toplamı",
            )
    else:
        # Aylık veri henüz çekilmemişse, çeyreklik kullan
        son_satir = df.iloc[-1]
        onceki_satir = df.iloc[-2] if len(df) >= 2 else None
        son12 = df.tail(4)[COL_CIH].sum()
        onceki12 = df.iloc[-8:-4][COL_CIH].sum() if len(df) >= 8 else 0

        col_k1, col_k2, col_k3 = st.columns(3)
        with col_k1:
            st.metric(
                "Son Çeyrek",
                f"{son_satir[COL_CIH]:,.0f} M$",
                f"{son_satir[COL_CIH] - onceki_satir[COL_CIH]:,.0f}" if onceki_satir is not None else None,
                delta_color="normal",
            )
        with col_k2:
            st.metric(
                f"{int(df['Yil'].max())} Yılı",
                f"{df[df['Yil'] == df['Yil'].max()][COL_CIH].sum():,.0f} M$",
                delta_color="normal",
            )
        with col_k3:
            st.metric(
                "Son 12 Ay",
                f"{son12:,.0f} M$",
                f"{son12 - onceki12:,.0f}" if onceki12 else None,
                delta_color="normal",
            )
        st.caption("ℹ️ Aylık veri için 'Aylık Bülten Tablosu' sekmesindeki butonu kullanın.")

    # Çeyreklik trend tabı için her halükarda hazır olsun
    son_satir = df.iloc[-1]
    son12 = df.tail(4)[COL_CIH].sum()

    st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

    # ─── Tablar ───
    tab_pdf, tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "📰 Aylık Bülten Tablosu",
        "📊 Çeyreklik Trend",
        "📅 Yıllık Toplam",
        "🔄 12 Ay Hareketli",
        "📋 Veri Tablosu",
        "📥 İndir",
    ])

    # ─── 0) PDF Bülten Tablosu ───
    with tab_pdf:
        st.subheader("Tablo 1. Cari İşlemler Dengesi (Milyon $)")
        odm_script = cari_dir / "odemeler_dengesi_fetch.py"
        odm_table_file = cari_dir / "odemeler_dengesi_tablo.xlsx"
        odm_raw_file = cari_dir / "odemeler_dengesi_raw.xlsx"

        col_pdf1, col_pdf2 = st.columns([1, 4])
        with col_pdf1:
            if st.button("🔄 Aylık Veriyi Çek", key="cari_aylik_update", use_container_width=True):
                if run_script(str(odm_script), timeout_sec=120):
                    st.cache_data.clear()
        with col_pdf2:
            if odm_table_file.exists():
                st.markdown(
                    f'<div class="update-info">Son güncelleme: {get_file_mod_time(odm_table_file)} '
                    f'&nbsp;·&nbsp; Kaynak: TCMB EVDS · bie_odeayrsunum6 (aylık)</div>',
                    unsafe_allow_html=True,
                )
            else:
                st.warning("Henüz aylık tablo verisi çekilmemiş. Yandaki butona tıklayın.")

        if odm_table_file.exists():
            tablo_df = pd.read_excel(odm_table_file)

            # Sayısal sütunları formatla: 268928 → "268.928"
            def fmt_num(v):
                if pd.isna(v):
                    return "—"
                try:
                    return f"{int(round(float(v))):,}".replace(",", ".")
                except (ValueError, TypeError):
                    return "—"

            num_cols = [c for c in tablo_df.columns if c not in ("Kalem", "_level")]
            display_df = tablo_df.copy()
            for c in num_cols:
                display_df[c] = display_df[c].apply(fmt_num)

            # Sade tablo stilleri (renksiz)
            def row_style(level, is_special=False):
                if is_special:
                    return 'font-weight:700; color:#0f172a; border-top:1px solid #cbd5e1;'
                if level == 0:
                    return 'font-weight:700; color:#0f172a;'
                elif level == 1:
                    return 'font-weight:600; color:#0f172a;'
                elif level == 2:
                    return 'font-weight:500; color:#1e293b;'
                return 'font-weight:400; color:#475569;'

            def cell_color(val_str, level):
                if not isinstance(val_str, str) or val_str == "—":
                    return "color:#cbd5e1;"
                weight = "700" if level <= 1 else ("500" if level == 2 else "400")
                return f"color:#0f172a; font-weight:{weight}; font-variant-numeric: tabular-nums;"

            display_cols = [c for c in display_df.columns if c != "_level"]

            # Custom CSS for the table
            html = ['''<style>
            .pdf-table {
                width: 100%;
                border-collapse: separate;
                border-spacing: 0;
                font-family: 'Inter', -apple-system, sans-serif;
                font-size: 0.78rem;
                background: white;
                border-radius: 10px;
                overflow: hidden;
                box-shadow: 0 1px 3px rgba(0,0,0,0.05), 0 4px 12px rgba(0,0,0,0.04);
                border: 1px solid #e2e8f0;
            }
            .pdf-table thead th {
                padding: 8px 10px;
                font-weight: 600;
                font-size: 0.72rem;
                color: #475569;
                background: #f8fafc;
                border-bottom: 1px solid #cbd5e1;
                white-space: nowrap;
            }
            .pdf-table thead th.col-name { text-align: left; }
            .pdf-table thead th.col-num { text-align: right; }
            .pdf-table tbody tr { transition: background 0.15s; }
            .pdf-table tbody tr:hover { background: #fafbfc !important; }
            .pdf-table td {
                padding: 5px 10px;
                border-bottom: 1px solid #f1f5f9;
                line-height: 1.3;
            }
            .pdf-table td.label { text-align: left; white-space: nowrap; }
            .pdf-table td.num { text-align: right; white-space: nowrap; }
            .pdf-table tbody tr:last-child td { border-bottom: none; }
            </style>''']

            html.append('<div style="overflow-x:auto; margin:0.5rem 0 1rem;">')
            html.append('<table class="pdf-table">')
            html.append('<thead><tr>')
            for col in display_cols:
                cls = "col-name" if col == "Kalem" else "col-num"
                html.append(f'<th class="{cls}">{col}</th>')
            html.append('</tr></thead><tbody>')

            for _, row in display_df.iterrows():
                level = int(row.get("_level", 0)) if pd.notna(row.get("_level", 0)) else 0
                label = str(row["Kalem"])
                # Özel satırı tespit et
                is_special = "Altın hariç" in label and "Cari" in label
                style = row_style(level, is_special)
                html.append(f'<tr style="{style}">')
                for col in display_cols:
                    val = str(row[col])
                    if col == "Kalem":
                        # Level 0 → sola dayalı (4px), her alt seviye 20px ek girinti
                        if level == 0:
                            indent_px = 4
                        else:
                            indent_px = 4 + level * 20
                        html.append(f'<td class="label" style="padding-left:{indent_px}px;">{val}</td>')
                    else:
                        cell_st = cell_color(val, level)
                        html.append(f'<td class="num" style="{cell_st}">{val}</td>')
                html.append('</tr>')
            html.append('</tbody></table></div>')
            st.markdown('\n'.join(html), unsafe_allow_html=True)

            st.caption(
                "** Birincil ve ikincil gelir dengelerinin toplamıdır. "
                "Negatif değerler açık, pozitif değerler fazlayı gösterir. "
                "Değerler milyon USD cinsindendir."
            )

            with st.expander("📥 Tablo ve Ham Veriyi İndir"):
                col_dl1, col_dl2 = st.columns(2)
                with col_dl1:
                    get_download_button(str(odm_table_file), "📥 Tablo (.xlsx)")
                with col_dl2:
                    if odm_raw_file.exists():
                        get_download_button(str(odm_raw_file), "📥 Ham Veri (.xlsx)")

    # ─── 1) Çeyreklik Trend ───
    with tab1:
        st.subheader("Çeyreklik Cari Açık / Finans Hesabı / Net Hata-Noksan")

        fig = go.Figure()
        fig.add_trace(go.Bar(
            x=df["Tarih"], y=df[COL_CIH],
            name=LABEL_CIH,
            marker_color=["#dc2626" if v < 0 else "#16a34a" for v in df[COL_CIH]],
        ))
        fig.add_trace(go.Scatter(
            x=df["Tarih"], y=df[COL_FH],
            name=LABEL_FH, mode="lines+markers",
            line=dict(color="#3b82f6", width=2),
            marker=dict(size=6),
        ))
        fig.add_trace(go.Scatter(
            x=df["Tarih"], y=df[COL_NHN],
            name=LABEL_NHN, mode="lines",
            line=dict(color="#f59e0b", width=2, dash="dot"),
        ))
        fig.add_hline(y=0, line_color="#94a3b8", line_width=1)
        fig.update_layout(
            height=480,
            barmode="overlay",
            xaxis_title="Dönem",
            yaxis_title="Milyon USD",
            hovermode="x unified",
        )
        styled_chart(fig, key="cari_quarterly")

        with st.expander("📖 Yorum"):
            son_caa = df.iloc[-1][COL_CIH]
            yorum = "**fazla**" if son_caa > 0 else "**açık**"
            st.markdown(
                f"- **{son_satir['Tarih']}** döneminde cari işlemler hesabı "
                f"**{son_caa:,.0f} milyon USD {yorum}** vermiştir.\n"
                f"- Son 4 çeyrek toplamı: **{son12:,.0f} milyon USD**\n"
                f"- Tüm dönem ortalaması: **{df[COL_CIH].mean():,.0f} M$**"
            )

    # ─── 2) Yıllık Toplam ───
    with tab2:
        st.subheader("Yıllık Toplam Cari İşlemler Hesabı")
        yillik = df.groupby("Yil").agg({
            COL_CIH: "sum",
            COL_FH: "sum",
            COL_NHN: "sum",
        }).reset_index()

        fig = go.Figure()
        fig.add_trace(go.Bar(
            x=yillik["Yil"], y=yillik[COL_CIH],
            name="Cari İşlemler",
            marker_color=["#dc2626" if v < 0 else "#16a34a" for v in yillik[COL_CIH]],
            text=yillik[COL_CIH].apply(lambda v: f"{v:,.0f}"),
            textposition="outside",
        ))
        fig.add_hline(y=0, line_color="#94a3b8")
        fig.update_layout(
            height=480,
            xaxis_title="Yıl",
            yaxis_title="Milyon USD",
        )
        styled_chart(fig, key="cari_yearly")

        # Yıllık karşılaştırma tablosu
        st.markdown("**Yıllık Toplamlar Tablosu**")
        yillik_disp = yillik.copy()
        yillik_disp = yillik_disp.rename(columns={
            COL_CIH: LABEL_CIH,
            COL_FH: LABEL_FH,
            COL_NHN: LABEL_NHN,
            "Yil": "Yıl",
        })
        for col in [LABEL_CIH, LABEL_FH, LABEL_NHN]:
            yillik_disp[col] = yillik_disp[col].apply(lambda v: f"{v:,.0f}")
        st.dataframe(yillik_disp, use_container_width=True, hide_index=True)

    # ─── 3) 12 Ay Hareketli (Yıllıklandırılmış) ───
    with tab3:
        st.subheader("12 Aylık Hareketli Toplam (Yıllıklandırılmış)")
        df_ma = df.copy()
        df_ma["12_Ay_CIH"] = df_ma[COL_CIH].rolling(window=4, min_periods=4).sum()
        df_ma["12_Ay_FH"] = df_ma[COL_FH].rolling(window=4, min_periods=4).sum()

        fig = go.Figure()
        fig.add_trace(go.Scatter(
            x=df_ma["Tarih"], y=df_ma["12_Ay_CIH"],
            name="12 Ay CİH (Cari İşlemler)",
            mode="lines",
            fill="tozeroy",
            line=dict(color="#dc2626", width=2.5),
            fillcolor="rgba(220,38,38,0.1)",
        ))
        fig.add_trace(go.Scatter(
            x=df_ma["Tarih"], y=df_ma["12_Ay_FH"],
            name="12 Ay Finans Hesabı",
            mode="lines",
            line=dict(color="#3b82f6", width=2),
        ))
        fig.add_hline(y=0, line_color="#94a3b8")
        fig.update_layout(
            height=480,
            xaxis_title="Dönem",
            yaxis_title="Milyon USD (12 ay yıllık)",
            hovermode="x unified",
        )
        styled_chart(fig, key="cari_ma12")

        st.caption("Hareketli toplam = son 4 çeyreğin (12 ay) kümülatif değeri")

    # ─── 4) Veri Tablosu ───
    with tab4:
        st.subheader("Tüm Dönemler")
        col_f1, col_f2 = st.columns(2)
        with col_f1:
            yillar = sorted(df["Yil"].unique(), reverse=True)
            secili_yillar = st.multiselect("Yıl filtresi", yillar, default=yillar[:5])
        with col_f2:
            ceyrek_filt = st.multiselect("Çeyrek filtresi", ["Q1", "Q2", "Q3", "Q4"],
                                          default=["Q1", "Q2", "Q3", "Q4"])

        df_filt = df[df["Yil"].isin(secili_yillar) & df["Ceyrek"].isin(ceyrek_filt)]
        disp = df_filt[["Tarih", COL_CIH, COL_FH, COL_NHN]].copy()
        disp = disp.rename(columns={COL_CIH: LABEL_CIH, COL_FH: LABEL_FH, COL_NHN: LABEL_NHN})
        for c in [LABEL_CIH, LABEL_FH, LABEL_NHN]:
            disp[c] = disp[c].apply(lambda v: f"{v:,.0f}" if pd.notna(v) else "-")

        st.dataframe(disp, use_container_width=True, hide_index=True, height=500)
        st.caption(f"Gösterilen: {len(df_filt)} kayıt / Toplam: {len(df)} kayıt")

    # ─── 5) İndir ───
    with tab5:
        st.subheader("Veriyi İndir")
        col_d1, col_d2 = st.columns(2)
        with col_d1:
            get_download_button(str(data_file), "📥 Excel (.xlsx)")
        with col_d2:
            csv_file = cari_dir / "cari_acik_son.csv"
            if csv_file.exists():
                get_download_button(str(csv_file), "📥 CSV")

        st.markdown("**Seri Bilgileri:**")
        st.markdown(
            "- `TP.ODEAYRSUNUM6.Q1` → Cari İşlemler Hesabı\n"
            "- `TP.ODEAYRSUNUM6.Q101` → Finans Hesabı\n"
            "- `TP.ODEAYRSUNUM6.Q210` → Net Hata ve Noksan\n"
            "- **Frekans:** Çeyreklik (3 aylık)\n"
            "- **Birim:** Milyon USD\n"
            "- **Kaynak:** TCMB EVDS API"
        )


# ══════════════════════════════════════════════════════════
# TCMB NET REZERV
# ══════════════════════════════════════════════════════════

elif selected == "net_rezerv":
    st.markdown('<div class="main-header">TCMB Uluslararası Rezervler</div>', unsafe_allow_html=True)

    rez_dir = BASE_DIR / "net rezerv"
    fetch_script = rez_dir / "net_rezerv_fetch.py"
    likidite_script = rez_dir / "likidite_fetch.py"
    data_file = rez_dir / "net_rezerv.xlsx"
    likidite_file = rez_dir / "likidite.xlsx"

    col_u1, col_u2 = st.columns([1.4, 4])
    with col_u1:
        if st.button("🔄 Güncelle (yerel + bulut)", key="rez_update", use_container_width=True):
            ok1 = run_script(str(fetch_script), timeout_sec=120)
            ok2 = run_script(str(likidite_script), timeout_sec=120)
            if ok1 or ok2:
                st.cache_data.clear()
                if _try_publish(["net rezerv/net_rezerv.xlsx", "net rezerv/likidite.xlsx"],
                                "Rezerv verisi guncellendi (otomatik yayin)"):
                    st.success("Güncellendi ve buluta yayınlandı ✓")
                else:
                    st.info("Yerel veri güncellendi. (Buluta yayın yalnızca yerel bilgisayardan yapılır.)")
    with col_u2:
        if data_file.exists():
            st.markdown(
                f'<div class="update-info">Son güncelleme: {get_file_mod_time(data_file)}</div>',
                unsafe_allow_html=True,
            )
        else:
            st.warning("Henüz veri çekilmemiş. Yandaki butona tıklayın.")

    if not data_file.exists():
        st.stop()

    @st.cache_data
    def load_rezerv(path, cache_key):
        d = pd.read_excel(path)
        d["tarih"] = pd.to_datetime(d["tarih"], errors="coerce")
        return d.dropna(subset=["tarih"]).sort_values("tarih").reset_index(drop=True)

    @st.cache_data
    def load_likidite(path, cache_key):
        d = pd.read_excel(path)
        d["tarih"] = pd.to_datetime(d["tarih"], errors="coerce")
        return d.dropna(subset=["tarih", "swap_toplam"]).sort_values("tarih").reset_index(drop=True)

    rez = load_rezerv(str(data_file), int(data_file.stat().st_mtime))
    if rez.empty:
        st.warning("Veri boş.")
        st.stop()
    if "net_ur" not in rez.columns:
        st.warning("Yeni rezerv serileri (Altın, Net UR) henüz çekilmemiş — '🔄 Güncelle'ye tıklayın.")
        st.stop()

    last = rez.iloc[-1]
    son_t = last["tarih"]

    # Türkçe milyar biçimi
    def _b1(v, sign=False):
        if v is None or pd.isna(v):
            return "—"
        fmt = "{:+,.1f}" if sign else "{:,.1f}"
        return fmt.format(v / 1000).replace(",", "\x00").replace(".", ",").replace("\x00", ".")

    # Haftalık (Cuma kapanışlı) değişimler — analist çerçevesi
    ri = rez.set_index("tarih")[["dis_varliklar", "altin", "brut_toplam", "net_ur"]]
    wd = ri.resample("W-FRI").last().dropna(how="all").diff()
    dW = wd.iloc[-1]
    ys = rez[rez["tarih"] >= pd.Timestamp(son_t.year, 1, 1)]
    ytd_nur = (float(last["net_ur"]) - float(ys.iloc[0]["net_ur"])) if len(ys) else None

    # Swap hariç: haftalık URDL likidite tablosuyla (II.2 + II.3)
    swap_haric = swap_tarih = swap_toplam = None
    if likidite_file.exists():
        try:
            lk = load_likidite(str(likidite_file), int(likidite_file.stat().st_mtime))
            if len(lk):
                Lk = lk.iloc[-1]
                es = rez[rez["tarih"] <= Lk["tarih"]]
                if len(es):
                    swap_toplam = float(Lk["swap_toplam"])
                    swap_haric = float(es.iloc[-1]["net_ur"]) + swap_toplam
                    swap_tarih = Lk["tarih"].strftime("%d.%m.%Y")
        except Exception:
            pass

    yon = "artışla" if dW["net_ur"] >= 0 else "azalışla"
    ozet = (f"📋 **{son_t.strftime('%d.%m.%Y')}** itibarıyla brüt döviz rezervleri "
            f"**{_b1(last['dis_varliklar'])} milyar USD** (haftalık {_b1(dW['dis_varliklar'], True)}), "
            f"altın **{_b1(last['altin'])} milyar USD** (haftalık {_b1(dW['altin'], True)}). "
            f"**Net uluslararası rezervler (swap dahil)** bir önceki haftaya göre "
            f"**{_b1(abs(dW['net_ur']))} milyar USD {yon} {_b1(last['net_ur'])} milyar USD** seviyesinde; "
            f"yıl başından beri {_b1(ytd_nur, True)} milyar USD.")
    if swap_haric is not None:
        ozet += (f" **Swap hariç net rezerv {_b1(swap_haric)} milyar USD** "
                 f"({swap_tarih} likidite tablosu).")
    st.info(ozet)

    st.markdown(f"#### {son_t.strftime('%d.%m.%Y')} itibarıyla  ·  USD/TRY = {last['usdtry']:.4f}  ·  Milyar USD")
    k1, k2, k3, k4, k5 = st.columns(5)
    with k1:
        st.metric("Brüt Döviz Rezervi", _b1(last["dis_varliklar"]), _b1(dW["dis_varliklar"], True))
    with k2:
        st.metric("Altın Rezervi", _b1(last["altin"]), _b1(dW["altin"], True))
    with k3:
        st.metric("Toplam Brüt Rezerv", _b1(last["brut_toplam"]), help="Döviz + altın.")
    with k4:
        st.metric("Net UR (Swap Dahil)", _b1(last["net_ur"]), _b1(dW["net_ur"], True),
                  help="TCMB haftalık vaziyet '2A Net Uluslararası Rezervler' kalemi (analitik bilanço).")
    with k5:
        st.metric("Net UR — Swap Hariç", "—" if swap_haric is None else _b1(swap_haric),
                  help=(f"Likidite tablosu ({swap_tarih}): Net UR − toplam swap/forward pozisyonu "
                        f"(URDL II.2 + II.3)." if swap_haric is not None
                        else "Likidite tablosu verisi için 'Güncelle'ye tıklayın."))
    st.caption("Δ = bir önceki Cuma kapanışına göre haftalık değişim (Milyar USD).")

    # ── Hesap kırılımı (formül şeffaflığı) ──
    if swap_haric is not None:
        with st.expander(f"🧮 Swap Hariç Hesabı ({swap_tarih} likidite tablosu)", expanded=False):
            es = rez[rez["tarih"] <= pd.to_datetime(swap_tarih, format="%d.%m.%Y")]
            nur_es = float(es.iloc[-1]["net_ur"]) if len(es) else None
            lk_son = load_likidite(str(likidite_file), int(likidite_file.stat().st_mtime)).iloc[-1]
            _f = lambda v: "—" if v is None or pd.isna(v) else f"{v:,.0f}".replace(",", ".")
            st.dataframe(pd.DataFrame([
                {"Kalem": "Net Uluslararası Rezervler (A20, aynı tarih)", "Değer (mn USD)": _f(nur_es)},
                {"Kalem": "II.2 Forward/Swap açık pozisyonları", "Değer (mn USD)": _f(lk_son.get("swap_forward"))},
                {"Kalem": "II.3 Diğer (repo vb.)", "Değer (mn USD)": _f(lk_son.get("diger"))},
                {"Kalem": "= Net UR — Swap Hariç", "Değer (mn USD)": _f(swap_haric)},
            ]), hide_index=True, use_container_width=True)

    # ── Grafikler ──
    st.subheader("Net Uluslararası Rezervler (Swap Dahil)")
    f1 = go.Figure()
    f1.add_scatter(x=rez["tarih"], y=rez["net_ur"] / 1000, mode="lines",
                   line=dict(color="#FF9E1B", width=2.5),
                   hovertemplate="%{x|%d.%m.%Y}<br>%{y:.1f} milyar USD<extra></extra>")
    f1.update_layout(height=400, separators=",.", showlegend=False)
    styled_chart(f1)

    st.subheader("Haftalık Değişim — Net UR (Milyar USD)")
    wser = wd["net_ur"].dropna() / 1000
    f2 = go.Figure()
    f2.add_bar(x=wser.index, y=wser.values,
               marker_color=["#FF5A5F" if v < 0 else "#26C281" for v in wser.values],
               hovertemplate="%{x|%d.%m.%Y} haftası<br>%{y:+.1f} milyar USD<extra></extra>")
    f2.update_layout(height=340, separators=",.", showlegend=False)
    styled_chart(f2)

    st.subheader("Brüt Rezervler — Döviz & Altın (Milyar USD)")
    f3 = go.Figure()
    f3.add_scatter(x=rez["tarih"], y=rez["dis_varliklar"] / 1000, mode="lines", name="Brüt Döviz",
                   line=dict(color="#4C9AFF", width=2.2),
                   hovertemplate="%{x|%d.%m.%Y}<br>%{y:.1f} milyar USD<extra>Döviz</extra>")
    f3.add_scatter(x=rez["tarih"], y=rez["altin"] / 1000, mode="lines", name="Altın",
                   line=dict(color="#FF9E1B", width=2),
                   hovertemplate="%{x|%d.%m.%Y}<br>%{y:.1f} milyar USD<extra>Altın</extra>")
    f3.update_layout(height=380, separators=",.", legend_title_text="")
    styled_chart(f3)

    col_d1, col_d2 = st.columns(2)
    with col_d1:
        get_download_button(str(data_file), "📥 Rezerv Verisi (.xlsx)")
    with col_d2:
        if likidite_file.exists():
            get_download_button(str(likidite_file), "📥 Likidite Tablosu (.xlsx)")

    st.markdown("**Seri Bilgileri:**")
    st.markdown(
        "- `TP.AB.A02` → Dış Varlıklar / brüt döviz (Bin TL)\n"
        "- `TP.AB.A18` → Altın (Bin TL)\n"
        "- `TP.AB.A20` → Net Uluslararası Rezervler, swap dahil (Bin TL — haftalık vaziyet '2A')\n"
        "- `TP.DK.USD.A.YTL` → USD/TRY alış kuru (günlük USD karşılıkları bu kurla)\n"
        "- **URDL haftalık şablonu** → toplam swap/forward pozisyonu (II.2 + II.3) — 'swap hariç' bu tabloyla hesaplanır\n"
        "- **Frekans:** Günlük (iş günü); likidite tablosu haftalık  ·  **Kaynak:** TCMB EVDS + TCMB URDL"
    )


# ══════════════════════════════════════════════════════════
# TÜFE ENFLASYON
# ══════════════════════════════════════════════════════════

elif selected == "enflasyon":
    st.markdown('<div class="main-header">TÜFE Enflasyon</div>', unsafe_allow_html=True)

    enf_dir = BASE_DIR / "enflasyon"
    fetch_script = enf_dir / "enflasyon_fetch.py"
    data_file = enf_dir / "enflasyon.xlsx"

    col_u1, col_u2 = st.columns([1, 4])
    with col_u1:
        if st.button("🔄 Verileri Güncelle", key="enf_update", use_container_width=True):
            if run_script(str(fetch_script), timeout_sec=120):
                st.cache_data.clear()
    with col_u2:
        if data_file.exists():
            st.markdown(
                f'<div class="update-info">Son güncelleme: {get_file_mod_time(data_file)}</div>',
                unsafe_allow_html=True,
            )
        else:
            st.warning("Henüz veri çekilmemiş. Yandaki butona tıklayın.")

    if not data_file.exists():
        st.stop()

    @st.cache_data
    def load_enflasyon(path):
        gen = pd.read_excel(path, sheet_name="Genel")
        gen["tarih"] = pd.to_datetime(gen["tarih"], errors="coerce")
        yil = pd.read_excel(path, sheet_name="AltKalem_Yillik")
        yil["tarih"] = pd.to_datetime(yil["tarih"], errors="coerce")
        end = pd.read_excel(path, sheet_name="AltKalem_Endeks")
        end["tarih"] = pd.to_datetime(end["tarih"], errors="coerce")
        ozet = pd.read_excel(path, sheet_name="AltKalem_Ozet")
        return (gen.dropna(subset=["tarih"]).reset_index(drop=True),
                yil.reset_index(drop=True), end.reset_index(drop=True), ozet)

    gen, alt_yillik, alt_endeks, ozet = load_enflasyon(str(data_file))
    if gen.empty:
        st.warning("Veri boş.")
        st.stop()

    last = gen.iloc[-1]
    prev = gen.iloc[-2] if len(gen) >= 2 else last
    AY = {1: "Ocak", 2: "Şubat", 3: "Mart", 4: "Nisan", 5: "Mayıs", 6: "Haziran",
          7: "Temmuz", 8: "Ağustos", 9: "Eylül", 10: "Ekim", 11: "Kasım", 12: "Aralık"}
    son_label = f"{AY[last['tarih'].month]} {last['tarih'].year}"

    # Türkçe 2 ondalık (1234.5 -> "1.234,50") ve alt kalem kısa etiketleri (COICOP 01..13)
    def _tr2(v):
        return "—" if pd.isna(v) else f"{v:,.2f}".replace(",", "\x00").replace(".", ",").replace("\x00", ".")
    _pct = lambda v: "—" if pd.isna(v) else f"{v:.2f}".replace(".", ",")
    _KISA = ["Gıda ve alkolsüz içecekler", "Alkollü içecekler ve tütün", "Giyim ve ayakkabı",
             "Konut", "Mobilya ve ev eşyası", "Sağlık", "Ulaştırma", "Bilgi ve iletişim",
             "Eğlence ve kültür", "Eğitim", "Lokanta ve konaklama",
             "Sigorta ve finansal hizmetler", "Çeşitli mal ve hizmetler"]
    alt_cols = [c for c in alt_yillik.columns if c != "tarih"]
    if len(alt_cols) == len(_KISA):
        label_map = dict(zip(alt_cols, _KISA))
    else:
        label_map = {c: re.sub(r"^\d+\.\s*", "", c) for c in alt_cols}
    _no_short = {f"{i+1:02d}": _KISA[i] for i in range(min(len(_KISA), 13))}

    def _kalem_kisa(row):
        return _no_short.get(str(row["No"]).split(".")[0].strip().zfill(2), row["Kalem"])

    # ── Dinamik özet (güncel duruma dair) ──
    _os = ozet.dropna(subset=["Yıllık %"]).sort_values("Yıllık %", ascending=False)
    _hi, _hi2, _lo = _os.iloc[0], _os.iloc[1], _os.iloc[-1]
    _d = last["yillik"] - prev["yillik"]
    _yon = "geriledi" if _d < 0 else ("yükseldi" if _d > 0 else "değişmedi")
    st.info(
        f"📊 **{son_label}** itibarıyla yıllık TÜFE enflasyonu **%{_pct(last['yillik'])}** — "
        f"bir önceki aya göre **{_pct(abs(_d))} puan {_yon}**. "
        f"Aylık artış **%{_pct(last['aylik'])}**, yılbaşından bu yana kümülatif **%{_pct(last['ytd'])}**; "
        f"genel endeks **{_pct(last['endeks'])}** (2025=100). "
        f"Ana harcama grupları içinde yıllık artışın en yükseği **{_kalem_kisa(_hi)}** "
        f"(%{_pct(_hi['Yıllık %'])}) ve **{_kalem_kisa(_hi2)}** (%{_pct(_hi2['Yıllık %'])}); "
        f"en düşüğü **{_kalem_kisa(_lo)}** (%{_pct(_lo['Yıllık %'])})."
    )

    st.markdown(f"#### {son_label} itibarıyla")

    k1, k2, k3, k4 = st.columns(4)
    with k1:
        st.metric("Yıllık Enflasyon", f"%{last['yillik']:.2f}",
                  f"{last['yillik'] - prev['yillik']:+.2f} puan",
                  delta_color="inverse",
                  help="12 ay öncesine göre. Δ = geçen aya göre oran değişimi (puan).")
    with k2:
        st.metric("Aylık Enflasyon", f"%{last['aylik']:.2f}",
                  f"{last['aylik'] - prev['aylik']:+.2f} puan",
                  delta_color="inverse",
                  help="Bir önceki aya göre. Δ = geçen ayki aylık orana göre değişim.")
    with k3:
        st.metric("Yılbaşından Beri", f"%{last['ytd']:.2f}",
                  help="Önceki yıl Aralık'a göre kümülatif.")
    with k4:
        st.metric("Endeks (2025=100)", f"{last['endeks']:.2f}",
                  f"{last['endeks'] - prev['endeks']:+.2f}",
                  help="TÜFE genel endeks düzeyi.")

    # ── Genel: yıllık enflasyon trendi ──
    st.subheader("Yıllık Enflasyon Trendi (Genel)")
    yil_min = int(gen["tarih"].dt.year.min())
    yil_max = int(gen["tarih"].dt.year.max())
    _default_start = max(yil_min, yil_max - 10)
    yr = st.slider("Başlangıç yılı", yil_min, yil_max, _default_start, key="enf_yr")
    gplot = gen[gen["tarih"].dt.year >= yr]
    fig = px.line(gplot, x="tarih", y="yillik",
                  labels={"tarih": "Tarih", "yillik": "Yıllık Enflasyon (%)"})
    fig.update_traces(line_color="#C0392B",
                      hovertemplate="%{x|%m.%Y}<br>Yıllık: %{y:.2f}<extra></extra>")
    fig.update_layout(height=380, template="plotly_white", separators=",.")
    styled_chart(fig)

    # ── Aylık enflasyon (son 24 ay) ──
    st.subheader("Aylık Enflasyon (Son 24 Ay)")
    g24 = gen.tail(24).copy()
    g24["ay_lbl"] = g24["tarih"].dt.strftime("%m.%Y")
    figm = px.bar(g24, x="ay_lbl", y="aylik",
                  labels={"ay_lbl": "Ay", "aylik": "Aylık Enflasyon (%)"},
                  color="aylik", color_continuous_scale="OrRd")
    figm.update_traces(texttemplate="%{y:.2f}", textposition="outside",
                       hovertemplate="%{x}<br>Aylık: %{y:.2f}<extra></extra>")
    figm.update_layout(height=380, template="plotly_white", coloraxis_showscale=False,
                       separators=",.")
    figm.update_xaxes(type="category")
    styled_chart(figm)

    # ── Alt kalemler: yıllık değişim (ay seçilebilir, TÜİK düzeni) ──
    st.subheader("Alt Kalemler — Yıllık Değişim (%)")
    _gsel = gen.dropna(subset=["yillik"]).copy()
    _gsel["lbl"] = _gsel["tarih"].apply(lambda d: f"{AY[d.month]} {d.year}")
    _labels_desc = _gsel["lbl"].tolist()[::-1]  # en yeni ay başta
    sel_lbl = st.selectbox("Ay seçin", _labels_desc, index=0, key="enf_ay_sec")
    sel_dt = _gsel[_gsel["lbl"] == sel_lbl]["tarih"].iloc[0]

    _gen_y = _gsel.loc[_gsel["tarih"] == sel_dt, "yillik"].iloc[0]
    _alt_row = alt_yillik[alt_yillik["tarih"] == sel_dt]
    cats = ["TÜFE"] + [label_map[c] for c in alt_cols]
    vals = [_gen_y] + [(_alt_row[c].iloc[0] if len(_alt_row) else None) for c in alt_cols]
    tip = ["TÜFE"] + ["Alt Kalem"] * len(alt_cols)
    bar_df = pd.DataFrame({"Kalem": cats, "Yıllık %": vals, "tip": tip})
    figh = px.bar(
        bar_df, x="Yıllık %", y="Kalem", orientation="h", color="tip",
        color_discrete_map={"TÜFE": "#C0392B", "Alt Kalem": "#2E6DB4"},
        text="Yıllık %", labels={"Yıllık %": "Yıllık Değişim (%)", "Kalem": ""},
    )
    figh.update_traces(texttemplate="%{x:.2f}", textposition="outside", cliponaxis=False,
                       hovertemplate="%{y}<br>%{x:.2f}<extra></extra>")
    figh.update_layout(height=560, template="plotly_white", separators=",.",
                       showlegend=False,
                       yaxis={"categoryorder": "array", "categoryarray": cats[::-1]})
    styled_chart(figh)

    # ── Alt kalem özet tablosu (seçili ay) ──
    with st.expander(f"📋 Alt Kalem Özet Tablosu — {sel_lbl}", expanded=False):
        _ei = alt_endeks.index[alt_endeks["tarih"] == sel_dt]
        _rows = []
        for i, c in enumerate(alt_cols):
            endeks_v = aylik_v = None
            if len(_ei):
                _e = _ei[0]
                endeks_v = alt_endeks.iloc[_e][c]
                if _e >= 1:
                    _pv = alt_endeks.iloc[_e - 1][c]
                    if pd.notna(_pv) and _pv:
                        aylik_v = (endeks_v / _pv - 1) * 100
            yillik_v = _alt_row[c].iloc[0] if len(_alt_row) else None
            _rows.append({"No": f"{i+1:02d}", "Kalem": label_map[c],
                          "Endeks": _tr2(endeks_v), "Aylık %": _tr2(aylik_v),
                          "Yıllık %": _tr2(yillik_v)})
        st.dataframe(pd.DataFrame(_rows), hide_index=True, use_container_width=True)

    # ── Alt kalem yıllık trend (seçmeli, Genel ile karşılaştır) ──
    st.subheader("Alt Kalem Yıllık Trendi")
    secili = st.multiselect("Kalem seçin (Genel ile karşılaştırın)", alt_cols,
                            default=alt_cols[:2], key="enf_kalem",
                            format_func=lambda c: label_map.get(c, c))
    comp = gen[["tarih", "yillik"]].rename(columns={"yillik": "Genel TÜFE"})
    merged = comp.copy()
    for k in secili:
        merged = merged.merge(
            alt_yillik[["tarih", k]].rename(columns={k: label_map.get(k, k)}),
            on="tarih", how="left")
    merged = merged[merged["tarih"].dt.year >= yr]
    long = merged.melt("tarih", var_name="Kalem", value_name="Yıllık %").dropna(subset=["Yıllık %"])
    figk = px.line(long, x="tarih", y="Yıllık %", color="Kalem", labels={"tarih": "Tarih"})
    figk.update_traces(hovertemplate="%{x|%m.%Y}<br>%{y:.2f}<extra></extra>")
    figk.update_layout(height=400, template="plotly_white", legend_title_text="", separators=",.")
    styled_chart(figk)

    get_download_button(str(data_file), "📥 Enflasyon Verisi (.xlsx)")

    st.markdown(
        "**Seri Bilgileri (EVDS):** `TP.TUKFIY2025.GENEL` (Genel) + `TP.TUKFIY2025.01…13` (alt kalemler)  ·  "
        "Baz **2025=100**  ·  Frekans **Aylık**  ·  Kaynak **TCMB EVDS**  ·  "
        "Alt kalem listesi `enflasyon alt kalemleri.xlsx` dosyasından okunur."
    )


# ══════════════════════════════════════════════════════════
# KREDİ FAİZLERİ
# ══════════════════════════════════════════════════════════

elif selected == "kredi":
    st.markdown('<div class="main-header">Kredi Faiz Oranları</div>', unsafe_allow_html=True)
    km_dir = BASE_DIR / "kredi mevduat"
    fetch_script = km_dir / "kredi_mevduat_fetch.py"
    data_file = km_dir / "kredi_mevduat.xlsx"

    _km_update_bar(fetch_script, data_file, "kredi_update")
    if not data_file.exists():
        st.stop()

    akim = _load_km_sheet(str(data_file), "Kredi_Akim")
    stok = _load_km_sheet(str(data_file), "Kredi_Stok")
    mev_akim = _load_km_sheet(str(data_file), "Mevduat_Akim")
    if akim.empty:
        st.warning("Veri boş.")
        st.stop()

    last = akim.iloc[-1]
    prev = akim.iloc[-2] if len(akim) >= 2 else last
    son_h = last["tarih"].strftime("%d.%m.%Y")
    ANA4 = ["İhtiyaç Kredisi", "Taşıt Kredisi", "Konut Kredisi", "Ticari Krediler"]

    # Dinamik özet
    _mev_last = mev_akim.iloc[-1]
    _makas = (last.get("Ticari Krediler") - _mev_last.get("Toplam")
              if "Ticari Krediler" in akim.columns and "Toplam" in mev_akim.columns else None)
    st.info(
        f"🏦 **{son_h}** haftası itibarıyla yeni açılan (akım) TL kredi faizleri: "
        f"İhtiyaç **%{_km_pct(last.get('İhtiyaç Kredisi'))}**, Taşıt **%{_km_pct(last.get('Taşıt Kredisi'))}**, "
        f"Konut **%{_km_pct(last.get('Konut Kredisi'))}**, Ticari **%{_km_pct(last.get('Ticari Krediler'))}**. "
        + (f"Ticari kredi–TL mevduat faiz makası **{_km_pct(_makas)} puan**." if _makas is not None else "")
    )

    st.markdown(f"#### Yeni Kredi Faizi (Akım) — {son_h} haftası")
    kcols = st.columns(len(ANA4))
    for col, name in zip(kcols, ANA4):
        with col:
            if name in akim.columns:
                d = last[name] - prev[name]
                st.metric(name, f"%{_km_pct(last[name])}",
                          f"{_km_pct(d)} puan" if pd.notna(d) else None, delta_color="off",
                          help="Yeni açılan kredilere uygulanan ortalama faiz (akım). Δ = geçen haftaya göre.")

    # Trend
    st.subheader("Kredi Faizi Trendi (Yeni / Akım)")
    kredi_cols = [c for c in akim.columns if c != "tarih"]
    sel = st.multiselect("Kredi türü", kredi_cols,
                         default=[c for c in ANA4 if c in kredi_cols], key="kredi_sel")
    yil_min, yil_max = int(akim["tarih"].dt.year.min()), int(akim["tarih"].dt.year.max())
    yr = st.slider("Başlangıç yılı", yil_min, yil_max, max(yil_min, yil_max - 5), key="kredi_yr")
    if sel:
        pl = akim[akim["tarih"].dt.year >= yr][["tarih"] + sel]
        long = pl.melt("tarih", var_name="Tür", value_name="Faiz").dropna(subset=["Faiz"])
        figt = px.line(long, x="tarih", y="Faiz", color="Tür", labels={"tarih": "Tarih", "Faiz": "Faiz (%)"})
        figt.update_traces(hovertemplate="%{x|%d.%m.%Y}<br>%{y:.2f}<extra></extra>")
        figt.update_layout(height=420, separators=",.", legend_title_text="")
        styled_chart(figt)

    # Akım vs Stok
    st.subheader("Yeni Kredi (Akım) vs Mevcut Portföy (Stok) Faizi")
    st.caption(f"Akım son hafta: {son_h}  ·  Stok son ay: {stok.iloc[-1]['tarih'].strftime('%m.%Y')}")
    stok_last = stok.iloc[-1]
    _cmp_cols = [c for c in (ANA4 + ["İhtiyaç Kredisi (KMH Dahil)"]) if c in akim.columns and c in stok.columns]
    crows = []
    for c in _cmp_cols:
        crows.append({"Tür": c, "Faiz": last[c], "Ölçüm": "Yeni (Akım)"})
        crows.append({"Tür": c, "Faiz": stok_last[c], "Ölçüm": "Mevcut (Stok)"})
    figc = px.bar(pd.DataFrame(crows), x="Faiz", y="Tür", color="Ölçüm", orientation="h", barmode="group",
                  color_discrete_map={"Yeni (Akım)": "#FF9E1B", "Mevcut (Stok)": "#4C9AFF"},
                  labels={"Faiz": "Faiz (%)", "Tür": ""}, text="Faiz")
    figc.update_traces(texttemplate="%{x:.2f}", textposition="outside", cliponaxis=False)
    figc.update_layout(height=380, separators=",.", legend_title_text="")
    styled_chart(figc)

    # Kredi-Mevduat makası
    if "Ticari Krediler" in akim.columns and "Toplam" in mev_akim.columns:
        st.subheader("Kredi–Mevduat Faiz Makası (Ticari Kredi − TL Mevduat Toplam)")
        mk = akim[["tarih", "Ticari Krediler"]].merge(mev_akim[["tarih", "Toplam"]], on="tarih", how="inner")
        mk["Makas"] = mk["Ticari Krediler"] - mk["Toplam"]
        mk = mk[mk["tarih"].dt.year >= yr]
        figm = px.area(mk, x="tarih", y="Makas", labels={"tarih": "Tarih", "Makas": "Makas (puan)"})
        figm.update_traces(line_color="#26C281", fillcolor="rgba(38,194,129,0.15)",
                           hovertemplate="%{x|%d.%m.%Y}<br>%{y:.2f} puan<extra></extra>")
        figm.update_layout(height=320, separators=",.")
        styled_chart(figm)

    # Tablo
    with st.expander("📋 Tüm Kredi Faizleri (son değerler)", expanded=False):
        trows = []
        for c in kredi_cols:
            trows.append({"Kredi Türü": c, "Yeni (Akım) %": _km_pct(last.get(c)),
                          "Mevcut (Stok) %": _km_pct(stok_last.get(c) if c in stok.columns else None)})
        st.dataframe(pd.DataFrame(trows), hide_index=True, use_container_width=True)

    get_download_button(str(data_file), "📥 Kredi & Mevduat Verisi (.xlsx)")
    st.markdown("**Kaynak:** TCMB EVDS  ·  Akım = yeni açılan kredi faizi (haftalık), "
                "Stok = mevcut portföy faizi (aylık)  ·  Seriler `kredi ve mevduat verileri.xlsx`'ten okunur.")


# ══════════════════════════════════════════════════════════
# MEVDUAT FAİZLERİ
# ══════════════════════════════════════════════════════════

elif selected == "mevduat":
    st.markdown('<div class="main-header">Mevduat Faiz Oranları</div>', unsafe_allow_html=True)
    km_dir = BASE_DIR / "kredi mevduat"
    fetch_script = km_dir / "kredi_mevduat_fetch.py"
    data_file = km_dir / "kredi_mevduat.xlsx"

    _km_update_bar(fetch_script, data_file, "mevduat_update")
    if not data_file.exists():
        st.stop()

    akim = _load_km_sheet(str(data_file), "Mevduat_Akim")
    stok = _load_km_sheet(str(data_file), "Mevduat_Stok")
    if akim.empty:
        st.warning("Veri boş.")
        st.stop()

    last = akim.iloc[-1]
    prev = akim.iloc[-2] if len(akim) >= 2 else last
    son_h = last["tarih"].strftime("%d.%m.%Y")
    VADELER = ["1 Aya Kadar Vadeli", "3 Aya Kadar Vadeli", "6 Aya Kadar Vadeli",
               "1 Yıla Kadar Vadeli", "1 Yıl ve Daha Uzun Vadeli"]

    st.info(
        f"💰 **{son_h}** haftası itibarıyla yeni açılan (akım) TL mevduat faizleri — "
        f"toplam **%{_km_pct(last.get('Toplam'))}**. Vadeye göre: 1 ay **%{_km_pct(last.get('1 Aya Kadar Vadeli'))}**, "
        f"3 ay **%{_km_pct(last.get('3 Aya Kadar Vadeli'))}**, 6 ay **%{_km_pct(last.get('6 Aya Kadar Vadeli'))}**, "
        f"1 yıl **%{_km_pct(last.get('1 Yıla Kadar Vadeli'))}**. "
        f"Döviz mevduatı: USD **%{_km_pct(last.get('Toplam (USD)'))}**, EUR **%{_km_pct(last.get('Toplam (EUR)'))}**."
    )

    st.markdown(f"#### Yeni TL Mevduat Faizi (Akım) — {son_h} haftası")
    kpi = ["1 Aya Kadar Vadeli", "3 Aya Kadar Vadeli", "6 Aya Kadar Vadeli", "1 Yıla Kadar Vadeli", "Toplam"]
    mcols = st.columns(len(kpi))
    for col, name in zip(mcols, kpi):
        with col:
            if name in akim.columns:
                d = last[name] - prev[name]
                st.metric(name.replace(" Kadar Vadeli", "").replace("1 Aya", "1 Ay").replace("3 Aya", "3 Ay").replace("6 Aya", "6 Ay").replace("1 Yıla", "1 Yıl"),
                          f"%{_km_pct(last[name])}",
                          f"{_km_pct(d)} puan" if pd.notna(d) else None, delta_color="off",
                          help="Yeni açılan TL mevduata verilen ortalama faiz (akım). Δ = geçen haftaya göre.")

    # Vade eğrisi (term structure)
    st.subheader("Mevduat Vade Eğrisi (Yeni / Akım, TL)")
    st.caption(f"{son_h} haftası — vadeye göre faiz")
    vcurve = [(v, last[v]) for v in VADELER if v in akim.columns and pd.notna(last[v])]
    if vcurve:
        cdf = pd.DataFrame(vcurve, columns=["Vade", "Faiz"])
        cdf["VadeK"] = cdf["Vade"].str.replace(" Kadar Vadeli", "", regex=False).str.replace("1 Yıl ve Daha Uzun Vadeli", "1 Yıl+", regex=False)
        figv = px.line(cdf, x="VadeK", y="Faiz", markers=True, labels={"VadeK": "Vade", "Faiz": "Faiz (%)"}, text="Faiz")
        figv.update_traces(line_color="#FF9E1B", texttemplate="%{y:.2f}", textposition="top center", cliponaxis=False)
        figv.update_layout(height=340, separators=",.")
        figv.update_xaxes(type="category")
        styled_chart(figv)

    # Trend TL vadeye göre
    st.subheader("Mevduat Faizi Trendi (Yeni / Akım)")
    mv_cols = [c for c in akim.columns if c != "tarih"]
    sel = st.multiselect("Kalem", mv_cols, default=[c for c in ["Toplam"] + VADELER[:4] if c in mv_cols], key="mev_sel")
    yil_min, yil_max = int(akim["tarih"].dt.year.min()), int(akim["tarih"].dt.year.max())
    yr = st.slider("Başlangıç yılı", yil_min, yil_max, max(yil_min, yil_max - 5), key="mev_yr")
    if sel:
        pl = akim[akim["tarih"].dt.year >= yr][["tarih"] + sel]
        long = pl.melt("tarih", var_name="Kalem", value_name="Faiz").dropna(subset=["Faiz"])
        figt = px.line(long, x="tarih", y="Faiz", color="Kalem", labels={"tarih": "Tarih", "Faiz": "Faiz (%)"})
        figt.update_traces(hovertemplate="%{x|%d.%m.%Y}<br>%{y:.2f}<extra></extra>")
        figt.update_layout(height=420, separators=",.", legend_title_text="")
        styled_chart(figt)

    # TL vs USD vs EUR (Toplam)
    st.subheader("Para Birimine Göre Toplam Mevduat Faizi")
    cur_map = {"Toplam": "TL", "Toplam (USD)": "USD", "Toplam (EUR)": "EUR"}
    have = {k: v for k, v in cur_map.items() if k in akim.columns}
    if have:
        cdf = akim[["tarih"] + list(have.keys())].rename(columns=have)
        cdf = cdf[cdf["tarih"].dt.year >= yr]
        long = cdf.melt("tarih", var_name="Para Birimi", value_name="Faiz").dropna(subset=["Faiz"])
        figx = px.line(long, x="tarih", y="Faiz", color="Para Birimi", labels={"tarih": "Tarih", "Faiz": "Faiz (%)"},
                       color_discrete_map={"TL": "#FF9E1B", "USD": "#4C9AFF", "EUR": "#26C281"})
        figx.update_traces(hovertemplate="%{x|%d.%m.%Y}<br>%{y:.2f}<extra></extra>")
        figx.update_layout(height=360, separators=",.", legend_title_text="")
        styled_chart(figx)

    with st.expander("📋 Tüm Mevduat Faizleri (son değerler)", expanded=False):
        stok_last = stok.iloc[-1]
        trows = []
        for c in mv_cols:
            trows.append({"Kalem": c, "Yeni (Akım) %": _km_pct(last.get(c)),
                          "Mevcut (Stok) %": _km_pct(stok_last.get(c) if c in stok.columns else None)})
        st.dataframe(pd.DataFrame(trows), hide_index=True, use_container_width=True)

    get_download_button(str(data_file), "📥 Kredi & Mevduat Verisi (.xlsx)")
    st.markdown("**Kaynak:** TCMB EVDS  ·  Akım = yeni açılan mevduat faizi (haftalık), "
                "Stok = mevcut mevduat faizi (aylık)  ·  Seriler `kredi ve mevduat verileri.xlsx`'ten okunur.")


# ══════════════════════════════════════════════════════════
# BÜTÇE DENGESİ (Merkezi Yönetim)
# ══════════════════════════════════════════════════════════

elif selected == "butce":
    st.markdown('<div class="main-header">Merkezi Yönetim Bütçe Dengesi</div>', unsafe_allow_html=True)

    butce_dir = BASE_DIR / "butce"
    fetch_script = butce_dir / "butce_fetch.py"
    data_file = butce_dir / "butce.xlsx"

    col_u1, col_u2 = st.columns([1.4, 4])
    with col_u1:
        if st.button("🔄 Güncelle (yerel + bulut)", key="butce_update", use_container_width=True):
            if run_script(str(fetch_script), timeout_sec=120):
                st.cache_data.clear()
                if _try_publish(["butce/butce.xlsx"], "Butce verisi guncellendi (otomatik yayin)"):
                    st.success("Güncellendi ve buluta yayınlandı ✓")
                else:
                    st.info("Yerel veri güncellendi. (Buluta yayın yalnızca yerel bilgisayardan yapılır.)")
    with col_u2:
        if data_file.exists():
            st.markdown(f'<div class="update-info">Son güncelleme: {get_file_mod_time(data_file)}</div>',
                        unsafe_allow_html=True)
        else:
            st.warning("Henüz veri çekilmemiş. Güncelle'ye tıklayın.")

    if not data_file.exists():
        st.stop()

    @st.cache_data
    def load_butce(path):
        d = pd.read_excel(path, sheet_name="Aylik")
        d["tarih"] = pd.to_datetime(d["tarih"], errors="coerce")
        return d.dropna(subset=["tarih"]).sort_values("tarih").reset_index(drop=True)

    b = load_butce(str(data_file))
    if b.empty:
        st.warning("Veri boş.")
        st.stop()

    AY = {1: "Ocak", 2: "Şubat", 3: "Mart", 4: "Nisan", 5: "Mayıs", 6: "Haziran",
          7: "Temmuz", 8: "Ağustos", 9: "Eylül", 10: "Ekim", 11: "Kasım", 12: "Aralık"}
    AY_KISA = ["Oca", "Şub", "Mar", "Nis", "May", "Haz", "Tem", "Ağu", "Eyl", "Eki", "Kas", "Ara"]
    last = b.iloc[-1]
    cy, cm = int(last["yil"]), int(last["ay"])
    son_label = f"{AY[cm]} {cy}"

    # Milyon TL -> Milyar TL, Türkçe 1 ondalık
    def _mr(v):
        return "—" if pd.isna(v) else f"{v/1000:,.1f}".replace(",", "\x00").replace(".", ",").replace("\x00", ".")

    ytd_now = b[(b["yil"] == cy) & (b["ay"] <= cm)]["denge"].sum()
    ytd_prev = b[(b["yil"] == cy - 1) & (b["ay"] <= cm)]["denge"].sum()

    _yon = "açık" if last["denge"] < 0 else "fazla"
    st.info(
        f"🏛️ **{son_label}** ayında merkezi yönetim bütçesi **{_mr(abs(last['denge']))} milyar TL {_yon}** verdi "
        f"(gelir {_mr(last['gelir'])}, gider {_mr(last['gider'])} milyar TL). "
        f"Faiz dışı denge **{_mr(last['faiz_disi_denge'])} milyar TL**, faiz gideri {_mr(last['faiz_gideri'])} milyar TL. "
        f"Yılbaşından beri kümülatif bütçe açığı **{_mr(abs(ytd_now))} milyar TL** — "
        f"geçen yıl aynı dönemde {_mr(abs(ytd_prev))} milyar TL idi."
    )

    st.markdown(f"#### {son_label} (Milyar TL)")
    k1, k2, k3, k4 = st.columns(4)
    with k1:
        st.metric("Bütçe Dengesi", _mr(last["denge"]), help="Gelir − Gider. Negatif = açık.")
    with k2:
        st.metric("Gelirler", _mr(last["gelir"]))
    with k3:
        st.metric("Giderler", _mr(last["gider"]))
    with k4:
        st.metric("Faiz Dışı Denge", _mr(last["faiz_disi_denge"]), help="Faiz gideri hariç bütçe dengesi.")

    yr = st.slider("Başlangıç yılı", int(b["yil"].min()), cy, max(int(b["yil"].min()), cy - 6), key="butce_yr")
    bp = b[b["yil"] >= yr].copy()

    # Aylık bütçe dengesi
    st.subheader("Aylık Bütçe Dengesi (Milyar TL)")
    bp["denge_mia"] = bp["denge"] / 1000
    bp["durum"] = bp["denge_mia"].apply(lambda v: "Fazla" if v >= 0 else "Açık")
    figd = px.bar(bp, x="tarih", y="denge_mia", color="durum",
                  color_discrete_map={"Açık": "#FF5A5F", "Fazla": "#26C281"},
                  labels={"tarih": "", "denge_mia": "Milyar TL", "durum": ""})
    figd.update_traces(hovertemplate="%{x|%m.%Y}<br>%{y:.1f} milyar TL<extra></extra>")
    figd.update_layout(height=380, separators=",.", legend_title_text="")
    styled_chart(figd)

    # Yılbaşından beri kümülatif denge (yıl karşılaştırması)
    st.subheader("Yılbaşından Beri Kümülatif Bütçe Dengesi (Milyar TL)")
    _yrs = sorted([int(y) for y in b["yil"].unique() if y >= cy - 4])
    crows = []
    for y in _yrs:
        yd = b[b["yil"] == y].sort_values("ay")
        cum = (yd["denge"].cumsum() / 1000).tolist()
        for a, c in zip(yd["ay"].tolist(), cum):
            crows.append({"Ay": int(a), "Yıl": str(y), "Kümülatif": c})
    cdf = pd.DataFrame(crows)
    figc = px.line(cdf, x="Ay", y="Kümülatif", color="Yıl", markers=True,
                   labels={"Kümülatif": "Milyar TL"})
    figc.update_traces(hovertemplate="%{y:.1f} milyar TL<extra></extra>")
    figc.update_layout(height=380, separators=",.", legend_title_text="")
    figc.update_xaxes(tickmode="array", tickvals=list(range(1, 13)), ticktext=AY_KISA)
    styled_chart(figc)

    # Gelir vs Gider
    st.subheader("Gelir vs Gider (Aylık, Milyar TL)")
    gg = bp[["tarih", "gelir", "gider"]].copy()
    gg["Gelir"] = gg["gelir"] / 1000
    gg["Gider"] = gg["gider"] / 1000
    glong = gg[["tarih", "Gelir", "Gider"]].melt("tarih", var_name="Kalem", value_name="Milyar TL")
    figg = px.line(glong, x="tarih", y="Milyar TL", color="Kalem",
                   color_discrete_map={"Gelir": "#26C281", "Gider": "#FF9E1B"}, labels={"tarih": ""})
    figg.update_traces(hovertemplate="%{x|%m.%Y}<br>%{y:.1f}<extra></extra>")
    figg.update_layout(height=340, separators=",.", legend_title_text="")
    styled_chart(figg)

    # Vergi kompozisyonu
    if "dolaysiz_vergi" in b.columns and "dolayli_vergi" in b.columns:
        st.subheader("Vergi Gelirleri Kompozisyonu (Aylık, Milyar TL)")
        vg = bp[["tarih", "dolaysiz_vergi", "dolayli_vergi"]].copy()
        vg["Dolaysız Vergiler"] = vg["dolaysiz_vergi"] / 1000
        vg["Dolaylı Vergiler"] = vg["dolayli_vergi"] / 1000
        vlong = vg[["tarih", "Dolaysız Vergiler", "Dolaylı Vergiler"]].melt("tarih", var_name="Tür", value_name="Milyar TL")
        figv = px.area(vlong, x="tarih", y="Milyar TL", color="Tür",
                       color_discrete_map={"Dolaysız Vergiler": "#4C9AFF", "Dolaylı Vergiler": "#B98AFF"},
                       labels={"tarih": ""})
        figv.update_layout(height=320, separators=",.", legend_title_text="")
        styled_chart(figv)

    # Yıllık özet
    with st.expander("📋 Yıllık Özet (Milyar TL) — cari yıl kümülatiftir", expanded=False):
        yt = b.groupby("yil").agg(
            Gelir=("gelir", "sum"), Gider=("gider", "sum"), Denge=("denge", "sum"),
            FaizDisi=("faiz_disi_denge", "sum"), Faiz=("faiz_gideri", "sum"),
        ).reset_index().sort_values("yil", ascending=False)
        disp = yt.copy()
        for c in ["Gelir", "Gider", "Denge", "FaizDisi", "Faiz"]:
            disp[c] = disp[c].apply(_mr)
        disp["yil"] = disp["yil"].astype(int).astype(str)
        disp.columns = ["Yıl", "Gelir", "Gider", "Bütçe Dengesi", "Faiz Dışı Denge", "Faiz Gideri"]
        st.dataframe(disp, hide_index=True, use_container_width=True)

    get_download_button(str(data_file), "📥 Bütçe Verisi (.xlsx)")
    st.markdown("**Kaynak:** HMB Kamu Finansmanı İstatistikleri → Merkezi Yönetim Bütçe Dengesi ve Finansmanı  ·  "
                "Birim: Milyon TL (grafiklerde Milyar TL)  ·  Aylık, 2006→bugün.")


# ══════════════════════════════════════════════════════════
# HAZİNE NAKİT GERÇEKLEŞMELERİ
# ══════════════════════════════════════════════════════════

elif selected == "nakit":
    st.markdown('<div class="main-header">Hazine Nakit Gerçekleşmeleri</div>', unsafe_allow_html=True)

    nakit_dir = BASE_DIR / "hazine nakit"
    fetch_script = nakit_dir / "nakit_fetch.py"
    data_file = nakit_dir / "nakit.xlsx"

    col_u1, col_u2 = st.columns([1.4, 4])
    with col_u1:
        if st.button("🔄 Güncelle (yerel + bulut)", key="nakit_update", use_container_width=True):
            if run_script(str(fetch_script), timeout_sec=120):
                st.cache_data.clear()
                if _try_publish(["hazine nakit/nakit.xlsx"], "Hazine nakit verisi guncellendi (otomatik yayin)"):
                    st.success("Güncellendi ve buluta yayınlandı ✓")
                else:
                    st.info("Yerel veri güncellendi. (Buluta yayın yalnızca yerel bilgisayardan yapılır.)")
    with col_u2:
        if data_file.exists():
            st.markdown(f'<div class="update-info">Son güncelleme: {get_file_mod_time(data_file)}</div>',
                        unsafe_allow_html=True)
        else:
            st.warning("Henüz veri çekilmemiş. Güncelle'ye tıklayın.")

    if not data_file.exists():
        st.stop()

    @st.cache_data
    def load_nakit(path):
        d = pd.read_excel(path, sheet_name="Aylik")
        d["tarih"] = pd.to_datetime(d["tarih"], errors="coerce")
        return d.dropna(subset=["tarih"]).sort_values("tarih").reset_index(drop=True)

    n = load_nakit(str(data_file))
    if n.empty:
        st.warning("Veri boş.")
        st.stop()

    AY = {1: "Ocak", 2: "Şubat", 3: "Mart", 4: "Nisan", 5: "Mayıs", 6: "Haziran",
          7: "Temmuz", 8: "Ağustos", 9: "Eylül", 10: "Ekim", 11: "Kasım", 12: "Aralık"}
    AY_KISA = ["Oca", "Şub", "Mar", "Nis", "May", "Haz", "Tem", "Ağu", "Eyl", "Eki", "Kas", "Ara"]
    last = n.iloc[-1]
    cy, cm = int(last["yil"]), int(last["ay"])
    son_label = f"{AY[cm]} {cy}"

    # Milyon TL -> Milyar TL, Türkçe 1 ondalık
    def _mr(v):
        return "—" if pd.isna(v) else f"{v/1000:,.1f}".replace(",", "\x00").replace(".", ",").replace("\x00", ".")

    ytd_now = n[(n["yil"] == cy) & (n["ay"] <= cm)]["nakit_denge"].sum()
    ytd_prev = n[(n["yil"] == cy - 1) & (n["ay"] <= cm)]["nakit_denge"].sum()

    _yon = "açık" if last["nakit_denge"] < 0 else "fazla"
    _yon_ytd = "açığı" if ytd_now < 0 else "fazlası"
    _yon_ytd_prev = "açık" if ytd_prev < 0 else "fazla"
    st.info(
        f"🪙 **{son_label}** ayında Hazine nakit dengesi **{_mr(abs(last['nakit_denge']))} milyar TL {_yon}** verdi "
        f"(gelir {_mr(last['gelir'])}, gider {_mr(last['gider'])} milyar TL). "
        f"Faiz dışı denge **{_mr(last['faiz_disi_denge'])} milyar TL**, faiz ödemesi {_mr(last['faiz_odemesi'])} milyar TL. "
        f"Yılbaşından beri kümülatif nakit {_yon_ytd} **{_mr(abs(ytd_now))} milyar TL** — "
        f"geçen yıl aynı dönemde {_mr(abs(ytd_prev))} milyar TL {_yon_ytd_prev} idi."
    )

    st.markdown(f"#### {son_label} (Milyar TL)")
    k1, k2, k3, k4 = st.columns(4)
    with k1:
        st.metric("Nakit Dengesi", _mr(last["nakit_denge"]),
                  help="Gelirler + Özelleştirme/Fon − Giderler. Negatif = nakit açığı.")
    with k2:
        st.metric("Gelirler", _mr(last["gelir"]))
    with k3:
        st.metric("Giderler", _mr(last["gider"]))
    with k4:
        st.metric("Faiz Dışı Denge", _mr(last["faiz_disi_denge"]), help="Faiz ödemeleri hariç nakit denge.")

    yr = st.slider("Başlangıç yılı", int(n["yil"].min()), cy, max(int(n["yil"].min()), cy - 6), key="nakit_yr")
    npf = n[n["yil"] >= yr].copy()

    # Aylık nakit dengesi
    st.subheader("Aylık Nakit Dengesi (Milyar TL)")
    npf["denge_mia"] = npf["nakit_denge"] / 1000
    npf["durum"] = npf["denge_mia"].apply(lambda v: "Fazla" if v >= 0 else "Açık")
    figd = px.bar(npf, x="tarih", y="denge_mia", color="durum",
                  color_discrete_map={"Açık": "#FF5A5F", "Fazla": "#26C281"},
                  labels={"tarih": "", "denge_mia": "Milyar TL", "durum": ""})
    figd.update_traces(hovertemplate="%{x|%m.%Y}<br>%{y:.1f} milyar TL<extra></extra>")
    figd.update_layout(height=380, separators=",.", legend_title_text="")
    styled_chart(figd)

    # Yılbaşından beri kümülatif nakit dengesi (yıl karşılaştırması)
    st.subheader("Yılbaşından Beri Kümülatif Nakit Dengesi (Milyar TL)")
    _yrs = sorted([int(y) for y in n["yil"].unique() if y >= cy - 4])
    crows = []
    for y in _yrs:
        yd = n[n["yil"] == y].sort_values("ay")
        cum = (yd["nakit_denge"].cumsum() / 1000).tolist()
        for a, c in zip(yd["ay"].tolist(), cum):
            crows.append({"Ay": int(a), "Yıl": str(y), "Kümülatif": c})
    cdf = pd.DataFrame(crows)
    figc = px.line(cdf, x="Ay", y="Kümülatif", color="Yıl", markers=True,
                   labels={"Kümülatif": "Milyar TL"})
    figc.update_traces(hovertemplate="%{y:.1f} milyar TL<extra></extra>")
    figc.update_layout(height=380, separators=",.", legend_title_text="")
    figc.update_xaxes(tickmode="array", tickvals=list(range(1, 13)), ticktext=AY_KISA)
    styled_chart(figc)

    # Nakit Gelir vs Gider
    st.subheader("Nakit Gelir vs Gider (Aylık, Milyar TL)")
    gg = npf[["tarih", "gelir", "gider"]].copy()
    gg["Gelir"] = gg["gelir"] / 1000
    gg["Gider"] = gg["gider"] / 1000
    glong = gg[["tarih", "Gelir", "Gider"]].melt("tarih", var_name="Kalem", value_name="Milyar TL")
    figg = px.line(glong, x="tarih", y="Milyar TL", color="Kalem",
                   color_discrete_map={"Gelir": "#26C281", "Gider": "#FF9E1B"}, labels={"tarih": ""})
    figg.update_traces(hovertemplate="%{x|%m.%Y}<br>%{y:.1f}<extra></extra>")
    figg.update_layout(height=340, separators=",.", legend_title_text="")
    styled_chart(figg)

    # Finansman — İç vs Dış Borçlanma (Net)
    if "ic_borclanma_net" in n.columns and "dis_borclanma_net" in n.columns:
        st.subheader("Finansman — Net Borçlanma (Aylık, Milyar TL)")
        fn = npf[["tarih", "ic_borclanma_net", "dis_borclanma_net"]].copy()
        fn["İç Borçlanma (Net)"] = fn["ic_borclanma_net"] / 1000
        fn["Dış Borçlanma (Net)"] = fn["dis_borclanma_net"] / 1000
        flong = fn[["tarih", "İç Borçlanma (Net)", "Dış Borçlanma (Net)"]].melt(
            "tarih", var_name="Tür", value_name="Milyar TL")
        figf = px.bar(flong, x="tarih", y="Milyar TL", color="Tür", barmode="group",
                      color_discrete_map={"İç Borçlanma (Net)": "#4C9AFF", "Dış Borçlanma (Net)": "#B98AFF"},
                      labels={"tarih": ""})
        figf.update_traces(hovertemplate="%{x|%m.%Y}<br>%{y:.1f} milyar TL<extra></extra>")
        figf.update_layout(height=340, separators=",.", legend_title_text="")
        styled_chart(figf)

    # Yıllık özet
    with st.expander("📋 Yıllık Özet (Milyar TL) — cari yıl kümülatiftir", expanded=False):
        yt = n.groupby("yil").agg(
            Gelir=("gelir", "sum"), Gider=("gider", "sum"), NakitDenge=("nakit_denge", "sum"),
            FaizDisi=("faiz_disi_denge", "sum"), Faiz=("faiz_odemesi", "sum"),
            NetBorclanma=("borclanma_net", "sum"),
        ).reset_index().sort_values("yil", ascending=False)
        disp = yt.copy()
        for c in ["Gelir", "Gider", "NakitDenge", "FaizDisi", "Faiz", "NetBorclanma"]:
            disp[c] = disp[c].apply(_mr)
        disp["yil"] = disp["yil"].astype(int).astype(str)
        disp.columns = ["Yıl", "Gelir", "Gider", "Nakit Dengesi", "Faiz Dışı Denge", "Faiz Ödemesi", "Net Borçlanma"]
        st.dataframe(disp, hide_index=True, use_container_width=True)

    get_download_button(str(data_file), "📥 Hazine Nakit Verisi (.xlsx)")
    st.markdown("**Kaynak:** HMB Kamu Finansmanı İstatistikleri → Hazine Nakit Gerçekleşmeleri  ·  "
                "Birim: Milyon TL (grafiklerde Milyar TL)  ·  Aylık, 2005→bugün.")


# ══════════════════════════════════════════════════════════
# YABANCI PARA HAREKETİ (DTH)
# ══════════════════════════════════════════════════════════

elif selected == "dth":
    st.markdown('<div class="main-header">Yabancı Para Hareketi</div>', unsafe_allow_html=True)

    dth_dir = BASE_DIR / "yabanci para hareketi"
    fetch_script = dth_dir / "dth_fetch.py"
    data_file = dth_dir / "dth.xlsx"

    col_u1, col_u2 = st.columns([1.4, 4])
    with col_u1:
        if st.button("🔄 Güncelle (yerel + bulut)", key="dth_update", use_container_width=True):
            if run_script(str(fetch_script), timeout_sec=120):
                st.cache_data.clear()
                if _try_publish(["yabanci para hareketi/dth.xlsx"], "DTH verisi guncellendi (otomatik yayin)"):
                    st.success("Güncellendi ve buluta yayınlandı ✓")
                else:
                    st.info("Yerel veri güncellendi. (Buluta yayın yalnızca yerel bilgisayardan yapılır.)")
    with col_u2:
        if data_file.exists():
            st.markdown(f'<div class="update-info">Son güncelleme: {get_file_mod_time(data_file)}</div>',
                        unsafe_allow_html=True)
        else:
            st.warning("Henüz veri çekilmemiş. Güncelle'ye tıklayın.")

    if not data_file.exists():
        st.stop()

    @st.cache_data
    def load_dth(path):
        d = pd.read_excel(path, sheet_name="Haftalik")
        d["tarih"] = pd.to_datetime(d["tarih"], errors="coerce")
        return d.dropna(subset=["tarih"]).sort_values("tarih").reset_index(drop=True)

    d = load_dth(str(data_file))
    if d.empty:
        st.warning("Veri boş.")
        st.stop()

    L = d.iloc[-1]
    hafta_lbl = L["tarih"].strftime("%d.%m.%Y")

    # Milyon USD -> Milyar USD, işaretli, Türkçe 1 ondalık
    def _bs(v, d_=1):
        if pd.isna(v):
            return "—"
        return f"{v/1000:+,.{d_}f}".replace(",", "\x00").replace(".", ",").replace("\x00", ".")

    # Mutlak değer (yön kelimesiyle birlikte kullanılır)
    def _ba(v, d_=1):
        if pd.isna(v):
            return "—"
        return f"{abs(v)/1000:,.{d_}f}".replace(",", "\x00").replace(".", ",").replace("\x00", ".")

    s4 = float(d["yerlesik_toplam"].tail(4).sum())
    ytd = float(d[d["tarih"].dt.year == int(L["tarih"].year)]["yerlesik_toplam"].sum())
    _yon = "arttı" if L["yerlesik_toplam"] >= 0 else "azaldı"
    _yon4 = "artış" if s4 >= 0 else "azalış"
    _yon_ytd = "artış (dolarizasyon)" if ytd >= 0 else "azalış (de-dolarizasyon)"
    st.info(
        f"💱 **{hafta_lbl}** haftasında yurt içi yerleşiklerin yabancı para mevduatı "
        f"(altın ve parite etkileri düzeltilmiş) **{_ba(L['yerlesik_toplam'])} milyar USD {_yon}**: "
        f"tüzel kişiler **{_bs(L['tuzel_kisiler'])}**, gerçek kişiler **{_bs(L['gercek_kisiler'])} milyar USD** "
        f"(altın {_bs(L['gk_altin'])}, döviz {_bs(L['gk_doviz'])}). "
        f"Son 4 haftada kümülatif {_ba(s4)} milyar USD {_yon4}; "
        f"yılbaşından beri **{_ba(ytd)} milyar USD {_yon_ytd}**."
    )

    st.markdown(f"#### {hafta_lbl} Haftası (Milyar USD)")
    k1, k2, k3, k4 = st.columns(4)
    with k1:
        st.metric("Yerleşikler Toplam", _bs(L["yerlesik_toplam"]),
                  help="Parite ve altın fiyat etkilerinden arındırılmış haftalık değişim.")
    with k2:
        st.metric("Gerçek Kişiler", _bs(L["gercek_kisiler"]))
    with k3:
        st.metric("Tüzel Kişiler", _bs(L["tuzel_kisiler"]))
    with k4:
        st.metric("Yılbaşından Beri", _bs(ytd), help="Cari yıl haftalık değişimlerin toplamı.")

    hafta_n = st.slider("Gösterilecek hafta sayısı", 12, 104, 27, key="dth_hafta")
    dp = d.tail(hafta_n).copy()

    # Son bar segment etiketi (ekran görüntüsü stili)
    def _son_bar_etiket(fig, x, comps):
        pos = neg = 0.0
        for v, color in comps:
            if pd.isna(v):
                continue
            if v >= 0:
                pos += v; y = pos; shift = 12
            else:
                neg += v; y = neg; shift = -14
            fig.add_annotation(x=x, y=y, text=f"{v:.1f}".replace(".", ","),
                               showarrow=False, yshift=shift,
                               font=dict(size=13, color=color))

    # 1) Gerçek Kişiler — Altın & Döviz
    st.subheader("Gerçek Kişiler DTH Hesapları Değer Değişimi")
    st.caption("Altın ve parite etkileri düzeltilmiş, milyar USD")
    g1 = dp[["tarih"]].copy()
    g1["Altın"] = dp["gk_altin"] / 1000
    g1["Döviz"] = dp["gk_doviz"] / 1000
    g1l = g1.melt("tarih", var_name="Kalem", value_name="Milyar USD")
    fig1 = px.bar(g1l, x="tarih", y="Milyar USD", color="Kalem",
                  color_discrete_map={"Altın": "#FF9E1B", "Döviz": "#26C281"}, labels={"tarih": ""})
    fig1.update_traces(hovertemplate="%{x|%d.%m.%Y}<br>%{y:.1f} milyar USD<extra>%{fullData.name}</extra>")
    fig1.update_layout(height=380, separators=",.", legend_title_text="", bargap=0.25)
    _son_bar_etiket(fig1, L["tarih"], [(L["gk_altin"] / 1000, "#FF9E1B"), (L["gk_doviz"] / 1000, "#26C281")])
    styled_chart(fig1)

    # 2) Yerleşikler — Tüzel & Gerçek
    st.subheader("Yerleşikler DTH Hesapları Değer Değişimi — Tüzel & Gerçek")
    st.caption("Altın ve parite etkileri düzeltilmiş, milyar USD")
    g2 = dp[["tarih"]].copy()
    g2["Tüzel Kişiler"] = dp["tuzel_kisiler"] / 1000
    g2["Gerçek Kişiler"] = dp["gercek_kisiler"] / 1000
    g2l = g2.melt("tarih", var_name="Kalem", value_name="Milyar USD")
    fig2 = px.bar(g2l, x="tarih", y="Milyar USD", color="Kalem",
                  color_discrete_map={"Tüzel Kişiler": "#4FC3F7", "Gerçek Kişiler": "#E64980"}, labels={"tarih": ""})
    fig2.update_traces(hovertemplate="%{x|%d.%m.%Y}<br>%{y:.1f} milyar USD<extra>%{fullData.name}</extra>")
    fig2.update_layout(height=380, separators=",.", legend_title_text="", bargap=0.25)
    _son_bar_etiket(fig2, L["tarih"], [(L["tuzel_kisiler"] / 1000, "#4FC3F7"), (L["gercek_kisiler"] / 1000, "#E64980")])
    styled_chart(fig2)

    # 3) Yerleşikler — Toplam
    st.subheader("Yerleşikler DTH Hesapları Değer Değişimi — Toplam")
    st.caption("Altın ve parite etkileri düzeltilmiş, milyar USD")
    g3 = dp[["tarih"]].copy()
    g3["Milyar USD"] = dp["yerlesik_toplam"] / 1000
    fig3 = px.bar(g3, x="tarih", y="Milyar USD", labels={"tarih": ""})
    fig3.update_traces(marker_color="#4C9AFF",
                       hovertemplate="%{x|%d.%m.%Y}<br>%{y:.1f} milyar USD<extra></extra>")
    fig3.update_layout(height=380, separators=",.", bargap=0.25)
    _son_bar_etiket(fig3, L["tarih"], [(L["yerlesik_toplam"] / 1000, "#4C9AFF")])
    styled_chart(fig3)

    # Detay tablo — son 8 hafta
    with st.expander("📋 Son 8 Hafta Detay (Milyar USD)", expanded=False):
        t8 = d.tail(8).iloc[::-1].copy()
        disp = pd.DataFrame({
            "Hafta": t8["tarih"].dt.strftime("%d.%m.%Y"),
            "Toplam": t8["yerlesik_toplam"].apply(_bs),
            "Gerçek Kişiler": t8["gercek_kisiler"].apply(_bs),
            "Tüzel Kişiler": t8["tuzel_kisiler"].apply(_bs),
            "GK Altın": t8["gk_altin"].apply(_bs),
            "GK Döviz": t8["gk_doviz"].apply(_bs),
        })
        st.dataframe(disp, hide_index=True, use_container_width=True)

    get_download_button(str(data_file), "📥 DTH Verisi (.xlsx)")
    st.markdown("**Kaynak:** TCMB EVDS → Haftalık Para ve Banka İstatistikleri, Tablo 5 (TP.HPBITABLO5) — "
                "parite ve kıymetli maden fiyat etkilerinden arındırılmış haftalık değişim  ·  "
                "Birim: Milyon USD (grafiklerde Milyar USD)  ·  Haftalık, 2015→bugün.")
