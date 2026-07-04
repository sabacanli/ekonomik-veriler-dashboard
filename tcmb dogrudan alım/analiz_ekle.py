import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from copy import copy

wb = openpyxl.load_workbook('tcmb_dogrudan_alim.xlsx')
data_ws = wb['Doğrudan Alım İşlemleri']
max_row = data_ws.max_row  # 1928
DATA_RANGE = f"$2:${max_row}"
SN = "'Doğrudan Alım İşlemleri'"  # Sheet name for formulas (quoted)

# ── Styles ──────────────────────────────────────────────────────
title_font = Font(bold=True, size=14, color="1F4E79", name="Calibri")
section_font = Font(bold=True, size=12, color="1F4E79", name="Calibri")
header_font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
data_font = Font(size=10, name="Calibri")
bold_font = Font(bold=True, size=10, name="Calibri")
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
sub_header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
total_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
even_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def style_header(ws, row, col_start, col_end, fill=None):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = fill or header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border


def style_data_cell(cell, fmt=None, bold=False, fill=None):
    cell.font = bold_font if bold else data_font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center')
    if fmt:
        cell.number_format = fmt
    if fill:
        cell.fill = fill


# ════════════════════════════════════════════════════════════════
# SAYFA 1: YILLIK BORÇLANMA ANALİZİ
# ════════════════════════════════════════════════════════════════
ws1 = wb.create_sheet("Yıllık Borçlanma Analizi")

# Title
ws1.merge_cells('A1:I1')
ws1.cell(1, 1, "YILLIK BORÇLANMA ANALİZİ (İŞLEM TARİHİNE GÖRE)").font = title_font

# Column refs in data sheet
# B=İşlem Tarihi, E=Vade, G=Teklif Nominal, H=Kazanan Nominal, I=Kazanan Net
# K=Ort Basit Faiz, N=Ort Bileşik Faiz

headers1 = [
    "Yıl", "İhale\nSayısı", "Toplam Teklif\n(Nominal, Bin TL)",
    "Toplam Borçlanma\n(Nominal, Bin TL)", "Toplam Borçlanma\n(Net, Bin TL)",
    "Karşılanma\nOranı (%)", "Ort. Basit\nFaiz (%)",
    "Ort. Bileşik\nFaiz (%)", "Ort. Vade\n(Gün)"
]

r = 3
for c, h in enumerate(headers1, 1):
    ws1.cell(r, c, h)
style_header(ws1, r, 1, len(headers1))

# Year rows: 2009-2026
years = list(range(2009, 2027))
for i, yr in enumerate(years):
    r = 4 + i
    yr_cell = ws1.cell(r, 1, yr)
    style_data_cell(yr_cell, bold=True)

    B = f"{SN}!$B{DATA_RANGE}"  # won't work directly - need actual range
    # Use full range references
    B_rng = f"{SN}!$B$2:$B${max_row}"
    G_rng = f"{SN}!$G$2:$G${max_row}"
    H_rng = f"{SN}!$H$2:$H${max_row}"
    I_rng = f"{SN}!$I$2:$I${max_row}"
    K_rng = f"{SN}!$K$2:$K${max_row}"
    N_rng = f"{SN}!$N$2:$N${max_row}"
    E_rng = f"{SN}!$E$2:$E${max_row}"

    yr_ref = f"$A${r}"

    # İhale Sayısı
    f_count = f'=COUNTIF({B_rng},">="&DATE({yr_ref},1,1))-COUNTIF({B_rng},">="&DATE({yr_ref}+1,1,1))'
    c2 = ws1.cell(r, 2, f_count)
    style_data_cell(c2, '#,##0')

    # Toplam Teklif Nominal
    f_teklif = f'=SUMPRODUCT((YEAR({B_rng})={yr_ref})*{G_rng})'
    c3 = ws1.cell(r, 3, f_teklif)
    style_data_cell(c3, '#,##0')

    # Toplam Borçlanma Nominal
    f_nominal = f'=SUMPRODUCT((YEAR({B_rng})={yr_ref})*{H_rng})'
    c4 = ws1.cell(r, 4, f_nominal)
    style_data_cell(c4, '#,##0')

    # Toplam Net
    f_net = f'=SUMPRODUCT((YEAR({B_rng})={yr_ref})*{I_rng})'
    c5 = ws1.cell(r, 5, f_net)
    style_data_cell(c5, '#,##0.00')

    # Karşılanma Oranı
    f_ratio = f'=IF(C{r}=0,"",D{r}/C{r}*100)'
    c6 = ws1.cell(r, 6, f_ratio)
    style_data_cell(c6, '#,##0.0')

    # Ort. Basit Faiz (ağırlıklı ortalama: nominal ile ağırlıklandır)
    f_basit = f'=IF(D{r}=0,"",SUMPRODUCT((YEAR({B_rng})={yr_ref})*{K_rng}*{H_rng})/SUMPRODUCT((YEAR({B_rng})={yr_ref})*{H_rng}))'
    c7 = ws1.cell(r, 7, f_basit)
    style_data_cell(c7, '#,##0.00')

    # Ort. Bileşik Faiz (ağırlıklı ortalama)
    f_bilesik = f'=IF(D{r}=0,"",SUMPRODUCT((YEAR({B_rng})={yr_ref})*{N_rng}*{H_rng})/SUMPRODUCT((YEAR({B_rng})={yr_ref})*{H_rng}))'
    c8 = ws1.cell(r, 8, f_bilesik)
    style_data_cell(c8, '#,##0.00')

    # Ort. Vade (gün) - valör ile vade arasındaki fark
    V_rng = f"{SN}!$D$2:$D${max_row}"  # Valör
    f_vade = f'=IF(D{r}=0,"",SUMPRODUCT((YEAR({B_rng})={yr_ref})*({E_rng}-{V_rng})*{H_rng})/SUMPRODUCT((YEAR({B_rng})={yr_ref})*{H_rng}))'
    c9 = ws1.cell(r, 9, f_vade)
    style_data_cell(c9, '#,##0')

    # Zebra
    if i % 2 == 0:
        for c in range(1, len(headers1) + 1):
            ws1.cell(r, c).fill = even_fill

# TOPLAM satırı
r_total = 4 + len(years)
ws1.cell(r_total, 1, "TOPLAM").font = bold_font
ws1.cell(r_total, 1).fill = total_fill
ws1.cell(r_total, 1).border = thin_border
ws1.cell(r_total, 1).alignment = Alignment(horizontal='center', vertical='center')

for c in range(2, len(headers1) + 1):
    cell = ws1.cell(r_total, c)
    cell.fill = total_fill
    cell.border = thin_border
    cell.font = bold_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

ws1.cell(r_total, 2, f'=SUM(B4:B{r_total-1})').number_format = '#,##0'
ws1.cell(r_total, 3, f'=SUM(C4:C{r_total-1})').number_format = '#,##0'
ws1.cell(r_total, 4, f'=SUM(D4:D{r_total-1})').number_format = '#,##0'
ws1.cell(r_total, 5, f'=SUM(E4:E{r_total-1})').number_format = '#,##0.00'
ws1.cell(r_total, 6, f'=IF(C{r_total}=0,"",D{r_total}/C{r_total}*100)').number_format = '#,##0.0'
# Ağırlıklı ort faizler toplam satırında
ws1.cell(r_total, 7, f'=SUMPRODUCT(G4:G{r_total-1},D4:D{r_total-1})/SUM(D4:D{r_total-1})').number_format = '#,##0.00'
ws1.cell(r_total, 8, f'=SUMPRODUCT(H4:H{r_total-1},D4:D{r_total-1})/SUM(D4:D{r_total-1})').number_format = '#,##0.00'
ws1.cell(r_total, 9, f'=SUMPRODUCT(I4:I{r_total-1},D4:D{r_total-1})/SUM(D4:D{r_total-1})').number_format = '#,##0'

# Column widths
col_widths1 = [8, 12, 22, 22, 22, 14, 14, 14, 12]
for i, w in enumerate(col_widths1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

ws1.freeze_panes = 'A4'

# ════════════════════════════════════════════════════════════════
# SAYFA 2: YILLIK İTFA PROFİLİ
# ════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Yıllık İtfa Profili")

ws2.merge_cells('A1:F1')
ws2.cell(1, 1, "YILLIK İTFA PROFİLİ (VADE TARİHİNE GÖRE)").font = title_font

headers2 = [
    "İtfa Yılı", "İtfa Olan\nNominal Tutar", "İtfa Olan\nNet Tutar",
    "İhale Sayısı", "Farklı ISIN\nSayısı", "Ort. Bileşik\nFaiz (%)"
]

r = 3
for c, h in enumerate(headers2, 1):
    ws2.cell(r, c, h)
style_header(ws2, r, 1, len(headers2))

# İtfa yılları: 2010-2034 (vadeler geleceğe uzanıyor)
itfa_years = list(range(2010, 2035))
E_rng = f"{SN}!$E$2:$E${max_row}"
H_rng = f"{SN}!$H$2:$H${max_row}"
I_rng = f"{SN}!$I$2:$I${max_row}"
N_rng = f"{SN}!$N$2:$N${max_row}"
F_rng = f"{SN}!$F$2:$F${max_row}"

for i, yr in enumerate(itfa_years):
    r = 4 + i
    yr_cell = ws2.cell(r, 1, yr)
    style_data_cell(yr_cell, bold=True)
    yr_ref = f"$A${r}"

    # İtfa Nominal
    f1 = f'=SUMPRODUCT((YEAR({E_rng})={yr_ref})*{H_rng})'
    c1 = ws2.cell(r, 2, f1)
    style_data_cell(c1, '#,##0')

    # İtfa Net
    f2 = f'=SUMPRODUCT((YEAR({E_rng})={yr_ref})*{I_rng})'
    c2 = ws2.cell(r, 3, f2)
    style_data_cell(c2, '#,##0.00')

    # İhale sayısı
    f3 = f'=COUNTIF({E_rng},">="&DATE({yr_ref},1,1))-COUNTIF({E_rng},">="&DATE({yr_ref}+1,1,1))'
    c3 = ws2.cell(r, 4, f3)
    style_data_cell(c3, '#,##0')

    # Farklı ISIN sayısı - SUMPRODUCT ile unique count
    f4 = f'=SUMPRODUCT((YEAR({E_rng})={yr_ref})/COUNTIF({F_rng},{F_rng})*(YEAR({E_rng})={yr_ref}))'
    c4 = ws2.cell(r, 5, f4)
    style_data_cell(c4, '#,##0')

    # Ort Bileşik Faiz (ağırlıklı)
    f5 = f'=IF(B{r}=0,"",SUMPRODUCT((YEAR({E_rng})={yr_ref})*{N_rng}*{H_rng})/SUMPRODUCT((YEAR({E_rng})={yr_ref})*{H_rng}))'
    c5 = ws2.cell(r, 6, f5)
    style_data_cell(c5, '#,##0.00')

    if i % 2 == 0:
        for c in range(1, len(headers2) + 1):
            ws2.cell(r, c).fill = even_fill

# Toplam
r_total2 = 4 + len(itfa_years)
ws2.cell(r_total2, 1, "TOPLAM").font = bold_font
ws2.cell(r_total2, 1).fill = total_fill
ws2.cell(r_total2, 1).border = thin_border
ws2.cell(r_total2, 1).alignment = Alignment(horizontal='center')
for c in range(2, len(headers2) + 1):
    cell = ws2.cell(r_total2, c)
    cell.fill = total_fill
    cell.border = thin_border
    cell.font = bold_font
    cell.alignment = Alignment(horizontal='center')

ws2.cell(r_total2, 2, f'=SUM(B4:B{r_total2-1})').number_format = '#,##0'
ws2.cell(r_total2, 3, f'=SUM(C4:C{r_total2-1})').number_format = '#,##0.00'
ws2.cell(r_total2, 4, f'=SUM(D4:D{r_total2-1})').number_format = '#,##0'

col_widths2 = [12, 22, 22, 14, 14, 16]
for i, w in enumerate(col_widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.freeze_panes = 'A4'

# ════════════════════════════════════════════════════════════════
# SAYFA 3: VADE DAĞILIM ANALİZİ
# ════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Vade Dağılım Analizi")

ws3.merge_cells('A1:H1')
ws3.cell(1, 1, "VADE DAĞILIM ANALİZİ (İŞLEM YILINA GÖRE)").font = title_font

headers3 = [
    "Yıl", "Toplam\nNominal",
    "0-1 Yıl\nNominal", "1-2 Yıl\nNominal", "2-3 Yıl\nNominal",
    "3-5 Yıl\nNominal", "5-7 Yıl\nNominal", "7+ Yıl\nNominal"
]

r = 3
for c, h in enumerate(headers3, 1):
    ws3.cell(r, c, h)
style_header(ws3, r, 1, len(headers3))

# Sub-header row for percentages
headers3b = [
    "", "",
    "0-1 Yıl\n(%)", "1-2 Yıl\n(%)", "2-3 Yıl\n(%)",
    "3-5 Yıl\n(%)", "5-7 Yıl\n(%)", "7+ Yıl\n(%)"
]

B_rng = f"{SN}!$B$2:$B${max_row}"
E_rng = f"{SN}!$E$2:$E${max_row}"
V_rng = f"{SN}!$D$2:$D${max_row}"  # Valör
H_rng = f"{SN}!$H$2:$H${max_row}"

for i, yr in enumerate(years):
    r_nom = 4 + i * 2      # Nominal row
    r_pct = 4 + i * 2 + 1  # Percentage row

    yr_cell = ws3.cell(r_nom, 1, yr)
    style_data_cell(yr_cell, bold=True)
    ws3.merge_cells(start_row=r_nom, start_column=1, end_row=r_pct, end_column=1)
    yr_cell.alignment = Alignment(horizontal='center', vertical='center')

    yr_ref = f"$A${r_nom}"

    # Toplam Nominal
    f_total = f'=SUMPRODUCT((YEAR({B_rng})={yr_ref})*{H_rng})'
    c_total = ws3.cell(r_nom, 2, f_total)
    style_data_cell(c_total, '#,##0')
    ws3.merge_cells(start_row=r_nom, start_column=2, end_row=r_pct, end_column=2)
    c_total.alignment = Alignment(horizontal='center', vertical='center')

    # Vade aralıkları (gün cinsinden: vade - valör)
    # 0-1Y: 0-365, 1-2Y: 366-730, 2-3Y: 731-1095, 3-5Y: 1096-1825, 5-7Y: 1826-2555, 7+: >2555
    vade_bounds = [(0, 365), (366, 730), (731, 1095), (1096, 1825), (1826, 2555), (2556, 99999)]

    for j, (lo, hi) in enumerate(vade_bounds):
        col = 3 + j
        # Nominal amount for this tenor bucket
        if hi == 99999:
            f_bucket = f'=SUMPRODUCT((YEAR({B_rng})={yr_ref})*(({E_rng}-{V_rng})>={lo})*{H_rng})'
        else:
            f_bucket = f'=SUMPRODUCT((YEAR({B_rng})={yr_ref})*(({E_rng}-{V_rng})>={lo})*(({E_rng}-{V_rng})<={hi})*{H_rng})'

        c_nom = ws3.cell(r_nom, col, f_bucket)
        style_data_cell(c_nom, '#,##0')

        # Percentage
        f_pct = f'=IF($B${r_nom}=0,"",{get_column_letter(col)}{r_nom}/$B${r_nom}*100)'
        c_pct = ws3.cell(r_pct, col, f_pct)
        style_data_cell(c_pct, '#,##0.0')
        c_pct.font = Font(size=9, italic=True, color="666666", name="Calibri")

    # Zebra
    if i % 2 == 0:
        for c in range(1, len(headers3) + 1):
            ws3.cell(r_nom, c).fill = even_fill
            ws3.cell(r_pct, c).fill = even_fill

col_widths3 = [8, 18, 16, 16, 16, 16, 16, 16]
for i, w in enumerate(col_widths3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

ws3.freeze_panes = 'A4'

# ════════════════════════════════════════════════════════════════
# SAYFA 4: BORÇLANMA vs İTFA KARŞILAŞTIRMASI
# ════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Borçlanma vs İtfa")

ws4.merge_cells('A1:G1')
ws4.cell(1, 1, "YILLIK BORÇLANMA vs İTFA KARŞILAŞTIRMASI").font = title_font

headers4 = [
    "Yıl", "Borçlanma\n(Nominal)", "İtfa\n(Nominal)",
    "Net Pozisyon\n(Borçlanma-İtfa)", "Borçlanma\n(Net Tutar)",
    "İtfa / Borçlanma\nOranı (%)", "Kümülatif\nNet Pozisyon"
]

r = 3
for c, h in enumerate(headers4, 1):
    ws4.cell(r, c, h)
style_header(ws4, r, 1, len(headers4))

B_rng = f"{SN}!$B$2:$B${max_row}"
E_rng = f"{SN}!$E$2:$E${max_row}"
H_rng = f"{SN}!$H$2:$H${max_row}"
I_rng = f"{SN}!$I$2:$I${max_row}"

compare_years = list(range(2009, 2035))
for i, yr in enumerate(compare_years):
    r = 4 + i
    yr_cell = ws4.cell(r, 1, yr)
    style_data_cell(yr_cell, bold=True)
    yr_ref = f"$A${r}"

    # Borçlanma Nominal
    f_borc = f'=SUMPRODUCT((YEAR({B_rng})={yr_ref})*{H_rng})'
    c1 = ws4.cell(r, 2, f_borc)
    style_data_cell(c1, '#,##0')

    # İtfa Nominal
    f_itfa = f'=SUMPRODUCT((YEAR({E_rng})={yr_ref})*{H_rng})'
    c2 = ws4.cell(r, 3, f_itfa)
    style_data_cell(c2, '#,##0')

    # Net Pozisyon
    f_net = f'=B{r}-C{r}'
    c3 = ws4.cell(r, 4, f_net)
    style_data_cell(c3, '#,##0')

    # Borçlanma Net Tutar
    f_net2 = f'=SUMPRODUCT((YEAR({B_rng})={yr_ref})*{I_rng})'
    c4 = ws4.cell(r, 5, f_net2)
    style_data_cell(c4, '#,##0.00')

    # İtfa/Borçlanma oranı
    f_oran = f'=IF(B{r}=0,"",C{r}/B{r}*100)'
    c5 = ws4.cell(r, 6, f_oran)
    style_data_cell(c5, '#,##0.0')

    # Kümülatif Net Pozisyon
    if i == 0:
        f_kum = f'=D{r}'
    else:
        f_kum = f'=G{r-1}+D{r}'
    c6 = ws4.cell(r, 7, f_kum)
    style_data_cell(c6, '#,##0')

    if i % 2 == 0:
        for c in range(1, len(headers4) + 1):
            ws4.cell(r, c).fill = even_fill

# Toplam
r_total4 = 4 + len(compare_years)
ws4.cell(r_total4, 1, "TOPLAM").font = bold_font
ws4.cell(r_total4, 1).fill = total_fill
ws4.cell(r_total4, 1).border = thin_border
ws4.cell(r_total4, 1).alignment = Alignment(horizontal='center')
for c in range(2, len(headers4) + 1):
    cell = ws4.cell(r_total4, c)
    cell.fill = total_fill
    cell.border = thin_border
    cell.font = bold_font
    cell.alignment = Alignment(horizontal='center')

ws4.cell(r_total4, 2, f'=SUM(B4:B{r_total4-1})').number_format = '#,##0'
ws4.cell(r_total4, 3, f'=SUM(C4:C{r_total4-1})').number_format = '#,##0'
ws4.cell(r_total4, 4, f'=SUM(D4:D{r_total4-1})').number_format = '#,##0'
ws4.cell(r_total4, 5, f'=SUM(E4:E{r_total4-1})').number_format = '#,##0.00'

col_widths4 = [8, 20, 20, 22, 20, 18, 20]
for i, w in enumerate(col_widths4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

ws4.freeze_panes = 'A4'

# ════════════════════════════════════════════════════════════════
# SAYFA 5: EN ÇOK KULLANILAN ISIN'LER
# ════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("ISIN Analizi")

ws5.merge_cells('A1:G1')
ws5.cell(1, 1, "EN ÇOK KULLANILAN ISIN ANALİZİ").font = title_font

# Bu sayfa için önce veriyi Python'da analiz edip ISIN listesini çıkaralım
# sonra her ISIN için formül yazalım
from collections import Counter

isins = []
for row in range(2, max_row + 1):
    isin = data_ws.cell(row, 6).value
    if isin:
        isins.append(isin)

isin_counts = Counter(isins)
top_isins = isin_counts.most_common(50)

headers5 = [
    "ISIN Kodu", "İhale\nSayısı", "Toplam Nominal\n(Bin TL)",
    "Toplam Net\n(Bin TL)", "İlk İşlem\nTarihi", "Son İşlem\nTarihi",
    "Vade Tarihi"
]

r = 3
for c, h in enumerate(headers5, 1):
    ws5.cell(r, c, h)
style_header(ws5, r, 1, len(headers5))

F_rng = f"{SN}!$F$2:$F${max_row}"
B_rng = f"{SN}!$B$2:$B${max_row}"
H_rng = f"{SN}!$H$2:$H${max_row}"
I_rng = f"{SN}!$I$2:$I${max_row}"
E_rng = f"{SN}!$E$2:$E${max_row}"

for i, (isin, count) in enumerate(top_isins):
    r = 4 + i
    isin_ref = f"$A${r}"

    ws5.cell(r, 1, isin)
    style_data_cell(ws5.cell(r, 1), bold=True)

    # İhale sayısı
    f1 = f'=COUNTIF({F_rng},{isin_ref})'
    ws5.cell(r, 2, f1)
    style_data_cell(ws5.cell(r, 2), '#,##0')

    # Toplam Nominal
    f2 = f'=SUMIF({F_rng},{isin_ref},{H_rng})'
    ws5.cell(r, 3, f2)
    style_data_cell(ws5.cell(r, 3), '#,##0')

    # Toplam Net
    f3 = f'=SUMIF({F_rng},{isin_ref},{I_rng})'
    ws5.cell(r, 4, f3)
    style_data_cell(ws5.cell(r, 4), '#,##0.00')

    # İlk İşlem Tarihi
    f4 = f'=MINIFS({B_rng},{F_rng},{isin_ref})'
    ws5.cell(r, 5, f4)
    style_data_cell(ws5.cell(r, 5), 'DD.MM.YYYY')

    # Son İşlem Tarihi
    f5 = f'=MAXIFS({B_rng},{F_rng},{isin_ref})'
    ws5.cell(r, 6, f5)
    style_data_cell(ws5.cell(r, 6), 'DD.MM.YYYY')

    # Vade Tarihi (hepsi aynı olmalı, max al)
    f6 = f'=MAXIFS({E_rng},{F_rng},{isin_ref})'
    ws5.cell(r, 7, f6)
    style_data_cell(ws5.cell(r, 7), 'DD.MM.YYYY')

    if i % 2 == 0:
        for c in range(1, len(headers5) + 1):
            ws5.cell(r, c).fill = even_fill

col_widths5 = [18, 12, 20, 20, 16, 16, 16]
for i, w in enumerate(col_widths5, 1):
    ws5.column_dimensions[get_column_letter(i)].width = w

ws5.freeze_panes = 'A4'

# ════════════════════════════════════════════════════════════════
# SAYFA 6: FAİZ TRENDİ
# ════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Aylık Faiz Trendi")

ws6.merge_cells('A1:F1')
ws6.cell(1, 1, "AYLIK AĞIRLIKLI ORTALAMA FAİZ TRENDİ").font = title_font

headers6 = [
    "Yıl", "Ay", "İhale\nSayısı", "Toplam\nNominal",
    "Ağırlıklı Ort.\nBasit Faiz (%)", "Ağırlıklı Ort.\nBileşik Faiz (%)"
]

r = 3
for c, h in enumerate(headers6, 1):
    ws6.cell(r, c, h)
style_header(ws6, r, 1, len(headers6))

B_rng = f"{SN}!$B$2:$B${max_row}"
H_rng = f"{SN}!$H$2:$H${max_row}"
K_rng = f"{SN}!$K$2:$K${max_row}"
N_rng = f"{SN}!$N$2:$N${max_row}"

row_idx = 4
for yr in range(2009, 2027):
    for mo in range(1, 13):
        # 2009 sadece Aralık'tan başlıyor, 2026 Nisan'a kadar
        if yr == 2009 and mo < 12:
            continue
        if yr == 2026 and mo > 4:
            continue

        ws6.cell(row_idx, 1, yr)
        style_data_cell(ws6.cell(row_idx, 1), bold=True)

        ws6.cell(row_idx, 2, mo)
        style_data_cell(ws6.cell(row_idx, 2))

        yr_ref = f"$A${row_idx}"
        mo_ref = f"$B${row_idx}"

        # İhale sayısı
        f1 = f'=SUMPRODUCT((YEAR({B_rng})={yr_ref})*(MONTH({B_rng})={mo_ref})*1)'
        ws6.cell(row_idx, 3, f1)
        style_data_cell(ws6.cell(row_idx, 3), '#,##0')

        # Toplam Nominal
        f2 = f'=SUMPRODUCT((YEAR({B_rng})={yr_ref})*(MONTH({B_rng})={mo_ref})*{H_rng})'
        ws6.cell(row_idx, 4, f2)
        style_data_cell(ws6.cell(row_idx, 4), '#,##0')

        # Ağırlıklı ort basit faiz
        f3 = f'=IF(D{row_idx}=0,"",SUMPRODUCT((YEAR({B_rng})={yr_ref})*(MONTH({B_rng})={mo_ref})*{K_rng}*{H_rng})/D{row_idx})'
        ws6.cell(row_idx, 5, f3)
        style_data_cell(ws6.cell(row_idx, 5), '#,##0.00')

        # Ağırlıklı ort bileşik faiz
        f4 = f'=IF(D{row_idx}=0,"",SUMPRODUCT((YEAR({B_rng})={yr_ref})*(MONTH({B_rng})={mo_ref})*{N_rng}*{H_rng})/D{row_idx})'
        ws6.cell(row_idx, 6, f4)
        style_data_cell(ws6.cell(row_idx, 6), '#,##0.00')

        if (row_idx - 4) % 2 == 0:
            for c in range(1, len(headers6) + 1):
                ws6.cell(row_idx, c).fill = even_fill

        row_idx += 1

col_widths6 = [8, 8, 12, 18, 20, 20]
for i, w in enumerate(col_widths6, 1):
    ws6.column_dimensions[get_column_letter(i)].width = w

ws6.freeze_panes = 'A4'

# ════════════════════════════════════════════════════════════════
# SAVE
# ════════════════════════════════════════════════════════════════
output = "/Users/sadettin/cowork/tcmb dogrudan alım/tcmb_dogrudan_alim.xlsx"
wb.save(output)
print(f"Analiz sayfaları eklendi: {output}")
print(f"Toplam sayfa: {len(wb.sheetnames)}")
print(f"Sayfalar: {wb.sheetnames}")
