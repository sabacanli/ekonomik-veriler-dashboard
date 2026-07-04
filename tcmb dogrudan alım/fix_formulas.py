import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from collections import Counter

wb = openpyxl.load_workbook('/Users/sadettin/cowork/tcmb dogrudan alım/tcmb_dogrudan_alim.xlsx')

# ══════════════════════════════════════════════════════════════
# ADIM 1: Veri sayfasındaki "-" text değerlerini temizle
# ══════════════════════════════════════════════════════════════
ds = wb['Doğrudan Alım İşlemleri']
max_row = ds.max_row
fixed = 0
for r in range(2, max_row + 1):
    for c in range(7, 16):  # G-O sütunları (sayısal olması gereken)
        v = ds.cell(r, c).value
        if v == '-' or v == '–' or v == '—':
            ds.cell(r, c).value = None  # Boş bırak, 0 değil
            fixed += 1

print(f"Veri sayfasında {fixed} adet '-' değeri temizlendi")

# ══════════════════════════════════════════════════════════════
# ADIM 2: Analiz sayfalarını sil ve yeniden oluştur
# ══════════════════════════════════════════════════════════════
sheets_to_remove = [s for s in wb.sheetnames if s != 'Doğrudan Alım İşlemleri']
for s in sheets_to_remove:
    del wb[s]

SN = "'Doğrudan Alım İşlemleri'"
MR = max_row

# ── Styles ──
title_font = Font(bold=True, size=14, color="1F4E79", name="Calibri")
header_font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
data_font = Font(size=10, name="Calibri")
bold_font = Font(bold=True, size=10, name="Calibri")
pct_font = Font(size=9, italic=True, color="666666", name="Calibri")
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
total_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
even_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def hdr(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

def dcell(ws, row, col, val, fmt=None, bold=False):
    cell = ws.cell(row, col, val)
    cell.font = bold_font if bold else data_font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center')
    if fmt:
        cell.number_format = fmt
    return cell

def total_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.fill = total_fill
        cell.border = thin_border
        cell.font = bold_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

def zebra(ws, row, cols, idx):
    if idx % 2 == 0:
        for c in range(1, cols + 1):
            ws.cell(row, c).fill = even_fill

# Range referansları - her formülde tekrar yazmamak için kısa isimler
B = f"{SN}!$B$2:$B${MR}"
D = f"{SN}!$D$2:$D${MR}"  # Valör
E = f"{SN}!$E$2:$E${MR}"  # Vade
F = f"{SN}!$F$2:$F${MR}"  # ISIN
G = f"{SN}!$G$2:$G${MR}"  # Teklif Nominal
H = f"{SN}!$H$2:$H${MR}"  # Kazanan Nominal
I = f"{SN}!$I$2:$I${MR}"  # Kazanan Net
K = f"{SN}!$K$2:$K${MR}"  # Ort Basit Faiz
N = f"{SN}!$N$2:$N${MR}"  # Ort Bileşik Faiz


# ════════════════════════════════════════════════════════════════
# SAYFA 1: YILLIK BORÇLANMA ANALİZİ
# ════════════════════════════════════════════════════════════════
ws1 = wb.create_sheet("Yıllık Borçlanma Analizi")
ws1.merge_cells('A1:I1')
ws1.cell(1, 1, "YILLIK BORÇLANMA ANALİZİ (İŞLEM TARİHİNE GÖRE)").font = title_font

headers1 = ["Yıl", "İhale\nSayısı", "Toplam Teklif\n(Nominal, Bin TL)",
            "Toplam Borçlanma\n(Nominal, Bin TL)", "Toplam Borçlanma\n(Net, Bin TL)",
            "Karşılanma\nOranı (%)", "Ağırlıklı Ort.\nBasit Faiz (%)",
            "Ağırlıklı Ort.\nBileşik Faiz (%)", "Ağırlıklı Ort.\nVade (Gün)"]
for c, h in enumerate(headers1, 1):
    ws1.cell(3, c, h)
hdr(ws1, 3, len(headers1))

years = list(range(2009, 2027))
for i, yr in enumerate(years):
    r = 4 + i
    a = f"$A${r}"
    dcell(ws1, r, 1, yr, bold=True)

    # İhale Sayısı - SUMPRODUCT ile (text safe)
    dcell(ws1, r, 2, f'=SUMPRODUCT((YEAR({B})={a})*1)', '#,##0')

    # Toplam Teklif Nominal - SUMPRODUCT, boş hücreler sorun olmaz
    dcell(ws1, r, 3, f'=SUMPRODUCT((YEAR({B})={a})*({G}<>"")*{G})', '#,##0')

    # Toplam Borçlanma Nominal
    dcell(ws1, r, 4, f'=SUMPRODUCT((YEAR({B})={a})*({H}<>"")*{H})', '#,##0')

    # Toplam Net
    dcell(ws1, r, 5, f'=SUMPRODUCT((YEAR({B})={a})*({I}<>"")*{I})', '#,##0.00')

    # Karşılanma Oranı
    dcell(ws1, r, 6, f'=IFERROR(D{r}/C{r}*100,"")', '#,##0.0')

    # Ağırlıklı Ort. Basit Faiz - sadece sayısal hücreleri al
    dcell(ws1, r, 7,
          f'=IFERROR(SUMPRODUCT((YEAR({B})={a})*ISNUMBER({K})*{K}*({H}<>"")*{H})/SUMPRODUCT((YEAR({B})={a})*ISNUMBER({K})*({H}<>"")*{H}),"")',
          '#,##0.00')

    # Ağırlıklı Ort. Bileşik Faiz
    dcell(ws1, r, 8,
          f'=IFERROR(SUMPRODUCT((YEAR({B})={a})*ISNUMBER({N})*{N}*({H}<>"")*{H})/SUMPRODUCT((YEAR({B})={a})*ISNUMBER({N})*({H}<>"")*{H}),"")',
          '#,##0.00')

    # Ağırlıklı Ort. Vade (gün)
    dcell(ws1, r, 9,
          f'=IFERROR(SUMPRODUCT((YEAR({B})={a})*({E}-{D})*({H}<>"")*{H})/SUMPRODUCT((YEAR({B})={a})*({H}<>"")*{H}),"")',
          '#,##0')

    zebra(ws1, r, len(headers1), i)

# Toplam satırı
rt = 4 + len(years)
total_row(ws1, rt, len(headers1))
dcell(ws1, rt, 1, "TOPLAM", bold=True)
ws1.cell(rt, 1).fill = total_fill
dcell(ws1, rt, 2, f'=SUM(B4:B{rt-1})', '#,##0')
ws1.cell(rt, 2).fill = total_fill
dcell(ws1, rt, 3, f'=SUM(C4:C{rt-1})', '#,##0')
ws1.cell(rt, 3).fill = total_fill
dcell(ws1, rt, 4, f'=SUM(D4:D{rt-1})', '#,##0')
ws1.cell(rt, 4).fill = total_fill
dcell(ws1, rt, 5, f'=SUM(E4:E{rt-1})', '#,##0.00')
ws1.cell(rt, 5).fill = total_fill
dcell(ws1, rt, 6, f'=IFERROR(D{rt}/C{rt}*100,"")', '#,##0.0')
ws1.cell(rt, 6).fill = total_fill
# Genel ağırlıklı ortalama
dcell(ws1, rt, 7, f'=IFERROR(SUMPRODUCT(G4:G{rt-1},D4:D{rt-1})/SUM(D4:D{rt-1}),"")', '#,##0.00')
ws1.cell(rt, 7).fill = total_fill
dcell(ws1, rt, 8, f'=IFERROR(SUMPRODUCT(H4:H{rt-1},D4:D{rt-1})/SUM(D4:D{rt-1}),"")', '#,##0.00')
ws1.cell(rt, 8).fill = total_fill
dcell(ws1, rt, 9, f'=IFERROR(SUMPRODUCT(I4:I{rt-1},D4:D{rt-1})/SUM(D4:D{rt-1}),"")', '#,##0')
ws1.cell(rt, 9).fill = total_fill

for i, w in enumerate([8, 12, 22, 22, 22, 14, 16, 16, 16], 1):
    ws1.column_dimensions[get_column_letter(i)].width = w
ws1.freeze_panes = 'A4'


# ════════════════════════════════════════════════════════════════
# SAYFA 2: YILLIK İTFA PROFİLİ
# ════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Yıllık İtfa Profili")
ws2.merge_cells('A1:E1')
ws2.cell(1, 1, "YILLIK İTFA PROFİLİ (VADE TARİHİNE GÖRE)").font = title_font

headers2 = ["İtfa Yılı", "İtfa Olan\nNominal Tutar", "İtfa Olan\nNet Tutar",
            "İhale Sayısı", "Ağırlıklı Ort.\nBileşik Faiz (%)"]
for c, h in enumerate(headers2, 1):
    ws2.cell(3, c, h)
hdr(ws2, 3, len(headers2))

itfa_years = list(range(2010, 2035))
for i, yr in enumerate(itfa_years):
    r = 4 + i
    a = f"$A${r}"
    dcell(ws2, r, 1, yr, bold=True)

    dcell(ws2, r, 2, f'=SUMPRODUCT((YEAR({E})={a})*({H}<>"")*{H})', '#,##0')
    dcell(ws2, r, 3, f'=SUMPRODUCT((YEAR({E})={a})*({I}<>"")*{I})', '#,##0.00')
    dcell(ws2, r, 4, f'=SUMPRODUCT((YEAR({E})={a})*1)', '#,##0')
    dcell(ws2, r, 5,
          f'=IFERROR(SUMPRODUCT((YEAR({E})={a})*ISNUMBER({N})*{N}*({H}<>"")*{H})/SUMPRODUCT((YEAR({E})={a})*ISNUMBER({N})*({H}<>"")*{H}),"")',
          '#,##0.00')

    zebra(ws2, r, len(headers2), i)

rt2 = 4 + len(itfa_years)
total_row(ws2, rt2, len(headers2))
dcell(ws2, rt2, 1, "TOPLAM", bold=True)
ws2.cell(rt2, 1).fill = total_fill
dcell(ws2, rt2, 2, f'=SUM(B4:B{rt2-1})', '#,##0')
ws2.cell(rt2, 2).fill = total_fill
dcell(ws2, rt2, 3, f'=SUM(C4:C{rt2-1})', '#,##0.00')
ws2.cell(rt2, 3).fill = total_fill
dcell(ws2, rt2, 4, f'=SUM(D4:D{rt2-1})', '#,##0')
ws2.cell(rt2, 4).fill = total_fill

for i, w in enumerate([12, 22, 22, 14, 18], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = 'A4'


# ════════════════════════════════════════════════════════════════
# SAYFA 3: VADE DAĞILIM ANALİZİ
# ════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Vade Dağılım Analizi")
ws3.merge_cells('A1:H1')
ws3.cell(1, 1, "VADE DAĞILIM ANALİZİ (İŞLEM YILINA GÖRE)").font = title_font

headers3 = ["Yıl", "Toplam\nNominal", "0-1 Yıl", "1-2 Yıl", "2-3 Yıl",
            "3-5 Yıl", "5-7 Yıl", "7+ Yıl"]
for c, h in enumerate(headers3, 1):
    ws3.cell(3, c, h)
hdr(ws3, 3, len(headers3))

vade_bounds = [(0, 365), (366, 730), (731, 1095), (1096, 1825), (1826, 2555), (2556, 99999)]

for i, yr in enumerate(years):
    r_nom = 4 + i * 2
    r_pct = 4 + i * 2 + 1
    a = f"$A${r_nom}"

    dcell(ws3, r_nom, 1, yr, bold=True)
    ws3.merge_cells(start_row=r_nom, start_column=1, end_row=r_pct, end_column=1)
    ws3.cell(r_nom, 1).alignment = Alignment(horizontal='center', vertical='center')

    dcell(ws3, r_nom, 2, f'=SUMPRODUCT((YEAR({B})={a})*({H}<>"")*{H})', '#,##0')
    ws3.merge_cells(start_row=r_nom, start_column=2, end_row=r_pct, end_column=2)
    ws3.cell(r_nom, 2).alignment = Alignment(horizontal='center', vertical='center')

    for j, (lo, hi) in enumerate(vade_bounds):
        col = 3 + j
        cl = get_column_letter(col)
        if hi == 99999:
            f_bucket = f'=SUMPRODUCT((YEAR({B})={a})*(({E}-{D})>={lo})*({H}<>"")*{H})'
        else:
            f_bucket = f'=SUMPRODUCT((YEAR({B})={a})*(({E}-{D})>={lo})*(({E}-{D})<={hi})*({H}<>"")*{H})'
        dcell(ws3, r_nom, col, f_bucket, '#,##0')

        f_pct = f'=IFERROR({cl}{r_nom}/$B${r_nom}*100,"")'
        c_pct = dcell(ws3, r_pct, col, f_pct, '#,##0.0')
        c_pct.font = pct_font

    if i % 2 == 0:
        for c in range(1, len(headers3) + 1):
            ws3.cell(r_nom, c).fill = even_fill
            ws3.cell(r_pct, c).fill = even_fill

for i, w in enumerate([8, 18, 14, 14, 14, 14, 14, 14], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.freeze_panes = 'A4'


# ════════════════════════════════════════════════════════════════
# SAYFA 4: BORÇLANMA vs İTFA KARŞILAŞTIRMASI
# ════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Borçlanma vs İtfa")
ws4.merge_cells('A1:G1')
ws4.cell(1, 1, "YILLIK BORÇLANMA vs İTFA KARŞILAŞTIRMASI").font = title_font

headers4 = ["Yıl", "Borçlanma\n(Nominal)", "İtfa\n(Nominal)",
            "Net Pozisyon\n(Borçlanma-İtfa)", "Borçlanma\n(Net Tutar)",
            "İtfa / Borçlanma\nOranı (%)", "Kümülatif\nNet Pozisyon"]
for c, h in enumerate(headers4, 1):
    ws4.cell(3, c, h)
hdr(ws4, 3, len(headers4))

compare_years = list(range(2009, 2035))
for i, yr in enumerate(compare_years):
    r = 4 + i
    a = f"$A${r}"
    dcell(ws4, r, 1, yr, bold=True)

    dcell(ws4, r, 2, f'=SUMPRODUCT((YEAR({B})={a})*({H}<>"")*{H})', '#,##0')
    dcell(ws4, r, 3, f'=SUMPRODUCT((YEAR({E})={a})*({H}<>"")*{H})', '#,##0')
    dcell(ws4, r, 4, f'=B{r}-C{r}', '#,##0')
    dcell(ws4, r, 5, f'=SUMPRODUCT((YEAR({B})={a})*({I}<>"")*{I})', '#,##0.00')
    dcell(ws4, r, 6, f'=IFERROR(C{r}/B{r}*100,"")', '#,##0.0')

    if i == 0:
        dcell(ws4, r, 7, f'=D{r}', '#,##0')
    else:
        dcell(ws4, r, 7, f'=G{r-1}+D{r}', '#,##0')

    zebra(ws4, r, len(headers4), i)

rt4 = 4 + len(compare_years)
total_row(ws4, rt4, len(headers4))
dcell(ws4, rt4, 1, "TOPLAM", bold=True)
ws4.cell(rt4, 1).fill = total_fill
dcell(ws4, rt4, 2, f'=SUM(B4:B{rt4-1})', '#,##0')
ws4.cell(rt4, 2).fill = total_fill
dcell(ws4, rt4, 3, f'=SUM(C4:C{rt4-1})', '#,##0')
ws4.cell(rt4, 3).fill = total_fill
dcell(ws4, rt4, 4, f'=SUM(D4:D{rt4-1})', '#,##0')
ws4.cell(rt4, 4).fill = total_fill
dcell(ws4, rt4, 5, f'=SUM(E4:E{rt4-1})', '#,##0.00')
ws4.cell(rt4, 5).fill = total_fill

for i, w in enumerate([8, 20, 20, 22, 20, 18, 20], 1):
    ws4.column_dimensions[get_column_letter(i)].width = w
ws4.freeze_panes = 'A4'


# ════════════════════════════════════════════════════════════════
# SAYFA 5: ISIN ANALİZİ
# ════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("ISIN Analizi")
ws5.merge_cells('A1:G1')
ws5.cell(1, 1, "EN ÇOK KULLANILAN ISIN ANALİZİ").font = title_font

# ISIN listesini Python ile çıkar (formül olarak yazmak pratik değil)
isins = []
for row in range(2, MR + 1):
    isin = ds.cell(row, 6).value
    if isin:
        isins.append(isin)
top_isins = Counter(isins).most_common(50)

headers5 = ["ISIN Kodu", "İhale\nSayısı", "Toplam Nominal\n(Bin TL)",
            "Toplam Net\n(Bin TL)", "İlk İşlem\nTarihi", "Son İşlem\nTarihi",
            "Vade Tarihi"]
for c, h in enumerate(headers5, 1):
    ws5.cell(3, c, h)
hdr(ws5, 3, len(headers5))

for i, (isin, _) in enumerate(top_isins):
    r = 4 + i
    a = f"$A${r}"
    dcell(ws5, r, 1, isin, bold=True)
    dcell(ws5, r, 2, f'=COUNTIF({F},{a})', '#,##0')
    dcell(ws5, r, 3, f'=SUMIF({F},{a},{H})', '#,##0')
    dcell(ws5, r, 4, f'=SUMIF({F},{a},{I})', '#,##0.00')
    dcell(ws5, r, 5, f'=MINIFS({B},{F},{a})', 'DD.MM.YYYY')
    dcell(ws5, r, 6, f'=MAXIFS({B},{F},{a})', 'DD.MM.YYYY')
    dcell(ws5, r, 7, f'=MAXIFS({E},{F},{a})', 'DD.MM.YYYY')
    zebra(ws5, r, len(headers5), i)

for i, w in enumerate([18, 12, 20, 20, 16, 16, 16], 1):
    ws5.column_dimensions[get_column_letter(i)].width = w
ws5.freeze_panes = 'A4'


# ════════════════════════════════════════════════════════════════
# SAYFA 6: AYLIK FAİZ TRENDİ
# ════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Aylık Faiz Trendi")
ws6.merge_cells('A1:F1')
ws6.cell(1, 1, "AYLIK AĞIRLIKLI ORTALAMA FAİZ TRENDİ").font = title_font

headers6 = ["Yıl", "Ay", "İhale\nSayısı", "Toplam\nNominal",
            "Ağırlıklı Ort.\nBasit Faiz (%)", "Ağırlıklı Ort.\nBileşik Faiz (%)"]
for c, h in enumerate(headers6, 1):
    ws6.cell(3, c, h)
hdr(ws6, 3, len(headers6))

row_idx = 4
for yr in range(2009, 2027):
    for mo in range(1, 13):
        if yr == 2009 and mo < 12:
            continue
        if yr == 2026 and mo > 4:
            continue

        a = f"$A${row_idx}"
        m = f"$B${row_idx}"

        dcell(ws6, row_idx, 1, yr, bold=True)
        dcell(ws6, row_idx, 2, mo)

        # İhale sayısı
        dcell(ws6, row_idx, 3,
              f'=SUMPRODUCT((YEAR({B})={a})*(MONTH({B})={m})*1)', '#,##0')

        # Toplam Nominal
        dcell(ws6, row_idx, 4,
              f'=SUMPRODUCT((YEAR({B})={a})*(MONTH({B})={m})*({H}<>"")*{H})', '#,##0')

        # Ağırlıklı ort basit faiz
        dcell(ws6, row_idx, 5,
              f'=IFERROR(SUMPRODUCT((YEAR({B})={a})*(MONTH({B})={m})*ISNUMBER({K})*{K}*({H}<>"")*{H})/D{row_idx},"")',
              '#,##0.00')

        # Ağırlıklı ort bileşik faiz
        dcell(ws6, row_idx, 6,
              f'=IFERROR(SUMPRODUCT((YEAR({B})={a})*(MONTH({B})={m})*ISNUMBER({N})*{N}*({H}<>"")*{H})/D{row_idx},"")',
              '#,##0.00')

        zebra(ws6, row_idx, len(headers6), row_idx - 4)
        row_idx += 1

for i, w in enumerate([8, 8, 12, 18, 20, 20], 1):
    ws6.column_dimensions[get_column_letter(i)].width = w
ws6.freeze_panes = 'A4'


# ════════════════════════════════════════════════════════════════
# KAYDET
# ════════════════════════════════════════════════════════════════
output = "/Users/sadettin/cowork/tcmb dogrudan alım/tcmb_dogrudan_alim.xlsx"
wb.save(output)
print(f"\nDosya kaydedildi: {output}")
print(f"Sayfalar: {wb.sheetnames}")
