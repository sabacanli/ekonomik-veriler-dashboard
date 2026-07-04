#!/usr/bin/env python3
"""
TCMB Doğrudan Alım Verileri - Güncelleme Scripti
==================================================
Bu scripti çalıştırınca:
  1. TCMB sitesinden güncel veriyi çeker
  2. Veri sayfasını sıfırdan yazar
  3. Analiz sayfalarını güncel satır aralığıyla yeniden oluşturur

Kullanım:
  python3 guncelle.py
"""

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from collections import Counter
from datetime import datetime, timedelta
import re
import os
import sys

# ══════════════════════════════════════════════════════════════
# AYARLAR
# ══════════════════════════════════════════════════════════════
URL = "https://www.tcmb.gov.tr/wps/wcm/connect/tr/tcmb+tr/main+page+site+area/acik+piyasa+islemleri/ihale+ile+gerceklestirilen+dogrudan+alim+islemleri+verileri"
OUTPUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "tcmb_dogrudan_alim.xlsx")

# ══════════════════════════════════════════════════════════════
# STİLLER
# ══════════════════════════════════════════════════════════════
title_font = Font(bold=True, size=14, color="1F4E79", name="Calibri")
header_font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
data_font = Font(size=10, name="Calibri")
bold_font = Font(bold=True, size=10, name="Calibri")
pct_font = Font(size=9, italic=True, color="666666", name="Calibri")
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
total_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
even_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def hdr(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border


def dcell(ws, row, col, val, fmt=None, bold=False):
    cell = ws.cell(row, col, val)
    cell.font = bold_font if bold else data_font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center')
    if fmt:
        cell.number_format = fmt
    return cell


def total_row_style(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.fill = total_fill
        cell.border = thin_border
        cell.font = bold_font
        cell.alignment = Alignment(horizontal='center', vertical='center')


def zebra(ws, row, cols, idx):
    if idx % 2 == 0:
        for c in range(1, cols + 1):
            ws.cell(row, c).fill = even_fill


# ══════════════════════════════════════════════════════════════
# ADIM 1: VERİYİ ÇEK
# ══════════════════════════════════════════════════════════════
print("📡 TCMB sitesinden veri çekiliyor...")
response = requests.get(URL, timeout=60)
response.encoding = 'utf-8'
if response.status_code != 200:
    print(f"❌ Hata: HTTP {response.status_code}")
    sys.exit(1)

soup = BeautifulSoup(response.text, 'html.parser')
tables = soup.find_all('table')
main_table = max(tables, key=lambda t: len(t.find_all('tr')))
rows = main_table.find_all('tr')

all_data = []
for row in rows:
    cells = row.find_all(['th', 'td'])
    row_data = [re.sub(r'\s+', ' ', cell.get_text(strip=True)) for cell in cells]
    if row_data and any(t for t in row_data):
        all_data.append(row_data)

# Türkçe/İngilizce başlık + not satırlarını filtrele
clean_data = []
for row in all_data:
    joined = ' '.join(row).upper()
    if any(kw in joined for kw in ['İHALE', 'İŞLEM', 'TEKLİF', 'FAİZ', 'AUCTION',
                                     'TRANSACTION', 'AMOUNT', 'RATE', 'METHOD',
                                     'YÖNTEM', 'GELENEKSEL=']):
        continue
    if row[0].startswith('*'):
        continue
    if len([c for c in row if c.strip()]) < 3:
        continue
    clean_data.append(row)

print(f"✅ {len(clean_data)} satır veri çekildi")

# ══════════════════════════════════════════════════════════════
# ADIM 2: VERİ SAYFASI OLUŞTUR
# ══════════════════════════════════════════════════════════════
print("📊 Excel dosyası oluşturuluyor...")
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Doğrudan Alım İşlemleri"

headers = [
    "İhale No", "İşlem Tarihi", "Yöntem", "Valör", "Vade",
    "Tanım (ISIN)", "Teklif Tutarı (Nominal)", "Kazanan Tutar (Nominal)",
    "Kazanan Tutar (Net)", "Minimum Basit Faiz", "Ortalama Basit Faiz",
    "Maksimum Basit Faiz", "Minimum Bileşik Faiz", "Ortalama Bileşik Faiz",
    "Maksimum Bileşik Faiz"
]

for c, h in enumerate(headers, 1):
    cell = ws.cell(1, c, h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

for row_idx, row_data in enumerate(clean_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        if col_idx > len(headers):
            break
        val = value.strip()
        cell = ws.cell(row_idx, col_idx)

        if val == '-' or val == '–' or val == '—':
            cell.value = None
            continue

        # Tarihler
        if col_idx in (2, 4, 5) and re.match(r'^\d{2}\.\d{2}\.\d{4}$', val):
            try:
                cell.value = datetime.strptime(val, '%d.%m.%Y')
                cell.number_format = 'DD.MM.YYYY'
            except:
                cell.value = val
        # Ondalıklı sayılar (86,471.88 veya 9.72)
        elif re.match(r'^-?[\d,]+\.\d+$', val):
            try:
                cell.value = float(val.replace(',', ''))
                cell.number_format = '#,##0.00'
            except:
                cell.value = val
        # Tam sayılar (558,000)
        elif re.match(r'^-?[\d,]+$', val) and ',' in val:
            try:
                cell.value = int(val.replace(',', ''))
                cell.number_format = '#,##0'
            except:
                cell.value = val
        elif re.match(r'^\d+$', val):
            cell.value = int(val)
            cell.number_format = '#,##0'
        else:
            cell.value = val

        cell.font = data_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if row_idx % 2 == 0:
            cell.fill = even_fill

col_widths = [12, 14, 14, 14, 14, 18, 20, 20, 20, 16, 16, 16, 16, 16, 16]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A2'
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

MR = ws.max_row
print(f"   Veri sayfası: {MR - 1} satır")

# ══════════════════════════════════════════════════════════════
# ADIM 3: ANALİZ SAYFALARI (formüller dinamik MR kullanır)
# ══════════════════════════════════════════════════════════════
SN = "'Doğrudan Alım İşlemleri'"
B = f"{SN}!$B$2:$B${MR}"
D = f"{SN}!$D$2:$D${MR}"
E = f"{SN}!$E$2:$E${MR}"
F = f"{SN}!$F$2:$F${MR}"
G = f"{SN}!$G$2:$G${MR}"
H = f"{SN}!$H$2:$H${MR}"
I = f"{SN}!$I$2:$I${MR}"
K = f"{SN}!$K$2:$K${MR}"
N = f"{SN}!$N$2:$N${MR}"

# Yıl aralıklarını veriden otomatik belirle
islem_yillari = set()
for row in clean_data:
    try:
        islem_yillari.add(int(row[1].split('.')[-1]))
    except:
        pass
min_year = min(islem_yillari)
max_year = max(islem_yillari)

# İtfa yılları: vade tarihlerinden
vade_years = set()
for row in clean_data:
    try:
        vy = int(row[4].split('.')[-1])
        vade_years.add(vy)
    except:
        pass
max_itfa_year = max(vade_years) if vade_years else max_year + 5

islem_years = list(range(min_year, max_year + 1))
itfa_years = list(range(min_year + 1, max_itfa_year + 1))
compare_years = list(range(min_year, max_itfa_year + 1))

# ────────────────────────────────────────────────────────────
# SAYFA 1: YILLIK BORÇLANMA ANALİZİ
# ────────────────────────────────────────────────────────────
print("   Yıllık Borçlanma Analizi...")
ws1 = wb.create_sheet("Yıllık Borçlanma Analizi")
ws1.merge_cells('A1:I1')
ws1.cell(1, 1, "YILLIK BORÇLANMA ANALİZİ (İŞLEM TARİHİNE GÖRE)").font = title_font

h1 = ["Yıl", "İhale\nSayısı", "Toplam Teklif\n(Nominal, Bin TL)",
      "Toplam Borçlanma\n(Nominal, Bin TL)", "Toplam Borçlanma\n(Net, Bin TL)",
      "Karşılanma\nOranı (%)", "Ağırlıklı Ort.\nBasit Faiz (%)",
      "Ağırlıklı Ort.\nBileşik Faiz (%)", "Ağırlıklı Ort.\nVade (Gün)"]
for c, h in enumerate(h1, 1):
    ws1.cell(3, c, h)
hdr(ws1, 3, len(h1))

for i, yr in enumerate(islem_years):
    r = 4 + i
    a = f"$A${r}"
    dcell(ws1, r, 1, yr, bold=True)
    dcell(ws1, r, 2, f'=SUMPRODUCT((YEAR({B})={a})*1)', '#,##0')
    dcell(ws1, r, 3, f'=SUMPRODUCT((YEAR({B})={a})*({G}<>"")*{G})', '#,##0')
    dcell(ws1, r, 4, f'=SUMPRODUCT((YEAR({B})={a})*({H}<>"")*{H})', '#,##0')
    dcell(ws1, r, 5, f'=SUMPRODUCT((YEAR({B})={a})*({I}<>"")*{I})', '#,##0.00')
    dcell(ws1, r, 6, f'=IFERROR(D{r}/C{r}*100,"")', '#,##0.0')
    dcell(ws1, r, 7, f'=IFERROR(SUMPRODUCT((YEAR({B})={a})*ISNUMBER({K})*{K}*({H}<>"")*{H})/SUMPRODUCT((YEAR({B})={a})*ISNUMBER({K})*({H}<>"")*{H}),"")', '#,##0.00')
    dcell(ws1, r, 8, f'=IFERROR(SUMPRODUCT((YEAR({B})={a})*ISNUMBER({N})*{N}*({H}<>"")*{H})/SUMPRODUCT((YEAR({B})={a})*ISNUMBER({N})*({H}<>"")*{H}),"")', '#,##0.00')
    dcell(ws1, r, 9, f'=IFERROR(SUMPRODUCT((YEAR({B})={a})*({E}-{D})*({H}<>"")*{H})/SUMPRODUCT((YEAR({B})={a})*({H}<>"")*{H}),"")', '#,##0')
    zebra(ws1, r, len(h1), i)

rt = 4 + len(islem_years)
total_row_style(ws1, rt, len(h1))
dcell(ws1, rt, 1, "TOPLAM", bold=True); ws1.cell(rt,1).fill = total_fill
for c, f in [(2,f'=SUM(B4:B{rt-1})'), (3,f'=SUM(C4:C{rt-1})'), (4,f'=SUM(D4:D{rt-1})'), (5,f'=SUM(E4:E{rt-1})')]:
    dcell(ws1, rt, c, f, '#,##0' if c != 5 else '#,##0.00'); ws1.cell(rt,c).fill = total_fill
dcell(ws1, rt, 6, f'=IFERROR(D{rt}/C{rt}*100,"")', '#,##0.0'); ws1.cell(rt,6).fill = total_fill
dcell(ws1, rt, 7, f'=IFERROR(SUMPRODUCT(G4:G{rt-1},D4:D{rt-1})/SUM(D4:D{rt-1}),"")', '#,##0.00'); ws1.cell(rt,7).fill = total_fill
dcell(ws1, rt, 8, f'=IFERROR(SUMPRODUCT(H4:H{rt-1},D4:D{rt-1})/SUM(D4:D{rt-1}),"")', '#,##0.00'); ws1.cell(rt,8).fill = total_fill
dcell(ws1, rt, 9, f'=IFERROR(SUMPRODUCT(I4:I{rt-1},D4:D{rt-1})/SUM(D4:D{rt-1}),"")', '#,##0'); ws1.cell(rt,9).fill = total_fill

for i, w in enumerate([8,12,22,22,22,14,16,16,16], 1):
    ws1.column_dimensions[get_column_letter(i)].width = w
ws1.freeze_panes = 'A4'

# ────────────────────────────────────────────────────────────
# SAYFA 2: YILLIK İTFA PROFİLİ
# ────────────────────────────────────────────────────────────
print("   Yıllık İtfa Profili...")
ws2 = wb.create_sheet("Yıllık İtfa Profili")
ws2.merge_cells('A1:E1')
ws2.cell(1, 1, "YILLIK İTFA PROFİLİ (VADE TARİHİNE GÖRE)").font = title_font

h2 = ["İtfa Yılı", "İtfa Olan\nNominal Tutar", "İtfa Olan\nNet Tutar",
      "İhale Sayısı", "Ağırlıklı Ort.\nBileşik Faiz (%)"]
for c, h in enumerate(h2, 1):
    ws2.cell(3, c, h)
hdr(ws2, 3, len(h2))

for i, yr in enumerate(itfa_years):
    r = 4 + i
    a = f"$A${r}"
    dcell(ws2, r, 1, yr, bold=True)
    dcell(ws2, r, 2, f'=SUMPRODUCT((YEAR({E})={a})*({H}<>"")*{H})', '#,##0')
    dcell(ws2, r, 3, f'=SUMPRODUCT((YEAR({E})={a})*({I}<>"")*{I})', '#,##0.00')
    dcell(ws2, r, 4, f'=SUMPRODUCT((YEAR({E})={a})*1)', '#,##0')
    dcell(ws2, r, 5, f'=IFERROR(SUMPRODUCT((YEAR({E})={a})*ISNUMBER({N})*{N}*({H}<>"")*{H})/SUMPRODUCT((YEAR({E})={a})*ISNUMBER({N})*({H}<>"")*{H}),"")', '#,##0.00')
    zebra(ws2, r, len(h2), i)

rt2 = 4 + len(itfa_years)
total_row_style(ws2, rt2, len(h2))
dcell(ws2, rt2, 1, "TOPLAM", bold=True); ws2.cell(rt2,1).fill = total_fill
dcell(ws2, rt2, 2, f'=SUM(B4:B{rt2-1})', '#,##0'); ws2.cell(rt2,2).fill = total_fill
dcell(ws2, rt2, 3, f'=SUM(C4:C{rt2-1})', '#,##0.00'); ws2.cell(rt2,3).fill = total_fill
dcell(ws2, rt2, 4, f'=SUM(D4:D{rt2-1})', '#,##0'); ws2.cell(rt2,4).fill = total_fill

for i, w in enumerate([12,22,22,14,18], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = 'A4'

# ────────────────────────────────────────────────────────────
# SAYFA 3: VADE DAĞILIM ANALİZİ
# ────────────────────────────────────────────────────────────
print("   Vade Dağılım Analizi...")
ws3 = wb.create_sheet("Vade Dağılım Analizi")
ws3.merge_cells('A1:H1')
ws3.cell(1, 1, "VADE DAĞILIM ANALİZİ (İŞLEM YILINA GÖRE)").font = title_font

h3 = ["Yıl", "Toplam\nNominal", "0-1 Yıl", "1-2 Yıl", "2-3 Yıl", "3-5 Yıl", "5-7 Yıl", "7+ Yıl"]
for c, h in enumerate(h3, 1):
    ws3.cell(3, c, h)
hdr(ws3, 3, len(h3))

vade_bounds = [(0,365),(366,730),(731,1095),(1096,1825),(1826,2555),(2556,99999)]
for i, yr in enumerate(islem_years):
    r_nom = 4 + i * 2
    r_pct = r_nom + 1
    a = f"$A${r_nom}"
    dcell(ws3, r_nom, 1, yr, bold=True)
    ws3.merge_cells(start_row=r_nom, start_column=1, end_row=r_pct, end_column=1)
    ws3.cell(r_nom, 1).alignment = Alignment(horizontal='center', vertical='center')
    dcell(ws3, r_nom, 2, f'=SUMPRODUCT((YEAR({B})={a})*({H}<>"")*{H})', '#,##0')
    ws3.merge_cells(start_row=r_nom, start_column=2, end_row=r_pct, end_column=2)
    ws3.cell(r_nom, 2).alignment = Alignment(horizontal='center', vertical='center')

    for j, (lo, hi) in enumerate(vade_bounds):
        col = 3 + j
        cl = get_column_letter(col)
        if hi == 99999:
            dcell(ws3, r_nom, col, f'=SUMPRODUCT((YEAR({B})={a})*(({E}-{D})>={lo})*({H}<>"")*{H})', '#,##0')
        else:
            dcell(ws3, r_nom, col, f'=SUMPRODUCT((YEAR({B})={a})*(({E}-{D})>={lo})*(({E}-{D})<={hi})*({H}<>"")*{H})', '#,##0')
        c_pct = dcell(ws3, r_pct, col, f'=IFERROR({cl}{r_nom}/$B${r_nom}*100,"")', '#,##0.0')
        c_pct.font = pct_font

    if i % 2 == 0:
        for c in range(1, len(h3) + 1):
            ws3.cell(r_nom, c).fill = even_fill
            ws3.cell(r_pct, c).fill = even_fill

for i, w in enumerate([8,18,14,14,14,14,14,14], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.freeze_panes = 'A4'

# ────────────────────────────────────────────────────────────
# SAYFA 4: BORÇLANMA vs İTFA
# ────────────────────────────────────────────────────────────
print("   Borçlanma vs İtfa...")
ws4 = wb.create_sheet("Borçlanma vs İtfa")
ws4.merge_cells('A1:G1')
ws4.cell(1, 1, "YILLIK BORÇLANMA vs İTFA KARŞILAŞTIRMASI").font = title_font

h4 = ["Yıl", "Borçlanma\n(Nominal)", "İtfa\n(Nominal)", "Net Pozisyon\n(Borçlanma-İtfa)",
      "Borçlanma\n(Net Tutar)", "İtfa / Borçlanma\nOranı (%)", "Kümülatif\nNet Pozisyon"]
for c, h in enumerate(h4, 1):
    ws4.cell(3, c, h)
hdr(ws4, 3, len(h4))

for i, yr in enumerate(compare_years):
    r = 4 + i
    a = f"$A${r}"
    dcell(ws4, r, 1, yr, bold=True)
    dcell(ws4, r, 2, f'=SUMPRODUCT((YEAR({B})={a})*({H}<>"")*{H})', '#,##0')
    dcell(ws4, r, 3, f'=SUMPRODUCT((YEAR({E})={a})*({H}<>"")*{H})', '#,##0')
    dcell(ws4, r, 4, f'=B{r}-C{r}', '#,##0')
    dcell(ws4, r, 5, f'=SUMPRODUCT((YEAR({B})={a})*({I}<>"")*{I})', '#,##0.00')
    dcell(ws4, r, 6, f'=IFERROR(C{r}/B{r}*100,"")', '#,##0.0')
    dcell(ws4, r, 7, f'=D{r}' if i == 0 else f'=G{r-1}+D{r}', '#,##0')
    zebra(ws4, r, len(h4), i)

rt4 = 4 + len(compare_years)
total_row_style(ws4, rt4, len(h4))
dcell(ws4, rt4, 1, "TOPLAM", bold=True); ws4.cell(rt4,1).fill = total_fill
for c, fmt in [(2,'#,##0'),(3,'#,##0'),(4,'#,##0'),(5,'#,##0.00')]:
    dcell(ws4, rt4, c, f'=SUM({chr(64+c)}4:{chr(64+c)}{rt4-1})', fmt); ws4.cell(rt4,c).fill = total_fill

for i, w in enumerate([8,20,20,22,20,18,20], 1):
    ws4.column_dimensions[get_column_letter(i)].width = w
ws4.freeze_panes = 'A4'

# ────────────────────────────────────────────────────────────
# SAYFA 5: ISIN ANALİZİ
# ────────────────────────────────────────────────────────────
print("   ISIN Analizi...")
ws5 = wb.create_sheet("ISIN Analizi")
ws5.merge_cells('A1:G1')
ws5.cell(1, 1, "EN ÇOK KULLANILAN ISIN ANALİZİ").font = title_font

isins = [ws.cell(r, 6).value for r in range(2, MR + 1) if ws.cell(r, 6).value]
top_isins = Counter(isins).most_common(50)

h5 = ["ISIN Kodu", "İhale\nSayısı", "Toplam Nominal\n(Bin TL)",
      "Toplam Net\n(Bin TL)", "İlk İşlem\nTarihi", "Son İşlem\nTarihi", "Vade Tarihi"]
for c, h in enumerate(h5, 1):
    ws5.cell(3, c, h)
hdr(ws5, 3, len(h5))

for i, (isin, _) in enumerate(top_isins):
    r = 4 + i
    a = f"$A${r}"
    dcell(ws5, r, 1, isin, bold=True)
    dcell(ws5, r, 2, f'=IFERROR(COUNTIF({F},{a}),0)', '#,##0')
    dcell(ws5, r, 3, f'=IFERROR(SUMIF({F},{a},{H}),0)', '#,##0')
    dcell(ws5, r, 4, f'=IFERROR(SUMIF({F},{a},{I}),0)', '#,##0.00')
    dcell(ws5, r, 5, f'=IFERROR(AGGREGATE(5,6,{B}/({F}={a}),1),"")', 'DD.MM.YYYY')
    dcell(ws5, r, 6, f'=IFERROR(AGGREGATE(14,6,{B}/({F}={a}),1),"")', 'DD.MM.YYYY')
    dcell(ws5, r, 7, f'=IFERROR(AGGREGATE(14,6,{E}/({F}={a}),1),"")', 'DD.MM.YYYY')
    zebra(ws5, r, len(h5), i)

for i, w in enumerate([18,12,20,20,16,16,16], 1):
    ws5.column_dimensions[get_column_letter(i)].width = w
ws5.freeze_panes = 'A4'

# ────────────────────────────────────────────────────────────
# SAYFA 6: AYLIK FAİZ TRENDİ
# ────────────────────────────────────────────────────────────
print("   Aylık Faiz Trendi...")
ws6 = wb.create_sheet("Aylık Faiz Trendi")
ws6.merge_cells('A1:F1')
ws6.cell(1, 1, "AYLIK AĞIRLIKLI ORTALAMA FAİZ TRENDİ").font = title_font

h6 = ["Yıl", "Ay", "İhale\nSayısı", "Toplam\nNominal",
      "Ağırlıklı Ort.\nBasit Faiz (%)", "Ağırlıklı Ort.\nBileşik Faiz (%)"]
for c, h in enumerate(h6, 1):
    ws6.cell(3, c, h)
hdr(ws6, 3, len(h6))

# Aktif ayları veriden belirle
active_months = set()
for row in clean_data:
    try:
        parts = row[1].split('.')
        active_months.add((int(parts[2]), int(parts[1])))
    except:
        pass

row_idx = 4
for yr in range(min_year, max_year + 1):
    for mo in range(1, 13):
        if (yr, mo) not in active_months:
            continue
        a = f"$A${row_idx}"
        m = f"$B${row_idx}"
        dcell(ws6, row_idx, 1, yr, bold=True)
        dcell(ws6, row_idx, 2, mo)
        dcell(ws6, row_idx, 3, f'=SUMPRODUCT((YEAR({B})={a})*(MONTH({B})={m})*1)', '#,##0')
        dcell(ws6, row_idx, 4, f'=SUMPRODUCT((YEAR({B})={a})*(MONTH({B})={m})*({H}<>"")*{H})', '#,##0')
        dcell(ws6, row_idx, 5, f'=IFERROR(SUMPRODUCT((YEAR({B})={a})*(MONTH({B})={m})*ISNUMBER({K})*{K}*({H}<>"")*{H})/D{row_idx},"")', '#,##0.00')
        dcell(ws6, row_idx, 6, f'=IFERROR(SUMPRODUCT((YEAR({B})={a})*(MONTH({B})={m})*ISNUMBER({N})*{N}*({H}<>"")*{H})/D{row_idx},"")', '#,##0.00')
        zebra(ws6, row_idx, len(h6), row_idx - 4)
        row_idx += 1

for i, w in enumerate([8,8,12,18,20,20], 1):
    ws6.column_dimensions[get_column_letter(i)].width = w
ws6.freeze_panes = 'A4'

# ────────────────────────────────────────────────────────────
# SAYFA 7: PORTFÖY ÖZETİ
# ────────────────────────────────────────────────────────────
print("   Portföy Özeti...")
ws7 = wb.create_sheet("Portföy Özeti", 0)  # İlk sayfa olarak ekle
ws7.merge_cells('A1:D1')
ws7.cell(1, 1, "TCMB DOĞRUDAN ALIM - PORTFÖY ÖZETİ").font = title_font

# Referans tarihi
dcell(ws7, 3, 1, "Referans Tarihi (Son İşlem):", bold=True)
ws7.cell(3, 1).alignment = Alignment(horizontal='right', vertical='center')
dcell(ws7, 3, 2, f'=MAX({B})', 'DD.MM.YYYY')
ws7.merge_cells('B3:D3')

# ── Bölüm 1: Mevcut Portföy ──
dcell(ws7, 5, 1, "MEVCUT PORTFÖY (Vadesi Gelmemiş Kağıtlar)", bold=True)
ws7.cell(5, 1).font = Font(bold=True, size=12, color="1F4E79", name="Calibri")
ws7.merge_cells('A5:D5')

labels_1 = [
    ("Toplam Stok (Nominal, Bin TL)",
     f'=SUMPRODUCT(({E}>B$3)*({H}<>"")*{H})'),
    ("Toplam Stok (Net, Bin TL)",
     f'=SUMPRODUCT(({E}>B$3)*({I}<>"")*{I})'),
    ("Toplam Stok (Nominal, Milyar TL)",
     '=B7/1000000'),
    ("Ağırlıklı Ort. Bileşik Getiri (%)",
     f'=IFERROR(SUMPRODUCT(({E}>B$3)*ISNUMBER({N})*{N}*({H}<>"")*{H})/SUMPRODUCT(({E}>B$3)*ISNUMBER({N})*({H}<>"")*{H}),"")'),
    ("Portföydeki İhale Sayısı",
     f'=SUMPRODUCT(({E}>B$3)*1)'),
]
for i, (label, formula) in enumerate(labels_1):
    r = 7 + i
    dcell(ws7, r, 1, label, bold=True)
    ws7.cell(r, 1).alignment = Alignment(horizontal='right', vertical='center')
    fmt = '#,##0' if 'Bin' in label else '#,##0.0' if 'Milyar' in label else '#,##0.00' if 'Getiri' in label else '#,##0'
    dcell(ws7, r, 2, formula, fmt)
    ws7.merge_cells(f'B{r}:D{r}')
    if i % 2 == 0:
        for c in range(1, 5):
            ws7.cell(r, c).fill = even_fill

# ── Bölüm 2: Yıllık Karşılaştırma ──
r_start = 14
dcell(ws7, r_start, 1, "YILLIK BORÇLANMA vs İTFA KARŞILAŞTIRMASI", bold=True)
ws7.cell(r_start, 1).font = Font(bold=True, size=12, color="1F4E79", name="Calibri")
ws7.merge_cells(f'A{r_start}:D{r_start}')

h7 = ["Yıl", "Borçlanma\n(Milyar TL)", "İtfa\n(Milyar TL)", "Net\n(Milyar TL)"]
r_h = r_start + 1
for c, h in enumerate(h7, 1):
    ws7.cell(r_h, c, h)
hdr(ws7, r_h, len(h7))

for i, yr in enumerate([2024, 2025, 2026]):
    r = r_start + 2 + i
    a = f"$A${r}"
    dcell(ws7, r, 1, yr, bold=True)
    dcell(ws7, r, 2, f'=IFERROR(SUMPRODUCT((YEAR({B})={a})*({H}<>"")*{H})/1000000,0)', '#,##0.0')
    dcell(ws7, r, 3, f'=IFERROR(SUMPRODUCT((YEAR({E})={a})*({H}<>"")*{H})/1000000,0)', '#,##0.0')
    dcell(ws7, r, 4, f'=B{r}-C{r}', '#,##0.0')
    zebra(ws7, r, len(h7), i)

# ── Bölüm 3: Özet Cümle (Python ile hesaplanıp statik metin olarak yazılır) ──
r_oz = r_start + 6
dcell(ws7, r_oz, 1, "ÖZET", bold=True)
ws7.cell(r_oz, 1).font = Font(bold=True, size=12, color="1F4E79", name="Calibri")
ws7.merge_cells(f'A{r_oz}:D{r_oz}')

# Yıl-satır eşlemesi: 2024=16, 2025=17, 2026=18
r24 = r_start + 2
r25 = r_start + 3
r26 = r_start + 4

# Değerleri Python'da hesapla (en güvenilir yol)
bugun = datetime.now()

# Stok: vadesi gelmemiş kağıtların nominal toplamı
stok_nom = 0
stok_net = 0
faiz_agirlik = 0
agirlik = 0
borc = {2024: 0, 2025: 0, 2026: 0}
itfa = {2024: 0, 2025: 0, 2026: 0}

for row in clean_data:
    try:
        islem_dt = datetime.strptime(row[1], '%d.%m.%Y')
        vade_dt = datetime.strptime(row[4], '%d.%m.%Y')
    except:
        continue

    nom_str = row[7].strip()
    net_str = row[8].strip()
    faiz_str = row[13].strip()  # Ort bileşik faiz

    if nom_str in ('-', '', '–'):
        nom = 0
    else:
        try:
            nom = float(nom_str.replace(',', ''))
        except:
            nom = 0

    if net_str in ('-', '', '–'):
        net = 0
    else:
        try:
            net = float(net_str.replace(',', ''))
        except:
            net = 0

    if faiz_str in ('-', '', '–'):
        faiz = None
    else:
        try:
            faiz = float(faiz_str.replace(',', ''))
        except:
            faiz = None

    # Stok
    if vade_dt > bugun and nom > 0:
        stok_nom += nom
        stok_net += net
        if faiz is not None:
            faiz_agirlik += faiz * nom
            agirlik += nom

    # Yıllık borçlanma/itfa
    yr = islem_dt.year
    if yr in borc:
        borc[yr] += nom
    vyr = vade_dt.year
    if vyr in itfa:
        itfa[vyr] += nom

son_tarih = max(
    (datetime.strptime(row[1], '%d.%m.%Y') for row in clean_data if len(row) > 1),
    default=bugun
)
ort_getiri = faiz_agirlik / agirlik if agirlik > 0 else 0

def fmt_milyar(val):
    """Bin TL -> Milyar TL, Türkçe format (nokta binlik, virgül ondalık)"""
    m = val / 1_000_000
    return f"{m:,.1f}".replace(",", "X").replace(".", ",").replace("X", ".")

def fmt_pct(val):
    """Yüzde, Türkçe format"""
    return f"{val:.2f}".replace(".", ",")

ozet_text = (
    f"TCMB'nin toplam menkul kıymet stoğu {son_tarih.strftime('%d.%m.%Y')} tarihi itibarıyla "
    f"{fmt_milyar(stok_nom)} milyar TL (nominal)'dir, portföy getirisi ise "
    f"%{fmt_pct(ort_getiri)}'dir. "
    f"2026 yılında {son_tarih.strftime('%d.%m.%Y')} tarihine kadar "
    f"{fmt_milyar(borc[2026])} milyar TL borçlanmıştır, "
    f"{fmt_milyar(itfa[2026])} milyar TL itfası olmuştur. "
    f"2025 yılında {fmt_milyar(itfa[2025])} milyar TL itfa, "
    f"{fmt_milyar(borc[2025])} milyar TL borçlanma iken "
    f"2024 yılında {fmt_milyar(itfa[2024])} milyar TL itfa, "
    f"{fmt_milyar(borc[2024])} milyar TL borçlanma şeklindedir."
)

r_c = r_oz + 1
ws7.cell(r_c, 1, ozet_text)
ws7.cell(r_c, 1).font = Font(size=11, name="Calibri")
ws7.cell(r_c, 1).alignment = Alignment(wrap_text=True, vertical='top')
ws7.merge_cells(f'A{r_c}:D{r_c + 2}')
ws7.row_dimensions[r_c].height = 80

# ── Bölüm 4: UYARILAR ──
print("   Uyarılar hesaplanıyor...")

# Parsed data listesi oluştur
from collections import defaultdict
parsed = []
for row in clean_data:
    try:
        islem_dt = datetime.strptime(row[1], '%d.%m.%Y')
        vade_dt = datetime.strptime(row[4], '%d.%m.%Y')
    except:
        continue
    nom_str = row[7].strip()
    net_str = row[8].strip()
    faiz_str = row[13].strip()
    nom = float(nom_str.replace(',', '')) if nom_str not in ('-', '', '–') else 0
    net = float(net_str.replace(',', '')) if net_str not in ('-', '', '–') else 0
    faiz = None
    if faiz_str not in ('-', '', '–'):
        try:
            faiz = float(faiz_str.replace(',', ''))
        except:
            pass
    teklif_str = row[6].strip()
    teklif = float(teklif_str.replace(',', '')) if teklif_str not in ('-', '', '–') else 0
    parsed.append({
        'tarih': islem_dt, 'vade': vade_dt, 'valor': datetime.strptime(row[3], '%d.%m.%Y'),
        'isin': row[5], 'teklif': teklif, 'nom': nom, 'net': net, 'faiz': faiz
    })

# Aylık aggregation
aylik = defaultdict(lambda: {'nom': 0, 'teklif': 0, 'count': 0, 'faiz_w': 0, 'w': 0, 'gunler': set()})
for d in parsed:
    if d['nom'] > 0:
        key = (d['tarih'].year, d['tarih'].month)
        aylik[key]['nom'] += d['nom']
        aylik[key]['teklif'] += d['teklif']
        aylik[key]['count'] += 1
        aylik[key]['gunler'].add(d['tarih'])
        if d['faiz'] is not None:
            aylik[key]['faiz_w'] += d['faiz'] * d['nom']
            aylik[key]['w'] += d['nom']

sorted_months = sorted(aylik.keys())

# ── Uyarı tespitleri ──
warn_fill_red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
warn_font_red = Font(bold=True, size=10, color="9C0006", name="Calibri")
warn_fill_orange = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
warn_font_orange = Font(bold=True, size=10, color="9C6500", name="Calibri")
warn_fill_blue = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
warn_font_blue = Font(bold=True, size=10, color="1F4E79", name="Calibri")

alerts = []  # (seviye, başlık, açıklama)

# 1. Aylık hacim sıçraması: son ay vs önceki 3 ay ortalaması
if len(sorted_months) >= 4:
    son_ay = sorted_months[-1]
    onceki_3 = sorted_months[-4:-1]
    son_hacim = aylik[son_ay]['nom']
    ort_onceki = sum(aylik[k]['nom'] for k in onceki_3) / 3
    if ort_onceki > 0:
        kat = son_hacim / ort_onceki
        if kat > 3:
            alerts.append(('KIRMIZI',
                f"HACIM SIČRAMASI: {son_ay[0]}-{son_ay[1]:02d}",
                f"Son ay {fmt_milyar(son_hacim)} milyar TL borçlanma, önceki 3 ay ortalaması {fmt_milyar(ort_onceki)} milyar TL. "
                f"{kat:.1f}x artış."))
        elif kat > 1.5:
            alerts.append(('TURUNCU',
                f"Hacim artışı: {son_ay[0]}-{son_ay[1]:02d}",
                f"Son ay {fmt_milyar(son_hacim)} milyar TL, önceki 3 ay ort. {fmt_milyar(ort_onceki)} milyar TL ({kat:.1f}x)."))

    # Son tam ay da kontrol et (son ay kısmi olabilir)
    if len(sorted_months) >= 5:
        tam_ay = sorted_months[-2]
        onceki_3_tam = sorted_months[-5:-2]
        tam_hacim = aylik[tam_ay]['nom']
        ort_onceki_tam = sum(aylik[k]['nom'] for k in onceki_3_tam) / 3
        if ort_onceki_tam > 0:
            kat_tam = tam_hacim / ort_onceki_tam
            if kat_tam > 3:
                alerts.append(('KIRMIZI',
                    f"HACIM SIČRAMASI: {tam_ay[0]}-{tam_ay[1]:02d}",
                    f"{fmt_milyar(tam_hacim)} milyar TL borçlanma, önceki 3 ay ortalaması {fmt_milyar(ort_onceki_tam)} milyar TL. "
                    f"{kat_tam:.1f}x artış."))

# 2. Faiz değişimi: son 3 ay vs önceki 3 ay
if len(sorted_months) >= 6:
    son3 = sorted_months[-3:]
    onceki3 = sorted_months[-6:-3]
    def wavg_faiz(months):
        tw, tf = 0, 0
        for k in months:
            tw += aylik[k]['w']
            tf += aylik[k]['faiz_w']
        return tf / tw if tw > 0 else 0
    f_son = wavg_faiz(son3)
    f_onceki = wavg_faiz(onceki3)
    fark = f_son - f_onceki
    if abs(fark) > 5:
        yonu = "ARTIŞ" if fark > 0 else "DÜŞÜŞ"
        alerts.append(('KIRMIZI',
            f"FAİZ {yonu}: {fark:+.1f} puan",
            f"Son 3 ay ağırlıklı ort. bileşik faiz: %{fmt_pct(f_son)}, "
            f"önceki 3 ay: %{fmt_pct(f_onceki)}. {abs(fark):.1f} puanlık {yonu.lower()}."))
    elif abs(fark) > 2:
        yonu = "artış" if fark > 0 else "düşüş"
        alerts.append(('TURUNCU',
            f"Faiz {yonu}: {fark:+.1f} puan",
            f"Son 3 ay ort: %{fmt_pct(f_son)}, önceki 3 ay: %{fmt_pct(f_onceki)}."))

# 3. Karşılanma oranı düşüşü
if len(sorted_months) >= 4:
    son_oran_months = sorted_months[-3:]
    onceki_oran_months = sorted_months[-6:-3] if len(sorted_months) >= 6 else sorted_months[:3]
    def avg_oran(months):
        t, k = 0, 0
        for m in months:
            t += aylik[m]['teklif']
            k += aylik[m]['nom']
        return k / t * 100 if t > 0 else 0
    oran_son = avg_oran(son_oran_months)
    oran_onceki = avg_oran(onceki_oran_months)
    if oran_son < 30:
        alerts.append(('TURUNCU',
            f"Düşük karşılanma oranı: %{oran_son:.0f}",
            f"Son 3 ayda tekliflerin yalnızca %{oran_son:.0f}'u karşılandı "
            f"(önceki 3 ay: %{oran_onceki:.0f}). Talep arzı aşıyor."))

# 4. İhale sıklığı artışı
if len(sorted_months) >= 2:
    son_ay_k = sorted_months[-2]  # Son tam ay
    ihale_per_gun = aylik[son_ay_k]['count'] / max(len(aylik[son_ay_k]['gunler']), 1)
    # Önceki yılın aynı ayı veya genel ortalama
    tum_freq = []
    for k in sorted_months[:-2]:
        gunler = len(aylik[k]['gunler'])
        if gunler > 0:
            tum_freq.append(aylik[k]['count'] / gunler)
    ort_freq = sum(tum_freq) / len(tum_freq) if tum_freq else 1
    if ihale_per_gun > ort_freq * 2.5:
        alerts.append(('TURUNCU',
            f"İhale sıklığı artışı: {ihale_per_gun:.1f} ihale/gün",
            f"{son_ay_k[0]}-{son_ay_k[1]:02d} ayında günde ortalama {ihale_per_gun:.1f} ihale yapıldı. "
            f"Tarihsel ortalama: {ort_freq:.1f} ihale/gün."))

# 5. Reddedilen ihaleler (teklif var ama kazanan 0)
reddedilen = [d for d in parsed if d['teklif'] > 0 and d['nom'] == 0
              and d['tarih'] >= bugun - timedelta(days=30)]
if reddedilen:
    alerts.append(('TURUNCU',
        f"Son 30 günde {len(reddedilen)} ihale reddedildi",
        "Teklif geldiği halde kabul edilmeyen ihaleler: " +
        ", ".join(f"{d['isin']} ({d['tarih'].strftime('%d.%m.%Y')})" for d in reddedilen[:5])))

# 6. Vade yapısı kayması
if len(sorted_months) >= 6:
    def vade_dagilim(months):
        buckets = {'kisa': 0, 'orta': 0, 'uzun': 0}
        total = 0
        for d in parsed:
            if (d['tarih'].year, d['tarih'].month) in months and d['nom'] > 0:
                gun = (d['vade'] - d['valor']).days
                total += d['nom']
                if gun <= 730:
                    buckets['kisa'] += d['nom']
                elif gun <= 1825:
                    buckets['orta'] += d['nom']
                else:
                    buckets['uzun'] += d['nom']
        if total > 0:
            return {k: v / total * 100 for k, v in buckets.items()}
        return buckets

    son_vade = vade_dagilim(sorted_months[-3:])
    onceki_vade = vade_dagilim(sorted_months[-6:-3])

    # Uzun vade artışı
    uzun_fark = son_vade.get('uzun', 0) - onceki_vade.get('uzun', 0)
    if uzun_fark > 10:
        alerts.append(('MAVI',
            f"Uzun vade ağırlığı artıyor: +{uzun_fark:.0f} puan",
            f"Son 3 ay uzun vadeli (5+ yıl) ağırlık: %{son_vade['uzun']:.0f}, "
            f"önceki 3 ay: %{onceki_vade['uzun']:.0f}. Vade uzatılıyor."))
    elif uzun_fark < -10:
        alerts.append(('MAVI',
            f"Kısa vadeye kayış: {uzun_fark:+.0f} puan",
            f"Uzun vadeli ağırlık %{onceki_vade['uzun']:.0f}'dan %{son_vade['uzun']:.0f}'a düştü."))

    # Kısa vade yoğunlaşması
    kisa_son = son_vade.get('kisa', 0)
    kisa_onceki = onceki_vade.get('kisa', 0)
    if kisa_son > 40 and kisa_son - kisa_onceki > 10:
        alerts.append(('TURUNCU',
            f"Kısa vadede yoğunlaşma: %{kisa_son:.0f}",
            f"Son 3 ayda borçlanmanın %{kisa_son:.0f}'i 0-2 yıl vadeli "
            f"(önceki 3 ay: %{kisa_onceki:.0f}). Roll-over riski artıyor."))

# 7. Tek ihaledeki yoğunlaşma (son 30 gün)
son_30 = [d for d in parsed if d['tarih'] >= bugun - timedelta(days=30) and d['nom'] > 0]
if son_30:
    toplam_son30 = sum(d['nom'] for d in son_30)
    max_ihale = max(son_30, key=lambda d: d['nom'])
    max_pay = max_ihale['nom'] / toplam_son30 * 100
    if max_pay > 25:
        alerts.append(('MAVI',
            f"Yoğunlaşma: Tek ihale toplam borçlanmanın %{max_pay:.0f}'i",
            f"{max_ihale['isin']} ({max_ihale['tarih'].strftime('%d.%m.%Y')}): "
            f"{fmt_milyar(max_ihale['nom'])} milyar TL - son 30 gün toplamının %{max_pay:.0f}'i."))

# ── Uyarıları sayfaya yaz ──
r_warn = r_c + 4
dcell(ws7, r_warn, 1, "UYARILAR VE TESPİTLER", bold=True)
ws7.cell(r_warn, 1).font = Font(bold=True, size=12, color="1F4E79", name="Calibri")
ws7.merge_cells(f'A{r_warn}:D{r_warn}')

if not alerts:
    r_w = r_warn + 1
    ws7.cell(r_w, 1, "Tespit edilen anormallik bulunmamaktadır.")
    ws7.cell(r_w, 1).font = data_font
    ws7.merge_cells(f'A{r_w}:D{r_w}')
else:
    for i, (seviye, baslik, aciklama) in enumerate(alerts):
        r_w = r_warn + 1 + i * 2

        # Seviye renkleri
        if seviye == 'KIRMIZI':
            fill, font_s = warn_fill_red, warn_font_red
            icon = "●"
        elif seviye == 'TURUNCU':
            fill, font_s = warn_fill_orange, warn_font_orange
            icon = "▲"
        else:
            fill, font_s = warn_fill_blue, warn_font_blue
            icon = "◆"

        # Başlık satırı
        ws7.cell(r_w, 1, f"{icon} {baslik}")
        ws7.cell(r_w, 1).font = font_s
        ws7.cell(r_w, 1).fill = fill
        ws7.cell(r_w, 1).border = thin_border
        ws7.cell(r_w, 1).alignment = Alignment(vertical='center')
        for c in range(2, 5):
            ws7.cell(r_w, c).fill = fill
            ws7.cell(r_w, c).border = thin_border
        ws7.merge_cells(f'A{r_w}:D{r_w}')

        # Açıklama satırı
        r_a = r_w + 1
        ws7.cell(r_a, 1, aciklama)
        ws7.cell(r_a, 1).font = Font(size=10, name="Calibri")
        ws7.cell(r_a, 1).alignment = Alignment(wrap_text=True, vertical='top')
        ws7.cell(r_a, 1).border = thin_border
        for c in range(2, 5):
            ws7.cell(r_a, c).border = thin_border
        ws7.merge_cells(f'A{r_a}:D{r_a}')
        ws7.row_dimensions[r_a].height = 35

    print(f"   {len(alerts)} uyarı tespit edildi")

# Sütun genişlikleri
ws7.column_dimensions['A'].width = 38
ws7.column_dimensions['B'].width = 22
ws7.column_dimensions['C'].width = 22
ws7.column_dimensions['D'].width = 22

# ══════════════════════════════════════════════════════════════
# KAYDET
# ══════════════════════════════════════════════════════════════
wb.save(OUTPUT)
print(f"\n✅ Dosya kaydedildi: {OUTPUT}")
print(f"   Sayfalar: {wb.sheetnames}")
print(f"   Veri: {MR - 1} satır ({min_year}-{max_year})")
print(f"   İtfa profili: {itfa_years[0]}-{itfa_years[-1]}")
print(f"   Son güncelleme: {datetime.now().strftime('%d.%m.%Y %H:%M')}")
