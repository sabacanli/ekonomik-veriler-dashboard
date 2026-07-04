import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import re
from datetime import datetime

url = "https://www.tcmb.gov.tr/wps/wcm/connect/tr/tcmb+tr/main+page+site+area/acik+piyasa+islemleri/ihale+ile+gerceklestirilen+dogrudan+alim+islemleri+verileri"

print("Sayfa indiriliyor...")
response = requests.get(url, timeout=60)
response.encoding = 'utf-8'
print(f"Status: {response.status_code}, Boyut: {len(response.text)} karakter")

soup = BeautifulSoup(response.text, 'html.parser')

# Find the main data table
tables = soup.find_all('table')
main_table = max(tables, key=lambda t: len(t.find_all('tr')))
rows = main_table.find_all('tr')
print(f"Tablodaki satır sayısı: {len(rows)}")

# Parse all rows
all_data = []
for row in rows:
    cells = row.find_all(['th', 'td'])
    row_data = []
    for cell in cells:
        text = cell.get_text(strip=True)
        text = re.sub(r'\s+', ' ', text)
        row_data.append(text)
    if row_data and any(t for t in row_data):
        all_data.append(row_data)

# Identify and skip English header rows, keep Turkish headers
# Row 0: Turkish header line 1
# Row 1: Turkish header line 2
# Row 2: Turkish header line 3 (NET, FAİZ etc)
# Row 3: English header line 1
# Row 4: English header line 2
# Row 5: English header line 3
# Find where English headers start (look for "TRANSACTION" or "AUCTION NO")
eng_start = None
eng_end = None
for i, row in enumerate(all_data):
    joined = ' '.join(row).upper()
    if 'TRANSACTION' in joined or 'AUCTION NO' in joined:
        if eng_start is None:
            eng_start = i
        eng_end = i

# Remove English header rows
if eng_start is not None:
    print(f"İngilizce başlık satırları kaldırılıyor: {eng_start}-{eng_end}")
    all_data = all_data[:eng_start] + all_data[eng_end+1:]

# Also remove any trailing English note row
all_data = [row for row in all_data if not any('METHOD' in cell.upper() and 'TRADITIONAL' in ' '.join(row).upper() for cell in row)]

# Determine number of Turkish header rows
tr_header_count = 0
for i, row in enumerate(all_data):
    joined = ' '.join(row).upper()
    if 'İHALE' in joined or 'TEKLİF' in joined or 'KAZANAN' in joined or 'FAİZ' in joined or 'NET' == row[-1].upper():
        tr_header_count = i + 1
    else:
        break

print(f"Türkçe başlık satır sayısı: {tr_header_count}")
print(f"Veri satır sayısı: {len(all_data) - tr_header_count}")

# Print sample
for i in range(min(tr_header_count + 3, len(all_data))):
    print(f"  Satır {i}: {all_data[i]}")

# Create combined header
# Merge multi-row headers into single clean headers
headers = [
    "İhale No",
    "İşlem Tarihi",
    "Yöntem",
    "Valör",
    "Vade",
    "Tanım (ISIN)",
    "Teklif Tutarı (Nominal)",
    "Kazanan Tutar (Nominal)",
    "Kazanan Tutar (Net)",
    "Minimum Basit Faiz",
    "Ortalama Basit Faiz",
    "Maksimum Basit Faiz",
    "Minimum Bileşik Faiz",
    "Ortalama Bileşik Faiz",
    "Maksimum Bileşik Faiz"
]

# Extract only data rows (skip header rows)
data_rows = all_data[tr_header_count:]

# Filter out any note/footer rows and remaining English header rows
clean_data = []
for row in data_rows:
    joined = ' '.join(row).upper()
    # Skip rows that look like notes
    if row[0].startswith('*') or row[0].startswith('Yöntemi'):
        continue
    # Skip remaining English header rows
    if 'RATE' in joined and 'NET' in joined and not any(c for c in row if re.match(r'\d{2}\.\d{2}\.\d{4}', c)):
        continue
    if 'AMOUNT' in joined or 'AUCTION' in joined or 'TRANSACTION' in joined:
        continue
    # Skip empty-looking rows
    if len([c for c in row if c.strip()]) < 3:
        continue
    clean_data.append(row)

print(f"Temizlenmiş veri satır sayısı: {len(clean_data)}")

# Create Excel workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Doğrudan Alım İşlemleri"

# Style definitions
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
data_font = Font(size=10, name="Calibri")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
even_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

# Write headers
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Write data rows
for row_idx, row_data in enumerate(clean_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        if col_idx > len(headers):
            break

        val = value.strip()
        cell = ws.cell(row=row_idx, column=col_idx)

        # Parse dates (dd.mm.yyyy)
        if col_idx in (2, 4, 5) and re.match(r'^\d{2}\.\d{2}\.\d{4}$', val):
            try:
                dt = datetime.strptime(val, '%d.%m.%Y')
                cell.value = dt
                cell.number_format = 'DD.MM.YYYY'
            except:
                cell.value = val
        # Parse numbers - site uses English format: comma=thousands, dot=decimal
        # e.g., "558,000" or "86,471.88" or "9.72"
        elif re.match(r'^-?[\d,]+\.\d+$', val):
            # Has decimal point (e.g., "86,471.88" or "9.72")
            try:
                numeric_val = float(val.replace(',', ''))
                cell.value = numeric_val
                if col_idx >= 10:  # Faiz sütunları
                    cell.number_format = '#,##0.00'
                else:
                    cell.number_format = '#,##0.00'
            except:
                cell.value = val
        elif re.match(r'^-?[\d,]+$', val) and ',' in val:
            # Integer with thousands separator (e.g., "558,000")
            try:
                numeric_val = int(val.replace(',', ''))
                cell.value = numeric_val
                cell.number_format = '#,##0'
            except:
                cell.value = val
        elif re.match(r'^\d+$', val):
            cell.value = int(val)
            cell.number_format = '#,##0'
        else:
            cell.value = val

        cell.font = data_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # Zebra striping
        if row_idx % 2 == 0:
            cell.fill = even_fill

# Set column widths
col_widths = [12, 14, 14, 14, 14, 18, 20, 20, 20, 16, 16, 16, 16, 16, 16]
for i, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# Freeze header row
ws.freeze_panes = 'A2'

# Auto-filter
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

output_path = "/Users/sadettin/cowork/tcmb dogrudan alım/tcmb_dogrudan_alim.xlsx"
wb.save(output_path)
print(f"\nExcel dosyası kaydedildi: {output_path}")
print(f"Toplam veri satırı: {ws.max_row - 1} (başlık hariç)")
print(f"Sütun sayısı: {len(headers)}")
